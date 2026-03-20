from __future__ import annotations

from io import BytesIO

from docx import Document
from openpyxl import load_workbook

from grantflow.exporters.excel_builder import build_xlsx_from_logframe
from grantflow.exporters.word_builder import build_docx_from_toc


def _eu_min_toc() -> dict:
    return {
        "toc": {
            "overall_objective": {"objective_id": "OO1", "title": "Digital governance", "rationale": "EU fit"},
            "specific_objectives": [{"objective_id": "SO1", "title": "Service quality", "rationale": "Delivery"}],
            "expected_outcomes": [
                {"outcome_id": "OUT1", "title": "Citizen trust", "expected_change": "Higher confidence"}
            ],
        }
    }


def test_docx_autofills_eu_safeguarding_annex_from_critic_findings():
    toc = _eu_min_toc()
    findings = [
        {
            "code": "EU_SAFE_001",
            "message": "Safeguarding protocol is missing for partner-facing activities.",
            "fix_hint": "Add referral pathway and incident-response SOP in annex.",
            "severity": "high",
        }
    ]

    doc = Document(BytesIO(build_docx_from_toc(toc, "eu", critic_findings=findings)))
    text = "\n".join(p.text for p in doc.paragraphs)

    assert "Risk & Safeguarding Annex" in text
    assert "referral pathway" in text.lower()


def test_xlsx_autofills_eu_safeguarding_annex_from_critic_findings():
    toc = _eu_min_toc()
    findings = [
        {
            "code": "EU_SAFE_002",
            "message": "Risk control plan is incomplete for field implementation.",
            "fix_suggestion": "Document do-no-harm controls and escalation owner in annex.",
            "severity": "medium",
        }
    ]

    content = build_xlsx_from_logframe({"indicators": []}, "eu", toc_draft=toc, critic_findings=findings)
    wb = load_workbook(BytesIO(content))
    rows = list(wb["EU_Assumptions_Risks"].iter_rows(values_only=True))

    assert any((r and r[0] == "Safeguarding Annex") for r in rows[1:])
    assert any((r and "do-no-harm" in str(r[1]).lower()) for r in rows[1:])
