# grantflow/tests/test_integration.py

import asyncio
import gzip
import io
import json
import time

import pytest
from fastapi.testclient import TestClient

import grantflow.api.app as api_app_module
from grantflow.api.app import app

client = TestClient(app)


def _wait_for_terminal_status(job_id: str, timeout_s: float = 3.0):
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        response = client.get(f"/status/{job_id}")
        assert response.status_code == 200
        body = response.json()
        if body["status"] in {"done", "error", "pending_hitl"}:
            return body
        time.sleep(0.05)
    raise AssertionError("Timed out waiting for job completion")


def _drain_hitl_to_done(job_id: str, *, initial_status: dict | None = None, max_cycles: int = 6) -> dict:
    status = initial_status or _wait_for_terminal_status(job_id)
    for _ in range(max_cycles):
        if status.get("status") == "done":
            return status
        assert status.get("status") == "pending_hitl"
        checkpoint_id = str(status.get("checkpoint_id") or "").strip()
        checkpoint_stage = str(status.get("checkpoint_stage") or "").strip().lower()
        assert checkpoint_id
        assert checkpoint_stage in {"toc", "logframe"}

        approve = client.post(
            "/hitl/approve",
            json={"checkpoint_id": checkpoint_id, "approved": True, "feedback": f"Auto approve {checkpoint_stage}"},
        )
        assert approve.status_code == 200

        resume = client.post(f"/resume/{job_id}", json={})
        assert resume.status_code == 200
        expected_resume_from = "mel" if checkpoint_stage == "toc" else "critic"
        assert resume.json()["resuming_from"] == expected_resume_from
        status = _wait_for_terminal_status(job_id)
    raise AssertionError("HITL pipeline did not reach done within max_cycles")


def test_health_endpoint():
    response = client.get("/health")
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "healthy"
    assert body["version"]
    diagnostics = body["diagnostics"]
    assert diagnostics["job_store"]["mode"] in {"inmem", "sqlite"}
    assert diagnostics["hitl_store"]["mode"] in {"inmem", "sqlite"}
    assert diagnostics["ingest_store"]["mode"] in {"inmem", "sqlite"}
    assert diagnostics["job_runner"]["mode"] in {"background_tasks", "inmemory_queue", "redis_queue"}
    assert isinstance(diagnostics["job_runner"]["queue_enabled"], bool)
    assert isinstance(diagnostics["job_runner"]["queue"]["queue_size"], int)
    assert isinstance(diagnostics["auth"]["api_key_configured"], bool)
    assert isinstance(diagnostics["auth"]["read_auth_required"], bool)
    assert diagnostics["vector_store"]["backend"] in {"chroma", "memory"}
    assert diagnostics["vector_store"]["collection_prefix"]
    preflight_policy = diagnostics["preflight_grounding_policy"]
    assert preflight_policy["mode"] in {"warn", "strict", "off"}
    thresholds = preflight_policy["thresholds"]
    assert 0.0 <= float(thresholds["high_risk_coverage_threshold"]) <= 1.0
    assert 0.0 <= float(thresholds["medium_risk_coverage_threshold"]) <= 1.0
    assert 0.0 <= float(thresholds["high_risk_depth_coverage_threshold"]) <= 1.0
    assert 0.0 <= float(thresholds["medium_risk_depth_coverage_threshold"]) <= 1.0
    assert int(thresholds["min_uploads"]) >= 1
    runtime_gate = diagnostics["runtime_grounded_quality_gate"]
    assert runtime_gate["mode"] in {"warn", "strict", "off"}
    runtime_thresholds = runtime_gate["thresholds"]
    assert int(runtime_thresholds["min_citations_for_gate"]) >= 0
    assert 0.0 <= float(runtime_thresholds["max_non_retrieval_citation_rate"]) <= 1.0
    assert int(runtime_thresholds["min_retrieval_grounded_citations"]) >= 0
    mel_policy = diagnostics["mel_grounding_policy"]
    assert mel_policy["mode"] in {"warn", "strict", "off"}
    mel_thresholds = mel_policy["thresholds"]
    assert int(mel_thresholds["min_mel_citations"]) >= 1
    assert 0.0 <= float(mel_thresholds["min_claim_support_rate"]) <= 1.0
    assert 0.0 <= float(mel_thresholds["min_traceability_complete_rate"]) <= 1.0
    assert 0.0 <= float(mel_thresholds["max_traceability_gap_rate"]) <= 1.0
    export_policy = diagnostics["export_grounding_policy"]
    assert export_policy["mode"] in {"warn", "strict", "off"}
    export_thresholds = export_policy["thresholds"]
    assert int(export_thresholds["min_architect_citations"]) >= 1
    assert 0.0 <= float(export_thresholds["min_claim_support_rate"]) <= 1.0
    assert 0.0 <= float(export_thresholds["min_traceability_complete_rate"]) <= 1.0
    assert 0.0 <= float(export_thresholds["max_traceability_gap_rate"]) <= 1.0
    export_contract_policy = diagnostics["export_contract_policy"]
    assert export_contract_policy["mode"] in {"warn", "strict", "off"}
    export_runtime_gate_policy = diagnostics["export_runtime_grounded_gate_policy"]
    assert isinstance(export_runtime_gate_policy["require_pass"], bool)
    assert isinstance(diagnostics.get("configuration_warnings"), list)


def test_store_backend_alignment_validation_detects_job_hitl_mismatch(monkeypatch):
    class FakeJobStore:
        db_path = "/tmp/grantflow_state.db"

    class FakeHitlManager:
        _use_sqlite = False

    monkeypatch.setattr(api_app_module, "JOB_STORE", FakeJobStore())
    monkeypatch.setattr(api_app_module, "hitl_manager", FakeHitlManager())
    with pytest.raises(RuntimeError, match="GRANTFLOW_JOB_STORE"):
        api_app_module._validate_store_backend_alignment()


def test_store_backend_alignment_validation_allows_matching_backends(monkeypatch):
    class FakeJobStore:
        db_path = "/tmp/grantflow_state.db"

    class FakeHitlManager:
        _use_sqlite = True

    monkeypatch.setattr(api_app_module, "JOB_STORE", FakeJobStore())
    monkeypatch.setattr(api_app_module, "hitl_manager", FakeHitlManager())
    api_app_module._validate_store_backend_alignment()


def test_store_backend_alignment_validation_runs_at_startup(monkeypatch):
    class FakeJobStore:
        db_path = "/tmp/grantflow_state.db"

    class FakeHitlManager:
        _use_sqlite = False

    monkeypatch.setattr(api_app_module, "JOB_STORE", FakeJobStore())
    monkeypatch.setattr(api_app_module, "hitl_manager", FakeHitlManager())

    async def _start_lifespan():
        async with api_app_module._app_lifespan(api_app_module.app):
            return None

    with pytest.raises(RuntimeError, match="GRANTFLOW_HITL_STORE"):
        asyncio.run(_start_lifespan())


def test_demo_console_page_loads():
    response = client.get("/demo")
    assert response.status_code == 200
    assert "text/html" in response.headers["content-type"]
    body = response.text
    assert "GrantFlow Demo Console" in body
    assert "generatePresetSelect" in body
    assert "applyPresetBtn" in body
    assert "clearPresetContextBtn" in body
    assert "generatePresetReadinessPill" in body
    assert "generatePresetReadinessText" in body
    assert "generatePresetReadinessHint" in body
    assert "syncGenerateReadinessBtn" in body
    assert "Sync Readiness Now" in body
    assert "Missing doc_family:" in body
    assert "skipZeroReadinessWarningCheckbox" in body
    assert "grantflow_demo_zero_readiness_warning_prefs" in body
    assert "Don't ask again for this preset" in body
    assert "RAG Readiness (for selected preset)" in body
    assert "Generate anyway?" in body
    assert "strictPreflight" in body
    assert "Strict Preflight" in body
    assert "Preflight risk" in body
    assert "Strict preflight" in body
    assert "Grounded gate" in body
    assert "inputContextJson" in body
    assert "usaid_gov_ai_kazakhstan" in body
    assert "worldbank_public_sector_uzbekistan" in body
    assert "ingestPresetSelect" in body
    assert "applyIngestPresetBtn" in body
    assert "syncIngestDonorBtn" in body
    assert "ingestFileInput" in body
    assert "ingestMetadataJson" in body
    assert "ingestBtn" in body
    assert "ingestPresetGuidanceList" in body
    assert "ingestChecklistSummary" in body
    assert "ingestChecklistProgressList" in body
    assert "resetIngestChecklistBtn" in body
    assert "syncIngestChecklistServerBtn" in body
    assert "copyIngestInventoryJsonBtn" in body
    assert "downloadIngestInventoryJsonBtn" in body
    assert "downloadIngestInventoryCsvBtn" in body
    assert "ingestInventoryJson" in body
    assert "grantflow_demo_ingest_checklist_progress" in body
    assert "doc_family=" in body
    assert "/ingest" in body
    assert "/ingest/inventory?" in body
    assert "/ingest/inventory/export?" in body
    assert "/status/${encodeURIComponent(jobId)}/metrics" in body
    assert "/status/${encodeURIComponent(jobId)}/quality" in body
    assert "/status/${encodeURIComponent(jobId)}/critic" in body
    assert "/status/${encodeURIComponent(jobId)}/export-payload" in body
    assert "/export" in body
    assert "criticSeverityFilter" in body
    assert "criticFindingStatusFilter" in body
    assert "criticCitationConfidenceFilter" in body
    assert "criticBulkTargetStatus" in body
    assert "criticBulkScope" in body
    assert "criticBulkApplyBtn" in body
    assert "criticBulkClearFiltersBtn" in body
    assert "criticBulkResultJson" in body
    assert "Apply Bulk Status" in body
    assert "Clear Critic Filters" in body
    assert "criticAdvisorySummaryList" in body
    assert "criticAdvisoryLabelsList" in body
    assert "criticAdvisoryNormalizationList" in body
    assert "Jump to Diff" in body
    assert "Create Comment" in body
    assert "criticContextList" in body
    assert "llm_advisory_diagnostics" in body
    assert "conf " in body
    assert "thr " in body
    assert "architect_threshold_hit_rate" in body
    assert "qualityBtn" in body
    assert "qualityCards" in body
    assert "qualityPreflightMetaLine" in body
    assert "qualityGroundedGatePill" in body
    assert "qualityGroundedGateExplainBtn" in body
    assert "qualityGroundedGateReasonsWrap" in body
    assert "qualityGroundedGateReasonsList" in body
    assert "Why blocked?" in body
    assert "qualityMelSummaryList" in body
    assert "qualityCitationTypeCountsList" in body
    assert "qualityArchitectCitationTypeCountsList" in body
    assert "Grounding KPI" in body
    assert "groundingKpiCards" in body
    assert "groundingKpiMetaLine" in body
    assert "groundingKpiCountsList" in body
    assert "groundingKpiPolicyReasonsList" in body
    assert "Claim-support" in body
    assert "Architect fallback" in body
    assert "MEL claim-support" in body
    assert "MEL fallback" in body
    assert "Fallback rate" in body
    assert "Traceability complete" in body
    assert "Traceability gap" in body
    assert "qualityJson" in body
    assert "exportPayloadBtn" in body
    assert "copyExportPayloadBtn" in body
    assert "exportZipFromPayloadBtn" in body
    assert "exportProductionZipFromPayloadBtn" in body
    assert "Production Export (enforced)" in body
    assert "exportContractPill" in body
    assert "exportContractPillText" in body
    assert "exportContractMetaLine" in body
    assert "exportContractWarningsList" in body
    assert "productionExportMode" in body
    assert "allowUnsafeExport" in body
    assert "exportPayloadJson" in body
    assert "production_export" in body
    assert "allow_unsafe_export" in body
    assert "commentsFilterStatus" in body
    assert "commentsFilterVersionId" in body
    assert "grantflow_demo_diff_section" in body
    assert "grantflow_demo_strict_preflight" in body
    assert "grantflow_demo_portfolio_warning_level" in body
    assert "grantflow_demo_portfolio_grounding_risk_level" in body
    assert "grantflow_demo_portfolio_finding_status" in body
    assert "grantflow_demo_portfolio_finding_severity" in body
    assert "grantflow_demo_portfolio_toc_text_risk_level" in body
    assert "grantflow_demo_portfolio_sla_hotspot_kind" in body
    assert "grantflow_demo_portfolio_sla_hotspot_severity" in body
    assert "grantflow_demo_portfolio_sla_min_overdue_hours" in body
    assert "grantflow_demo_portfolio_sla_top_limit" in body
    assert "grantflow_demo_export_gzip_enabled" in body
    assert "grantflow_demo_production_export_mode" in body
    assert "grantflow_demo_allow_unsafe_export" in body
    assert "generatePreflightAlert" in body
    assert "generatePreflightAlertTitle" in body
    assert "generatePreflightAlertBody" in body
    assert "portfolioBtn" in body
    assert "portfolioClearBtn" in body
    assert "clearPortfolioToCTextRiskBtn" in body
    assert "portfolioWarningLevelFilter" in body
    assert "portfolioGroundingRiskLevelFilter" in body
    assert "portfolioFindingStatusFilter" in body
    assert "portfolioFindingSeverityFilter" in body
    assert "portfolioToCTextRiskLevelFilter" in body
    assert "/portfolio/quality" in body
    assert "/portfolio/metrics/export" in body
    assert "/portfolio/quality/export" in body
    assert "/portfolio/review-workflow" in body
    assert "/portfolio/review-workflow/export" in body
    assert "/portfolio/review-workflow/sla" in body
    assert "/portfolio/review-workflow/sla/export" in body
    assert "/portfolio/review-workflow/sla/hotspots" in body
    assert "/portfolio/review-workflow/sla/hotspots/export" in body
    assert "/portfolio/review-workflow/sla/hotspots/trends" in body
    assert "/portfolio/review-workflow/sla/hotspots/trends/export" in body
    assert "/portfolio/review-workflow/trends" in body
    assert "/portfolio/review-workflow/trends/export" in body
    assert "/portfolio/review-workflow/sla/trends" in body
    assert "/portfolio/review-workflow/sla/trends/export" in body
    assert "portfolioMetricsCards" in body
    assert "portfolioQualityCards" in body
    assert "portfolioReviewWorkflowBtn" in body
    assert "portfolioReviewWorkflowSummaryLine" in body
    assert "portfolioReviewWorkflowList" in body
    assert "portfolioReviewWorkflowJson" in body
    assert "copyPortfolioReviewWorkflowJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowCsvBtn" in body
    assert "exportPortfolioReviewWorkflowJsonBtn" in body
    assert "exportPortfolioReviewWorkflowCsvBtn" in body
    assert "portfolioReviewWorkflowSlaBtn" in body
    assert "portfolioReviewWorkflowSlaSummaryLine" in body
    assert "portfolioReviewWorkflowSlaList" in body
    assert "portfolioReviewWorkflowSlaJson" in body
    assert "portfolioSlaHotspotKindFilter" in body
    assert "portfolioSlaHotspotSeverityFilter" in body
    assert "portfolioSlaMinOverdueHoursFilter" in body
    assert "portfolioSlaTopLimitFilter" in body
    assert "portfolioReviewWorkflowSlaHotspotsBtn" in body
    assert "portfolioReviewWorkflowSlaHotspotsSummaryLine" in body
    assert "portfolioReviewWorkflowSlaHotspotsList" in body
    assert "portfolioReviewWorkflowSlaHotspotsJson" in body
    assert "portfolioReviewWorkflowSlaHotspotsTrendsBtn" in body
    assert "portfolioReviewWorkflowSlaHotspotsTrendsSummaryLine" in body
    assert "portfolioReviewWorkflowSlaHotspotsTrendSparkline" in body
    assert "portfolioReviewWorkflowSlaHotspotsTrendsList" in body
    assert "portfolioReviewWorkflowSlaHotspotsTrendsJson" in body
    assert "copyPortfolioReviewWorkflowSlaJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowSlaJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowSlaCsvBtn" in body
    assert "copyPortfolioReviewWorkflowSlaHotspotsJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowSlaHotspotsJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowSlaHotspotsCsvBtn" in body
    assert "copyPortfolioReviewWorkflowSlaHotspotsTrendsJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowSlaHotspotsTrendsJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowSlaHotspotsTrendsCsvBtn" in body
    assert "exportPortfolioReviewWorkflowSlaJsonBtn" in body
    assert "exportPortfolioReviewWorkflowSlaCsvBtn" in body
    assert "exportPortfolioReviewWorkflowSlaHotspotsJsonBtn" in body
    assert "exportPortfolioReviewWorkflowSlaHotspotsCsvBtn" in body
    assert "exportPortfolioReviewWorkflowSlaHotspotsTrendsJsonBtn" in body
    assert "exportPortfolioReviewWorkflowSlaHotspotsTrendsCsvBtn" in body
    assert "portfolioReviewWorkflowTrendsBtn" in body
    assert "portfolioReviewWorkflowTrendsSummaryLine" in body
    assert "portfolioReviewWorkflowTrendSparkline" in body
    assert "portfolioReviewWorkflowTrendsList" in body
    assert "portfolioReviewWorkflowTrendsJson" in body
    assert "copyPortfolioReviewWorkflowTrendsJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowTrendsJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowTrendsCsvBtn" in body
    assert "exportPortfolioReviewWorkflowTrendsJsonBtn" in body
    assert "exportPortfolioReviewWorkflowTrendsCsvBtn" in body
    assert "portfolioReviewWorkflowSlaTrendsBtn" in body
    assert "portfolioReviewWorkflowSlaTrendsSummaryLine" in body
    assert "portfolioReviewWorkflowSlaTrendSparkline" in body
    assert "portfolioReviewWorkflowSlaTrendsList" in body
    assert "portfolioReviewWorkflowSlaTrendsJson" in body
    assert "copyPortfolioReviewWorkflowSlaTrendsJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowSlaTrendsJsonBtn" in body
    assert "downloadPortfolioReviewWorkflowSlaTrendsCsvBtn" in body
    assert "exportPortfolioReviewWorkflowSlaTrendsJsonBtn" in body
    assert "exportPortfolioReviewWorkflowSlaTrendsCsvBtn" in body
    assert "portfolioMetricsWarningLevelsList" in body
    assert "portfolioMetricsGroundingRiskLevelsList" in body
    assert "portfolioQualityRiskList" in body
    assert "portfolioQualityOpenFindingsList" in body
    assert "portfolioQualityWarningLevelsList" in body
    assert "portfolioQualityGroundingRiskLevelsList" in body
    assert "portfolioQualityCitationTypeCountsList" in body
    assert "portfolioQualityArchitectCitationTypeCountsList" in body
    assert "portfolioQualityMelCitationTypeCountsList" in body
    assert "portfolioQualityFindingStatusList" in body
    assert "portfolioQualityFindingSeverityList" in body
    assert "portfolioQualityToCTextRiskList" in body
    assert "portfolioQualityGroundingRiskList" in body
    assert "portfolioQualityGroundedGateSectionsList" in body
    assert "portfolioQualityGroundedGateReasonsList" in body
    assert "portfolioQualityGroundedGateDonorsList" in body
    assert "portfolioQualityPrioritySignalsList" in body
    assert "portfolioQualityWeightedDonorsList" in body
    assert "% High-warning Jobs" in body
    assert "% Medium-warning Jobs" in body
    assert "% Low-warning Jobs" in body
    assert "% No-warning Jobs" in body
    assert "Fallback Dominance" in body
    assert "Claim-support Avg" in body
    assert "High-Risk Donors" in body
    assert "% High ToC-text Risk" in body
    assert "ToC Text Issues" in body
    assert "% Grounded Gate Block" in body
    assert "Grounded Gate Blocks" in body
    assert "% Grounded Gate Pass (present)" in body
    assert "ToC Text Risk Levels" in body
    assert "Clear ToC Risk Filter" in body
    assert "portfolioWarningMetaLine" in body
    assert "qualityLlmFindingLabelsList" in body
    assert "qualityAdvisoryBadgeList" in body
    assert "qualityReadinessWarningLevelPill" in body
    assert "portfolioQualityLlmLabelCountsList" in body
    assert "portfolioQualityTopDonorLlmLabelCountsList" in body
    assert "portfolioQualityTopDonorAdvisoryRejectedReasonsList" in body
    assert "portfolioQualityTopDonorAdvisoryAppliedList" in body
    assert "portfolioQualityFocusedDonorSummaryList" in body
    assert "portfolioQualityFocusedDonorAdvisoryPill" in body
    assert "portfolioQualityFocusedDonorAdvisoryPillText" in body
    assert "portfolioQualityFocusedDonorLlmLabelCountsList" in body
    assert "portfolioQualityFocusedDonorAdvisoryRejectedReasonsList" in body
    assert "portfolioQualityFocusedDonorAdvisoryAppliedLabelCountsList" in body
    assert "portfolioQualityFocusedDonorAdvisoryRejectedLabelCountsList" in body
    assert "portfolioQualityAdvisoryAppliedList" in body
    assert "portfolioQualityAdvisoryRejectedReasonsList" in body
    assert "portfolioQualityJson" in body
    assert 'params.set("finding_status",' in body
    assert 'params.set("finding_severity",' in body
    assert 'params.set("toc_text_risk_level",' in body
    assert "clearPortfolioToCTextRiskFilter" in body
    assert "copyPortfolioMetricsJsonBtn" in body
    assert "downloadPortfolioMetricsJsonBtn" in body
    assert "downloadPortfolioMetricsCsvBtn" in body
    assert "exportGzipEnabled" in body
    assert "exportInventoryJsonBtn" in body
    assert "exportInventoryCsvBtn" in body
    assert "exportPortfolioMetricsJsonBtn" in body
    assert "exportPortfolioMetricsCsvBtn" in body
    assert "exportPortfolioQualityJsonBtn" in body
    assert "exportPortfolioQualityCsvBtn" in body
    assert "Export (Server-side)" in body
    assert 'params.set("gzip", "true")' in body
    assert "grantflow_portfolio_metrics.csv" in body
    assert "copyPortfolioQualityJsonBtn" in body
    assert "downloadPortfolioQualityJsonBtn" in body
    assert "downloadPortfolioQualityCsvBtn" in body
    assert "grantflow_portfolio_quality.csv" in body
    assert "portfolioStatusCountsList" in body
    assert "portfolioDonorCountsList" in body
    assert "Click to filter" in body
    assert "clearFiltersBtn" in body
    assert "Acknowledge" in body
    assert "Resolve Finding" in body
    assert "Reopen Finding" in body
    assert "linkedFindingId" in body
    assert "/critic/findings/bulk-status" in body
    assert "reviewWorkflowBtn" in body
    assert "reviewWorkflowSummaryLine" in body
    assert "reviewWorkflowTrendsSummaryLine" in body
    assert "reviewWorkflowTimelineList" in body
    assert "reviewWorkflowTrendSparkline" in body
    assert "reviewWorkflowTrendsList" in body
    assert "reviewWorkflowJson" in body
    assert "reviewWorkflowTrendsJson" in body
    assert "reviewWorkflowEventTypeFilter" in body
    assert "reviewWorkflowFindingIdFilter" in body
    assert "reviewWorkflowFindingCodeFilter" in body
    assert "reviewWorkflowFindingSectionFilter" in body
    assert "reviewWorkflowCommentStatusFilter" in body
    assert "reviewWorkflowStateFilter" in body
    assert "reviewWorkflowOverdueHoursFilter" in body
    assert "reviewWorkflowClearFiltersBtn" in body
    assert "reviewWorkflowExportJsonBtn" in body
    assert "reviewWorkflowExportCsvBtn" in body
    assert "reviewWorkflowTrendsBtn" in body
    assert "reviewWorkflowTrendsExportJsonBtn" in body
    assert "reviewWorkflowTrendsExportCsvBtn" in body
    assert "reviewWorkflowSlaBtn" in body
    assert "reviewWorkflowSlaHotspotsBtn" in body
    assert "reviewWorkflowSlaHotspotsTrendsBtn" in body
    assert "reviewWorkflowSlaTrendsBtn" in body
    assert "reviewWorkflowSlaProfileBtn" in body
    assert "reviewWorkflowSlaRecomputeBtn" in body
    assert "reviewWorkflowSlaExportJsonBtn" in body
    assert "reviewWorkflowSlaExportCsvBtn" in body
    assert "reviewWorkflowSlaTrendsExportJsonBtn" in body
    assert "reviewWorkflowSlaTrendsExportCsvBtn" in body
    assert "reviewWorkflowSlaHotspotsExportJsonBtn" in body
    assert "reviewWorkflowSlaHotspotsExportCsvBtn" in body
    assert "reviewWorkflowSlaHotspotsTrendsExportJsonBtn" in body
    assert "reviewWorkflowSlaHotspotsTrendsExportCsvBtn" in body
    assert "copyReviewWorkflowSlaHotspotsJsonBtn" in body
    assert "downloadReviewWorkflowSlaHotspotsJsonBtn" in body
    assert "downloadReviewWorkflowSlaHotspotsCsvBtn" in body
    assert "copyReviewWorkflowSlaHotspotsTrendsJsonBtn" in body
    assert "downloadReviewWorkflowSlaHotspotsTrendsJsonBtn" in body
    assert "downloadReviewWorkflowSlaHotspotsTrendsCsvBtn" in body
    assert "reviewWorkflowSlaHighHours" in body
    assert "reviewWorkflowSlaMediumHours" in body
    assert "reviewWorkflowSlaLowHours" in body
    assert "reviewWorkflowSlaCommentDefaultHours" in body
    assert "reviewWorkflowSlaUseSavedProfile" in body
    assert "reviewWorkflowSlaHotspotKindFilter" in body
    assert "reviewWorkflowSlaHotspotSeverityFilter" in body
    assert "reviewWorkflowSlaMinOverdueHoursFilter" in body
    assert "reviewWorkflowSlaTopLimitFilter" in body
    assert "reviewWorkflowSlaSummaryLine" in body
    assert "reviewWorkflowSlaTrendsSummaryLine" in body
    assert "reviewWorkflowSlaHotspotsList" in body
    assert "reviewWorkflowSlaHotspotsSummaryLine" in body
    assert "reviewWorkflowSlaTrendSparkline" in body
    assert "reviewWorkflowSlaTrendsList" in body
    assert "reviewWorkflowSlaHotspotsTrendSparkline" in body
    assert "reviewWorkflowSlaHotspotsTrendsSummaryLine" in body
    assert "reviewWorkflowSlaHotspotsTrendsList" in body
    assert "reviewWorkflowSlaJson" in body
    assert "reviewWorkflowSlaTrendsJson" in body
    assert "reviewWorkflowSlaHotspotsJson" in body
    assert "reviewWorkflowSlaHotspotsTrendsJson" in body
    assert "reviewWorkflowSlaProfileSummaryLine" in body
    assert "reviewWorkflowSlaProfileJson" in body
    assert "grantflow_demo_review_workflow_event_type" in body
    assert "grantflow_demo_review_workflow_finding_id" in body
    assert "grantflow_demo_review_workflow_finding_code" in body
    assert "grantflow_demo_review_workflow_finding_section" in body
    assert "grantflow_demo_review_workflow_comment_status" in body
    assert "grantflow_demo_review_workflow_state" in body
    assert "grantflow_demo_review_workflow_overdue_hours" in body
    assert "grantflow_demo_review_workflow_sla_hotspot_kind" in body
    assert "grantflow_demo_review_workflow_sla_hotspot_severity" in body
    assert "grantflow_demo_review_workflow_sla_min_overdue_hours" in body
    assert "grantflow_demo_review_workflow_sla_top_limit" in body
    assert "grantflow_demo_review_workflow_sla_use_saved_profile" in body
    assert 'params.set("finding_code",' in body
    assert 'params.set("finding_section",' in body
    assert 'params.set("workflow_state",' in body
    assert 'params.set("overdue_after_hours",' in body
    assert "buildReviewWorkflowSlaFilterQueryString" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/trends" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/sla" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/sla/hotspots" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/sla/hotspots/trends" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/sla/trends" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/sla/export" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/sla/hotspots/export" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/sla/hotspots/trends/export" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/sla/trends/export" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/sla/profile" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/sla/recompute" in body
    assert "finding_sla_hours" in body
    assert "default_comment_sla_hours" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/export" in body
    assert "/status/${encodeURIComponent(jobId)}/review/workflow/trends/export" in body


def test_ready_endpoint():
    response = client.get("/ready")
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "ready"
    checks = body["checks"]
    assert checks["vector_store"]["backend"] in {"chroma", "memory"}
    assert checks["vector_store"]["ready"] is True
    assert checks["job_runner"]["mode"] in {"background_tasks", "inmemory_queue", "redis_queue"}
    assert isinstance(checks["job_runner"]["ready"], bool)
    assert isinstance(checks["job_runner"]["queue"]["queue_size"], int)
    preflight_policy = checks["preflight_grounding_policy"]
    assert preflight_policy["mode"] in {"warn", "strict", "off"}
    thresholds = preflight_policy["thresholds"]
    assert 0.0 <= float(thresholds["high_risk_coverage_threshold"]) <= 1.0
    assert 0.0 <= float(thresholds["medium_risk_coverage_threshold"]) <= 1.0
    assert int(thresholds["min_uploads"]) >= 1
    runtime_gate = checks["runtime_grounded_quality_gate"]
    assert runtime_gate["mode"] in {"warn", "strict", "off"}
    runtime_thresholds = runtime_gate["thresholds"]
    assert int(runtime_thresholds["min_citations_for_gate"]) >= 0
    assert 0.0 <= float(runtime_thresholds["max_non_retrieval_citation_rate"]) <= 1.0
    assert int(runtime_thresholds["min_retrieval_grounded_citations"]) >= 0
    mel_policy = checks["mel_grounding_policy"]
    assert mel_policy["mode"] in {"warn", "strict", "off"}
    mel_thresholds = mel_policy["thresholds"]
    assert int(mel_thresholds["min_mel_citations"]) >= 1
    assert 0.0 <= float(mel_thresholds["min_claim_support_rate"]) <= 1.0
    assert 0.0 <= float(mel_thresholds["min_traceability_complete_rate"]) <= 1.0
    assert 0.0 <= float(mel_thresholds["max_traceability_gap_rate"]) <= 1.0
    export_policy = checks["export_grounding_policy"]
    assert export_policy["mode"] in {"warn", "strict", "off"}
    export_thresholds = export_policy["thresholds"]
    assert int(export_thresholds["min_architect_citations"]) >= 1
    assert 0.0 <= float(export_thresholds["min_claim_support_rate"]) <= 1.0
    assert 0.0 <= float(export_thresholds["min_traceability_complete_rate"]) <= 1.0
    assert 0.0 <= float(export_thresholds["max_traceability_gap_rate"]) <= 1.0
    export_contract_policy = checks["export_contract_policy"]
    assert export_contract_policy["mode"] in {"warn", "strict", "off"}
    export_runtime_gate_policy = checks["export_runtime_grounded_gate_policy"]
    assert isinstance(export_runtime_gate_policy["require_pass"], bool)
    assert isinstance(checks.get("configuration_warnings"), list)


def test_ready_endpoint_redis_dispatcher_mode_without_local_consumer(monkeypatch):
    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "redis_queue")
    monkeypatch.setattr(
        api_app_module,
        "JOB_RUNNER",
        type(
            "_Runner",
            (),
            {
                "diagnostics": staticmethod(
                    lambda: {
                        "backend": "redis",
                        "consumer_enabled": False,
                        "running": False,
                        "redis_available": True,
                        "queue_size": 0,
                    }
                )
            },
        )(),
    )
    monkeypatch.setattr(api_app_module, "_vector_store_readiness", lambda: {"ready": True, "backend": "memory"})

    response = client.get("/ready")
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "ready"
    checks = body["checks"]
    assert checks["job_runner"]["mode"] == "redis_queue"
    assert checks["job_runner"]["ready"] is True


def test_ready_endpoint_reports_dead_letter_alert_when_threshold_exceeded(monkeypatch):
    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "redis_queue")
    monkeypatch.setattr(api_app_module.config.job_runner, "dead_letter_alert_threshold", 1)
    monkeypatch.setattr(api_app_module.config.job_runner, "dead_letter_alert_blocking", False)
    monkeypatch.setattr(
        api_app_module,
        "JOB_RUNNER",
        type(
            "_Runner",
            (),
            {
                "diagnostics": staticmethod(
                    lambda: {
                        "backend": "redis",
                        "consumer_enabled": False,
                        "running": False,
                        "redis_available": True,
                        "queue_size": 0,
                        "dead_letter_queue_size": 3,
                    }
                )
            },
        )(),
    )
    monkeypatch.setattr(api_app_module, "_vector_store_readiness", lambda: {"ready": True, "backend": "memory"})

    response = client.get("/ready")
    assert response.status_code == 200
    body = response.json()
    checks = body["checks"]
    assert checks["job_runner"]["ready"] is True
    alert = checks["job_runner"]["dead_letter_alert"]
    assert alert["triggered"] is True
    assert alert["blocking"] is False
    alerts = checks["job_runner"]["alerts"]
    assert alerts and alerts[0]["code"] == "DEAD_LETTER_QUEUE_THRESHOLD_EXCEEDED"


def test_ready_endpoint_can_block_when_dead_letter_alert_is_blocking(monkeypatch):
    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "redis_queue")
    monkeypatch.setattr(api_app_module.config.job_runner, "dead_letter_alert_threshold", 1)
    monkeypatch.setattr(api_app_module.config.job_runner, "dead_letter_alert_blocking", True)
    monkeypatch.setattr(
        api_app_module,
        "JOB_RUNNER",
        type(
            "_Runner",
            (),
            {
                "diagnostics": staticmethod(
                    lambda: {
                        "backend": "redis",
                        "consumer_enabled": False,
                        "running": False,
                        "redis_available": True,
                        "queue_size": 0,
                        "dead_letter_queue_size": 2,
                    }
                )
            },
        )(),
    )
    monkeypatch.setattr(api_app_module, "_vector_store_readiness", lambda: {"ready": True, "backend": "memory"})

    response = client.get("/ready")
    assert response.status_code == 503
    detail = response.json()["detail"]
    assert detail["status"] == "degraded"
    checks = detail["checks"]
    assert checks["job_runner"]["ready"] is False
    alert = checks["job_runner"]["dead_letter_alert"]
    assert alert["triggered"] is True
    assert alert["blocking"] is True


def test_ready_endpoint_returns_503_when_vector_store_unavailable(monkeypatch):
    class BrokenClient:
        def heartbeat(self):
            raise RuntimeError("chroma unavailable")

    monkeypatch.setattr(api_app_module.vector_store, "client", BrokenClient())

    response = client.get("/ready")
    assert response.status_code == 503
    body = response.json()
    assert body["detail"]["status"] == "degraded"
    assert body["detail"]["checks"]["vector_store"]["backend"] == "chroma"
    assert body["detail"]["checks"]["vector_store"]["ready"] is False
    assert "chroma unavailable" in body["detail"]["checks"]["vector_store"]["error"]
    preflight_policy = body["detail"]["checks"]["preflight_grounding_policy"]
    assert preflight_policy["mode"] in {"warn", "strict", "off"}
    runtime_gate = body["detail"]["checks"]["runtime_grounded_quality_gate"]
    assert runtime_gate["mode"] in {"warn", "strict", "off"}
    mel_policy = body["detail"]["checks"]["mel_grounding_policy"]
    assert mel_policy["mode"] in {"warn", "strict", "off"}
    export_policy = body["detail"]["checks"]["export_grounding_policy"]
    assert export_policy["mode"] in {"warn", "strict", "off"}
    export_contract_policy = body["detail"]["checks"]["export_contract_policy"]
    assert export_contract_policy["mode"] in {"warn", "strict", "off"}
    export_runtime_gate_policy = body["detail"]["checks"]["export_runtime_grounded_gate_policy"]
    assert isinstance(export_runtime_gate_policy["require_pass"], bool)


def test_ready_endpoint_reflects_preflight_grounding_threshold_overrides(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "grounding_gate_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "preflight_grounding_policy_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "preflight_grounding_high_risk_coverage_threshold", 0.42)
    monkeypatch.setattr(api_app_module.config.graph, "preflight_grounding_medium_risk_coverage_threshold", 0.91)
    monkeypatch.setattr(api_app_module.config.graph, "preflight_grounding_min_uploads", 7)
    monkeypatch.setattr(api_app_module.config.graph, "preflight_grounding_min_key_claim_coverage_rate", 0.61)
    monkeypatch.setattr(api_app_module.config.graph, "preflight_grounding_max_fallback_claim_ratio", 0.72)
    monkeypatch.setattr(api_app_module.config.graph, "preflight_grounding_max_traceability_gap_rate", 0.33)
    monkeypatch.setattr(api_app_module.config.graph, "preflight_grounding_min_threshold_hit_rate", 0.44)

    response = client.get("/ready")
    assert response.status_code == 200
    body = response.json()
    policy = body["checks"]["preflight_grounding_policy"]
    assert policy["mode"] == "strict"
    thresholds = policy["thresholds"]
    assert thresholds["high_risk_coverage_threshold"] == 0.42
    assert thresholds["medium_risk_coverage_threshold"] == 0.91
    assert thresholds["min_uploads"] == 7
    assert thresholds["min_key_claim_coverage_rate"] == 0.61
    assert thresholds["max_fallback_claim_ratio"] == 0.72
    assert thresholds["max_traceability_gap_rate"] == 0.33
    assert thresholds["min_threshold_hit_rate"] == 0.44


def test_ready_endpoint_preflight_policy_mode_can_differ_from_pipeline_mode(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "grounding_gate_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "preflight_grounding_policy_mode", "warn")

    response = client.get("/ready")
    assert response.status_code == 200
    body = response.json()
    assert body["checks"]["preflight_grounding_policy"]["mode"] == "warn"


def test_ready_endpoint_reflects_runtime_grounded_quality_gate_overrides(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "runtime_grounded_quality_gate_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "runtime_grounded_quality_gate_min_citations", 9)
    monkeypatch.setattr(
        api_app_module.config.graph,
        "runtime_grounded_quality_gate_max_non_retrieval_citation_rate",
        0.22,
    )
    monkeypatch.setattr(
        api_app_module.config.graph,
        "runtime_grounded_quality_gate_min_retrieval_grounded_citations",
        4,
    )

    response = client.get("/ready")
    assert response.status_code == 200
    body = response.json()
    policy = body["checks"]["runtime_grounded_quality_gate"]
    assert policy["mode"] == "strict"
    thresholds = policy["thresholds"]
    assert thresholds["min_citations_for_gate"] == 9
    assert thresholds["max_non_retrieval_citation_rate"] == 0.22
    assert thresholds["min_retrieval_grounded_citations"] == 4


def test_ready_endpoint_reflects_mel_grounding_policy_overrides(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_policy_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_min_mel_citations", 5)
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_min_claim_support_rate", 0.61)
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_min_traceability_complete_rate", 0.72)
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_max_traceability_gap_rate", 0.28)

    response = client.get("/ready")
    assert response.status_code == 200
    body = response.json()
    mel_policy = body["checks"]["mel_grounding_policy"]
    assert mel_policy["mode"] == "strict"
    thresholds = mel_policy["thresholds"]
    assert thresholds["min_mel_citations"] == 5
    assert thresholds["min_claim_support_rate"] == 0.61
    assert thresholds["min_traceability_complete_rate"] == 0.72
    assert thresholds["max_traceability_gap_rate"] == 0.28


def test_ready_endpoint_reflects_export_grounding_policy_overrides(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_policy_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_min_architect_citations", 7)
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_min_claim_support_rate", 0.66)
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_min_traceability_complete_rate", 0.74)
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_max_traceability_gap_rate", 0.26)

    response = client.get("/ready")
    assert response.status_code == 200
    body = response.json()
    export_policy = body["checks"]["export_grounding_policy"]
    assert export_policy["mode"] == "strict"
    thresholds = export_policy["thresholds"]
    assert thresholds["min_architect_citations"] == 7
    assert thresholds["min_claim_support_rate"] == 0.66
    assert thresholds["min_traceability_complete_rate"] == 0.74
    assert thresholds["max_traceability_gap_rate"] == 0.26


def test_ready_endpoint_reflects_export_contract_policy_override(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "export_contract_policy_mode", "strict")

    response = client.get("/ready")
    assert response.status_code == 200
    body = response.json()
    contract_policy = body["checks"]["export_contract_policy"]
    assert contract_policy["mode"] == "strict"


def test_ready_endpoint_reflects_export_runtime_grounded_gate_policy_override(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "export_require_grounded_gate_pass", True)

    response = client.get("/ready")
    assert response.status_code == 200
    body = response.json()
    policy = body["checks"]["export_runtime_grounded_gate_policy"]
    assert policy["require_pass"] is True


def test_health_endpoint_reports_chroma_port_conflict_warning(monkeypatch):
    monkeypatch.setattr(api_app_module.vector_store, "_chroma_host", "127.0.0.1")
    monkeypatch.setattr(api_app_module.vector_store, "_chroma_port", 8000)

    response = client.get("/health")
    assert response.status_code == 200
    body = response.json()
    warnings = body["diagnostics"].get("configuration_warnings") or []
    assert any(w.get("code") == "CHROMA_PORT_MAY_CONFLICT_WITH_API_DEFAULT" for w in warnings if isinstance(w, dict))


def test_ready_endpoint_reports_chroma_port_conflict_warning(monkeypatch):
    monkeypatch.setattr(api_app_module.vector_store, "_chroma_host", "127.0.0.1")
    monkeypatch.setattr(api_app_module.vector_store, "_chroma_port", 8000)

    response = client.get("/ready")
    assert response.status_code == 200
    body = response.json()
    warnings = body["checks"].get("configuration_warnings") or []
    assert any(w.get("code") == "CHROMA_PORT_MAY_CONFLICT_WITH_API_DEFAULT" for w in warnings if isinstance(w, dict))


def test_list_donors():
    response = client.get("/donors")
    assert response.status_code == 200
    donors = response.json()["donors"]
    donor_ids = [d["id"] for d in donors]
    assert "usaid" in donor_ids
    assert "eu" in donor_ids
    assert "worldbank" in donor_ids


def test_generate_preflight_reports_high_risk_when_namespace_empty():
    api_app_module.INGEST_AUDIT_STORE.clear()

    response = client.post("/generate/preflight", json={"donor_id": "usaid"})
    assert response.status_code == 200
    body = response.json()
    assert body["donor_id"] == "usaid"
    assert isinstance(body.get("retrieval_query_terms"), list)
    assert "development objectives" in body.get("retrieval_query_terms")
    assert body["namespace_empty"] is True
    assert body["risk_level"] == "high"
    assert body["grounding_risk_level"] == "high"
    grounding_policy = body.get("grounding_policy") or {}
    assert grounding_policy["mode"] in {"warn", "strict", "off"}
    assert grounding_policy["risk_level"] == "high"
    assert grounding_policy["go_ahead"] in {True, False}
    assert isinstance(grounding_policy.get("reasons"), list)
    thresholds = grounding_policy.get("thresholds") or {}
    assert 0.0 <= float(thresholds.get("high_risk_coverage_threshold") or 0.0) <= 1.0
    assert 0.0 <= float(thresholds.get("medium_risk_coverage_threshold") or 0.0) <= 1.0
    assert 0.0 <= float(thresholds.get("high_risk_depth_coverage_threshold") or 0.0) <= 1.0
    assert 0.0 <= float(thresholds.get("medium_risk_depth_coverage_threshold") or 0.0) <= 1.0
    assert float(thresholds.get("medium_risk_coverage_threshold") or 0.0) >= float(
        thresholds.get("high_risk_coverage_threshold") or 0.0
    )
    assert float(thresholds.get("medium_risk_depth_coverage_threshold") or 0.0) >= float(
        thresholds.get("high_risk_depth_coverage_threshold") or 0.0
    )
    assert int(thresholds.get("min_uploads") or 0) >= 1
    assert 0.0 <= float(thresholds.get("min_key_claim_coverage_rate") or 0.0) <= 1.0
    assert 0.0 <= float(thresholds.get("max_fallback_claim_ratio") or 0.0) <= 1.0
    assert 0.0 <= float(thresholds.get("max_traceability_gap_rate") or 0.0) <= 1.0
    assert 0.0 <= float(thresholds.get("min_threshold_hit_rate") or 0.0) <= 1.0
    architect_claims = body.get("architect_claims") or {}
    assert isinstance(architect_claims, dict)
    assert architect_claims.get("available") is False
    assert architect_claims.get("reason") == "input_context_missing"
    assert isinstance(body.get("doc_family_min_uploads"), dict)
    assert isinstance(body.get("depth_ready_doc_families"), list)
    assert isinstance(body.get("depth_gap_doc_families"), list)
    assert body.get("depth_ready_count") == 0
    assert body.get("depth_gap_count") == int(body.get("expected_count") or 0)
    assert body.get("depth_coverage_rate") == 0.0


def test_generate_preflight_accepts_architect_rag_toggle():
    api_app_module.INGEST_AUDIT_STORE.clear()

    response = client.post(
        "/generate/preflight",
        json={"donor_id": "usaid", "architect_rag_enabled": False},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["architect_rag_enabled"] is False
    architect_claims = body.get("architect_claims") or {}
    assert architect_claims.get("retrieval_expected") is False


def test_generate_preflight_reports_depth_coverage_warning_when_family_min_not_met():
    api_app_module.INGEST_AUDIT_STORE.clear()
    api_app_module.INGEST_AUDIT_STORE.append(
        {
            "event_id": "pref-depth-1",
            "ts": "2026-03-03T10:00:00+00:00",
            "donor_id": "usaid",
            "tenant_id": None,
            "namespace": "usaid_ads201",
            "filename": "ads201.pdf",
            "content_type": "application/pdf",
            "metadata": {"doc_family": "donor_policy", "source_type": "donor_guidance"},
            "result": {"chunks_ingested": 10},
        }
    )
    api_app_module.INGEST_AUDIT_STORE.append(
        {
            "event_id": "pref-depth-2",
            "ts": "2026-03-03T10:05:00+00:00",
            "donor_id": "usaid",
            "tenant_id": None,
            "namespace": "usaid_ads201",
            "filename": "country.pdf",
            "content_type": "application/pdf",
            "metadata": {"doc_family": "country_context", "source_type": "country_context"},
            "result": {"chunks_ingested": 8},
        }
    )

    response = client.post("/generate/preflight", json={"donor_id": "usaid"})
    assert response.status_code == 200
    body = response.json()
    assert body["donor_id"] == "usaid"
    assert body["namespace_empty"] is False
    assert body["coverage_rate"] == 0.6667
    assert body["depth_ready_count"] == 1
    assert body["depth_gap_count"] == 2
    assert body["depth_coverage_rate"] == 0.3333
    warnings = body.get("warnings") or []
    warning_codes = {str(row.get("code") or "") for row in warnings if isinstance(row, dict)}
    assert "LOW_DOC_DEPTH_COVERAGE" in warning_codes
    grounding_policy = body.get("grounding_policy") or {}
    assert grounding_policy["risk_level"] in {"medium", "high"}
    assert "depth_coverage_below_medium_threshold" in (grounding_policy.get("reasons") or [])


def test_ingest_readiness_reports_preflight_payload_and_expected_families_override():
    api_app_module.INGEST_AUDIT_STORE.clear()
    api_app_module.INGEST_AUDIT_STORE.append(
        {
            "event_id": "ing-read-1",
            "ts": "2026-03-03T10:00:00+00:00",
            "donor_id": "usaid",
            "tenant_id": None,
            "namespace": "usaid_ads201",
            "filename": "ads201-a.pdf",
            "content_type": "application/pdf",
            "metadata": {"doc_family": "donor_policy", "source_type": "donor_guidance"},
            "result": {"chunks_ingested": 7},
        }
    )
    api_app_module.INGEST_AUDIT_STORE.append(
        {
            "event_id": "ing-read-2",
            "ts": "2026-03-03T10:01:00+00:00",
            "donor_id": "usaid",
            "tenant_id": None,
            "namespace": "usaid_ads201",
            "filename": "ads201-b.pdf",
            "content_type": "application/pdf",
            "metadata": {"doc_family": "donor_policy", "source_type": "donor_guidance"},
            "result": {"chunks_ingested": 6},
        }
    )
    api_app_module.INGEST_AUDIT_STORE.append(
        {
            "event_id": "ing-read-3",
            "ts": "2026-03-03T10:02:00+00:00",
            "donor_id": "usaid",
            "tenant_id": None,
            "namespace": "usaid_ads201",
            "filename": "country.pdf",
            "content_type": "application/pdf",
            "metadata": {"doc_family": "country_context", "source_type": "country_context"},
            "result": {"chunks_ingested": 5},
        }
    )

    response = client.post(
        "/ingest/readiness",
        json={
            "donor_id": "usaid",
            "expected_doc_families": ["donor_policy", "country_context"],
            "input_context": {"project": "AI training", "country": "Kazakhstan"},
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["donor_id"] == "usaid"
    assert isinstance(body.get("retrieval_query_terms"), list)
    assert "development objectives" in body.get("retrieval_query_terms")
    assert body["expected_doc_families"] == ["donor_policy", "country_context"]
    assert body["present_doc_families"] == ["donor_policy", "country_context"]
    assert body["missing_doc_families"] == []
    assert body["coverage_rate"] == 1.0
    assert body["depth_coverage_rate"] == 1.0
    assert body["risk_level"] in {"none", "low"}
    assert body["namespace_empty"] is False
    claims = body.get("architect_claims") or {}
    assert claims.get("available") is True


def test_generate_with_tenant_id_uses_tenant_scoped_namespace():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "tenant_id": "Acme Gov",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] in {"done", "error", "pending_hitl"}
    state = status.get("state") or {}
    assert state.get("tenant_id") == "acme_gov"
    assert state.get("rag_namespace") == "acme_gov/usaid_ads201"
    retrieval = state.get("architect_retrieval") or {}
    assert retrieval.get("namespace") == "acme_gov/usaid_ads201"
    assert retrieval.get("namespace_normalized") == "acme_gov_usaid_ads201"


def test_generate_requires_allowed_tenant_when_tenant_authz_enabled(monkeypatch):
    monkeypatch.setenv("GRANTFLOW_TENANT_AUTHZ_ENABLED", "true")
    monkeypatch.setenv("GRANTFLOW_ALLOWED_TENANTS", "tenant_alpha")

    missing_tenant = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert missing_tenant.status_code == 403
    assert "tenant_id" in str(missing_tenant.json().get("detail") or "").lower()

    not_allowed = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "tenant_id": "tenant_beta",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert not_allowed.status_code == 403

    allowed = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "tenant_id": "tenant_alpha",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert allowed.status_code == 200


def test_status_endpoint_blocks_cross_tenant_when_tenant_authz_enabled(monkeypatch):
    monkeypatch.setenv("GRANTFLOW_TENANT_AUTHZ_ENABLED", "true")
    monkeypatch.setenv("GRANTFLOW_ALLOWED_TENANTS", "tenant_red,tenant_blue")

    generated = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "tenant_id": "tenant_red",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert generated.status_code == 200
    job_id = generated.json()["job_id"]

    forbidden = client.get(f"/status/{job_id}", headers={"X-Tenant-ID": "tenant_blue"})
    assert forbidden.status_code == 403

    allowed = client.get(f"/status/{job_id}", headers={"X-Tenant-ID": "tenant_red"})
    assert allowed.status_code == 200


def test_portfolio_metrics_filters_by_tenant_when_tenant_authz_enabled(monkeypatch):
    monkeypatch.setenv("GRANTFLOW_TENANT_AUTHZ_ENABLED", "true")
    monkeypatch.setenv("GRANTFLOW_ALLOWED_TENANTS", "tenant_portfolio_a,tenant_portfolio_b")

    a_job = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "tenant_id": "tenant_portfolio_a",
            "input_context": {"project": "Gov Services A", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert a_job.status_code == 200

    b_job = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "tenant_id": "tenant_portfolio_b",
            "input_context": {"project": "Gov Services B", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert b_job.status_code == 200

    a_metrics = client.get("/portfolio/metrics", headers={"X-Tenant-ID": "tenant_portfolio_a"})
    assert a_metrics.status_code == 200
    a_body = a_metrics.json()
    assert a_body["job_count"] >= 1
    assert a_body["donor_counts"].get("usaid", 0) >= 1

    b_metrics = client.get("/portfolio/metrics", headers={"X-Tenant-ID": "tenant_portfolio_b"})
    assert b_metrics.status_code == 200
    b_body = b_metrics.json()
    assert b_body["job_count"] >= 1
    assert b_body["donor_counts"].get("usaid", 0) >= 1


def test_hitl_pending_filters_by_tenant_when_tenant_authz_enabled(monkeypatch):
    monkeypatch.setenv("GRANTFLOW_TENANT_AUTHZ_ENABLED", "true")
    monkeypatch.setenv("GRANTFLOW_ALLOWED_TENANTS", "tenant_hitl_a,tenant_hitl_b")

    a_job = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "tenant_id": "tenant_hitl_a",
            "input_context": {"project": "HITL A", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
        },
    )
    assert a_job.status_code == 200
    a_job_id = a_job.json()["job_id"]
    deadline = time.time() + 3.0
    while time.time() < deadline:
        a_status = client.get(f"/status/{a_job_id}", headers={"X-Tenant-ID": "tenant_hitl_a"})
        assert a_status.status_code == 200
        if a_status.json()["status"] == "pending_hitl":
            break
        time.sleep(0.05)
    else:
        raise AssertionError("Timed out waiting for tenant A HITL pending state")

    b_job = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "tenant_id": "tenant_hitl_b",
            "input_context": {"project": "HITL B", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
        },
    )
    assert b_job.status_code == 200
    b_job_id = b_job.json()["job_id"]
    deadline = time.time() + 3.0
    while time.time() < deadline:
        b_status = client.get(f"/status/{b_job_id}", headers={"X-Tenant-ID": "tenant_hitl_b"})
        assert b_status.status_code == 200
        if b_status.json()["status"] == "pending_hitl":
            break
        time.sleep(0.05)
    else:
        raise AssertionError("Timed out waiting for tenant B HITL pending state")

    tenant_a_pending = client.get("/hitl/pending", headers={"X-Tenant-ID": "tenant_hitl_a"})
    assert tenant_a_pending.status_code == 200
    tenant_a_body = tenant_a_pending.json()
    assert tenant_a_body["pending_count"] >= 1
    assert all((cp.get("tenant_id") == "tenant_hitl_a") for cp in tenant_a_body["checkpoints"])

    tenant_b_pending = client.get("/hitl/pending", headers={"X-Tenant-ID": "tenant_hitl_b"})
    assert tenant_b_pending.status_code == 200
    tenant_b_body = tenant_b_pending.json()
    assert tenant_b_body["pending_count"] >= 1
    assert all((cp.get("tenant_id") == "tenant_hitl_b") for cp in tenant_b_body["checkpoints"])


def test_status_write_endpoints_block_cross_tenant_when_tenant_authz_enabled(monkeypatch):
    monkeypatch.setenv("GRANTFLOW_TENANT_AUTHZ_ENABLED", "true")
    monkeypatch.setenv("GRANTFLOW_ALLOWED_TENANTS", "tenant_write_a,tenant_write_b")

    generated = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "tenant_id": "tenant_write_a",
            "input_context": {"project": "Cross tenant writes", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert generated.status_code == 200
    job_id = generated.json()["job_id"]

    deadline = time.time() + 3.0
    while time.time() < deadline:
        status = client.get(f"/status/{job_id}", headers={"X-Tenant-ID": "tenant_write_a"})
        assert status.status_code == 200
        if status.json()["status"] in {"done", "error", "pending_hitl"}:
            break
        time.sleep(0.05)
    else:
        raise AssertionError("Timed out waiting for tenant A terminal status")

    cross_tenant_headers = {"X-Tenant-ID": "tenant_write_b"}
    cross_tenant_requests = [
        ("POST", f"/cancel/{job_id}", {}),
        (
            "POST",
            f"/status/{job_id}/comments",
            {"section": "general", "message": "cross tenant create"},
        ),
        ("POST", f"/status/{job_id}/comments/nonexistent/resolve", {}),
        ("POST", f"/status/{job_id}/comments/nonexistent/reopen", {}),
        ("POST", f"/status/{job_id}/critic/findings/nonexistent/ack", {}),
        ("POST", f"/status/{job_id}/critic/findings/nonexistent/open", {}),
        ("POST", f"/status/{job_id}/critic/findings/nonexistent/resolve", {}),
        (
            "POST",
            f"/status/{job_id}/critic/findings/bulk-status",
            {"next_status": "acknowledged", "apply_to_all": True},
        ),
        ("POST", f"/status/{job_id}/review/workflow/sla/recompute", {}),
    ]
    for method, path, payload in cross_tenant_requests:
        response = client.request(method, path, headers=cross_tenant_headers, json=payload)
        assert response.status_code == 403

    owner_comment = client.post(
        f"/status/{job_id}/comments",
        headers={"X-Tenant-ID": "tenant_write_a"},
        json={"section": "general", "message": "owner comment"},
    )
    assert owner_comment.status_code == 200


def test_hitl_write_endpoints_block_cross_tenant_when_tenant_authz_enabled(monkeypatch):
    monkeypatch.setenv("GRANTFLOW_TENANT_AUTHZ_ENABLED", "true")
    monkeypatch.setenv("GRANTFLOW_ALLOWED_TENANTS", "tenant_hitl_write_a,tenant_hitl_write_b")

    generated = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "tenant_id": "tenant_hitl_write_a",
            "input_context": {"project": "HITL tenant writes", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
        },
    )
    assert generated.status_code == 200
    job_id = generated.json()["job_id"]

    deadline = time.time() + 3.0
    checkpoint_id = None
    while time.time() < deadline:
        status = client.get(f"/status/{job_id}", headers={"X-Tenant-ID": "tenant_hitl_write_a"})
        assert status.status_code == 200
        body = status.json()
        if body["status"] == "pending_hitl":
            checkpoint_id = body.get("checkpoint_id")
            break
        time.sleep(0.05)
    else:
        raise AssertionError("Timed out waiting for tenant HITL pending state")
    assert checkpoint_id

    cross_headers = {"X-Tenant-ID": "tenant_hitl_write_b"}
    cross_approve = client.post(
        "/hitl/approve",
        headers=cross_headers,
        json={"checkpoint_id": checkpoint_id, "approved": True, "feedback": "cross tenant"},
    )
    assert cross_approve.status_code == 403

    cross_resume = client.post(f"/resume/{job_id}", headers=cross_headers, json={})
    assert cross_resume.status_code == 403

    owner_approve = client.post(
        "/hitl/approve",
        headers={"X-Tenant-ID": "tenant_hitl_write_a"},
        json={"checkpoint_id": checkpoint_id, "approved": True, "feedback": "owner approved"},
    )
    assert owner_approve.status_code == 200

    owner_resume = client.post(f"/resume/{job_id}", headers={"X-Tenant-ID": "tenant_hitl_write_a"}, json={})
    assert owner_resume.status_code == 200


def test_generate_response_and_status_include_preflight_payload():
    api_app_module.INGEST_AUDIT_STORE.clear()

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "accepted"
    preflight = data.get("preflight")
    assert isinstance(preflight, dict)
    assert preflight["donor_id"] == "usaid"
    assert preflight["architect_rag_enabled"] is True
    assert preflight["risk_level"] == "high"
    assert preflight["grounding_risk_level"] == "high"
    assert isinstance(preflight.get("grounding_policy"), dict)

    status = _wait_for_terminal_status(data["job_id"])
    assert status["generate_preflight"] == preflight
    assert status["state"]["generate_preflight"] == preflight
    assert status["state"]["architect_rag_enabled"] is True
    assert status["strict_preflight"] is False
    assert status["state"]["strict_preflight"] is False


def test_generate_respects_architect_rag_enabled_flag():
    api_app_module.INGEST_AUDIT_STORE.clear()

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "architect_rag_enabled": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    data = response.json()
    assert data["preflight"]["architect_rag_enabled"] is False
    job_id = data["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["state"]["architect_rag_enabled"] is False


def test_generate_strict_preflight_blocks_when_risk_is_high():
    api_app_module.INGEST_AUDIT_STORE.clear()

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
            "strict_preflight": True,
        },
    )
    assert response.status_code == 409
    detail = response.json()["detail"]
    assert detail["reason"] == "preflight_high_risk_block"
    assert detail["preflight"]["risk_level"] == "high"
    assert detail["preflight"]["grounding_risk_level"] == "high"
    assert "strict_reasons" in detail
    assert detail["preflight"]["go_ahead"] is False


def test_generate_strict_preflight_allows_when_risk_is_not_high(monkeypatch):
    monkeypatch.setattr(
        api_app_module,
        "_build_generate_preflight",
        lambda donor_id, strategy, client_metadata, **kwargs: {
            "donor_id": donor_id,
            "risk_level": "medium",
            "go_ahead": True,
            "warning_count": 1,
            "retrieval_namespace": "usaid_ads201",
            "namespace_empty": False,
            "warnings": [{"code": "LOW_DOC_COVERAGE", "severity": "medium"}],
        },
    )
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
            "strict_preflight": True,
        },
    )
    assert response.status_code == 200
    assert response.json()["status"] == "accepted"


def test_generate_strict_preflight_blocks_when_grounding_risk_is_high(monkeypatch):
    monkeypatch.setattr(
        api_app_module,
        "_build_generate_preflight",
        lambda donor_id, strategy, client_metadata, **kwargs: {
            "donor_id": donor_id,
            "risk_level": "medium",
            "grounding_risk_level": "high",
            "go_ahead": True,
            "warning_count": 1,
            "retrieval_namespace": "usaid_ads201",
            "namespace_empty": False,
            "warnings": [{"code": "LOW_DOC_COVERAGE", "severity": "medium"}],
            "grounding_policy": {
                "mode": "warn",
                "risk_level": "high",
                "blocking": False,
                "go_ahead": True,
                "summary": "namespace_empty_or_low_coverage",
                "reasons": ["coverage_below_50pct"],
            },
        },
    )
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
            "strict_preflight": True,
        },
    )
    assert response.status_code == 409
    detail = response.json()["detail"]
    assert detail["reason"] == "preflight_high_risk_block"
    assert "grounding_risk_high" in (detail.get("strict_reasons") or [])
    assert detail["preflight"]["risk_level"] == "medium"
    assert detail["preflight"]["grounding_risk_level"] == "high"


def test_generate_require_grounded_generation_blocks_llm_when_grounding_risk_is_high(monkeypatch):
    monkeypatch.setattr(
        api_app_module,
        "_build_generate_preflight",
        lambda donor_id, strategy, client_metadata, **kwargs: {
            "donor_id": donor_id,
            "risk_level": "medium",
            "grounding_risk_level": "high",
            "go_ahead": True,
            "warning_count": 1,
            "retrieval_namespace": "usaid_ads201",
            "namespace_empty": False,
            "warnings": [{"code": "LOW_DOC_COVERAGE", "severity": "medium"}],
            "grounding_policy": {
                "mode": "warn",
                "risk_level": "high",
                "blocking": False,
                "go_ahead": True,
                "summary": "namespace_empty_or_low_coverage",
                "reasons": ["coverage_below_50pct"],
            },
        },
    )
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": True,
            "require_grounded_generation": True,
            "hitl_enabled": False,
            "strict_preflight": False,
        },
    )
    assert response.status_code == 409
    detail = response.json()["detail"]
    assert detail["reason"] == "llm_grounded_generation_block"
    assert "grounding_risk_high" in (detail.get("strict_reasons") or [])
    assert detail["preflight"]["grounding_risk_level"] == "high"


def test_generate_blocks_when_preflight_grounding_policy_is_strict(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "grounding_gate_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "preflight_grounding_policy_mode", "strict")
    api_app_module.INGEST_AUDIT_STORE.clear()

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
            "strict_preflight": False,
        },
    )
    assert response.status_code == 409
    detail = response.json()["detail"]
    assert detail["reason"] == "preflight_grounding_policy_block"
    assert detail["preflight"]["grounding_policy"]["mode"] == "strict"
    assert detail["preflight"]["grounding_policy"]["blocking"] is True
    assert detail["preflight"]["go_ahead"] is False


def test_generate_basic_async_job_flow():
    payload = {
        "donor_id": "usaid",
        "input_context": {"project": "Water Sanitation", "country": "Kenya"},
        "llm_mode": False,
        "hitl_enabled": False,
    }
    response = client.post("/generate", json=payload)
    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "accepted"
    assert "job_id" in data

    status = _wait_for_terminal_status(data["job_id"])
    assert status["status"] == "done"
    state = status["state"]
    assert state["toc_draft"]
    assert state["logframe_draft"]
    assert state["quality_score"] >= 0
    critic_notes = state.get("critic_notes") or {}
    assert isinstance(critic_notes.get("fatal_flaws"), list)
    if critic_notes.get("fatal_flaws"):
        flaw = critic_notes["fatal_flaws"][0]
        assert isinstance(flaw, dict)
        assert "code" in flaw
        assert "section" in flaw
        assert "message" in flaw
    assert isinstance(critic_notes.get("rule_checks"), list)


def test_generate_llm_mode_false_uses_non_llm_toc_engine():
    payload = {
        "donor_id": "usaid",
        "input_context": {"project": "Public Administration", "country": "Kazakhstan"},
        "llm_mode": False,
        "hitl_enabled": False,
    }
    response = client.post("/generate", json=payload)
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "done"
    state = status["state"]
    generation_meta = state.get("toc_generation_meta") or {}
    assert generation_meta.get("llm_used") is False
    assert str(generation_meta.get("engine") or "").startswith("deterministic:")
    assert generation_meta.get("llm_requested") is False
    assert generation_meta.get("llm_attempted") is False
    assert generation_meta.get("fallback_used") is False
    assert generation_meta.get("fallback_class") == "deterministic_mode"


def test_strict_grounding_gate_blocks_job_finalization(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "grounding_gate_mode", "strict")
    monkeypatch.setattr(
        api_app_module,
        "_build_generate_preflight",
        lambda donor_id, strategy, client_metadata, **kwargs: {
            "donor_id": donor_id,
            "risk_level": "low",
            "grounding_risk_level": "low",
            "go_ahead": True,
            "warning_count": 0,
            "warnings": [],
            "retrieval_namespace": "usaid_ads201",
            "namespace_empty": False,
            "grounding_policy": {
                "mode": "strict",
                "risk_level": "low",
                "blocking": False,
                "go_ahead": True,
                "summary": "readiness_signals_ok",
                "reasons": ["sufficient_readiness_signals"],
            },
        },
    )
    monkeypatch.setattr(
        api_app_module.vector_store,
        "query",
        lambda namespace, query_texts, n_results=5, where=None, top_k=None: {
            "documents": [[]],
            "metadatas": [[]],
            "ids": [[]],
        },
    )

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Gov services modernization", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    terminal = _wait_for_terminal_status(job_id)
    assert terminal["status"] == "error"
    assert "Grounding gate (strict) blocked finalization" in str(terminal.get("error") or "")
    state = terminal.get("state") or {}
    gate = state.get("grounding_gate") or {}
    assert gate.get("mode") == "strict"
    assert gate.get("blocking") is True
    assert gate.get("passed") is False


def test_runtime_grounded_quality_gate_blocks_llm_when_non_retrieval_signals_dominate(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "grounding_gate_mode", "warn")
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_policy_mode", "warn")
    monkeypatch.setattr(api_app_module.config.graph, "runtime_grounded_quality_gate_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "runtime_grounded_quality_gate_min_citations", 1)
    monkeypatch.setattr(
        api_app_module.config.graph, "runtime_grounded_quality_gate_max_non_retrieval_citation_rate", 0.2
    )
    monkeypatch.setattr(
        api_app_module.config.graph, "runtime_grounded_quality_gate_min_retrieval_grounded_citations", 1
    )
    monkeypatch.setattr(
        api_app_module,
        "_build_generate_preflight",
        lambda donor_id, strategy, client_metadata, **kwargs: {
            "donor_id": donor_id,
            "risk_level": "low",
            "grounding_risk_level": "low",
            "go_ahead": True,
            "warning_count": 0,
            "warnings": [],
            "retrieval_namespace": "usaid_ads201",
            "namespace_empty": False,
            "grounding_policy": {
                "mode": "warn",
                "risk_level": "low",
                "blocking": False,
                "go_ahead": True,
                "summary": "readiness_signals_ok",
                "reasons": ["sufficient_readiness_signals"],
            },
        },
    )
    monkeypatch.setattr(
        api_app_module.vector_store,
        "query",
        lambda namespace, query_texts, n_results=5, where=None, top_k=None: {
            "documents": [[]],
            "metadatas": [[]],
            "ids": [[]],
        },
    )

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Gov workflow modernization", "country": "Kenya"},
            "llm_mode": True,
            "hitl_enabled": False,
            "architect_rag_enabled": True,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    terminal = _wait_for_terminal_status(job_id)
    assert terminal["status"] == "error"
    assert "Grounded quality gate (strict) blocked finalization" in str(terminal.get("error") or "")
    state = terminal.get("state") or {}
    gate = state.get("grounded_quality_gate") or {}
    assert gate.get("mode") == "strict"
    assert gate.get("applicable") is True
    assert gate.get("blocking") is True
    assert gate.get("passed") is False
    reasons = gate.get("reasons") if isinstance(gate.get("reasons"), list) else []
    assert "non_retrieval_citation_rate_above_max" in reasons
    reason_details = gate.get("reason_details") if isinstance(gate.get("reason_details"), list) else []
    assert reason_details
    detail_codes = {str(item.get("code") or "") for item in reason_details if isinstance(item, dict)}
    assert "non_retrieval_citation_rate_above_max" in detail_codes
    section_signals = gate.get("section_signals") if isinstance(gate.get("section_signals"), dict) else {}
    assert "toc" in section_signals
    assert "logframe" in section_signals
    flaws = (state.get("critic_notes") or {}).get("fatal_flaws") or []
    flaw_codes = {str(item.get("code") or "") for item in flaws if isinstance(item, dict)}
    assert "RUNTIME_GROUNDED_QUALITY_GATE_BLOCK" in flaw_codes
    runtime_gate_flaw = next(
        (
            item
            for item in flaws
            if isinstance(item, dict) and str(item.get("code") or "") == "RUNTIME_GROUNDED_QUALITY_GATE_BLOCK"
        ),
        {},
    )
    related_sections = runtime_gate_flaw.get("related_sections")
    assert isinstance(related_sections, list)
    assert any(section in {"toc", "logframe"} for section in related_sections)


def test_runtime_grounded_quality_gate_skips_when_architect_rag_disabled(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "grounding_gate_mode", "warn")
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_policy_mode", "warn")
    monkeypatch.setattr(api_app_module.config.graph, "runtime_grounded_quality_gate_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "runtime_grounded_quality_gate_min_citations", 1)
    monkeypatch.setattr(
        api_app_module.config.graph, "runtime_grounded_quality_gate_max_non_retrieval_citation_rate", 0.0
    )
    monkeypatch.setattr(
        api_app_module.config.graph, "runtime_grounded_quality_gate_min_retrieval_grounded_citations", 99
    )
    monkeypatch.setattr(
        api_app_module,
        "_build_generate_preflight",
        lambda donor_id, strategy, client_metadata, **kwargs: {
            "donor_id": donor_id,
            "risk_level": "low",
            "grounding_risk_level": "low",
            "go_ahead": True,
            "warning_count": 0,
            "warnings": [],
            "retrieval_namespace": "usaid_ads201",
            "namespace_empty": False,
            "grounding_policy": {
                "mode": "warn",
                "risk_level": "low",
                "blocking": False,
                "go_ahead": True,
                "summary": "readiness_signals_ok",
                "reasons": ["sufficient_readiness_signals"],
            },
        },
    )
    monkeypatch.setattr(
        api_app_module.vector_store,
        "query",
        lambda namespace, query_texts, n_results=5, where=None, top_k=None: {
            "documents": [[]],
            "metadatas": [[]],
            "ids": [[]],
        },
    )

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Gov workflow modernization", "country": "Kenya"},
            "llm_mode": True,
            "hitl_enabled": False,
            "architect_rag_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    terminal = _wait_for_terminal_status(job_id)
    assert terminal["status"] == "done"
    state = terminal.get("state") or {}
    gate = state.get("grounded_quality_gate") or {}
    assert gate.get("mode") == "strict"
    assert gate.get("applicable") is False
    assert gate.get("blocking") is False
    assert gate.get("passed") is True


def test_runtime_grounded_quality_gate_passes_with_seeded_corpus_grounded_llm(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "grounding_gate_mode", "warn")
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_policy_mode", "warn")
    monkeypatch.setattr(api_app_module.config.graph, "runtime_grounded_quality_gate_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "runtime_grounded_quality_gate_min_citations", 5)
    monkeypatch.setattr(
        api_app_module.config.graph, "runtime_grounded_quality_gate_max_non_retrieval_citation_rate", 0.35
    )
    monkeypatch.setattr(
        api_app_module.config.graph, "runtime_grounded_quality_gate_min_retrieval_grounded_citations", 2
    )
    monkeypatch.setattr(
        api_app_module,
        "_build_generate_preflight",
        lambda donor_id, strategy, client_metadata, **kwargs: {
            "donor_id": donor_id,
            "risk_level": "low",
            "grounding_risk_level": "low",
            "go_ahead": True,
            "warning_count": 0,
            "warnings": [],
            "retrieval_namespace": "usaid_ads201",
            "namespace_empty": False,
            "grounding_policy": {
                "mode": "warn",
                "risk_level": "low",
                "blocking": False,
                "go_ahead": True,
                "summary": "readiness_signals_ok",
                "reasons": ["sufficient_readiness_signals"],
            },
        },
    )

    api_app_module.INGEST_AUDIT_STORE.clear()
    api_app_module.INGEST_AUDIT_STORE.append(
        {
            "event_id": "seed-gate-pass-1",
            "ts": "2026-03-01T09:00:00+00:00",
            "donor_id": "usaid",
            "namespace": "usaid_ads201",
            "filename": "ads-policy-seed.pdf",
            "content_type": "application/pdf",
            "metadata": {"doc_family": "donor_policy", "source_type": "donor_guidance"},
            "result": {"chunks_ingested": 10},
        }
    )
    api_app_module.INGEST_AUDIT_STORE.append(
        {
            "event_id": "seed-gate-pass-2",
            "ts": "2026-03-01T09:01:00+00:00",
            "donor_id": "usaid",
            "namespace": "usaid_ads201",
            "filename": "responsible-ai-seed.pdf",
            "content_type": "application/pdf",
            "metadata": {"doc_family": "responsible_ai_guidance", "source_type": "reference_guidance"},
            "result": {"chunks_ingested": 8},
        }
    )
    api_app_module.INGEST_AUDIT_STORE.append(
        {
            "event_id": "seed-gate-pass-3",
            "ts": "2026-03-01T09:02:00+00:00",
            "donor_id": "usaid",
            "namespace": "usaid_ads201",
            "filename": "country-context-seed.pdf",
            "content_type": "application/pdf",
            "metadata": {"doc_family": "country_context", "source_type": "country_context"},
            "result": {"chunks_ingested": 7},
        }
    )

    def fake_query(namespace, query_texts, n_results=5, where=None, top_k=None):
        query_rows = query_texts if isinstance(query_texts, list) and query_texts else [str(query_texts or "")]
        docs: list[list[str]] = []
        metadatas: list[list[dict[str, object]]] = []
        ids: list[list[str]] = []
        distances: list[list[float]] = []
        limit = max(1, min(int(n_results or top_k or 5), 4))
        for q_idx, query in enumerate(query_rows):
            docs_row: list[str] = []
            metas_row: list[dict[str, object]] = []
            ids_row: list[str] = []
            distances_row: list[float] = []
            for hit_idx in range(limit):
                doc_id = f"{namespace}-seed-doc-{q_idx + 1}-{hit_idx + 1}"
                chunk_id = f"{doc_id}-chunk"
                docs_row.append(
                    f"Seed evidence for grounded drafting. Query={str(query)[:120]}. "
                    f"Reference #{hit_idx + 1} focuses on USAID-style compliance and MEL traceability."
                )
                metas_row.append(
                    {
                        "doc_id": doc_id,
                        "chunk_id": chunk_id,
                        "source": f"seed-source-{hit_idx + 1}.pdf",
                        "page": hit_idx + 1,
                        "doc_family": "donor_policy" if hit_idx == 0 else "responsible_ai_guidance",
                    }
                )
                ids_row.append(chunk_id)
                distances_row.append(0.05 + (hit_idx * 0.02))
            docs.append(docs_row)
            metadatas.append(metas_row)
            ids.append(ids_row)
            distances.append(distances_row)
        return {"documents": docs, "metadatas": metadatas, "ids": ids, "distances": distances}

    monkeypatch.setattr(api_app_module.vector_store, "query", fake_query)

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {
                "project": "AI civil service capability and governance modernization",
                "country": "Kazakhstan",
                "timeframe": "2026-2027",
            },
            "llm_mode": True,
            "hitl_enabled": False,
            "architect_rag_enabled": True,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    terminal = _wait_for_terminal_status(job_id, timeout_s=5.0)
    assert terminal["status"] == "done"
    state = terminal.get("state") or {}
    gate = state.get("grounded_quality_gate") or {}
    assert gate.get("mode") == "strict"
    assert gate.get("applicable") is True
    assert gate.get("passed") is True
    assert gate.get("blocking") is False
    assert float(gate.get("non_retrieval_citation_rate") or 0.0) <= 0.35
    assert int(gate.get("retrieval_grounded_citation_count") or 0) >= 2
    section_signals = gate.get("section_signals") if isinstance(gate.get("section_signals"), dict) else {}
    assert "toc" in section_signals
    assert "logframe" in section_signals

    grounding_gate_response = client.get(f"/status/{job_id}/grounding-gate")
    assert grounding_gate_response.status_code == 200
    grounding_gate_body = grounding_gate_response.json()
    runtime_gate = grounding_gate_body.get("grounded_gate") or {}
    assert runtime_gate.get("passed") is True
    assert runtime_gate.get("blocking") is False


def test_strict_mel_grounding_policy_blocks_job_finalization(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_policy_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_min_mel_citations", 2)
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_min_claim_support_rate", 0.75)
    monkeypatch.setattr(api_app_module.config.graph, "grounding_gate_mode", "warn")
    monkeypatch.setattr(
        api_app_module,
        "_build_generate_preflight",
        lambda donor_id, strategy, client_metadata, **kwargs: {
            "donor_id": donor_id,
            "risk_level": "low",
            "grounding_risk_level": "low",
            "go_ahead": True,
            "warning_count": 0,
            "warnings": [],
            "retrieval_namespace": "usaid_ads201",
            "namespace_empty": False,
            "grounding_policy": {
                "mode": "warn",
                "risk_level": "low",
                "blocking": False,
                "go_ahead": True,
                "summary": "readiness_signals_ok",
                "reasons": ["sufficient_readiness_signals"],
            },
        },
    )
    monkeypatch.setattr(
        api_app_module.vector_store,
        "query",
        lambda namespace, query_texts, n_results=5, where=None, top_k=None: {
            "documents": [[]],
            "metadatas": [[]],
            "ids": [[]],
            "distances": [[]],
        },
    )

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Gov services modernization", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    terminal = _wait_for_terminal_status(job_id)
    assert terminal["status"] == "error"
    assert "MEL grounding policy (strict) blocked finalization" in str(terminal.get("error") or "")
    state = terminal.get("state") or {}
    mel_policy = state.get("mel_grounding_policy") or {}
    assert mel_policy.get("mode") == "strict"
    assert mel_policy.get("blocking") is True
    assert mel_policy.get("passed") is False
    assert "mel_claim_support_rate_below_min" in (mel_policy.get("reasons") or [])


def test_strict_mel_grounding_policy_blocks_after_hitl_resume(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_policy_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_min_mel_citations", 2)
    monkeypatch.setattr(api_app_module.config.graph, "mel_grounding_min_claim_support_rate", 0.75)
    monkeypatch.setattr(api_app_module.config.graph, "grounding_gate_mode", "warn")
    monkeypatch.setattr(
        api_app_module,
        "_build_generate_preflight",
        lambda donor_id, strategy, client_metadata, **kwargs: {
            "donor_id": donor_id,
            "risk_level": "low",
            "grounding_risk_level": "low",
            "go_ahead": True,
            "warning_count": 0,
            "warnings": [],
            "retrieval_namespace": "usaid_ads201",
            "namespace_empty": False,
            "grounding_policy": {
                "mode": "warn",
                "risk_level": "low",
                "blocking": False,
                "go_ahead": True,
                "summary": "readiness_signals_ok",
                "reasons": ["sufficient_readiness_signals"],
            },
        },
    )
    monkeypatch.setattr(
        api_app_module.vector_store,
        "query",
        lambda namespace, query_texts, n_results=5, where=None, top_k=None: {
            "documents": [[]],
            "metadatas": [[]],
            "ids": [[]],
            "distances": [[]],
        },
    )

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Gov services modernization", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    terminal = None
    resume_targets = []
    # GrantFlow may re-enter HITL checkpoints across revision iterations.
    # Keep approving and resuming until the job reaches a terminal non-HITL status.
    for _ in range(10):
        status = _wait_for_terminal_status(job_id, timeout_s=5.0)
        if status["status"] != "pending_hitl":
            terminal = status
            break
        cp_id = status["checkpoint_id"]
        approve = client.post("/hitl/approve", json={"checkpoint_id": cp_id, "approved": True, "feedback": "ok"})
        assert approve.status_code == 200
        resume = client.post(f"/resume/{job_id}", json={})
        assert resume.status_code == 200
        resume_from = str(resume.json().get("resuming_from") or "")
        resume_targets.append(resume_from)
    assert terminal is not None, "Timed out waiting for job finalization with repeated HITL resumes"
    assert "mel" in resume_targets
    assert "critic" in resume_targets
    assert terminal["status"] == "error"
    assert "MEL grounding policy (strict) blocked finalization" in str(terminal.get("error") or "")
    state = terminal.get("state") or {}
    mel_policy = state.get("mel_grounding_policy") or {}
    assert mel_policy.get("blocking") is True


def test_export_endpoint_blocks_when_strict_grounding_gate_failed_unless_overridden():
    payload = {
        "state": {
            "donor_id": "usaid",
            "toc_draft": {"toc": {"brief": "Sample ToC"}},
            "logframe_draft": {"indicators": []},
            "grounding_gate": {
                "mode": "strict",
                "passed": False,
                "blocking": True,
                "summary": "fallback_or_low_rag_citations_dominate",
                "reasons": ["fallback_or_low_rag_citations_dominate"],
            },
        }
    }

    blocked = client.post("/export", json={"payload": payload, "format": "docx"})
    assert blocked.status_code == 409
    detail = blocked.json()["detail"]
    assert detail["reason"] == "grounding_gate_strict_block"

    allowed = client.post("/export", json={"payload": payload, "format": "docx", "allow_unsafe_export": True})
    assert allowed.status_code == 200
    assert allowed.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


def test_export_endpoint_blocks_when_runtime_grounded_gate_policy_requires_pass(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "export_require_grounded_gate_pass", True)
    payload = {
        "state": {
            "donor_id": "usaid",
            "toc_draft": {"toc": {"brief": "Sample ToC"}},
            "logframe_draft": {"indicators": []},
            "grounded_quality_gate": {
                "mode": "strict",
                "passed": False,
                "blocking": True,
                "summary": "runtime_non_retrieval_signals_high",
                "reasons": ["non_retrieval_rate_above_max"],
                "failed_sections": ["toc"],
            },
        }
    }

    blocked = client.post("/export", json={"payload": payload, "format": "docx"})
    assert blocked.status_code == 409
    detail = blocked.json()["detail"]
    assert detail["reason"] == "runtime_grounded_quality_gate_block"
    grounded_gate = detail["grounded_gate"]
    assert grounded_gate["blocking"] is True
    assert grounded_gate["passed"] is False

    allowed = client.post("/export", json={"payload": payload, "format": "docx", "allow_unsafe_export": True})
    assert allowed.status_code == 200
    assert allowed.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


def test_export_endpoint_blocks_production_export_when_contract_policy_strict_and_missing_sections(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "export_contract_policy_mode", "strict")
    payload = {
        "state": {
            "donor_id": "eu",
            "toc_draft": {
                "toc": {
                    "overall_objective": {
                        "objective_id": "OO1",
                        "title": "Digital governance",
                        "rationale": "EU fit",
                    }
                }
            },
            "logframe_draft": {"indicators": []},
        }
    }

    blocked = client.post("/export", json={"payload": payload, "format": "docx", "production_export": True})
    assert blocked.status_code == 409
    detail = blocked.json()["detail"]
    assert detail["reason"] == "export_contract_policy_block"
    gate = detail["export_contract_gate"]
    assert gate["mode"] == "strict"
    assert gate["blocking"] is True
    assert gate["status"] == "warning"
    assert "specific_objectives" in (gate.get("missing_required_sections") or [])
    assert "missing_required_toc_sections" in (gate.get("reasons") or [])

    non_production = client.post("/export", json={"payload": payload, "format": "docx"})
    assert non_production.status_code == 200
    assert non_production.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    assert non_production.headers.get("x-grantflow-export-contract-mode") == "strict"
    assert non_production.headers.get("x-grantflow-export-contract-status") == "warning"

    overridden = client.post(
        "/export",
        json={"payload": payload, "format": "docx", "production_export": True, "allow_unsafe_export": True},
    )
    assert overridden.status_code == 200
    assert overridden.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


def test_export_endpoint_blocks_production_xlsx_when_primary_headers_missing(monkeypatch):
    from openpyxl import Workbook

    monkeypatch.setattr(api_app_module.config.graph, "export_contract_policy_mode", "strict")

    def _stub_build_xlsx_from_logframe(*_args, **_kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "LogFrame"
        ws.append(["Indicator ID", "Name", "Justification", "Citation", "Baseline", "Target"])
        usaid_sheet = wb.create_sheet("USAID_RF")
        usaid_sheet.append(["DO ID", "DO Description"])
        wb.create_sheet("Template Meta")
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.read()

    monkeypatch.setattr(api_app_module, "build_xlsx_from_logframe", _stub_build_xlsx_from_logframe)

    payload = {
        "state": {
            "donor_id": "usaid",
            "toc_draft": {
                "toc": {
                    "project_goal": "Improve civic services",
                    "development_objectives": [
                        {"do_id": "DO1", "description": "Improved digital delivery", "intermediate_results": []}
                    ],
                }
            },
            "logframe_draft": {"indicators": []},
        }
    }

    blocked = client.post("/export", json={"payload": payload, "format": "xlsx", "production_export": True})
    assert blocked.status_code == 409
    detail = blocked.json()["detail"]
    assert detail["reason"] == "export_contract_policy_block"
    gate = detail["export_contract_gate"]
    assert gate["mode"] == "strict"
    assert gate["blocking"] is True
    assert gate["status"] == "warning"
    assert "missing_required_primary_sheet_headers" in (gate.get("reasons") or [])
    assert "IR ID" in (gate.get("missing_required_primary_sheet_headers") or [])

    non_production = client.post("/export", json={"payload": payload, "format": "xlsx"})
    assert non_production.status_code == 200
    assert non_production.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert non_production.headers.get("x-grantflow-export-contract-status") == "warning"


def test_export_endpoint_blocks_when_export_claim_support_policy_strict_and_below_threshold(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_policy_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_min_architect_citations", 3)
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_min_claim_support_rate", 0.75)

    payload = {
        "state": {
            "donor_id": "usaid",
            "toc_draft": {"toc": {"brief": "Sample ToC"}},
            "logframe_draft": {"indicators": []},
            "citations": [
                {"stage": "architect", "citation_type": "rag_claim_support", "citation_confidence": 0.9},
                {"stage": "architect", "citation_type": "fallback_namespace", "citation_confidence": 0.1},
                {"stage": "architect", "citation_type": "rag_low_confidence", "citation_confidence": 0.2},
            ],
        }
    }

    blocked = client.post("/export", json={"payload": payload, "format": "docx"})
    assert blocked.status_code == 409
    detail = blocked.json()["detail"]
    assert detail["reason"] == "export_grounding_policy_block"
    policy = detail["export_grounding_policy"]
    assert policy["mode"] == "strict"
    assert policy["blocking"] is True
    assert policy["architect_citation_count"] == 3
    assert policy["architect_claim_support_citation_count"] == 1
    assert policy["architect_claim_support_rate"] == 0.3333
    assert "claim_support_rate_below_min" in (policy.get("reasons") or [])

    allowed = client.post("/export", json={"payload": payload, "format": "docx", "allow_unsafe_export": True})
    assert allowed.status_code == 200
    assert allowed.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


def test_export_endpoint_blocks_when_export_traceability_policy_strict_and_below_threshold(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_policy_mode", "strict")
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_min_architect_citations", 1)
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_min_claim_support_rate", 0.0)
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_min_traceability_complete_rate", 1.0)
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_max_traceability_gap_rate", 0.0)

    payload = {
        "state": {
            "donor_id": "usaid",
            "toc_draft": {"toc": {"brief": "Sample ToC"}},
            "logframe_draft": {"indicators": []},
            "citations": [
                {"stage": "architect", "citation_type": "rag_claim_support", "citation_confidence": 0.95},
                {"stage": "architect", "citation_type": "rag_claim_support", "citation_confidence": 0.91},
            ],
        }
    }

    blocked = client.post("/export", json={"payload": payload, "format": "docx"})
    assert blocked.status_code == 409
    detail = blocked.json()["detail"]
    assert detail["reason"] == "export_grounding_policy_block"
    policy = detail["export_grounding_policy"]
    assert policy["mode"] == "strict"
    assert policy["blocking"] is True
    assert policy["architect_traceability_complete_rate"] == 0.0
    assert policy["architect_traceability_gap_rate"] == 1.0
    assert "traceability_complete_rate_below_min" in (policy.get("reasons") or [])
    assert "traceability_gap_rate_above_max" in (policy.get("reasons") or [])

    allowed = client.post("/export", json={"payload": payload, "format": "docx", "allow_unsafe_export": True})
    assert allowed.status_code == 200
    assert allowed.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


def test_export_endpoint_warn_mode_does_not_block_on_low_claim_support(monkeypatch):
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_policy_mode", "warn")
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_min_architect_citations", 3)
    monkeypatch.setattr(api_app_module.config.graph, "export_grounding_min_claim_support_rate", 0.75)

    payload = {
        "state": {
            "donor_id": "usaid",
            "toc_draft": {"toc": {"brief": "Sample ToC"}},
            "logframe_draft": {"indicators": []},
            "citations": [
                {"stage": "architect", "citation_type": "fallback_namespace", "citation_confidence": 0.1},
                {"stage": "architect", "citation_type": "rag_low_confidence", "citation_confidence": 0.2},
                {"stage": "architect", "citation_type": "rag_low_confidence", "citation_confidence": 0.22},
            ],
        }
    }

    export = client.post("/export", json={"payload": payload, "format": "docx"})
    assert export.status_code == 200
    assert export.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


def test_status_redacts_internal_strategy_objects():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Governance Support", "country": "Ukraine"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "done"
    state = status["state"]
    assert "donor_strategy" not in state
    assert "strategy" not in state
    assert state["donor_id"] == "usaid"
    assert state["toc_draft"]


def test_status_critic_endpoint_returns_typed_payload():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Maternal Health", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    terminal = _wait_for_terminal_status(job_id)
    assert terminal["status"] == "done"

    critic_resp = client.get(f"/status/{job_id}/critic")
    assert critic_resp.status_code == 200
    body = critic_resp.json()
    assert body["job_id"] == job_id
    assert body["status"] == "done"
    assert isinstance(body["fatal_flaw_count"], int)
    assert isinstance(body["fatal_flaws"], list)
    assert isinstance(body["fatal_flaw_messages"], list)
    assert isinstance(body["rule_check_count"], int)
    assert isinstance(body["rule_checks"], list)
    assert body["rule_check_count"] == len(body["rule_checks"])
    assert body["fatal_flaw_count"] == len(body["fatal_flaws"])
    if body["rule_checks"]:
        check = body["rule_checks"][0]
        assert "code" in check
        assert "status" in check
        assert "section" in check
    if body["fatal_flaws"]:
        flaw = body["fatal_flaws"][0]
        assert "id" in flaw and flaw["id"]
        assert "finding_id" in flaw and flaw["finding_id"]
        assert flaw["id"] == flaw["finding_id"]
        assert flaw.get("status") in {"open", "acknowledged", "resolved"}
        assert "code" in flaw
        assert "severity" in flaw
        assert "section" in flaw
        assert "message" in flaw


def test_status_critic_findings_can_be_acknowledged_resolved_and_linked_to_comments():
    job_id = "critic-findings-linkage-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "quality_score": 6.5,
                "critic_score": 6.5,
                "needs_revision": True,
                "critic_notes": {
                    "engine": "rules",
                    "rule_score": 6.5,
                    "llm_score": None,
                    "llm_advisory_diagnostics": {
                        "llm_finding_count": 1,
                        "candidate_label_counts": {"TOY_LABEL": 1},
                        "advisory_candidate_count": 0,
                        "advisory_candidate_labels": [],
                        "advisory_applies": False,
                        "advisory_rejected_reason": "non_advisory_llm_finding_present",
                    },
                    "revision_instructions": "Fix ToC and indicators.",
                    "fatal_flaws": [
                        {
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "general",
                            "message": "Theory of Change draft does not match expected donor schema.",
                            "fix_hint": "Regenerate ToC with valid structured output.",
                            "source": "rules",
                        }
                    ],
                    "fatal_flaw_messages": ["Theory of Change draft does not match expected donor schema."],
                    "rule_checks": [{"code": "TOC_SCHEMA_VALID", "status": "fail", "section": "toc"}],
                },
            },
            "review_comments": [],
        },
    )

    critic_resp = client.get(f"/status/{job_id}/critic")
    assert critic_resp.status_code == 200
    critic_body = critic_resp.json()
    assert critic_body["fatal_flaws"]
    assert critic_body["llm_advisory_diagnostics"]["llm_finding_count"] == 1
    assert critic_body["llm_advisory_diagnostics"]["advisory_rejected_reason"] == "non_advisory_llm_finding_present"
    finding = critic_body["fatal_flaws"][0]
    finding_id = finding["finding_id"]
    assert isinstance(finding.get("due_at"), str) and finding["due_at"]
    assert int(finding.get("sla_hours") or 0) == 24

    ack_resp = client.post(
        f"/status/{job_id}/critic/findings/{finding_id}/ack",
        headers={"X-Reviewer": "qa_reviewer"},
    )
    assert ack_resp.status_code == 200
    ack_body = ack_resp.json()
    assert ack_body["id"] == finding_id
    assert ack_body["finding_id"] == finding_id
    assert ack_body["status"] == "acknowledged"
    assert ack_body.get("acknowledged_at")
    assert ack_body.get("updated_at")
    assert ack_body.get("updated_by") == "qa_reviewer"
    assert ack_body.get("acknowledged_by") == "qa_reviewer"
    assert isinstance(ack_body.get("due_at"), str) and ack_body["due_at"]

    comment_resp = client.post(
        f"/status/{job_id}/comments",
        json={
            "section": ack_body["section"],
            "message": "Acknowledged by reviewer; will revise next iteration.",
            "linked_finding_id": finding_id,
        },
    )
    assert comment_resp.status_code == 200
    comment_body = comment_resp.json()
    assert comment_body["linked_finding_id"] == finding_id
    assert int(comment_body.get("sla_hours") or 0) == 24
    assert isinstance(comment_body.get("due_at"), str) and comment_body["due_at"]

    critic_resp_2 = client.get(f"/status/{job_id}/critic")
    assert critic_resp_2.status_code == 200
    critic_body_2 = critic_resp_2.json()
    linked_finding = next(f for f in critic_body_2["fatal_flaws"] if f["finding_id"] == finding_id)
    assert comment_body["comment_id"] in (linked_finding.get("linked_comment_ids") or [])

    resolve_resp = client.post(
        f"/status/{job_id}/critic/findings/{finding_id}/resolve",
        headers={"X-User": "qa_resolver"},
    )
    assert resolve_resp.status_code == 200
    resolve_body = resolve_resp.json()
    assert resolve_body["id"] == finding_id
    assert resolve_body["finding_id"] == finding_id
    assert resolve_body["status"] == "resolved"
    assert resolve_body.get("resolved_at")
    assert resolve_body.get("updated_at")
    assert resolve_body.get("updated_by") == "qa_resolver"
    assert resolve_body.get("acknowledged_by") == "qa_reviewer"
    assert resolve_body.get("resolved_by") == "qa_resolver"

    reopen_resp = client.post(
        f"/status/{job_id}/critic/findings/{finding_id}/open",
        headers={"X-Actor": "qa_reopener"},
    )
    assert reopen_resp.status_code == 200
    reopen_body = reopen_resp.json()
    assert reopen_body["id"] == finding_id
    assert reopen_body["finding_id"] == finding_id
    assert reopen_body["status"] == "open"
    assert reopen_body.get("updated_at")
    assert reopen_body.get("updated_by") == "qa_reopener"
    assert reopen_body.get("acknowledged_at") is None
    assert reopen_body.get("acknowledged_by") is None
    assert reopen_body.get("resolved_at") is None
    assert reopen_body.get("resolved_by") is None
    assert isinstance(reopen_body.get("due_at"), str) and reopen_body["due_at"]
    assert reopen_body["due_at"] != ack_body["due_at"]


def test_status_critic_normalizes_legacy_string_findings_into_entities():
    job_id = "critic-findings-legacy-string-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "quality_score": 5.0,
                "critic_score": 5.0,
                "needs_revision": True,
                "critic_notes": {
                    "engine": "rules",
                    "fatal_flaws": [
                        "Missing indicator baseline and target for key outcome.",
                    ],
                    "fatal_flaw_messages": ["Missing indicator baseline and target for key outcome."],
                    "rule_checks": [],
                },
            },
        },
    )

    critic_resp = client.get(f"/status/{job_id}/critic")
    assert critic_resp.status_code == 200
    critic_body = critic_resp.json()
    assert critic_body["fatal_flaw_count"] == 1
    flaw = critic_body["fatal_flaws"][0]
    assert flaw["id"] == flaw["finding_id"]
    assert flaw["finding_id"]
    assert flaw["code"] == "LEGACY_UNSTRUCTURED_FINDING"
    assert flaw["status"] == "open"
    assert flaw["source"] == "rules"
    assert flaw["message"].startswith("Missing indicator baseline")
    assert flaw.get("fix_suggestion")

    ack_resp = client.post(f"/status/{job_id}/critic/findings/{flaw['finding_id']}/ack")
    assert ack_resp.status_code == 200
    ack_body = ack_resp.json()
    assert ack_body["status"] == "acknowledged"
    assert ack_body.get("acknowledged_at")


def test_status_critic_accepts_id_only_finding_entities():
    job_id = "critic-findings-id-only-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "quality_score": 6.0,
                "critic_score": 6.0,
                "needs_revision": True,
                "critic_notes": {
                    "engine": "rules",
                    "fatal_flaws": [
                        {
                            "id": "finding-id-only-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "Theory of Change draft does not match expected donor schema.",
                            "fix_suggestion": "Regenerate ToC with valid structured output.",
                            "source": "rules",
                        }
                    ],
                    "fatal_flaw_messages": ["Theory of Change draft does not match expected donor schema."],
                    "rule_checks": [{"code": "TOC_SCHEMA_VALID", "status": "fail", "section": "toc"}],
                },
            },
            "review_comments": [],
        },
    )

    critic_resp = client.get(f"/status/{job_id}/critic")
    assert critic_resp.status_code == 200
    flaw = critic_resp.json()["fatal_flaws"][0]
    assert flaw["id"] == "finding-id-only-1"
    assert flaw["finding_id"] == "finding-id-only-1"

    ack_resp = client.post(f"/status/{job_id}/critic/findings/{flaw['id']}/ack")
    assert ack_resp.status_code == 200
    ack_body = ack_resp.json()
    assert ack_body["id"] == "finding-id-only-1"
    assert ack_body["finding_id"] == "finding-id-only-1"
    assert ack_body["status"] == "acknowledged"


def test_status_critic_single_status_dry_run_previews_without_persisting():
    job_id = "critic-findings-single-dry-run-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "single-dry-f1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC invalid.",
                            "source": "rules",
                        }
                    ]
                }
            },
        },
    )

    dry_resp = client.post(
        f"/status/{job_id}/critic/findings/single-dry-f1/ack",
        params={"dry_run": "true"},
        headers={"X-Reviewer": "qa_reviewer"},
    )
    assert dry_resp.status_code == 200
    dry_body = dry_resp.json()
    assert dry_body["status"] == "acknowledged"
    assert dry_body["dry_run"] is True
    assert dry_body["persisted"] is False
    assert dry_body["changed"] is True
    assert dry_body["updated_by"] == "qa_reviewer"

    critic_after_dry = client.get(f"/status/{job_id}/critic")
    assert critic_after_dry.status_code == 200
    flaw_after_dry = critic_after_dry.json()["fatal_flaws"][0]
    assert flaw_after_dry["status"] == "open"

    commit_resp = client.post(
        f"/status/{job_id}/critic/findings/single-dry-f1/ack",
        headers={"X-Reviewer": "qa_reviewer"},
    )
    assert commit_resp.status_code == 200
    commit_body = commit_resp.json()
    assert commit_body["dry_run"] is False
    assert commit_body["persisted"] is True
    assert commit_body["changed"] is True
    assert commit_body["status"] == "acknowledged"

    critic_after_commit = client.get(f"/status/{job_id}/critic")
    assert critic_after_commit.status_code == 200
    flaw_after_commit = critic_after_commit.json()["fatal_flaws"][0]
    assert flaw_after_commit["status"] == "acknowledged"


def test_status_critic_single_status_if_match_returns_conflict_without_changes():
    job_id = "critic-findings-single-if-match-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "single-if-f1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "acknowledged",
                            "message": "ToC invalid.",
                            "source": "rules",
                        }
                    ]
                }
            },
        },
    )

    conflict_resp = client.post(
        f"/status/{job_id}/critic/findings/single-if-f1/resolve",
        params={"if_match_status": "open"},
        headers={"X-Actor": "qa_reviewer"},
    )
    assert conflict_resp.status_code == 409
    conflict_body = conflict_resp.json()
    detail = conflict_body.get("detail") if isinstance(conflict_body, dict) else {}
    assert detail.get("reason") == "finding_status_conflict"
    assert detail.get("if_match_status") == "open"
    assert detail.get("current_status") == "acknowledged"

    critic_after_conflict = client.get(f"/status/{job_id}/critic")
    assert critic_after_conflict.status_code == 200
    flaw_after_conflict = critic_after_conflict.json()["fatal_flaws"][0]
    assert flaw_after_conflict["status"] == "acknowledged"

    commit_resp = client.post(
        f"/status/{job_id}/critic/findings/single-if-f1/resolve",
        params={"if_match_status": "acknowledged"},
        headers={"X-Actor": "qa_reviewer"},
    )
    assert commit_resp.status_code == 200
    commit_body = commit_resp.json()
    assert commit_body["status"] == "resolved"
    assert commit_body["persisted"] is True
    assert commit_body.get("if_match_status") == "acknowledged"


def test_status_critic_single_status_request_id_is_idempotent():
    job_id = "critic-findings-single-request-id-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "single-rid-f1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC invalid.",
                            "source": "rules",
                        }
                    ]
                }
            },
        },
    )

    first = client.post(
        f"/status/{job_id}/critic/findings/single-rid-f1/ack",
        params={"request_id": "rid-single-1"},
        headers={"X-Actor": "qa_reviewer"},
    )
    assert first.status_code == 200
    first_body = first.json()
    assert first_body["status"] == "acknowledged"
    assert first_body.get("request_id") == "rid-single-1"
    assert first_body.get("idempotent_replay") is not True

    replay = client.post(
        f"/status/{job_id}/critic/findings/single-rid-f1/ack",
        params={"request_id": "rid-single-1"},
        headers={"X-Actor": "qa_reviewer"},
    )
    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["status"] == "acknowledged"
    assert replay_body["request_id"] == "rid-single-1"
    assert replay_body["idempotent_replay"] is True

    mismatch = client.post(
        f"/status/{job_id}/critic/findings/single-rid-f1/resolve",
        params={"request_id": "rid-single-1"},
        headers={"X-Actor": "qa_reviewer"},
    )
    assert mismatch.status_code == 409
    mismatch_detail = mismatch.json().get("detail") or {}
    assert mismatch_detail.get("reason") == "request_id_reused_with_different_payload"

    job = api_app_module.JOB_STORE.get(job_id) or {}
    events = job.get("job_events") if isinstance(job.get("job_events"), list) else []
    finding_events = [e for e in events if isinstance(e, dict) and e.get("type") == "critic_finding_status_changed"]
    assert len(finding_events) == 1


def test_status_critic_supports_legacy_state_critic_fatal_flaws_alias():
    job_id = "critic-findings-legacy-alias-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "quality_score": 6.0,
                "critic_score": 6.0,
                "needs_revision": True,
                "critic_fatal_flaws": [
                    {
                        "code": "TOC_SCHEMA_INVALID",
                        "severity": "high",
                        "section": "toc",
                        "message": "Legacy alias finding should still be normalized.",
                        "source": "rules",
                    }
                ],
            },
            "review_comments": [],
        },
    )

    critic_resp = client.get(f"/status/{job_id}/critic")
    assert critic_resp.status_code == 200
    body = critic_resp.json()
    assert body["fatal_flaw_count"] == 1
    flaw = body["fatal_flaws"][0]
    assert flaw["id"] == flaw["finding_id"]
    assert flaw["status"] == "open"

    ack_resp = client.post(f"/status/{job_id}/critic/findings/{flaw['finding_id']}/ack")
    assert ack_resp.status_code == 200
    assert ack_resp.json()["status"] == "acknowledged"


def test_status_critic_bulk_status_updates_support_filters_and_apply_all():
    job_id = "critic-findings-bulk-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "quality_score": 6.0,
                "critic_score": 6.0,
                "needs_revision": True,
                "critic_notes": {
                    "engine": "rules",
                    "fatal_flaws": [
                        {
                            "finding_id": "bulk-f1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC invalid.",
                            "source": "rules",
                        },
                        {
                            "finding_id": "bulk-f2",
                            "code": "LOGFRAME_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "open",
                            "message": "Missing baseline.",
                            "source": "rules",
                        },
                        {
                            "finding_id": "bulk-f3",
                            "code": "TOC_ASSUMPTIONS_WEAK",
                            "severity": "low",
                            "section": "toc",
                            "status": "acknowledged",
                            "message": "Assumptions weak.",
                            "source": "rules",
                        },
                    ],
                    "rule_checks": [],
                },
            },
        },
    )

    no_selector_resp = client.post(
        f"/status/{job_id}/critic/findings/bulk-status",
        json={"next_status": "resolved"},
    )
    assert no_selector_resp.status_code == 400
    assert "selector" in str(no_selector_resp.json().get("detail") or "")

    filtered_resp = client.post(
        f"/status/{job_id}/critic/findings/bulk-status",
        headers={"X-Actor": "bulk_reviewer"},
        json={
            "next_status": "acknowledged",
            "section": "toc",
            "severity": "high",
            "finding_status": "open",
        },
    )
    assert filtered_resp.status_code == 200
    filtered_body = filtered_resp.json()
    assert filtered_body["requested_status"] == "acknowledged"
    assert filtered_body["actor"] == "bulk_reviewer"
    assert filtered_body["matched_count"] == 1
    assert filtered_body["changed_count"] == 1
    assert filtered_body["unchanged_count"] == 0
    assert filtered_body["filters"]["section"] == "toc"
    assert filtered_body["filters"]["severity"] == "high"
    assert filtered_body["filters"]["finding_status"] == "open"
    assert filtered_body["updated_findings"][0]["finding_id"] == "bulk-f1"
    assert filtered_body["updated_findings"][0]["status"] == "acknowledged"

    id_filtered_resp = client.post(
        f"/status/{job_id}/critic/findings/bulk-status",
        json={
            "next_status": "resolved",
            "finding_ids": ["bulk-f2", "bulk-f404"],
        },
    )
    assert id_filtered_resp.status_code == 200
    id_filtered_body = id_filtered_resp.json()
    assert id_filtered_body["matched_count"] == 1
    assert id_filtered_body["changed_count"] == 1
    assert id_filtered_body["not_found_finding_ids"] == ["bulk-f404"]
    assert id_filtered_body["updated_findings"][0]["finding_id"] == "bulk-f2"
    assert id_filtered_body["updated_findings"][0]["status"] == "resolved"

    apply_all_resp = client.post(
        f"/status/{job_id}/critic/findings/bulk-status",
        json={"next_status": "resolved", "apply_to_all": True},
    )
    assert apply_all_resp.status_code == 200
    apply_all_body = apply_all_resp.json()
    assert apply_all_body["matched_count"] == 3
    assert apply_all_body["changed_count"] >= 1
    assert apply_all_body["filters"]["apply_to_all"] is True

    critic_resp = client.get(f"/status/{job_id}/critic")
    assert critic_resp.status_code == 200
    critic_flaws = critic_resp.json()["fatal_flaws"]
    assert len(critic_flaws) == 3
    assert all(str(item.get("status") or "") == "resolved" for item in critic_flaws)


def test_status_critic_bulk_status_dry_run_previews_without_persisting():
    job_id = "critic-findings-bulk-dry-run-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "quality_score": 6.0,
                "critic_score": 6.0,
                "needs_revision": True,
                "critic_notes": {
                    "engine": "rules",
                    "fatal_flaws": [
                        {
                            "finding_id": "dry-f1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC invalid.",
                            "source": "rules",
                        },
                        {
                            "finding_id": "dry-f2",
                            "code": "LOGFRAME_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "open",
                            "message": "Missing baseline.",
                            "source": "rules",
                        },
                    ],
                    "rule_checks": [],
                },
            },
        },
    )

    dry_run_resp = client.post(
        f"/status/{job_id}/critic/findings/bulk-status",
        headers={"X-Actor": "bulk_previewer"},
        json={"next_status": "resolved", "apply_to_all": True, "dry_run": True},
    )
    assert dry_run_resp.status_code == 200
    dry_run_body = dry_run_resp.json()
    assert dry_run_body["dry_run"] is True
    assert dry_run_body["persisted"] is False
    assert dry_run_body["matched_count"] == 2
    assert dry_run_body["changed_count"] == 2
    assert all(str(item.get("status") or "") == "resolved" for item in dry_run_body["updated_findings"])

    critic_after_preview_resp = client.get(f"/status/{job_id}/critic")
    assert critic_after_preview_resp.status_code == 200
    critic_after_preview = critic_after_preview_resp.json()["fatal_flaws"]
    assert len(critic_after_preview) == 2
    assert all(str(item.get("status") or "") == "open" for item in critic_after_preview)

    commit_resp = client.post(
        f"/status/{job_id}/critic/findings/bulk-status",
        headers={"X-Actor": "bulk_reviewer"},
        json={"next_status": "resolved", "apply_to_all": True},
    )
    assert commit_resp.status_code == 200
    commit_body = commit_resp.json()
    assert commit_body["dry_run"] is False
    assert commit_body["persisted"] is True
    assert commit_body["changed_count"] == 2

    critic_after_commit_resp = client.get(f"/status/{job_id}/critic")
    assert critic_after_commit_resp.status_code == 200
    critic_after_commit = critic_after_commit_resp.json()["fatal_flaws"]
    assert len(critic_after_commit) == 2
    assert all(str(item.get("status") or "") == "resolved" for item in critic_after_commit)


def test_status_critic_bulk_status_if_match_returns_conflict_without_changes():
    job_id = "critic-findings-bulk-if-match-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "bulk-if-f1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC invalid.",
                            "source": "rules",
                        },
                        {
                            "finding_id": "bulk-if-f2",
                            "code": "LOGFRAME_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "acknowledged",
                            "message": "Missing baseline.",
                            "source": "rules",
                        },
                    ]
                }
            },
        },
    )

    conflict_resp = client.post(
        f"/status/{job_id}/critic/findings/bulk-status",
        json={"next_status": "resolved", "apply_to_all": True, "if_match_status": "open"},
        headers={"X-Actor": "bulk_reviewer"},
    )
    assert conflict_resp.status_code == 409
    conflict_body = conflict_resp.json()
    detail = conflict_body.get("detail") if isinstance(conflict_body, dict) else {}
    assert detail.get("reason") == "finding_status_conflict"
    assert detail.get("if_match_status") == "open"
    assert int(detail.get("conflict_count") or 0) >= 1

    critic_after_conflict_resp = client.get(f"/status/{job_id}/critic")
    assert critic_after_conflict_resp.status_code == 200
    critic_after_conflict = critic_after_conflict_resp.json()["fatal_flaws"]
    by_id = {item["finding_id"]: item for item in critic_after_conflict}
    assert by_id["bulk-if-f1"]["status"] == "open"
    assert by_id["bulk-if-f2"]["status"] == "acknowledged"

    success_resp = client.post(
        f"/status/{job_id}/critic/findings/bulk-status",
        json={"next_status": "resolved", "finding_ids": ["bulk-if-f2"], "if_match_status": "acknowledged"},
        headers={"X-Actor": "bulk_reviewer"},
    )
    assert success_resp.status_code == 200
    success_body = success_resp.json()
    assert success_body["changed_count"] == 1
    assert success_body["filters"]["if_match_status"] == "acknowledged"
    assert success_body["updated_findings"][0]["finding_id"] == "bulk-if-f2"
    assert success_body["updated_findings"][0]["status"] == "resolved"


def test_status_critic_bulk_status_request_id_is_idempotent():
    job_id = "critic-findings-bulk-request-id-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "bulk-rid-f1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC invalid.",
                            "source": "rules",
                        },
                        {
                            "finding_id": "bulk-rid-f2",
                            "code": "LOGFRAME_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "open",
                            "message": "Missing baseline.",
                            "source": "rules",
                        },
                    ]
                }
            },
        },
    )

    first = client.post(
        f"/status/{job_id}/critic/findings/bulk-status",
        json={"next_status": "resolved", "apply_to_all": True, "request_id": "rid-bulk-1"},
        headers={"X-Actor": "bulk_reviewer"},
    )
    assert first.status_code == 200
    first_body = first.json()
    assert first_body["changed_count"] == 2
    assert first_body["request_id"] == "rid-bulk-1"
    assert first_body.get("idempotent_replay") is not True

    replay = client.post(
        f"/status/{job_id}/critic/findings/bulk-status",
        json={"next_status": "resolved", "apply_to_all": True, "request_id": "rid-bulk-1"},
        headers={"X-Actor": "bulk_reviewer"},
    )
    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["request_id"] == "rid-bulk-1"
    assert replay_body["idempotent_replay"] is True
    assert replay_body["changed_count"] == 2

    mismatch = client.post(
        f"/status/{job_id}/critic/findings/bulk-status",
        json={"next_status": "acknowledged", "apply_to_all": True, "request_id": "rid-bulk-1"},
        headers={"X-Actor": "bulk_reviewer"},
    )
    assert mismatch.status_code == 409
    mismatch_detail = mismatch.json().get("detail") or {}
    assert mismatch_detail.get("reason") == "request_id_reused_with_different_payload"


def test_status_critic_findings_list_and_detail_support_filters():
    job_id = "critic-findings-list-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "quality_score": 6.0,
                "critic_score": 6.0,
                "needs_revision": True,
                "critic_notes": {
                    "engine": "rules",
                    "fatal_flaws": [
                        {
                            "finding_id": "f-open",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "version_id": "toc_v2",
                            "message": "ToC invalid.",
                            "source": "rules",
                        },
                        {
                            "finding_id": "f-ack",
                            "code": "LOGFRAME_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "acknowledged",
                            "version_id": "logframe_v1",
                            "acknowledged_at": "2026-02-27T09:30:00+00:00",
                            "message": "Missing baseline.",
                            "source": "rules",
                        },
                        {
                            "finding_id": "f-resolved",
                            "code": "GENERAL_REVIEW_NOTE",
                            "severity": "low",
                            "section": "general",
                            "status": "resolved",
                            "version_id": None,
                            "resolved_at": "2026-02-27T10:40:00+00:00",
                            "message": "General review note.",
                            "source": "rules",
                        },
                    ],
                    "rule_checks": [],
                },
            },
        },
    )

    list_resp = client.get(f"/status/{job_id}/critic/findings")
    assert list_resp.status_code == 200
    list_body = list_resp.json()
    assert list_body["job_id"] == job_id
    assert list_body["summary"]["finding_count"] == 3
    assert len(list_body["findings"]) == 3

    open_resp = client.get(f"/status/{job_id}/critic/findings", params={"status": "open"})
    assert open_resp.status_code == 200
    open_body = open_resp.json()
    assert open_body["filters"]["status"] == "open"
    assert open_body["summary"]["finding_count"] == 1
    assert open_body["findings"][0]["finding_id"] == "f-open"

    unresolved_resp = client.get(f"/status/{job_id}/critic/findings", params={"include_resolved": "false"})
    assert unresolved_resp.status_code == 200
    unresolved_body = unresolved_resp.json()
    assert unresolved_body["filters"]["include_resolved"] is False
    assert unresolved_body["summary"]["finding_count"] == 2
    assert all(str(item.get("status") or "") != "resolved" for item in unresolved_body["findings"])

    resolved_workflow_resp = client.get(
        f"/status/{job_id}/critic/findings",
        params={"workflow_state": "resolved"},
    )
    assert resolved_workflow_resp.status_code == 200
    resolved_workflow_body = resolved_workflow_resp.json()
    assert resolved_workflow_body["summary"]["finding_count"] == 1
    assert resolved_workflow_body["findings"][0]["finding_id"] == "f-resolved"
    assert resolved_workflow_body["findings"][0]["workflow_state"] == "resolved"

    detail_resp = client.get(f"/status/{job_id}/critic/findings/f-ack")
    assert detail_resp.status_code == 200
    detail_body = detail_resp.json()
    assert detail_body["finding_id"] == "f-ack"
    assert detail_body["status"] == "acknowledged"

    invalid_status_resp = client.get(f"/status/{job_id}/critic/findings", params={"status": "invalid"})
    assert invalid_status_resp.status_code == 400
    assert "Unsupported finding status filter" in str(invalid_status_resp.json().get("detail") or "")


def test_status_export_payload_endpoint_returns_review_ready_payload():
    job_id = "export-payload-1"
    api_app_module.INGEST_AUDIT_STORE.clear()
    api_app_module.INGEST_AUDIT_STORE.append(
        {
            "event_id": "ing-exp-1",
            "ts": "2026-02-24T09:00:00+00:00",
            "donor_id": "usaid",
            "namespace": "usaid_ads201",
            "filename": "ads.pdf",
            "content_type": "application/pdf",
            "metadata": {"doc_family": "donor_policy", "source_type": "donor_guidance"},
            "result": {"chunks_ingested": 4},
        }
    )
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "strict_preflight": True,
            "generate_preflight": {
                "donor_id": "usaid",
                "risk_level": "medium",
                "go_ahead": True,
                "warning_count": 1,
                "warnings": [{"code": "LOW_DOC_COVERAGE", "severity": "medium"}],
            },
            "client_metadata": {
                "demo_generate_preset_key": "usaid_gov_ai_kazakhstan",
                "rag_readiness": {
                    "expected_doc_families": ["donor_policy", "country_context"],
                    "donor_id": "usaid",
                },
            },
            "state": {
                "donor_id": "usaid",
                "toc_draft": {"toc": {"brief": "Sample ToC"}},
                "logframe_draft": {"indicators": [{"indicator_id": "IND_001"}]},
                "citations": [{"stage": "mel", "label": "USAID ADS 201 p.12"}],
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-1",
                            "status": "acknowledged",
                            "severity": "high",
                            "section": "toc",
                            "code": "TOC_SCHEMA_INVALID",
                            "message": "Schema mismatch",
                            "fix_hint": "Fix objective structure",
                            "source": "rules",
                        }
                    ]
                },
                "strategy": object(),
                "donor_strategy": object(),
            },
            "review_comments": [
                {
                    "comment_id": "comment-1",
                    "status": "open",
                    "section": "toc",
                    "message": "Please revise objective hierarchy",
                    "linked_finding_id": "finding-1",
                    "version_id": "toc_v1",
                }
            ],
        },
    )

    response = client.get(f"/status/{job_id}/export-payload")
    assert response.status_code == 200
    body = response.json()
    assert body["job_id"] == job_id
    assert body["status"] == "done"
    payload = body["payload"]
    assert isinstance(payload, dict)
    assert isinstance(payload["state"], dict)
    assert "strategy" not in payload["state"]
    assert "donor_strategy" not in payload["state"]
    assert payload["state"]["toc_draft"]["toc"]["brief"] == "Sample ToC"
    assert isinstance(payload["critic_findings"], list)
    assert payload["critic_findings"][0]["id"] == "finding-1"
    assert payload["critic_findings"][0]["finding_id"] == "finding-1"
    assert payload["critic_findings"][0]["linked_comment_ids"] == ["comment-1"]
    assert isinstance(payload["review_comments"], list)
    assert payload["review_comments"][0]["linked_finding_id"] == "finding-1"
    assert payload["readiness"]["preset_key"] == "usaid_gov_ai_kazakhstan"
    assert payload["readiness"]["expected_doc_families"] == ["donor_policy", "country_context"]
    assert payload["readiness"]["present_doc_families"] == ["donor_policy"]
    assert payload["readiness"]["missing_doc_families"] == ["country_context"]
    assert payload["readiness"]["coverage_rate"] == 0.5
    export_contract = payload["export_contract"]
    assert export_contract["template_key"] == "usaid"
    assert export_contract["status"] == "warning"
    assert "missing_required_toc_sections" in (export_contract.get("reasons") or [])


def test_status_includes_citations_traceability(monkeypatch):
    def fake_query(namespace, query_texts, n_results=5, where=None, top_k=None):
        return {
            "documents": [["Official indicator guidance excerpt"]],
            "ids": [["usaid_ads201_p12_c0"]],
            "distances": [[0.18]],
            "metadatas": [
                [
                    {
                        "source": "/tmp/usaid_guide.pdf",
                        "uploaded_filename": "usaid_guide.pdf",
                        "page": 12,
                        "page_start": 12,
                        "page_end": 12,
                        "chunk": 3,
                        "chunk_id": "usaid_ads201_p12_c0",
                        "indicator_id": "EG.3.2-1",
                        "citation": "USAID ADS 201 p.12",
                        "name": "Households with improved access",
                    }
                ]
            ],
        }

    monkeypatch.setattr(api_app_module.vector_store, "query", fake_query)

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Water Sanitation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "done"
    state = status["state"]
    citations = state.get("citations")
    assert isinstance(citations, list)
    assert citations

    mel_citations = [c for c in citations if c.get("stage") == "mel"]
    assert mel_citations
    citation = mel_citations[0]
    assert citation["citation_type"] == "rag_result"
    assert citation["namespace"] == "usaid_ads201"
    assert citation["namespace_normalized"] == "usaid_ads201"
    assert citation["source"] == "usaid_guide.pdf"
    assert citation["page"] == 12
    assert citation["chunk"] == 3
    assert citation["doc_id"] == "usaid_ads201_p12_c0"
    assert citation["chunk_id"] == "usaid_ads201_p12_c0"
    assert citation["used_for"] == "EG.3.2-1"
    assert citation["retrieval_rank"] == 1
    assert citation["retrieval_confidence"] > 0.0
    assert citation["retrieval_distance"] == 0.18
    assert "excerpt" in citation and citation["excerpt"]

    architect_citations = [c for c in citations if c.get("stage") == "architect"]
    assert architect_citations
    assert any(c.get("statement_path") for c in architect_citations)
    assert any(
        c.get("citation_type") in {"rag_claim_support", "rag_low_confidence", "fallback_namespace"}
        for c in architect_citations
    )
    assert any(c.get("doc_id") == "usaid_ads201_p12_c0" for c in architect_citations)
    assert all(c.get("namespace_normalized") == "usaid_ads201" for c in architect_citations)
    assert any(c.get("retrieval_rank") == 1 for c in architect_citations)
    for c in citations:
        source = str(c.get("source") or "")
        assert "grantflow_ingest_" not in source

    citations_response = client.get(f"/status/{job_id}/citations")
    assert citations_response.status_code == 200
    citations_body = citations_response.json()
    assert citations_body["job_id"] == job_id
    assert citations_body["status"] == "done"
    assert citations_body["citation_count"] >= 1
    assert isinstance(citations_body["citations"], list)
    assert citations_body["citations"][0]["stage"]
    assert any("statement_path" in c for c in citations_body["citations"])
    assert any("citation_confidence" in c for c in citations_body["citations"])
    assert any("doc_id" in c for c in citations_body["citations"])
    assert any("retrieval_rank" in c for c in citations_body["citations"])
    for c in citations_body["citations"]:
        if c.get("citation_confidence") is not None:
            assert 0.0 <= float(c["citation_confidence"]) <= 1.0
        if c.get("retrieval_confidence") is not None:
            assert 0.0 <= float(c["retrieval_confidence"]) <= 1.0
        if c.get("retrieval_distance") is not None:
            assert float(c["retrieval_distance"]) >= 0.0


def test_status_includes_draft_versions_traceability():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Livelihoods", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "done"
    state = status["state"]
    assert state.get("toc_validation", {}).get("valid") is True
    assert state.get("toc_validation", {}).get("schema_name")
    assert state.get("architect_retrieval", {}).get("namespace")
    assert state.get("architect_retrieval", {}).get("namespace_normalized")
    assert "hits_count" in (state.get("architect_retrieval") or {})
    retrieval_hits = (state.get("architect_retrieval") or {}).get("hits")
    if isinstance(retrieval_hits, list) and retrieval_hits:
        assert "doc_id" in retrieval_hits[0]
        assert "retrieval_rank" in retrieval_hits[0]
        assert "retrieval_confidence" in retrieval_hits[0]
        assert "retrieval_distance" in retrieval_hits[0]
    assert state.get("toc_generation_meta", {}).get("engine")
    assert isinstance(state.get("toc_draft", {}).get("validation"), dict)
    assert isinstance(state.get("toc_draft", {}).get("architect_retrieval"), dict)
    versions = state.get("draft_versions")
    assert isinstance(versions, list)
    assert len(versions) >= 2
    sections = {v.get("section") for v in versions}
    assert "toc" in sections
    assert "logframe" in sections
    assert "job_events" not in status

    events_resp = client.get(f"/status/{job_id}/events")
    assert events_resp.status_code == 200
    events_body = events_resp.json()
    assert events_body["job_id"] == job_id
    assert events_body["event_count"] >= 3
    event_types = [e["type"] for e in events_body["events"]]
    assert "status_changed" in event_types
    statuses = [e.get("to_status") for e in events_body["events"] if e["type"] == "status_changed"]
    assert "accepted" in statuses
    assert "running" in statuses
    assert "done" in statuses


def test_versions_and_diff_endpoints():
    job_id = "test-job-versions-1"
    api_app_module._set_job(
        job_id,
        {
            "status": "done",
            "hitl_enabled": False,
            "state": {
                "draft_versions": [
                    {
                        "version_id": "toc_v1",
                        "sequence": 1,
                        "section": "toc",
                        "node": "architect",
                        "iteration": 1,
                        "content": {"toc": {"brief": "v1", "project": "Water"}},
                        "content_hash": "hash1",
                    },
                    {
                        "version_id": "toc_v2",
                        "sequence": 2,
                        "section": "toc",
                        "node": "architect",
                        "iteration": 2,
                        "content": {"toc": {"brief": "v2", "project": "Water"}},
                        "content_hash": "hash2",
                    },
                    {
                        "version_id": "logframe_v1",
                        "sequence": 3,
                        "section": "logframe",
                        "node": "mel_specialist",
                        "iteration": 2,
                        "content": {"indicators": [{"indicator_id": "IND_001"}]},
                        "content_hash": "hash3",
                    },
                ]
            },
        },
    )

    versions_resp = client.get(f"/status/{job_id}/versions")
    assert versions_resp.status_code == 200
    versions_body = versions_resp.json()
    assert versions_body["job_id"] == job_id
    assert versions_body["version_count"] == 3
    assert [v["version_id"] for v in versions_body["versions"]] == ["toc_v1", "toc_v2", "logframe_v1"]

    toc_versions_resp = client.get(f"/status/{job_id}/versions", params={"section": "toc"})
    assert toc_versions_resp.status_code == 200
    assert toc_versions_resp.json()["version_count"] == 2

    diff_resp = client.get(f"/status/{job_id}/diff", params={"section": "toc"})
    assert diff_resp.status_code == 200
    diff_body = diff_resp.json()
    assert diff_body["job_id"] == job_id
    assert diff_body["section"] == "toc"
    assert diff_body["from_version_id"] == "toc_v1"
    assert diff_body["to_version_id"] == "toc_v2"
    assert diff_body["has_diff"] is True
    assert "toc_v1" in diff_body["diff_text"]
    assert "toc_v2" in diff_body["diff_text"]
    assert any(line.startswith("-") or line.startswith("+") for line in diff_body["diff_lines"])


def test_status_comments_endpoints_create_list_and_filter():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "WASH", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]
    _wait_for_terminal_status(job_id)

    versions_resp = client.get(f"/status/{job_id}/versions", params={"section": "toc"})
    assert versions_resp.status_code == 200
    versions = versions_resp.json()["versions"]
    assert versions
    toc_version_id = versions[-1]["version_id"]

    create_resp = client.post(
        f"/status/{job_id}/comments",
        json={
            "section": "toc",
            "message": "Please tighten assumptions and clarify beneficiary targeting.",
            "author": "reviewer-1",
            "version_id": toc_version_id,
        },
    )
    assert create_resp.status_code == 200
    created = create_resp.json()
    assert created["section"] == "toc"
    assert created["status"] == "open"
    assert created["version_id"] == toc_version_id
    assert created["author"] == "reviewer-1"
    assert isinstance(created.get("due_at"), str) and created["due_at"]
    assert int(created.get("sla_hours") or 0) >= 1

    comments_resp = client.get(f"/status/{job_id}/comments")
    assert comments_resp.status_code == 200
    comments_body = comments_resp.json()
    assert comments_body["job_id"] == job_id
    assert comments_body["comment_count"] >= 1
    assert any(c["comment_id"] == created["comment_id"] for c in comments_body["comments"])

    resolve_resp = client.post(f"/status/{job_id}/comments/{created['comment_id']}/resolve")
    assert resolve_resp.status_code == 200
    resolved = resolve_resp.json()
    assert resolved["comment_id"] == created["comment_id"]
    assert resolved["status"] == "resolved"
    assert isinstance(resolved.get("due_at"), str) and resolved["due_at"]

    resolved_filter_resp = client.get(f"/status/{job_id}/comments", params={"status": "resolved"})
    assert resolved_filter_resp.status_code == 200
    resolved_filter_body = resolved_filter_resp.json()
    assert any(c["comment_id"] == created["comment_id"] for c in resolved_filter_body["comments"])

    reopen_resp = client.post(f"/status/{job_id}/comments/{created['comment_id']}/reopen")
    assert reopen_resp.status_code == 200
    reopened = reopen_resp.json()
    assert reopened["comment_id"] == created["comment_id"]
    assert reopened["status"] == "open"
    assert isinstance(reopened.get("due_at"), str) and reopened["due_at"]
    assert reopened["due_at"] != resolved["due_at"]

    filtered_resp = client.get(
        f"/status/{job_id}/comments",
        params={"section": "toc", "status": "open", "version_id": toc_version_id},
    )
    assert filtered_resp.status_code == 200
    filtered_body = filtered_resp.json()
    assert filtered_body["comment_count"] >= 1
    assert all(c["section"] == "toc" for c in filtered_body["comments"])
    assert all(c["status"] == "open" for c in filtered_body["comments"])
    assert all((c.get("version_id") or "") == toc_version_id for c in filtered_body["comments"])

    status_resp = client.get(f"/status/{job_id}")
    assert status_resp.status_code == 200
    assert "review_comments" not in status_resp.json()


def test_status_comments_create_request_id_is_idempotent():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Education", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]
    _wait_for_terminal_status(job_id)

    first = client.post(
        f"/status/{job_id}/comments",
        json={
            "section": "toc",
            "message": "Clarify outcome assumptions.",
            "author": "reviewer-idem",
            "request_id": "rid-comment-create-1",
        },
    )
    assert first.status_code == 200
    first_body = first.json()
    assert first_body["request_id"] == "rid-comment-create-1"
    assert first_body.get("idempotent_replay") is not True

    replay = client.post(
        f"/status/{job_id}/comments",
        json={
            "section": "toc",
            "message": "Clarify outcome assumptions.",
            "author": "reviewer-idem",
            "request_id": "rid-comment-create-1",
        },
    )
    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["request_id"] == "rid-comment-create-1"
    assert replay_body["idempotent_replay"] is True
    assert replay_body["comment_id"] == first_body["comment_id"]

    mismatch = client.post(
        f"/status/{job_id}/comments",
        json={
            "section": "toc",
            "message": "Different payload for same request id.",
            "author": "reviewer-idem",
            "request_id": "rid-comment-create-1",
        },
    )
    assert mismatch.status_code == 409
    mismatch_detail = mismatch.json().get("detail") or {}
    assert mismatch_detail.get("reason") == "request_id_reused_with_different_payload"

    job = api_app_module.JOB_STORE.get(job_id) or {}
    events = job.get("job_events") if isinstance(job.get("job_events"), list) else []
    comment_events = [e for e in events if isinstance(e, dict) and e.get("type") == "review_comment_added"]
    assert len(comment_events) == 1
    assert comment_events[0].get("request_id") == "rid-comment-create-1"


def test_status_comments_status_request_id_is_idempotent():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Governance", "country": "Jordan"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]
    _wait_for_terminal_status(job_id)

    create_resp = client.post(
        f"/status/{job_id}/comments",
        json={"section": "general", "message": "Needs tighter risks section."},
    )
    assert create_resp.status_code == 200
    comment_id = create_resp.json()["comment_id"]

    first = client.post(
        f"/status/{job_id}/comments/{comment_id}/resolve",
        params={"request_id": "rid-comment-status-1"},
    )
    assert first.status_code == 200
    first_body = first.json()
    assert first_body["status"] == "resolved"
    assert first_body["request_id"] == "rid-comment-status-1"
    assert first_body.get("idempotent_replay") is not True

    replay = client.post(
        f"/status/{job_id}/comments/{comment_id}/resolve",
        params={"request_id": "rid-comment-status-1"},
    )
    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["status"] == "resolved"
    assert replay_body["request_id"] == "rid-comment-status-1"
    assert replay_body["idempotent_replay"] is True

    mismatch = client.post(
        f"/status/{job_id}/comments/{comment_id}/reopen",
        params={"request_id": "rid-comment-status-1"},
    )
    assert mismatch.status_code == 409
    mismatch_detail = mismatch.json().get("detail") or {}
    assert mismatch_detail.get("reason") == "request_id_reused_with_different_payload"

    job = api_app_module.JOB_STORE.get(job_id) or {}
    events = job.get("job_events") if isinstance(job.get("job_events"), list) else []
    status_events = [e for e in events if isinstance(e, dict) and e.get("type") == "review_comment_status_changed"]
    matching = [e for e in status_events if e.get("comment_id") == comment_id]
    assert len(matching) == 1
    assert matching[0].get("request_id") == "rid-comment-status-1"


def test_status_comments_endpoint_rejects_invalid_section_or_version():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Nutrition", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]
    _wait_for_terminal_status(job_id)

    bad_section = client.post(
        f"/status/{job_id}/comments",
        json={"section": "budget", "message": "Budget narrative mismatch"},
    )
    assert bad_section.status_code == 400

    bad_version = client.post(
        f"/status/{job_id}/comments",
        json={"section": "toc", "message": "Version mismatch", "version_id": "missing-version"},
    )
    assert bad_version.status_code == 400


def test_status_review_workflow_endpoint_aggregates_findings_comments_and_timeline():
    job_id = "review-workflow-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "resolved",
                            "message": "ToC schema mismatch.",
                            "resolved_at": "2026-02-27T10:12:00+00:00",
                        },
                        {
                            "finding_id": "finding-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "low",
                            "section": "logframe",
                            "status": "open",
                            "message": "Baseline is missing for one indicator.",
                        },
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-1",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "section": "toc",
                    "status": "resolved",
                    "message": "Reviewed and addressed in v2.",
                    "author": "reviewer-1",
                    "linked_finding_id": "finding-1",
                },
                {
                    "comment_id": "comment-2",
                    "ts": "2026-02-27T10:01:00+00:00",
                    "section": "general",
                    "status": "open",
                    "message": "Need sign-off from program manager.",
                    "author": "reviewer-2",
                    "linked_finding_id": "missing-finding",
                },
            ],
            "job_events": [
                {
                    "event_id": "rwf-1",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "comment-1",
                    "section": "toc",
                    "author": "reviewer-1",
                },
                {
                    "event_id": "rwf-2",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-1",
                    "status": "acknowledged",
                    "section": "toc",
                    "severity": "high",
                    "actor": "qa-reviewer",
                },
                {
                    "event_id": "rwf-3",
                    "ts": "2026-02-27T10:10:00+00:00",
                    "type": "review_comment_status_changed",
                    "comment_id": "comment-1",
                    "status": "resolved",
                    "section": "toc",
                    "actor": "qa-reviewer",
                },
                {
                    "event_id": "rwf-4",
                    "ts": "2026-02-27T10:12:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-1",
                    "status": "resolved",
                    "section": "toc",
                    "severity": "high",
                    "actor": "qa-reviewer",
                },
                {
                    "event_id": "rwf-ignore",
                    "ts": "2026-02-27T10:15:00+00:00",
                    "type": "status_changed",
                    "to_status": "done",
                },
            ],
        },
    )

    response = client.get(f"/status/{job_id}/review/workflow")
    assert response.status_code == 200
    body = response.json()
    assert body["job_id"] == job_id
    assert body["status"] == "done"
    assert body.get("filters", {}).get("event_type") is None
    assert body.get("filters", {}).get("finding_id") is None
    assert body.get("filters", {}).get("finding_code") is None
    assert body.get("filters", {}).get("finding_section") is None
    assert body.get("filters", {}).get("comment_status") is None
    assert body.get("filters", {}).get("workflow_state") is None
    assert body.get("filters", {}).get("overdue_after_hours") == 48
    assert len(body["findings"]) == 2
    assert len(body["comments"]) == 2

    summary = body["summary"]
    assert summary["finding_count"] == 2
    assert summary["comment_count"] == 2
    assert summary["linked_comment_count"] == 2
    assert summary["orphan_linked_comment_count"] == 1
    assert summary["open_finding_count"] == 1
    assert summary["acknowledged_finding_count"] == 0
    assert summary["resolved_finding_count"] == 1
    assert summary["pending_finding_count"] == 1
    assert summary["overdue_finding_count"] == 0
    assert summary["open_comment_count"] == 1
    assert summary["resolved_comment_count"] == 1
    assert summary["pending_comment_count"] == 1
    assert summary["overdue_comment_count"] == 0
    assert summary["finding_status_counts"]["open"] == 1
    assert summary["finding_status_counts"]["resolved"] == 1
    assert summary["finding_severity_counts"]["high"] == 1
    assert summary["finding_severity_counts"]["low"] == 1
    assert summary["comment_status_counts"]["open"] == 1
    assert summary["comment_status_counts"]["resolved"] == 1
    assert summary["timeline_event_count"] == 4
    assert summary["last_activity_at"] == "2026-02-27T10:12:00+00:00"

    timeline = body["timeline"]
    assert len(timeline) == 4
    assert timeline[0]["ts"] == "2026-02-27T10:12:00+00:00"
    assert timeline[-1]["ts"] == "2026-02-27T10:00:00+00:00"
    timeline_types = {row["type"] for row in timeline}
    assert timeline_types == {
        "critic_finding_status_changed",
        "review_comment_added",
        "review_comment_status_changed",
    }

    finding_filtered = client.get(f"/status/{job_id}/review/workflow", params={"finding_id": "finding-1"})
    assert finding_filtered.status_code == 200
    finding_filtered_body = finding_filtered.json()
    assert finding_filtered_body["filters"]["finding_id"] == "finding-1"
    assert len(finding_filtered_body["findings"]) == 1
    assert all((row.get("finding_id") or row.get("id")) == "finding-1" for row in finding_filtered_body["findings"])
    assert all(str(row.get("linked_finding_id") or "") == "finding-1" for row in finding_filtered_body["comments"])

    finding_code_filtered = client.get(
        f"/status/{job_id}/review/workflow",
        params={"finding_code": "mel_baseline_missing"},
    )
    assert finding_code_filtered.status_code == 200
    finding_code_filtered_body = finding_code_filtered.json()
    assert finding_code_filtered_body["filters"]["finding_code"] == "MEL_BASELINE_MISSING"
    assert {row["finding_id"] for row in finding_code_filtered_body["findings"]} == {"finding-2"}
    assert finding_code_filtered_body["comments"] == []
    assert finding_code_filtered_body["timeline"] == []

    finding_section_filtered = client.get(
        f"/status/{job_id}/review/workflow",
        params={"finding_section": "toc"},
    )
    assert finding_section_filtered.status_code == 200
    finding_section_filtered_body = finding_section_filtered.json()
    assert finding_section_filtered_body["filters"]["finding_section"] == "toc"
    assert {row["finding_id"] for row in finding_section_filtered_body["findings"]} == {"finding-1"}
    assert {row["comment_id"] for row in finding_section_filtered_body["comments"]} == {"comment-1"}
    assert finding_section_filtered_body["timeline"]
    assert all(
        str(row.get("section") or "").lower() == "toc"
        or str(row.get("finding_id") or "") == "finding-1"
        or str(row.get("comment_id") or "") == "comment-1"
        for row in finding_section_filtered_body["timeline"]
    )

    comment_section_only = client.get(
        f"/status/{job_id}/review/workflow",
        params={"finding_section": "general"},
    )
    assert comment_section_only.status_code == 200
    comment_section_only_body = comment_section_only.json()
    assert comment_section_only_body["filters"]["finding_section"] == "general"
    assert comment_section_only_body["findings"] == []
    assert {row["comment_id"] for row in comment_section_only_body["comments"]} == {"comment-2"}
    assert comment_section_only_body["timeline"] == []

    event_type_filtered = client.get(
        f"/status/{job_id}/review/workflow",
        params={"event_type": "review_comment_status_changed"},
    )
    assert event_type_filtered.status_code == 200
    event_type_filtered_body = event_type_filtered.json()
    assert event_type_filtered_body["filters"]["event_type"] == "review_comment_status_changed"
    assert event_type_filtered_body["timeline"]
    assert all(row["type"] == "review_comment_status_changed" for row in event_type_filtered_body["timeline"])

    comment_status_filtered = client.get(
        f"/status/{job_id}/review/workflow",
        params={"comment_status": "resolved"},
    )
    assert comment_status_filtered.status_code == 200
    comment_status_filtered_body = comment_status_filtered.json()
    assert comment_status_filtered_body["filters"]["comment_status"] == "resolved"
    assert len(comment_status_filtered_body["comments"]) == 1
    assert all(str(row.get("status") or "") == "resolved" for row in comment_status_filtered_body["comments"])


def test_status_review_workflow_supports_pending_overdue_filters_and_validation():
    job_id = "review-workflow-pending-overdue-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-a",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC mismatch.",
                            "updated_at": "2026-02-27T06:00:00+00:00",
                        },
                        {
                            "finding_id": "finding-b",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "acknowledged",
                            "message": "Baseline is incomplete.",
                            "acknowledged_at": "2026-02-27T09:30:00+00:00",
                        },
                        {
                            "finding_id": "finding-c",
                            "code": "CITATION_GAP",
                            "severity": "low",
                            "section": "general",
                            "status": "resolved",
                            "message": "Citation evidence now linked.",
                            "resolved_at": "2026-02-27T10:40:00+00:00",
                        },
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-a",
                    "ts": "2026-02-27T06:30:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "finding-a",
                },
                {
                    "comment_id": "comment-b",
                    "ts": "2026-02-27T10:20:00+00:00",
                    "section": "logframe",
                    "status": "open",
                    "message": "Indicator wording updated.",
                    "linked_finding_id": "finding-b",
                },
                {
                    "comment_id": "comment-c",
                    "ts": "2026-02-27T10:30:00+00:00",
                    "section": "general",
                    "status": "resolved",
                    "message": "Resolved in reviewer pass.",
                    "linked_finding_id": "finding-c",
                },
            ],
            "job_events": [
                {
                    "event_id": "rwf-po-1",
                    "ts": "2026-02-27T06:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-a",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "rwf-po-2",
                    "ts": "2026-02-27T06:30:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "comment-a",
                    "section": "toc",
                },
                {
                    "event_id": "rwf-po-3",
                    "ts": "2026-02-27T09:30:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-b",
                    "status": "acknowledged",
                    "section": "logframe",
                    "severity": "medium",
                },
                {
                    "event_id": "rwf-po-4",
                    "ts": "2026-02-27T10:20:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "comment-b",
                    "section": "logframe",
                },
                {
                    "event_id": "rwf-po-5",
                    "ts": "2026-02-27T10:40:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-c",
                    "status": "resolved",
                    "section": "general",
                    "severity": "low",
                },
            ],
        },
    )

    invalid_filter = client.get(f"/status/{job_id}/review/workflow", params={"workflow_state": "inbox"})
    assert invalid_filter.status_code == 400
    invalid_section_filter = client.get(f"/status/{job_id}/review/workflow", params={"finding_section": "budget"})
    assert invalid_section_filter.status_code == 400

    all_items = client.get(f"/status/{job_id}/review/workflow", params={"overdue_after_hours": 2})
    assert all_items.status_code == 200
    all_body = all_items.json()
    assert all_body["summary"]["pending_finding_count"] == 1
    assert all_body["summary"]["overdue_finding_count"] == 1
    assert all_body["summary"]["pending_comment_count"] == 1
    assert all_body["summary"]["overdue_comment_count"] == 1

    pending_only = client.get(
        f"/status/{job_id}/review/workflow",
        params={"workflow_state": "pending", "overdue_after_hours": 2},
    )
    assert pending_only.status_code == 200
    pending_body = pending_only.json()
    assert pending_body["filters"]["workflow_state"] == "pending"
    assert pending_body["filters"]["overdue_after_hours"] == 2
    assert {row["finding_id"] for row in pending_body["findings"]} == {"finding-b"}
    assert {row["comment_id"] for row in pending_body["comments"]} == {"comment-b"}
    assert all(row.get("workflow_state") == "pending" for row in pending_body["findings"])
    assert all(row.get("workflow_state") == "pending" for row in pending_body["comments"])

    overdue_only = client.get(
        f"/status/{job_id}/review/workflow",
        params={"workflow_state": "overdue", "overdue_after_hours": 2},
    )
    assert overdue_only.status_code == 200
    overdue_body = overdue_only.json()
    assert overdue_body["filters"]["workflow_state"] == "overdue"
    assert overdue_body["filters"]["overdue_after_hours"] == 2
    assert {row["finding_id"] for row in overdue_body["findings"]} == {"finding-a"}
    assert {row["comment_id"] for row in overdue_body["comments"]} == {"comment-a"}
    assert all(row.get("workflow_state") == "overdue" for row in overdue_body["findings"])
    assert all(row.get("workflow_state") == "overdue" for row in overdue_body["comments"])


def test_status_review_workflow_trends_endpoint_returns_bucketed_series():
    job_id = "review-workflow-trends-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-wt-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC mismatch.",
                            "updated_at": "2026-02-26T11:00:00+00:00",
                        },
                        {
                            "finding_id": "finding-wt-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "acknowledged",
                            "message": "MEL baseline missing.",
                            "updated_at": "2026-02-27T11:00:00+00:00",
                        },
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-wt-1",
                    "ts": "2026-02-26T10:00:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Needs stronger assumptions.",
                    "linked_finding_id": "finding-wt-1",
                },
                {
                    "comment_id": "comment-wt-2",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "section": "logframe",
                    "status": "resolved",
                    "message": "Resolved in review pass.",
                    "linked_finding_id": "finding-wt-2",
                },
            ],
            "job_events": [
                {
                    "event_id": "rwf-tr-1",
                    "ts": "2026-02-26T10:00:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "comment-wt-1",
                    "section": "toc",
                },
                {
                    "event_id": "rwf-tr-2",
                    "ts": "2026-02-26T11:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-wt-1",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "rwf-tr-3",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "type": "review_comment_status_changed",
                    "comment_id": "comment-wt-2",
                    "status": "resolved",
                    "section": "logframe",
                },
                {
                    "event_id": "rwf-tr-4",
                    "ts": "2026-02-27T11:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-wt-2",
                    "status": "acknowledged",
                    "section": "logframe",
                    "severity": "medium",
                },
            ],
        },
    )

    resp = client.get(f"/status/{job_id}/review/workflow/trends", params={"overdue_after_hours": 2})
    assert resp.status_code == 200
    body = resp.json()
    assert body["job_id"] == job_id
    assert body["status"] == "done"
    assert body["filters"] == {"overdue_after_hours": 2}
    assert body["bucket_granularity"] == "day"
    assert body["bucket_count"] == 2
    assert body["time_window_start"] == "2026-02-26"
    assert body["time_window_end"] == "2026-02-27"
    assert body["timeline_event_count"] == 4
    assert body["top_event_type"] == "critic_finding_status_changed"
    assert body["top_event_type_count"] == 2
    assert body["total_series"] == [
        {"bucket": "2026-02-26", "count": 2},
        {"bucket": "2026-02-27", "count": 2},
    ]
    assert body["event_type_series"]["review_comment_added"] == [{"bucket": "2026-02-26", "count": 1}]
    assert body["event_type_series"]["critic_finding_status_changed"] == [
        {"bucket": "2026-02-26", "count": 1},
        {"bucket": "2026-02-27", "count": 1},
    ]

    code_filtered = client.get(
        f"/status/{job_id}/review/workflow/trends",
        params={"finding_code": "TOC_SCHEMA_INVALID", "overdue_after_hours": 2},
    )
    assert code_filtered.status_code == 200
    code_payload = code_filtered.json()
    assert code_payload["filters"]["finding_code"] == "TOC_SCHEMA_INVALID"
    assert code_payload["bucket_count"] == 1
    assert code_payload["timeline_event_count"] == 2
    assert code_payload["total_series"] == [{"bucket": "2026-02-26", "count": 2}]


def test_status_review_workflow_sla_endpoint_aggregates_overdue_hotspots():
    job_id = "review-workflow-sla-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-a",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC mismatch.",
                            "updated_at": "2026-02-27T07:00:00+00:00",
                            "due_at": "2026-02-27T08:00:00+00:00",
                            "sla_hours": 24,
                        },
                        {
                            "finding_id": "finding-b",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "acknowledged",
                            "message": "Baseline missing.",
                            "acknowledged_at": "2026-02-27T09:30:00+00:00",
                            "due_at": "2026-02-27T12:00:00+00:00",
                            "sla_hours": 72,
                        },
                        {
                            "finding_id": "finding-c",
                            "code": "CITATION_GAP",
                            "severity": "low",
                            "section": "general",
                            "status": "resolved",
                            "message": "Citation linked.",
                            "resolved_at": "2026-02-27T10:30:00+00:00",
                            "due_at": "2026-02-28T10:00:00+00:00",
                            "sla_hours": 120,
                        },
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-a",
                    "ts": "2026-02-27T06:30:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "finding-a",
                    "due_at": "2026-02-27T09:00:00+00:00",
                    "sla_hours": 24,
                },
                {
                    "comment_id": "comment-b",
                    "ts": "2026-02-27T10:20:00+00:00",
                    "section": "logframe",
                    "status": "open",
                    "message": "Indicator wording update pending.",
                    "linked_finding_id": "finding-b",
                    "due_at": "2026-02-27T12:30:00+00:00",
                    "sla_hours": 72,
                },
                {
                    "comment_id": "comment-c",
                    "ts": "2026-02-27T10:35:00+00:00",
                    "section": "general",
                    "status": "resolved",
                    "message": "Resolved by reviewer.",
                    "linked_finding_id": "finding-c",
                    "due_at": "2026-02-27T13:00:00+00:00",
                    "sla_hours": 72,
                },
            ],
            "job_events": [
                {
                    "event_id": "rwf-sla-1",
                    "ts": "2026-02-27T07:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-a",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "rwf-sla-2",
                    "ts": "2026-02-27T09:30:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-b",
                    "status": "acknowledged",
                    "section": "logframe",
                    "severity": "medium",
                },
                {
                    "event_id": "rwf-sla-3",
                    "ts": "2026-02-27T10:40:00+00:00",
                    "type": "review_comment_status_changed",
                    "comment_id": "comment-c",
                    "status": "resolved",
                    "section": "general",
                },
            ],
        },
    )

    resp = client.get(f"/status/{job_id}/review/workflow/sla", params={"overdue_after_hours": 2})
    assert resp.status_code == 200
    body = resp.json()
    assert body["job_id"] == job_id
    assert body["status"] == "done"
    assert body["filters"] == {"overdue_after_hours": 2}
    assert body["overdue_after_hours"] == 2
    assert body["finding_total"] == 3
    assert body["comment_total"] == 3
    assert body["unresolved_finding_count"] == 2
    assert body["unresolved_comment_count"] == 2
    assert body["unresolved_total"] == 4
    assert body["overdue_finding_count"] == 1
    assert body["overdue_comment_count"] == 1
    assert body["overdue_total"] == 2
    assert body["breach_rate"] == 0.5
    assert body["overdue_by_section"]["toc"] == 2
    assert body["overdue_by_severity"]["high"] == 2
    assert body["oldest_overdue"]["kind"] == "finding"
    assert body["oldest_overdue"]["id"] == "finding-a"
    assert len(body["top_overdue"]) == 2
    assert body["top_overdue"][0]["id"] == "finding-a"
    assert body["top_overdue"][1]["id"] == "comment-a"

    section_filtered = client.get(
        f"/status/{job_id}/review/workflow/sla",
        params={"finding_section": "logframe", "overdue_after_hours": 2},
    )
    assert section_filtered.status_code == 200
    section_payload = section_filtered.json()
    assert section_payload["filters"]["finding_section"] == "logframe"
    assert section_payload["filters"]["overdue_after_hours"] == 2
    assert section_payload["finding_total"] == 1
    assert section_payload["comment_total"] == 1
    assert section_payload["overdue_total"] == 0
    assert section_payload["top_overdue"] == []

    code_filtered = client.get(
        f"/status/{job_id}/review/workflow/sla",
        params={"finding_code": "TOC_SCHEMA_INVALID", "overdue_after_hours": 2},
    )
    assert code_filtered.status_code == 200
    code_payload = code_filtered.json()
    assert code_payload["filters"]["finding_code"] == "TOC_SCHEMA_INVALID"
    assert code_payload["filters"]["overdue_after_hours"] == 2
    assert code_payload["finding_total"] == 1
    assert code_payload["comment_total"] == 1
    assert code_payload["overdue_total"] == 2
    assert {row["id"] for row in code_payload["top_overdue"]} == {"finding-a", "comment-a"}

    workflow_filtered = client.get(
        f"/status/{job_id}/review/workflow/sla",
        params={"workflow_state": "pending", "overdue_after_hours": 2},
    )
    assert workflow_filtered.status_code == 200
    workflow_payload = workflow_filtered.json()
    assert workflow_payload["filters"]["workflow_state"] == "pending"
    assert workflow_payload["filters"]["overdue_after_hours"] == 2
    assert workflow_payload["overdue_total"] == 0
    assert workflow_payload["finding_total"] == 1
    assert workflow_payload["comment_total"] == 1

    invalid_state = client.get(
        f"/status/{job_id}/review/workflow/sla",
        params={"workflow_state": "inbox", "overdue_after_hours": 2},
    )
    assert invalid_state.status_code == 400
    invalid_section = client.get(
        f"/status/{job_id}/review/workflow/sla",
        params={"finding_section": "budget", "overdue_after_hours": 2},
    )
    assert invalid_section.status_code == 400


def test_status_review_workflow_sla_trends_endpoint_returns_bucketed_series():
    job_id = "review-workflow-sla-trends-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-t1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC mismatch.",
                            "updated_at": "2026-02-27T14:00:00+00:00",
                            "due_at": "2026-02-27T08:00:00+00:00",
                            "sla_hours": 24,
                        },
                        {
                            "finding_id": "finding-t2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "acknowledged",
                            "message": "Baseline missing.",
                            "updated_at": "2026-02-27T14:30:00+00:00",
                            "due_at": "2026-02-26T08:00:00+00:00",
                            "sla_hours": 72,
                        },
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-t1",
                    "ts": "2026-02-27T13:00:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "finding-t1",
                    "due_at": "2026-02-27T09:00:00+00:00",
                    "sla_hours": 24,
                },
                {
                    "comment_id": "comment-t2",
                    "ts": "2026-02-27T13:05:00+00:00",
                    "section": "logframe",
                    "status": "open",
                    "message": "Indicator wording update pending.",
                    "linked_finding_id": "finding-t2",
                    "due_at": "2026-02-26T09:00:00+00:00",
                    "sla_hours": 72,
                },
            ],
            "job_events": [
                {
                    "event_id": "rwf-sla-trend-1",
                    "ts": "2026-02-27T15:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-t2",
                    "status": "acknowledged",
                    "section": "logframe",
                    "severity": "medium",
                }
            ],
        },
    )

    resp = client.get(f"/status/{job_id}/review/workflow/sla/trends", params={"overdue_after_hours": 2})
    assert resp.status_code == 200
    body = resp.json()
    assert body["job_id"] == job_id
    assert body["status"] == "done"
    assert body["filters"] == {"overdue_after_hours": 2}
    assert body["bucket_granularity"] == "day"
    assert body["bucket_count"] == 2
    assert body["time_window_start"] == "2026-02-26"
    assert body["time_window_end"] == "2026-02-27"
    assert body["overdue_total"] == 4
    assert body["total_series"] == [
        {"bucket": "2026-02-26", "count": 2},
        {"bucket": "2026-02-27", "count": 2},
    ]
    assert body["severity_series"]["high"] == [{"bucket": "2026-02-27", "count": 2}]
    assert body["severity_series"]["medium"] == [{"bucket": "2026-02-26", "count": 2}]
    assert body["section_series"]["toc"] == [{"bucket": "2026-02-27", "count": 2}]
    assert body["section_series"]["logframe"] == [{"bucket": "2026-02-26", "count": 2}]

    code_filtered = client.get(
        f"/status/{job_id}/review/workflow/sla/trends",
        params={"finding_code": "TOC_SCHEMA_INVALID", "overdue_after_hours": 2},
    )
    assert code_filtered.status_code == 200
    code_payload = code_filtered.json()
    assert code_payload["filters"]["finding_code"] == "TOC_SCHEMA_INVALID"
    assert code_payload["overdue_total"] == 2
    assert code_payload["bucket_count"] == 1
    assert code_payload["total_series"] == [{"bucket": "2026-02-27", "count": 2}]


def test_status_review_workflow_sla_hotspots_endpoints_filter_and_aggregate():
    job_id = "review-workflow-sla-hotspots-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-h1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC mismatch.",
                            "updated_at": "2026-02-27T14:00:00+00:00",
                            "due_at": "2026-02-27T08:00:00+00:00",
                            "sla_hours": 24,
                        },
                        {
                            "finding_id": "finding-h2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "acknowledged",
                            "message": "Baseline missing.",
                            "updated_at": "2026-02-27T14:30:00+00:00",
                            "due_at": "2026-02-26T08:00:00+00:00",
                            "sla_hours": 72,
                        },
                        {
                            "finding_id": "finding-h3",
                            "code": "CITATION_GAP",
                            "severity": "low",
                            "section": "general",
                            "status": "resolved",
                            "message": "Citation fixed.",
                            "updated_at": "2026-02-27T10:00:00+00:00",
                            "due_at": "2026-02-26T10:00:00+00:00",
                            "sla_hours": 72,
                        },
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-h1",
                    "ts": "2026-02-27T13:00:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "finding-h1",
                    "due_at": "2026-02-27T09:00:00+00:00",
                    "sla_hours": 24,
                },
                {
                    "comment_id": "comment-h2",
                    "ts": "2026-02-27T13:05:00+00:00",
                    "section": "logframe",
                    "status": "open",
                    "message": "Indicator wording update pending.",
                    "linked_finding_id": "finding-h2",
                    "due_at": "2026-02-26T09:00:00+00:00",
                    "sla_hours": 72,
                },
                {
                    "comment_id": "comment-h3",
                    "ts": "2026-02-27T13:06:00+00:00",
                    "section": "general",
                    "status": "resolved",
                    "message": "Resolved by reviewer.",
                    "linked_finding_id": "finding-h3",
                    "due_at": "2026-02-26T11:00:00+00:00",
                    "sla_hours": 72,
                },
            ],
            "job_events": [
                {
                    "event_id": "rwf-sla-hotspots-1",
                    "ts": "2026-02-27T15:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-h2",
                    "status": "acknowledged",
                    "section": "logframe",
                    "severity": "medium",
                }
            ],
        },
    )

    hotspots_resp = client.get(f"/status/{job_id}/review/workflow/sla/hotspots", params={"overdue_after_hours": 2})
    assert hotspots_resp.status_code == 200
    hotspots = hotspots_resp.json()
    assert hotspots["job_id"] == job_id
    assert hotspots["status"] == "done"
    assert hotspots["filters"]["overdue_after_hours"] == 2
    assert hotspots["filters"]["top_limit"] == 10
    assert hotspots["total_overdue_items"] == 4
    assert hotspots["hotspot_count"] == 4
    assert hotspots["top_overdue"][0]["id"] == "finding-h2"
    assert hotspots["hotspot_kind_counts"]["finding"] == 2
    assert hotspots["hotspot_kind_counts"]["comment"] == 2
    assert hotspots["hotspot_severity_counts"]["high"] == 2
    assert hotspots["hotspot_severity_counts"]["medium"] == 2

    kind_filtered = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots",
        params={"hotspot_kind": "comment", "overdue_after_hours": 2},
    )
    assert kind_filtered.status_code == 200
    kind_payload = kind_filtered.json()
    assert kind_payload["filters"]["hotspot_kind"] == "comment"
    assert kind_payload["total_overdue_items"] == 2
    assert kind_payload["hotspot_kind_counts"]["comment"] == 2

    severity_filtered = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots",
        params={"hotspot_severity": "high", "overdue_after_hours": 2},
    )
    assert severity_filtered.status_code == 200
    severity_payload = severity_filtered.json()
    assert severity_payload["filters"]["hotspot_severity"] == "high"
    assert severity_payload["total_overdue_items"] == 2
    assert all(str(item.get("severity") or "") == "high" for item in severity_payload["top_overdue"])

    min_overdue_filtered = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots",
        params={"min_overdue_hours": 31, "overdue_after_hours": 2},
    )
    assert min_overdue_filtered.status_code == 200
    min_overdue_payload = min_overdue_filtered.json()
    assert min_overdue_payload["filters"]["min_overdue_hours"] == 31.0
    assert min_overdue_payload["total_overdue_items"] == 1
    assert min_overdue_payload["top_overdue"][0]["id"] == "finding-h2"

    invalid_kind = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots",
        params={"hotspot_kind": "timeline", "overdue_after_hours": 2},
    )
    assert invalid_kind.status_code == 400

    trends_resp = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots/trends",
        params={"overdue_after_hours": 2},
    )
    assert trends_resp.status_code == 200
    trends = trends_resp.json()
    assert trends["job_id"] == job_id
    assert trends["status"] == "done"
    assert trends["hotspot_count_total"] == 4
    assert trends["bucket_count"] == 2
    assert trends["time_window_start"] == "2026-02-26"
    assert trends["time_window_end"] == "2026-02-27"
    assert trends["total_series"] == [
        {"bucket": "2026-02-26", "count": 2},
        {"bucket": "2026-02-27", "count": 2},
    ]
    assert trends["kind_series"]["finding"] == [
        {"bucket": "2026-02-26", "count": 1},
        {"bucket": "2026-02-27", "count": 1},
    ]
    assert trends["kind_series"]["comment"] == [
        {"bucket": "2026-02-26", "count": 1},
        {"bucket": "2026-02-27", "count": 1},
    ]

    trends_filtered = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots/trends",
        params={"hotspot_kind": "finding", "overdue_after_hours": 2},
    )
    assert trends_filtered.status_code == 200
    trends_filtered_payload = trends_filtered.json()
    assert trends_filtered_payload["filters"]["hotspot_kind"] == "finding"
    assert trends_filtered_payload["hotspot_count_total"] == 2
    assert trends_filtered_payload["bucket_count"] == 2
    assert trends_filtered_payload["total_series"] == [
        {"bucket": "2026-02-26", "count": 1},
        {"bucket": "2026-02-27", "count": 1},
    ]


def test_status_review_workflow_sla_hotspots_export_supports_csv_json_and_gzip():
    job_id = "review-workflow-sla-hotspots-export-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-h-exp-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC mismatch.",
                            "updated_at": "2026-02-27T14:00:00+00:00",
                            "due_at": "2026-02-27T08:00:00+00:00",
                            "sla_hours": 24,
                        },
                        {
                            "finding_id": "finding-h-exp-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "acknowledged",
                            "message": "Baseline missing.",
                            "updated_at": "2026-02-27T14:30:00+00:00",
                            "due_at": "2026-02-26T08:00:00+00:00",
                            "sla_hours": 72,
                        },
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-h-exp-1",
                    "ts": "2026-02-27T13:00:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "finding-h-exp-1",
                    "due_at": "2026-02-27T09:00:00+00:00",
                    "sla_hours": 24,
                },
                {
                    "comment_id": "comment-h-exp-2",
                    "ts": "2026-02-27T13:05:00+00:00",
                    "section": "logframe",
                    "status": "open",
                    "message": "Indicator wording update pending.",
                    "linked_finding_id": "finding-h-exp-2",
                    "due_at": "2026-02-26T09:00:00+00:00",
                    "sla_hours": 72,
                },
            ],
            "job_events": [
                {
                    "event_id": "rwf-sla-hotspots-exp-1",
                    "ts": "2026-02-27T15:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-h-exp-2",
                    "status": "acknowledged",
                    "section": "logframe",
                    "severity": "medium",
                }
            ],
        },
    )

    csv_resp = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots/export",
        params={"hotspot_kind": "finding", "overdue_after_hours": 2, "format": "csv"},
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert f"grantflow_review_workflow_sla_hotspots_{job_id}.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert "filters.hotspot_kind,finding" in csv_text
    assert "filters.overdue_after_hours,2" in csv_text
    assert "total_overdue_items,2" in csv_text

    json_resp = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots/trends/export",
        params={"hotspot_severity": "high", "overdue_after_hours": 2, "format": "json"},
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_payload = json_resp.json()
    assert json_payload["filters"]["hotspot_severity"] == "high"
    assert json_payload["hotspot_count_total"] == 2
    assert json_payload["bucket_count"] == 1

    gzip_resp = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots/trends/export",
        params={"overdue_after_hours": 2, "format": "json", "gzip": "true"},
    )
    assert gzip_resp.status_code == 200
    assert gzip_resp.headers["content-type"].startswith("application/gzip")
    gzip_disposition = gzip_resp.headers.get("content-disposition", "")
    assert f"grantflow_review_workflow_sla_hotspots_trends_{job_id}.json.gz" in gzip_disposition
    gzip_payload = json.loads(gzip.decompress(gzip_resp.content).decode("utf-8"))
    assert gzip_payload["bucket_count"] == 2
    assert gzip_payload["hotspot_count_total"] == 4


def test_status_review_workflow_sla_trends_export_supports_csv_json_and_gzip():
    job_id = "review-workflow-sla-trends-export-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-t-exp-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC mismatch.",
                            "updated_at": "2026-02-27T14:00:00+00:00",
                            "due_at": "2026-02-27T08:00:00+00:00",
                            "sla_hours": 24,
                        },
                        {
                            "finding_id": "finding-t-exp-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "acknowledged",
                            "message": "Baseline missing.",
                            "updated_at": "2026-02-27T14:30:00+00:00",
                            "due_at": "2026-02-26T08:00:00+00:00",
                            "sla_hours": 72,
                        },
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-t-exp-1",
                    "ts": "2026-02-27T13:00:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "finding-t-exp-1",
                    "due_at": "2026-02-27T09:00:00+00:00",
                    "sla_hours": 24,
                },
                {
                    "comment_id": "comment-t-exp-2",
                    "ts": "2026-02-27T13:05:00+00:00",
                    "section": "logframe",
                    "status": "open",
                    "message": "Indicator wording update pending.",
                    "linked_finding_id": "finding-t-exp-2",
                    "due_at": "2026-02-26T09:00:00+00:00",
                    "sla_hours": 72,
                },
            ],
            "job_events": [
                {
                    "event_id": "rwf-sla-trend-exp-1",
                    "ts": "2026-02-27T15:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-t-exp-2",
                    "status": "acknowledged",
                    "section": "logframe",
                    "severity": "medium",
                }
            ],
        },
    )

    csv_resp = client.get(
        f"/status/{job_id}/review/workflow/sla/trends/export",
        params={"finding_section": "toc", "overdue_after_hours": 2, "format": "csv"},
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert f"grantflow_review_workflow_sla_trends_{job_id}.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert "filters.finding_section,toc" in csv_text
    assert "filters.overdue_after_hours,2" in csv_text
    assert "overdue_total,2" in csv_text
    assert "total_series[0].bucket,2026-02-27" in csv_text

    json_resp = client.get(
        f"/status/{job_id}/review/workflow/sla/trends/export",
        params={"finding_code": "MEL_BASELINE_MISSING", "overdue_after_hours": 2, "format": "json"},
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_payload = json_resp.json()
    assert json_payload["filters"]["finding_code"] == "MEL_BASELINE_MISSING"
    assert json_payload["bucket_count"] == 1
    assert json_payload["overdue_total"] == 2

    gzip_resp = client.get(
        f"/status/{job_id}/review/workflow/sla/trends/export",
        params={"overdue_after_hours": 2, "format": "json", "gzip": "true"},
    )
    assert gzip_resp.status_code == 200
    assert gzip_resp.headers["content-type"].startswith("application/gzip")
    gzip_disposition = gzip_resp.headers.get("content-disposition", "")
    assert f"grantflow_review_workflow_sla_trends_{job_id}.json.gz" in gzip_disposition
    gzip_payload = json.loads(gzip.decompress(gzip_resp.content).decode("utf-8"))
    assert gzip_payload["bucket_count"] == 2
    assert gzip_payload["overdue_total"] == 4


def test_status_review_workflow_sla_export_supports_csv_json_and_gzip():
    job_id = "review-workflow-sla-export-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-sla-exp-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC mismatch.",
                            "updated_at": "2026-02-27T07:00:00+00:00",
                            "due_at": "2026-02-27T08:00:00+00:00",
                            "sla_hours": 24,
                        },
                        {
                            "finding_id": "finding-sla-exp-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "acknowledged",
                            "message": "Baseline missing.",
                            "acknowledged_at": "2026-02-27T09:30:00+00:00",
                            "due_at": "2026-02-27T12:00:00+00:00",
                            "sla_hours": 72,
                        },
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-sla-exp-1",
                    "ts": "2026-02-27T06:30:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "finding-sla-exp-1",
                    "due_at": "2026-02-27T09:00:00+00:00",
                    "sla_hours": 24,
                },
                {
                    "comment_id": "comment-sla-exp-2",
                    "ts": "2026-02-27T10:20:00+00:00",
                    "section": "logframe",
                    "status": "resolved",
                    "message": "Resolved in reviewer pass.",
                    "linked_finding_id": "finding-sla-exp-2",
                    "due_at": "2026-02-27T12:30:00+00:00",
                    "sla_hours": 72,
                },
            ],
            "job_events": [
                {
                    "event_id": "rwf-sla-exp-1",
                    "ts": "2026-02-27T07:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-sla-exp-1",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "rwf-sla-exp-2",
                    "ts": "2026-02-27T09:30:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-sla-exp-2",
                    "status": "acknowledged",
                    "section": "logframe",
                    "severity": "medium",
                },
            ],
        },
    )

    csv_resp = client.get(
        f"/status/{job_id}/review/workflow/sla/export",
        params={
            "finding_section": "toc",
            "finding_code": "toc_schema_invalid",
            "overdue_after_hours": 2,
            "format": "csv",
        },
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert f"grantflow_review_workflow_sla_{job_id}.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert "filters.finding_section,toc" in csv_text
    assert "filters.finding_code,TOC_SCHEMA_INVALID" in csv_text
    assert "filters.overdue_after_hours,2" in csv_text
    assert "overdue_total,2" in csv_text

    json_resp = client.get(
        f"/status/{job_id}/review/workflow/sla/export",
        params={"finding_section": "logframe", "overdue_after_hours": 2, "format": "json"},
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_payload = json_resp.json()
    assert json_payload["filters"]["finding_section"] == "logframe"
    assert json_payload["filters"]["overdue_after_hours"] == 2
    assert json_payload["finding_total"] == 1
    assert json_payload["comment_total"] == 1
    assert json_payload["overdue_total"] == 0

    gzip_resp = client.get(
        f"/status/{job_id}/review/workflow/sla/export",
        params={"finding_code": "MEL_BASELINE_MISSING", "overdue_after_hours": 2, "format": "json", "gzip": "true"},
    )
    assert gzip_resp.status_code == 200
    assert gzip_resp.headers["content-type"].startswith("application/gzip")
    gzip_disposition = gzip_resp.headers.get("content-disposition", "")
    assert f"grantflow_review_workflow_sla_{job_id}.json.gz" in gzip_disposition
    gzip_payload = json.loads(gzip.decompress(gzip_resp.content).decode("utf-8"))
    assert gzip_payload["filters"]["finding_code"] == "MEL_BASELINE_MISSING"
    assert gzip_payload["finding_total"] == 1


def test_status_review_workflow_sla_profile_endpoint_returns_saved_or_default_profile():
    saved_job_id = "review-workflow-sla-profile-saved-job-1"
    api_app_module.JOB_STORE.set(
        saved_job_id,
        {
            "status": "done",
            "client_metadata": {
                "sla_profile": {
                    "finding_sla_hours": {"high": 11, "medium": 33, "low": 77},
                    "default_comment_sla_hours": 42,
                    "updated_at": "2026-02-27T12:00:00+00:00",
                    "updated_by": "qa-user",
                }
            },
            "state": {"critic_notes": {"fatal_flaws": []}},
            "review_comments": [],
        },
    )
    saved_resp = client.get(f"/status/{saved_job_id}/review/workflow/sla/profile")
    assert saved_resp.status_code == 200
    saved_body = saved_resp.json()
    assert saved_body["source"] == "saved"
    assert saved_body["saved_profile_available"] is True
    assert saved_body["saved_profile_valid"] is True
    assert saved_body["finding_sla_hours"] == {"high": 11, "medium": 33, "low": 77}
    assert int(saved_body["default_comment_sla_hours"]) == 42
    assert saved_body["saved_profile_updated_at"] == "2026-02-27T12:00:00+00:00"
    assert saved_body["saved_profile_updated_by"] == "qa-user"

    default_job_id = "review-workflow-sla-profile-default-job-1"
    api_app_module.JOB_STORE.set(
        default_job_id,
        {
            "status": "done",
            "state": {"critic_notes": {"fatal_flaws": []}},
            "review_comments": [],
        },
    )
    default_resp = client.get(f"/status/{default_job_id}/review/workflow/sla/profile")
    assert default_resp.status_code == 200
    default_body = default_resp.json()
    assert default_body["source"] == "default"
    assert default_body["saved_profile_available"] is False
    assert default_body["saved_profile_valid"] is True
    assert default_body["finding_sla_hours"] == {"high": 24, "medium": 72, "low": 120}
    assert int(default_body["default_comment_sla_hours"]) == 72


def test_status_review_workflow_sla_recompute_rewrites_due_dates_and_emits_event():
    job_id = "review-workflow-sla-recompute-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-r1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC mismatch.",
                            "updated_at": "2026-02-27T08:00:00+00:00",
                            "due_at": "2000-01-01T00:00:00+00:00",
                            "sla_hours": 24,
                        }
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-r1",
                    "ts": "2026-02-27T08:30:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "finding-r1",
                    "due_at": "2000-01-01T00:00:00+00:00",
                    "sla_hours": 24,
                },
                {
                    "comment_id": "comment-r2",
                    "ts": "2026-02-27T09:00:00+00:00",
                    "section": "general",
                    "status": "open",
                    "message": "General review queue.",
                    "due_at": "2000-01-01T00:00:00+00:00",
                    "sla_hours": 72,
                },
            ],
            "job_events": [],
        },
    )

    resp = client.post(
        f"/status/{job_id}/review/workflow/sla/recompute",
        json={
            "finding_sla_hours": {"high": 12, "medium": 48, "low": 96},
            "default_comment_sla_hours": 36,
        },
        headers={"X-Reviewer": "qa-recompute"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["job_id"] == job_id
    assert body["status"] == "done"
    assert body["actor"] == "qa-recompute"
    assert body["use_saved_profile"] is False
    assert body["applied_finding_sla_hours"] == {"high": 12, "medium": 48, "low": 96}
    assert body["applied_default_comment_sla_hours"] == 36
    assert body["finding_checked_count"] == 1
    assert body["comment_checked_count"] == 2
    assert body["finding_updated_count"] == 1
    assert body["comment_updated_count"] == 2
    assert body["total_updated_count"] == 3
    assert body["sla"]["job_id"] == job_id

    critic_resp = client.get(f"/status/{job_id}/critic")
    assert critic_resp.status_code == 200
    critic_finding = critic_resp.json()["fatal_flaws"][0]
    assert critic_finding["due_at"] == "2026-02-27T20:00:00+00:00"
    assert int(critic_finding.get("sla_hours") or 0) == 12

    comments_resp = client.get(f"/status/{job_id}/comments")
    assert comments_resp.status_code == 200
    comments_by_id = {row["comment_id"]: row for row in comments_resp.json()["comments"]}
    assert comments_by_id["comment-r1"]["due_at"] == "2026-02-27T20:30:00+00:00"
    assert int(comments_by_id["comment-r1"].get("sla_hours") or 0) == 12
    assert comments_by_id["comment-r2"]["due_at"] == "2026-02-28T21:00:00+00:00"
    assert int(comments_by_id["comment-r2"].get("sla_hours") or 0) == 36

    events_resp = client.get(f"/status/{job_id}/events")
    assert events_resp.status_code == 200
    event_rows = [e for e in events_resp.json()["events"] if e.get("type") == "review_workflow_sla_recomputed"]
    assert event_rows
    last = event_rows[-1]
    assert last["actor"] == "qa-recompute"
    assert int(last["total_updated_count"]) == 3
    assert last["use_saved_profile"] is False
    assert last["applied_finding_sla_hours"] == {"high": 12, "medium": 48, "low": 96}
    assert int(last["applied_default_comment_sla_hours"]) == 36

    stored_job = api_app_module.JOB_STORE.get(job_id)
    assert isinstance(stored_job, dict)
    stored_meta = stored_job.get("client_metadata")
    assert isinstance(stored_meta, dict)
    saved_profile = stored_meta.get("sla_profile")
    assert isinstance(saved_profile, dict)
    assert saved_profile["finding_sla_hours"] == {"high": 12, "medium": 48, "low": 96}
    assert int(saved_profile["default_comment_sla_hours"]) == 36


def test_status_review_workflow_sla_recompute_can_use_saved_profile():
    job_id = "review-workflow-sla-recompute-use-saved-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-saved-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC mismatch.",
                            "updated_at": "2026-02-27T08:00:00+00:00",
                            "due_at": "2000-01-01T00:00:00+00:00",
                            "sla_hours": 24,
                        }
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-saved-1",
                    "ts": "2026-02-27T09:00:00+00:00",
                    "section": "general",
                    "status": "open",
                    "message": "General queue.",
                    "due_at": "2000-01-01T00:00:00+00:00",
                    "sla_hours": 72,
                }
            ],
            "client_metadata": {
                "sla_profile": {
                    "finding_sla_hours": {"high": 10, "medium": 20, "low": 30},
                    "default_comment_sla_hours": 16,
                }
            },
            "job_events": [],
        },
    )

    resp = client.post(
        f"/status/{job_id}/review/workflow/sla/recompute",
        json={"use_saved_profile": True},
        headers={"X-Reviewer": "qa-saved"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["use_saved_profile"] is True
    assert body["applied_finding_sla_hours"] == {"high": 10, "medium": 20, "low": 30}
    assert int(body["applied_default_comment_sla_hours"]) == 16

    critic_resp = client.get(f"/status/{job_id}/critic")
    assert critic_resp.status_code == 200
    critic_finding = critic_resp.json()["fatal_flaws"][0]
    assert critic_finding["due_at"] == "2026-02-27T18:00:00+00:00"
    assert int(critic_finding.get("sla_hours") or 0) == 10

    comments_resp = client.get(f"/status/{job_id}/comments")
    assert comments_resp.status_code == 200
    comment = comments_resp.json()["comments"][0]
    assert comment["due_at"] == "2026-02-28T01:00:00+00:00"
    assert int(comment.get("sla_hours") or 0) == 16


def test_status_review_workflow_sla_recompute_rejects_invalid_profile():
    job_id = "review-workflow-sla-recompute-invalid-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {"critic_notes": {"fatal_flaws": []}},
            "review_comments": [],
        },
    )

    bad_key = client.post(
        f"/status/{job_id}/review/workflow/sla/recompute",
        json={"finding_sla_hours": {"critical": 6}},
    )
    assert bad_key.status_code == 400
    assert "Unsupported SLA severity key" in bad_key.json()["detail"]

    bad_value = client.post(
        f"/status/{job_id}/review/workflow/sla/recompute",
        json={"default_comment_sla_hours": 0},
    )
    assert bad_value.status_code == 400
    assert "default_comment_sla_hours must be within" in bad_value.json()["detail"]


def test_status_review_workflow_export_supports_csv_json_and_gzip():
    job_id = "review-workflow-export-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-export-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC schema mismatch.",
                        },
                        {
                            "finding_id": "finding-export-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "open",
                            "message": "MEL baseline is missing.",
                        },
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-export-1",
                    "ts": "2026-02-27T11:00:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need update before submission.",
                    "linked_finding_id": "finding-export-1",
                },
                {
                    "comment_id": "comment-export-2",
                    "ts": "2026-02-27T11:02:00+00:00",
                    "section": "logframe",
                    "status": "resolved",
                    "message": "Baseline evidence added.",
                    "linked_finding_id": "finding-export-2",
                },
            ],
            "job_events": [
                {
                    "event_id": "rwf-exp-1",
                    "ts": "2026-02-27T11:00:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "comment-export-1",
                    "section": "toc",
                },
                {
                    "event_id": "rwf-exp-2",
                    "ts": "2026-02-27T11:01:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-export-1",
                    "status": "acknowledged",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "rwf-exp-3",
                    "ts": "2026-02-27T11:02:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "comment-export-2",
                    "section": "logframe",
                },
                {
                    "event_id": "rwf-exp-4",
                    "ts": "2026-02-27T11:03:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-export-2",
                    "status": "open",
                    "section": "logframe",
                    "severity": "medium",
                },
            ],
        },
    )

    csv_resp = client.get(
        f"/status/{job_id}/review/workflow/export",
        params={
            "finding_id": "finding-export-1",
            "finding_code": "toc_schema_invalid",
            "finding_section": "toc",
            "comment_status": "open",
            "format": "csv",
        },
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert f"grantflow_review_workflow_{job_id}.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert "filters.finding_id,finding-export-1" in csv_text
    assert "filters.finding_code,TOC_SCHEMA_INVALID" in csv_text
    assert "filters.finding_section,toc" in csv_text
    assert "filters.comment_status,open" in csv_text
    assert "summary.finding_count,1" in csv_text

    json_resp = client.get(
        f"/status/{job_id}/review/workflow/export",
        params={"finding_section": "logframe", "format": "json"},
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_payload = json_resp.json()
    assert json_payload["filters"]["finding_section"] == "logframe"
    assert {row["finding_id"] for row in json_payload["findings"]} == {"finding-export-2"}
    assert {row["comment_id"] for row in json_payload["comments"]} == {"comment-export-2"}
    assert json_payload["timeline"]
    assert all(
        str(row.get("section") or "").lower() == "logframe"
        or str(row.get("finding_id") or "") == "finding-export-2"
        or str(row.get("comment_id") or "") == "comment-export-2"
        for row in json_payload["timeline"]
    )

    gzip_resp = client.get(
        f"/status/{job_id}/review/workflow/export",
        params={"finding_code": "MEL_BASELINE_MISSING", "format": "json", "gzip": "true"},
    )
    assert gzip_resp.status_code == 200
    assert gzip_resp.headers["content-type"].startswith("application/gzip")
    gzip_disposition = gzip_resp.headers.get("content-disposition", "")
    assert f"grantflow_review_workflow_{job_id}.json.gz" in gzip_disposition
    gzip_payload = json.loads(gzip.decompress(gzip_resp.content).decode("utf-8"))
    assert gzip_payload["filters"]["finding_code"] == "MEL_BASELINE_MISSING"
    assert {row["finding_id"] for row in gzip_payload["findings"]} == {"finding-export-2"}


def test_status_review_workflow_trends_export_supports_csv_json_and_gzip():
    job_id = "review-workflow-trends-export-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "finding-wt-exp-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "ToC mismatch.",
                            "updated_at": "2026-02-26T11:00:00+00:00",
                        },
                        {
                            "finding_id": "finding-wt-exp-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "logframe",
                            "status": "acknowledged",
                            "message": "MEL baseline missing.",
                            "updated_at": "2026-02-27T11:00:00+00:00",
                        },
                    ]
                }
            },
            "review_comments": [
                {
                    "comment_id": "comment-wt-exp-1",
                    "ts": "2026-02-26T10:00:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Needs stronger assumptions.",
                    "linked_finding_id": "finding-wt-exp-1",
                },
                {
                    "comment_id": "comment-wt-exp-2",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "section": "logframe",
                    "status": "resolved",
                    "message": "Resolved in review pass.",
                    "linked_finding_id": "finding-wt-exp-2",
                },
            ],
            "job_events": [
                {
                    "event_id": "rwf-tr-exp-1",
                    "ts": "2026-02-26T10:00:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "comment-wt-exp-1",
                    "section": "toc",
                },
                {
                    "event_id": "rwf-tr-exp-2",
                    "ts": "2026-02-26T11:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-wt-exp-1",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "rwf-tr-exp-3",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "type": "review_comment_status_changed",
                    "comment_id": "comment-wt-exp-2",
                    "status": "resolved",
                    "section": "logframe",
                },
                {
                    "event_id": "rwf-tr-exp-4",
                    "ts": "2026-02-27T11:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "finding-wt-exp-2",
                    "status": "acknowledged",
                    "section": "logframe",
                    "severity": "medium",
                },
            ],
        },
    )

    csv_resp = client.get(
        f"/status/{job_id}/review/workflow/trends/export",
        params={"finding_section": "toc", "overdue_after_hours": 2, "format": "csv"},
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert f"grantflow_review_workflow_trends_{job_id}.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert "filters.finding_section,toc" in csv_text
    assert "filters.overdue_after_hours,2" in csv_text
    assert "total_series[0].bucket,2026-02-26" in csv_text

    json_resp = client.get(
        f"/status/{job_id}/review/workflow/trends/export",
        params={"finding_code": "MEL_BASELINE_MISSING", "overdue_after_hours": 2, "format": "json"},
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_payload = json_resp.json()
    assert json_payload["filters"]["finding_code"] == "MEL_BASELINE_MISSING"
    assert json_payload["bucket_count"] == 1
    assert json_payload["timeline_event_count"] == 2

    gzip_resp = client.get(
        f"/status/{job_id}/review/workflow/trends/export",
        params={"overdue_after_hours": 2, "format": "json", "gzip": "true"},
    )
    assert gzip_resp.status_code == 200
    assert gzip_resp.headers["content-type"].startswith("application/gzip")
    gzip_disposition = gzip_resp.headers.get("content-disposition", "")
    assert f"grantflow_review_workflow_trends_{job_id}.json.gz" in gzip_disposition
    gzip_payload = json.loads(gzip.decompress(gzip_resp.content).decode("utf-8"))
    assert gzip_payload["bucket_count"] == 2
    assert gzip_payload["timeline_event_count"] == 4


def test_metrics_endpoint_derives_timeline_metrics_from_events():
    job_id = "test-job-metrics-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "hitl_enabled": True,
            "job_events": [
                {
                    "event_id": "e1",
                    "ts": "2026-02-24T10:00:00+00:00",
                    "type": "status_changed",
                    "from_status": None,
                    "to_status": "accepted",
                    "status": "accepted",
                },
                {
                    "event_id": "e2",
                    "ts": "2026-02-24T10:00:10+00:00",
                    "type": "status_changed",
                    "from_status": "accepted",
                    "to_status": "running",
                    "status": "running",
                },
                {
                    "event_id": "e3",
                    "ts": "2026-02-24T10:01:00+00:00",
                    "type": "status_changed",
                    "from_status": "running",
                    "to_status": "pending_hitl",
                    "status": "pending_hitl",
                },
                {
                    "event_id": "e4",
                    "ts": "2026-02-24T10:05:00+00:00",
                    "type": "resume_requested",
                    "status": "accepted",
                    "resuming_from": "mel",
                },
                {
                    "event_id": "e5",
                    "ts": "2026-02-24T10:05:05+00:00",
                    "type": "status_changed",
                    "from_status": "pending_hitl",
                    "to_status": "accepted",
                    "status": "accepted",
                },
                {
                    "event_id": "e6",
                    "ts": "2026-02-24T10:05:06+00:00",
                    "type": "status_changed",
                    "from_status": "accepted",
                    "to_status": "running",
                    "status": "running",
                },
                {
                    "event_id": "e7",
                    "ts": "2026-02-24T10:06:00+00:00",
                    "type": "status_changed",
                    "from_status": "running",
                    "to_status": "done",
                    "status": "done",
                },
            ],
        },
    )

    response = client.get(f"/status/{job_id}/metrics")
    assert response.status_code == 200
    body = response.json()
    assert body["job_id"] == job_id
    assert body["status"] == "done"
    assert body["event_count"] == 7
    assert body["status_change_count"] == 6
    assert body["pause_count"] == 1
    assert body["resume_count"] == 1
    assert body["created_at"] == "2026-02-24T10:00:00+00:00"
    assert body["started_at"] == "2026-02-24T10:00:10+00:00"
    assert body["first_pending_hitl_at"] == "2026-02-24T10:01:00+00:00"
    assert body["terminal_at"] == "2026-02-24T10:06:00+00:00"
    assert body["terminal_status"] == "done"
    assert body["time_to_first_draft_seconds"] == 60.0
    assert body["time_to_terminal_seconds"] == 360.0
    assert body["time_in_pending_hitl_seconds"] == 245.0
    assert body["retrieval_expected"] is True
    assert body["grounding_risk_level"] == "unknown"
    assert body["citation_count"] == 0
    assert body["fallback_namespace_citation_count"] == 0
    assert body["strategy_reference_citation_count"] == 0
    assert body["retrieval_grounded_citation_count"] == 0
    assert body["non_retrieval_citation_count"] == 0
    assert body.get("retrieval_grounded_citation_rate") is None
    assert body.get("non_retrieval_citation_rate") is None


def test_quality_summary_endpoint_aggregates_quality_signals():
    job_id = "quality-job-1"
    api_app_module.INGEST_AUDIT_STORE.clear()
    api_app_module.INGEST_AUDIT_STORE.append(
        {
            "event_id": "ing-q-1",
            "ts": "2026-02-24T09:55:00+00:00",
            "donor_id": "usaid",
            "namespace": "usaid_ads201",
            "filename": "ads-guidance.pdf",
            "content_type": "application/pdf",
            "metadata": {"doc_family": "donor_policy", "source_type": "donor_guidance"},
            "result": {"chunks_ingested": 10},
        }
    )
    api_app_module.INGEST_AUDIT_STORE.append(
        {
            "event_id": "ing-q-2",
            "ts": "2026-02-24T09:56:00+00:00",
            "donor_id": "usaid",
            "namespace": "usaid_ads201",
            "filename": "kz-context.pdf",
            "content_type": "application/pdf",
            "metadata": {"doc_family": "country_context", "source_type": "country_context"},
            "result": {"chunks_ingested": 8},
        }
    )
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "strict_preflight": True,
            "generate_preflight": {
                "donor_id": "usaid",
                "risk_level": "medium",
                "go_ahead": True,
                "warning_count": 1,
                "warnings": [{"code": "LOW_DOC_COVERAGE", "severity": "medium"}],
            },
            "client_metadata": {
                "demo_generate_preset_key": "usaid_gov_ai_kazakhstan",
                "rag_readiness": {
                    "expected_doc_families": [
                        "donor_policy",
                        "responsible_ai_guidance",
                        "country_context",
                    ],
                    "donor_id": "usaid",
                },
            },
            "state": {
                "donor_id": "usaid",
                "toc_draft": {"toc": {"brief": "Sample ToC"}},
                "quality_score": 9.1,
                "critic_score": 8.9,
                "needs_revision": False,
                "toc_validation": {"valid": True, "schema_name": "USAID_TOC"},
                "toc_generation_meta": {
                    "engine": "fallback:contract_synthesizer",
                    "llm_used": False,
                    "retrieval_used": True,
                    "citation_policy": {"threshold_mode": "donor_section"},
                },
                "architect_retrieval": {"enabled": True, "hits_count": 3, "namespace": "usaid_ads201"},
                "grounded_quality_gate": {
                    "mode": "strict",
                    "applicable": True,
                    "passed": True,
                    "blocking": False,
                    "summary": "runtime_grounded_signals_ok",
                    "reasons": [],
                },
                "critic_notes": {
                    "engine": "rules",
                    "rule_score": 8.9,
                    "llm_score": None,
                    "rule_checks": [
                        {"code": "TOC_SCHEMA_VALID", "status": "pass", "section": "toc"},
                        {"code": "TOC_CLAIM_CITATIONS", "status": "warn", "section": "toc"},
                        {"code": "LOGFRAME_CITATIONS_PRESENT", "status": "fail", "section": "logframe"},
                    ],
                    "fatal_flaws": [
                        {
                            "finding_id": "f1",
                            "status": "open",
                            "severity": "high",
                            "section": "toc",
                            "code": "X",
                            "message": "m",
                        },
                        {
                            "finding_id": "f2",
                            "status": "acknowledged",
                            "severity": "medium",
                            "section": "logframe",
                            "code": "Y",
                            "message": "m",
                            "source": "llm",
                            "label": "BASELINE_TARGET_MISSING",
                        },
                        {
                            "finding_id": "f3",
                            "status": "resolved",
                            "severity": "low",
                            "section": "toc",
                            "code": "Z",
                            "message": "m",
                        },
                    ],
                },
                "citations": [
                    {
                        "stage": "architect",
                        "citation_type": "rag_claim_support",
                        "citation_confidence": 0.81,
                        "confidence_threshold": 0.42,
                    },
                    {
                        "stage": "architect",
                        "citation_type": "rag_low_confidence",
                        "citation_confidence": 0.22,
                        "confidence_threshold": 0.42,
                    },
                    {
                        "stage": "architect",
                        "citation_type": "fallback_namespace",
                        "citation_confidence": 0.1,
                        "confidence_threshold": 0.42,
                    },
                    {"stage": "mel", "citation_type": "rag_result", "citation_confidence": 0.73},
                ],
            },
            "job_events": [
                {
                    "event_id": "q1",
                    "ts": "2026-02-24T10:00:00+00:00",
                    "type": "status_changed",
                    "to_status": "accepted",
                    "status": "accepted",
                },
                {
                    "event_id": "q2",
                    "ts": "2026-02-24T10:00:05+00:00",
                    "type": "status_changed",
                    "to_status": "running",
                    "status": "running",
                },
                {
                    "event_id": "q3",
                    "ts": "2026-02-24T10:01:00+00:00",
                    "type": "status_changed",
                    "to_status": "done",
                    "status": "done",
                },
            ],
        },
    )

    response = client.get(f"/status/{job_id}/quality")
    assert response.status_code == 200
    body = response.json()
    assert body["job_id"] == job_id
    assert body["status"] == "done"
    assert body["quality_score"] == 9.1
    assert body["critic_score"] == 8.9
    assert body["needs_revision"] is False
    assert body["strict_preflight"] is True
    assert body["terminal_status"] == "done"
    assert body["critic"]["fatal_flaw_count"] == 3
    assert body["critic"]["open_finding_count"] == 1
    assert body["critic"]["acknowledged_finding_count"] == 1
    assert body["critic"]["resolved_finding_count"] == 1
    assert body["critic"]["high_severity_fatal_flaw_count"] == 1
    assert body["critic"]["version_bindable_finding_count"] == 3
    assert body["critic"]["version_bound_finding_count"] == 0
    assert body["critic"]["version_binding_rate"] == 0.0
    assert body["critic"]["failed_rule_check_count"] == 1
    assert body["critic"]["warned_rule_check_count"] == 1
    assert body["critic"]["llm_finding_label_counts"]["BASELINE_TARGET_MISSING"] == 1
    assert body["citations"]["citation_count"] == 4
    assert body["citations"]["architect_citation_count"] == 3
    assert body["citations"]["mel_citation_count"] == 1
    assert body["citations"]["citation_type_counts"]["rag_claim_support"] == 1
    assert body["citations"]["citation_type_counts"]["rag_low_confidence"] == 1
    assert body["citations"]["citation_type_counts"]["fallback_namespace"] == 1
    assert body["citations"]["citation_type_counts"]["rag_result"] == 1
    assert body["citations"]["architect_citation_type_counts"]["rag_claim_support"] == 1
    assert body["citations"]["architect_citation_type_counts"]["rag_low_confidence"] == 1
    assert body["citations"]["architect_citation_type_counts"]["fallback_namespace"] == 1
    assert body["citations"]["mel_citation_type_counts"]["rag_result"] == 1
    assert body["citations"]["architect_claim_support_citation_count"] == 1
    assert body["citations"]["architect_claim_support_rate"] == 0.3333
    assert body["citations"]["architect_fallback_namespace_citation_count"] == 1
    assert body["citations"]["architect_fallback_namespace_citation_rate"] == 0.3333
    assert body["citations"]["mel_claim_support_citation_count"] == 1
    assert body["citations"]["mel_claim_support_rate"] == 1.0
    assert body["citations"]["mel_fallback_namespace_citation_count"] == 0
    assert body["citations"]["mel_fallback_namespace_citation_rate"] == 0.0
    assert body["citations"]["high_confidence_citation_count"] == 2
    assert body["citations"]["low_confidence_citation_count"] == 2
    assert body["citations"]["architect_rag_low_confidence_citation_count"] == 1
    assert body["citations"]["mel_rag_low_confidence_citation_count"] == 0
    assert body["citations"]["rag_low_confidence_citation_count"] == 1
    assert body["citations"]["fallback_namespace_citation_count"] == 1
    assert body["citations"]["fallback_namespace_citation_rate"] == 0.25
    assert body["citations"]["strategy_reference_citation_count"] == 0
    assert body["citations"]["strategy_reference_citation_rate"] == 0.0
    assert body["citations"]["retrieval_grounded_citation_count"] == 3
    assert body["citations"]["retrieval_grounded_citation_rate"] == 0.75
    assert body["citations"]["doc_id_present_citation_count"] == 0
    assert body["citations"]["doc_id_present_citation_rate"] == 0.0
    assert body["citations"]["retrieval_rank_present_citation_count"] == 0
    assert body["citations"]["retrieval_rank_present_citation_rate"] == 0.0
    assert body["citations"]["retrieval_confidence_present_citation_count"] == 4
    assert body["citations"]["retrieval_confidence_present_citation_rate"] == 1.0
    assert body["citations"]["retrieval_metadata_complete_citation_count"] == 0
    assert body["citations"]["retrieval_metadata_complete_citation_rate"] == 0.0
    assert body["citations"]["architect_doc_id_present_citation_count"] == 0
    assert body["citations"]["architect_doc_id_present_citation_rate"] == 0.0
    assert body["citations"]["architect_retrieval_rank_present_citation_count"] == 0
    assert body["citations"]["architect_retrieval_rank_present_citation_rate"] == 0.0
    assert body["citations"]["architect_retrieval_confidence_present_citation_count"] == 3
    assert body["citations"]["architect_retrieval_confidence_present_citation_rate"] == 1.0
    assert body["citations"]["architect_retrieval_metadata_complete_citation_count"] == 0
    assert body["citations"]["architect_retrieval_metadata_complete_citation_rate"] == 0.0
    assert body["citations"]["mel_doc_id_present_citation_count"] == 0
    assert body["citations"]["mel_doc_id_present_citation_rate"] == 0.0
    assert body["citations"]["mel_retrieval_rank_present_citation_count"] == 0
    assert body["citations"]["mel_retrieval_rank_present_citation_rate"] == 0.0
    assert body["citations"]["mel_retrieval_confidence_present_citation_count"] == 1
    assert body["citations"]["mel_retrieval_confidence_present_citation_rate"] == 1.0
    assert body["citations"]["mel_retrieval_metadata_complete_citation_count"] == 0
    assert body["citations"]["mel_retrieval_metadata_complete_citation_rate"] == 0.0
    assert body["citations"]["non_retrieval_citation_count"] == 1
    assert body["citations"]["non_retrieval_citation_rate"] == 0.25
    assert body["citations"]["retrieval_expected"] is True
    assert body["citations"]["grounding_risk_level"] == "low"
    assert body["citations"]["traceability_complete_citation_count"] == 0
    assert body["citations"]["traceability_partial_citation_count"] == 0
    assert body["citations"]["traceability_missing_citation_count"] == 4
    assert body["citations"]["traceability_gap_citation_count"] == 4
    assert body["citations"]["traceability_gap_citation_rate"] == 1.0
    assert body["citations"]["architect_threshold_hit_rate"] == 0.3333
    assert body["architect_claims"]["claim_citation_count"] == 0
    assert body["architect_claims"]["claims_total"] == 0
    assert body["architect_claims"]["key_claims_total"] == 0
    assert body["architect_claims"]["fallback_claim_count"] == 0
    assert body["architect_claims"]["traceability_gap_citation_count"] == 0
    assert body["architect_claims"].get("threshold_hit_rate") is None
    assert body["architect_claims"].get("claim_coverage_ratio") is None
    assert body["architect_claims"].get("key_claim_coverage_ratio") is None
    assert body["architect_claims"].get("fallback_claim_ratio") is None
    assert body["architect"]["retrieval_enabled"] is True
    assert body["architect"]["retrieval_hits_count"] == 3
    assert body["architect"]["toc_schema_valid"] is True
    assert body["architect"]["citation_policy"]["threshold_mode"] == "donor_section"
    mel_summary = body.get("mel") or {}
    assert mel_summary.get("engine") in {"deterministic:retrieval_template", "llm:stub-mel-model", None}
    assert mel_summary.get("retrieval_namespace") in {None, "usaid_ads201"}
    assert mel_summary.get("retrieval_hits_count") in {None, 3}
    assert "mel_grounding_policy" in body
    export_contract = body.get("export_contract") or {}
    assert export_contract["template_key"] == "usaid"
    assert export_contract["status"] == "warning"
    assert "missing_required_toc_sections" in (export_contract.get("reasons") or [])
    grounded_gate = body.get("grounded_gate") or {}
    assert grounded_gate["mode"] == "strict"
    assert grounded_gate["blocking"] is False
    assert grounded_gate["passed"] is True
    assert body["preflight"]["risk_level"] == "medium"
    assert body["preflight"]["warning_count"] == 1
    assert body["preflight"]["warnings"][0]["code"] == "LOW_DOC_COVERAGE"
    assert body["readiness"]["preset_key"] == "usaid_gov_ai_kazakhstan"
    assert body["readiness"]["donor_id"] == "usaid"
    assert body["readiness"]["expected_doc_families"] == [
        "donor_policy",
        "responsible_ai_guidance",
        "country_context",
    ]
    assert body["readiness"]["present_doc_families"] == ["donor_policy", "country_context"]
    assert body["readiness"]["missing_doc_families"] == ["responsible_ai_guidance"]
    assert body["readiness"]["expected_count"] == 3
    assert body["readiness"]["loaded_count"] == 2
    assert body["readiness"]["coverage_rate"] == 0.6667
    assert body["readiness"]["inventory_total_uploads"] == 2
    assert body["readiness"]["inventory_family_count"] == 2


def test_status_grounding_gate_endpoint_returns_runtime_and_preflight_policies():
    job_id = "grounding-gate-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "error",
            "generate_preflight": {
                "risk_level": "high",
                "grounding_policy": {
                    "mode": "strict",
                    "blocking": True,
                    "go_ahead": False,
                    "summary": "preflight_blocked",
                    "reasons": ["namespace_empty"],
                },
            },
            "state": {
                "donor_id": "usaid",
                "grounded_quality_gate": {
                    "mode": "strict",
                    "applicable": True,
                    "passed": False,
                    "blocking": True,
                    "go_ahead": False,
                    "summary": "non_retrieval_citation_rate_above_max",
                    "reasons": ["non_retrieval_citation_rate_above_max"],
                    "reason_details": [
                        {
                            "code": "non_retrieval_citation_rate_above_max",
                            "message": "Non-retrieval citations exceed allowed maximum rate.",
                            "section": "overall",
                            "observed": 0.5,
                            "threshold": 0.35,
                        },
                        {
                            "code": "section_non_retrieval_citation_rate_above_max",
                            "message": "toc non-retrieval citation rate exceeds allowed maximum.",
                            "section": "toc",
                            "observed": 1.0,
                            "threshold": 0.35,
                        },
                    ],
                    "section_signals": {
                        "toc": {
                            "citation_count": 4,
                            "non_retrieval_citation_count": 4,
                            "retrieval_grounded_citation_count": 0,
                            "non_retrieval_citation_rate": 1.0,
                        },
                        "logframe": {
                            "citation_count": 4,
                            "non_retrieval_citation_count": 0,
                            "retrieval_grounded_citation_count": 4,
                            "non_retrieval_citation_rate": 0.0,
                        },
                    },
                    "failed_sections": ["toc"],
                    "evidence": {
                        "sample_citations_by_section": {
                            "toc": [
                                {
                                    "stage": "architect",
                                    "citation_type": "fallback_namespace",
                                    "statement_path": "development_objectives.0",
                                    "doc_id": "strategy::usaid_ads201::development_objectives.0",
                                }
                            ]
                        },
                        "failed_sections": ["toc"],
                    },
                    "citation_count": 8,
                },
                "mel_grounding_policy": {
                    "mode": "warn",
                    "blocking": False,
                    "passed": True,
                    "summary": "mel_grounding_signals_ok",
                },
            },
        },
    )

    response = client.get(f"/status/{job_id}/grounding-gate")
    assert response.status_code == 200
    body = response.json()
    assert body["job_id"] == job_id
    assert body["status"] == "error"
    runtime_gate = body.get("grounded_gate") or {}
    assert runtime_gate["mode"] == "strict"
    assert runtime_gate["blocking"] is True
    assert runtime_gate["passed"] is False
    assert runtime_gate["citation_count"] == 8
    assert runtime_gate["reasons"] == ["non_retrieval_citation_rate_above_max"]
    assert runtime_gate["reason_details"][0]["code"] == "non_retrieval_citation_rate_above_max"
    assert runtime_gate["section_signals"]["toc"]["non_retrieval_citation_rate"] == 1.0
    assert runtime_gate["failed_sections"] == ["toc"]
    assert runtime_gate["evidence"]["sample_citations_by_section"]["toc"][0]["citation_type"] == "fallback_namespace"
    preflight_gate = body.get("preflight_grounding_policy") or {}
    assert preflight_gate["mode"] == "strict"
    assert preflight_gate["blocking"] is True
    assert preflight_gate["summary"] == "preflight_blocked"
    mel_policy = body.get("mel_grounding_policy") or {}
    assert mel_policy["mode"] == "warn"
    assert mel_policy["passed"] is True


def test_quality_summary_readiness_emits_namespace_empty_and_low_coverage_warnings():
    job_id = "quality-job-readiness-warn"
    api_app_module.INGEST_AUDIT_STORE.clear()
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "client_metadata": {
                "demo_generate_preset_key": "usaid_gov_ai_kazakhstan",
                "rag_readiness": {
                    "expected_doc_families": [
                        "donor_policy",
                        "responsible_ai_guidance",
                        "country_context",
                    ],
                    "donor_id": "usaid",
                },
            },
            "state": {
                "donor_id": "usaid",
                "quality_score": 8.0,
                "critic_score": 8.0,
                "needs_revision": False,
                "toc_validation": {"valid": True, "schema_name": "USAID_TOC"},
                "toc_generation_meta": {"engine": "fallback:contract_synthesizer"},
                "architect_retrieval": {"enabled": True, "hits_count": 0, "namespace": "usaid_ads201"},
                "critic_notes": {"fatal_flaws": [], "rule_checks": []},
                "citations": [],
            },
            "job_events": [
                {
                    "event_id": "rw1",
                    "ts": "2026-02-24T10:00:00+00:00",
                    "type": "status_changed",
                    "to_status": "accepted",
                    "status": "accepted",
                },
                {
                    "event_id": "rw2",
                    "ts": "2026-02-24T10:00:05+00:00",
                    "type": "status_changed",
                    "to_status": "running",
                    "status": "running",
                },
                {
                    "event_id": "rw3",
                    "ts": "2026-02-24T10:01:00+00:00",
                    "type": "status_changed",
                    "to_status": "done",
                    "status": "done",
                },
            ],
        },
    )

    response = client.get(f"/status/{job_id}/quality")
    assert response.status_code == 200
    body = response.json()
    readiness = body.get("readiness") or {}
    assert readiness["namespace_empty"] is True
    assert readiness["low_doc_coverage"] is True
    assert readiness["architect_retrieval_enabled"] is True
    assert readiness["architect_retrieval_hits_count"] == 0
    assert readiness["retrieval_namespace"] == "usaid_ads201"
    assert readiness["warning_count"] >= 2
    assert readiness["warning_level"] in {"high", "medium"}
    warning_codes = {str(row.get("code") or "") for row in (readiness.get("warnings") or []) if isinstance(row, dict)}
    assert "NAMESPACE_EMPTY" in warning_codes
    assert "LOW_DOC_COVERAGE" in warning_codes
    assert "ARCHITECT_RETRIEVAL_NO_HITS" in warning_codes


def test_quality_summary_exposes_toc_text_quality_aggregate():
    job_id = "quality-job-toc-text"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {
                "donor_id": "usaid",
                "quality_score": 8.2,
                "critic_score": 8.2,
                "needs_revision": False,
                "toc_draft": {"toc": {"brief": "Sample ToC"}},
                "critic_notes": {
                    "engine": "rules",
                    "rule_score": 8.2,
                    "llm_score": None,
                    "rule_checks": [
                        {"code": "TOC_TEXT_COMPLETENESS", "status": "warn", "section": "toc"},
                        {"code": "TOC_NARRATIVE_DIVERSITY", "status": "warn", "section": "toc"},
                    ],
                    "fatal_flaws": [
                        {
                            "finding_id": "tt1",
                            "status": "open",
                            "severity": "low",
                            "section": "toc",
                            "code": "TOC_PLACEHOLDER_CONTENT",
                            "message": "ToC contains placeholder text.",
                        },
                        {
                            "finding_id": "tt2",
                            "status": "open",
                            "severity": "low",
                            "section": "toc",
                            "code": "TOC_BOILERPLATE_REPETITION",
                            "message": "ToC contains repeated boilerplate narrative.",
                        },
                    ],
                },
                "citations": [],
            },
            "job_events": [
                {
                    "event_id": "tq1",
                    "ts": "2026-02-24T10:00:00+00:00",
                    "type": "status_changed",
                    "to_status": "accepted",
                    "status": "accepted",
                },
                {
                    "event_id": "tq2",
                    "ts": "2026-02-24T10:00:05+00:00",
                    "type": "status_changed",
                    "to_status": "running",
                    "status": "running",
                },
                {
                    "event_id": "tq3",
                    "ts": "2026-02-24T10:01:00+00:00",
                    "type": "status_changed",
                    "to_status": "done",
                    "status": "done",
                },
            ],
        },
    )

    response = client.get(f"/status/{job_id}/quality")
    assert response.status_code == 200
    body = response.json()
    toc_text_quality = body.get("toc_text_quality") or {}
    assert toc_text_quality["risk_level"] == "medium"
    assert toc_text_quality["issues_total"] == 2
    assert toc_text_quality["placeholder_finding_count"] == 1
    assert toc_text_quality["repetition_finding_count"] == 1
    assert toc_text_quality["placeholder_check_status"] == "warn"
    assert toc_text_quality["repetition_check_status"] == "warn"


def test_portfolio_metrics_endpoint_aggregates_jobs_and_filters():
    api_app_module.JOB_STORE.set(
        "portfolio-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "medium", "risk_level": "medium"},
            "state": {
                "donor_id": "usaid",
                "critic_notes": {
                    "rule_checks": [{"code": "TOC_TEXT_COMPLETENESS", "status": "fail"}],
                },
                "citations": [
                    {"citation_type": "fallback_namespace"},
                    {"citation_type": "fallback_namespace"},
                    {"citation_type": "fallback_namespace"},
                ],
            },
            "job_events": [
                {
                    "event_id": "a1",
                    "ts": "2026-02-24T10:00:00+00:00",
                    "type": "status_changed",
                    "to_status": "accepted",
                    "status": "accepted",
                },
                {
                    "event_id": "a2",
                    "ts": "2026-02-24T10:00:05+00:00",
                    "type": "status_changed",
                    "to_status": "running",
                    "status": "running",
                },
                {
                    "event_id": "a3",
                    "ts": "2026-02-24T10:00:20+00:00",
                    "type": "status_changed",
                    "to_status": "pending_hitl",
                    "status": "pending_hitl",
                },
                {"event_id": "a4", "ts": "2026-02-24T10:01:00+00:00", "type": "resume_requested", "status": "accepted"},
                {
                    "event_id": "a5",
                    "ts": "2026-02-24T10:01:02+00:00",
                    "type": "status_changed",
                    "to_status": "accepted",
                    "status": "accepted",
                },
                {
                    "event_id": "a6",
                    "ts": "2026-02-24T10:01:03+00:00",
                    "type": "status_changed",
                    "to_status": "running",
                    "status": "running",
                },
                {
                    "event_id": "a7",
                    "ts": "2026-02-24T10:02:00+00:00",
                    "type": "status_changed",
                    "to_status": "done",
                    "status": "done",
                },
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-job-2",
        {
            "status": "error",
            "hitl_enabled": False,
            "generate_preflight": {"warning_level": "high", "risk_level": "high"},
            "state": {
                "donor_id": "eu",
                "critic_notes": {
                    "rule_checks": [{"code": "TOC_TEXT_COMPLETENESS", "status": "pass"}],
                },
                "citations": [
                    {"citation_type": "rag_claim_support"},
                    {"citation_type": "rag_claim_support"},
                    {"citation_type": "rag_support"},
                    {"citation_type": "rag_support"},
                ],
            },
            "job_events": [
                {
                    "event_id": "b1",
                    "ts": "2026-02-24T11:00:00+00:00",
                    "type": "status_changed",
                    "to_status": "accepted",
                    "status": "accepted",
                },
                {
                    "event_id": "b2",
                    "ts": "2026-02-24T11:00:03+00:00",
                    "type": "status_changed",
                    "to_status": "running",
                    "status": "running",
                },
                {
                    "event_id": "b3",
                    "ts": "2026-02-24T11:00:30+00:00",
                    "type": "status_changed",
                    "to_status": "error",
                    "status": "error",
                },
            ],
        },
    )

    response = client.get("/portfolio/metrics")
    assert response.status_code == 200
    body = response.json()
    assert body["job_count"] >= 2
    assert body["status_counts"]["done"] >= 1
    assert body["status_counts"]["error"] >= 1
    assert body["donor_counts"]["usaid"] >= 1
    assert body["donor_counts"]["eu"] >= 1
    assert body["warning_level_counts"]["medium"] >= 1
    assert body["warning_level_counts"]["high"] >= 1
    assert body["warning_level_job_counts"]["high"] >= 1
    assert body["warning_level_job_counts"]["medium"] >= 1
    assert "low" in body["warning_level_job_counts"]
    assert "none" in body["warning_level_job_counts"]
    assert sum(int(v or 0) for v in body["warning_level_job_counts"].values()) == body["job_count"]
    assert body["warning_level_job_rates"]["high"] is not None
    assert body["warning_level_job_rates"]["medium"] is not None
    assert body["warning_level_job_rates"]["low"] is not None
    assert body["warning_level_job_rates"]["none"] is not None
    assert body["grounding_risk_counts"]["high"] >= 1
    assert body["grounding_risk_counts"]["low"] >= 1
    assert body["grounding_risk_job_counts"]["high"] >= 1
    assert body["grounding_risk_job_counts"]["low"] >= 1
    assert body["grounding_risk_job_counts"]["unknown"] >= 0
    assert sum(int(v or 0) for v in body["grounding_risk_job_counts"].values()) == body["job_count"]
    assert body["grounding_risk_job_rates"]["high"] is not None
    assert body["grounding_risk_job_rates"]["medium"] is not None
    assert body["grounding_risk_job_rates"]["low"] is not None
    assert body["grounding_risk_job_rates"]["unknown"] is not None
    assert body["terminal_job_count"] >= 2
    assert body["hitl_job_count"] >= 1
    assert body["total_pause_count"] >= 1
    assert body["total_resume_count"] >= 1
    assert body["avg_time_to_terminal_seconds"] is not None

    filtered = client.get("/portfolio/metrics", params={"donor_id": "usaid", "status": "done", "hitl_enabled": "true"})
    assert filtered.status_code == 200
    filtered_body = filtered.json()
    assert filtered_body["filters"]["donor_id"] == "usaid"
    assert filtered_body["filters"]["status"] == "done"
    assert filtered_body["filters"]["hitl_enabled"] is True
    assert filtered_body["filters"].get("toc_text_risk_level") is None
    assert filtered_body["job_count"] >= 1
    assert "error" not in filtered_body["status_counts"]
    assert filtered_body["warning_level_counts"]["medium"] >= 1
    assert filtered_body["warning_level_job_counts"]["medium"] >= 1

    warning_filtered = client.get("/portfolio/metrics", params={"warning_level": "high"})
    assert warning_filtered.status_code == 200
    warning_filtered_body = warning_filtered.json()
    assert warning_filtered_body["filters"]["warning_level"] == "high"
    assert warning_filtered_body["job_count"] >= 1
    assert warning_filtered_body["warning_level_counts"] == {"high": warning_filtered_body["job_count"]}
    assert warning_filtered_body["warning_level_job_counts"]["high"] == warning_filtered_body["job_count"]
    assert warning_filtered_body["warning_level_job_counts"]["medium"] == 0

    grounding_filtered = client.get("/portfolio/metrics", params={"grounding_risk_level": "high"})
    assert grounding_filtered.status_code == 200
    grounding_filtered_body = grounding_filtered.json()
    assert grounding_filtered_body["filters"]["grounding_risk_level"] == "high"
    assert grounding_filtered_body["job_count"] >= 1
    assert grounding_filtered_body["grounding_risk_counts"] == {"high": grounding_filtered_body["job_count"]}
    assert grounding_filtered_body["grounding_risk_job_counts"]["high"] == grounding_filtered_body["job_count"]
    assert grounding_filtered_body["grounding_risk_job_counts"]["low"] == 0

    toc_text_filtered = client.get("/portfolio/metrics", params={"toc_text_risk_level": "high"})
    assert toc_text_filtered.status_code == 200
    toc_text_filtered_body = toc_text_filtered.json()
    assert toc_text_filtered_body["filters"]["toc_text_risk_level"] == "high"
    assert toc_text_filtered_body["job_count"] >= 1
    assert toc_text_filtered_body["status_counts"].get("done", 0) >= 1
    assert toc_text_filtered_body["status_counts"].get("error", 0) == 0


def test_portfolio_quality_endpoint_aggregates_quality_signals():
    api_app_module.JOB_STORE.set(
        "portfolio-quality-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "medium", "risk_level": "medium"},
            "state": {
                "donor_id": "usaid",
                "quality_score": 8.5,
                "critic_score": 8.0,
                "needs_revision": True,
                "toc_validation": {"valid": True, "schema_name": "UsaidTocSchema"},
                "toc_generation_meta": {
                    "engine": "fallback:contract_synthesizer",
                    "citation_policy": {"threshold_mode": "donor_section"},
                },
                "architect_retrieval": {"enabled": True, "hits_count": 2, "namespace": "usaid_ads201"},
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "f1",
                            "severity": "high",
                            "status": "open",
                            "section": "toc",
                            "code": "CAUSAL_LINK_DETAIL",
                            "message": "Causal links are underspecified for core pathway.",
                            "source": "llm",
                            "label": "CAUSAL_LINK_DETAIL",
                        },
                        {
                            "finding_id": "f2",
                            "severity": "low",
                            "status": "resolved",
                            "section": "general",
                            "code": "GENERAL_NOTE",
                            "message": "Minor editorial improvement.",
                        },
                    ],
                    "rule_checks": [
                        {"code": "toc.complete", "status": "fail"},
                        {"code": "toc.assumptions", "status": "warn"},
                    ],
                    "llm_advisory_diagnostics": {
                        "advisory_applies": True,
                        "advisory_candidate_count": 1,
                        "candidate_label_counts": {"CAUSAL_LINK_DETAIL": 1},
                    },
                },
                "citations": [
                    {
                        "stage": "architect",
                        "citation_type": "rag_claim_support",
                        "citation_confidence": 0.9,
                        "confidence_threshold": 0.7,
                    },
                    {
                        "stage": "architect",
                        "citation_type": "rag_low_confidence",
                        "citation_confidence": 0.2,
                        "confidence_threshold": 0.5,
                    },
                    {"stage": "mel", "citation_type": "rag_support", "citation_confidence": 0.6},
                ],
            },
            "job_events": [
                {
                    "event_id": "qa1",
                    "ts": "2026-02-24T10:00:00+00:00",
                    "type": "status_changed",
                    "to_status": "accepted",
                },
                {
                    "event_id": "qa2",
                    "ts": "2026-02-24T10:00:05+00:00",
                    "type": "status_changed",
                    "to_status": "running",
                },
                {
                    "event_id": "qa3",
                    "ts": "2026-02-24T10:00:20+00:00",
                    "type": "status_changed",
                    "to_status": "pending_hitl",
                },
                {"event_id": "qa4", "ts": "2026-02-24T10:01:20+00:00", "type": "resume_requested"},
                {"event_id": "qa5", "ts": "2026-02-24T10:01:30+00:00", "type": "status_changed", "to_status": "done"},
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-quality-job-2",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "low", "risk_level": "low"},
            "state": {
                "donor_id": "usaid",
                "quality_score": 6.5,
                "critic_score": 6.0,
                "needs_revision": False,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "f3",
                            "severity": "medium",
                            "status": "acknowledged",
                            "section": "logframe",
                            "code": "BASELINE_TARGET_MISSING",
                            "message": "Indicators are missing complete baseline and target fields.",
                            "source": "llm",
                            "label": "BASELINE_TARGET_MISSING",
                        }
                    ],
                    "rule_checks": [{"code": "logframe.complete", "status": "pass"}],
                    "llm_advisory_diagnostics": {
                        "advisory_applies": False,
                        "advisory_candidate_count": 1,
                        "advisory_rejected_reason": "grounding_threshold_not_met",
                        "candidate_label_counts": {"BASELINE_TARGET_MISSING": 1},
                    },
                },
                "citations": [
                    {
                        "stage": "architect",
                        "citation_type": "rag_claim_support",
                        "citation_confidence": 0.8,
                        "confidence_threshold": 0.7,
                    }
                ],
            },
            "job_events": [
                {
                    "event_id": "qb1",
                    "ts": "2026-02-24T11:00:00+00:00",
                    "type": "status_changed",
                    "to_status": "accepted",
                },
                {
                    "event_id": "qb2",
                    "ts": "2026-02-24T11:00:03+00:00",
                    "type": "status_changed",
                    "to_status": "running",
                },
                {"event_id": "qb3", "ts": "2026-02-24T11:00:25+00:00", "type": "status_changed", "to_status": "done"},
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-quality-job-3",
        {
            "status": "error",
            "hitl_enabled": False,
            "generate_preflight": {"warning_level": "high", "risk_level": "high"},
            "state": {"donor_id": "eu", "quality_score": 4.0, "critic_score": 3.0, "needs_revision": True},
            "job_events": [
                {
                    "event_id": "qc1",
                    "ts": "2026-02-24T12:00:00+00:00",
                    "type": "status_changed",
                    "to_status": "accepted",
                },
                {
                    "event_id": "qc2",
                    "ts": "2026-02-24T12:00:02+00:00",
                    "type": "status_changed",
                    "to_status": "running",
                },
                {"event_id": "qc3", "ts": "2026-02-24T12:00:10+00:00", "type": "status_changed", "to_status": "error"},
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-quality-job-4",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "medium", "risk_level": "medium"},
            "state": {
                "donor_id": "usaid",
                "quality_score": 6.0,
                "critic_score": 5.5,
                "needs_revision": True,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "f4",
                            "severity": "medium",
                            "status": "open",
                            "section": "toc",
                            "code": "GROUNDING_WEAK",
                            "message": "Grounding evidence is mostly fallback.",
                        }
                    ],
                    "rule_checks": [
                        {"code": "toc.grounding", "status": "warn"},
                        {"code": "TOC_TEXT_COMPLETENESS", "status": "fail"},
                    ],
                },
                "citations": [
                    {"stage": "architect", "citation_type": "fallback_namespace", "citation_confidence": 0.2},
                    {"stage": "mel", "citation_type": "fallback_namespace", "citation_confidence": 0.2},
                ],
            },
            "job_events": [
                {"event_id": "qd1", "ts": "2026-02-24T13:00:00+00:00", "type": "status_changed", "to_status": "done"}
            ],
        },
    )

    response = client.get("/portfolio/quality", params={"donor_id": "usaid", "status": "done", "hitl_enabled": "true"})
    assert response.status_code == 200
    body = response.json()
    assert body["filters"]["donor_id"] == "usaid"
    assert body["filters"]["status"] == "done"
    assert body["filters"]["hitl_enabled"] is True
    assert body["filters"].get("finding_status") is None
    assert body["filters"].get("finding_severity") is None
    assert body["filters"].get("toc_text_risk_level") is None
    assert body["warning_level_counts"]["medium"] >= 1
    assert body["warning_level_counts"]["low"] >= 1
    assert body["warning_level_high_job_count"] >= 0
    assert body["warning_level_medium_job_count"] >= 1
    assert body["warning_level_low_job_count"] >= 1
    assert body["warning_level_none_job_count"] >= 0
    assert body["warning_level_high_rate"] is not None
    assert body["warning_level_medium_rate"] is not None
    assert body["warning_level_low_rate"] is not None
    assert body["warning_level_none_rate"] is not None
    assert body["warning_level_job_counts"]["high"] >= 0
    assert body["warning_level_job_counts"]["medium"] >= 1
    assert body["warning_level_job_counts"]["low"] >= 1
    assert body["warning_level_job_counts"]["none"] >= 0
    assert sum(int(v or 0) for v in body["warning_level_job_counts"].values()) == body["job_count"]
    assert body["warning_level_job_rates"]["medium"] is not None
    assert body["warning_level_job_rates"]["low"] is not None
    assert body["grounding_risk_counts"]["high"] >= 1
    assert body["grounding_risk_counts"]["low"] >= 1
    assert body["grounding_risk_job_counts"]["high"] >= 1
    assert body["grounding_risk_job_counts"]["low"] >= 1
    assert sum(int(v or 0) for v in body["grounding_risk_job_counts"].values()) == body["job_count"]
    assert body["grounding_risk_job_rates"]["high"] is not None
    assert body["donor_grounding_risk_counts"]["high"] >= 0
    assert body["donor_grounding_risk_counts"]["medium"] >= 0
    assert body["donor_grounding_risk_counts"]["low"] >= 0
    assert body["donor_grounding_risk_counts"]["unknown"] >= 0
    assert body["high_grounding_risk_donor_count"] >= 0
    assert body["medium_grounding_risk_donor_count"] >= 0
    assert body["job_count"] >= 2
    assert body["avg_quality_score"] is not None
    assert body["avg_critic_score"] is not None
    assert body["severity_weighted_risk_score"] >= 1
    assert body["high_priority_signal_count"] >= 1
    assert body["critic"]["open_findings_total"] >= 1
    assert body["critic"]["high_severity_findings_total"] >= 1
    assert body["critic"]["needs_revision_job_count"] >= 1
    assert body["critic"]["needs_revision_rate"] is not None
    assert body["critic"]["llm_finding_label_counts"]["CAUSAL_LINK_DETAIL"] >= 1
    assert body["critic"]["llm_finding_label_counts"]["BASELINE_TARGET_MISSING"] >= 1
    assert body["critic"]["llm_advisory_diagnostics_job_count"] >= 2
    assert body["critic"]["llm_advisory_applied_job_count"] >= 1
    assert body["critic"]["llm_advisory_applied_rate"] is not None
    assert body["critic"]["llm_advisory_candidate_finding_count"] >= 2
    assert body["critic"]["llm_advisory_rejected_reason_counts"]["grounding_threshold_not_met"] >= 1
    assert body["citations"]["citation_count_total"] >= 4
    assert body["citations"]["architect_citation_count_total"] >= 3
    assert body["citations"]["architect_claim_support_citation_count"] >= 1
    assert body["citations"]["architect_claim_support_rate"] is not None
    assert body["citations"]["citation_confidence_avg"] is not None
    assert body["citations"]["citation_type_counts_total"]["rag_claim_support"] >= 1
    assert body["citations"]["citation_type_counts_total"]["fallback_namespace"] >= 1

    assert body["citations"]["architect_citation_type_counts_total"]["rag_claim_support"] >= 1
    assert body["citations"]["mel_citation_type_counts_total"]["fallback_namespace"] >= 1
    assert body["citations"]["low_confidence_citation_count"] >= 1
    assert body["citations"]["architect_rag_low_confidence_citation_count"] >= 1
    assert "mel_rag_low_confidence_citation_count" in body["citations"]
    assert body["citations"]["rag_low_confidence_citation_count"] >= 1
    assert "architect_rag_low_confidence_citation_rate" in body["citations"]
    assert "mel_rag_low_confidence_citation_rate" in body["citations"]
    assert "fallback_namespace_citation_count" in body["citations"]
    assert "fallback_namespace_citation_rate" in body["citations"]
    assert "doc_id_present_citation_count" in body["citations"]
    assert "doc_id_present_citation_rate" in body["citations"]
    assert "retrieval_rank_present_citation_count" in body["citations"]
    assert "retrieval_rank_present_citation_rate" in body["citations"]
    assert "retrieval_confidence_present_citation_count" in body["citations"]
    assert "retrieval_confidence_present_citation_rate" in body["citations"]
    assert "retrieval_metadata_complete_citation_count" in body["citations"]
    assert "retrieval_metadata_complete_citation_rate" in body["citations"]
    assert "architect_doc_id_present_citation_count" in body["citations"]
    assert "architect_doc_id_present_citation_rate" in body["citations"]
    assert "architect_retrieval_rank_present_citation_count" in body["citations"]
    assert "architect_retrieval_rank_present_citation_rate" in body["citations"]
    assert "architect_retrieval_confidence_present_citation_count" in body["citations"]
    assert "architect_retrieval_confidence_present_citation_rate" in body["citations"]
    assert "architect_retrieval_metadata_complete_citation_count" in body["citations"]
    assert "architect_retrieval_metadata_complete_citation_rate" in body["citations"]
    assert "mel_doc_id_present_citation_count" in body["citations"]
    assert "mel_doc_id_present_citation_rate" in body["citations"]
    assert "mel_retrieval_rank_present_citation_count" in body["citations"]
    assert "mel_retrieval_rank_present_citation_rate" in body["citations"]
    assert "mel_retrieval_confidence_present_citation_count" in body["citations"]
    assert "mel_retrieval_confidence_present_citation_rate" in body["citations"]
    assert "mel_retrieval_metadata_complete_citation_count" in body["citations"]
    assert "mel_retrieval_metadata_complete_citation_rate" in body["citations"]
    assert body["citations"]["grounding_risk_level"] in {"high", "medium", "low", "unknown"}
    assert "traceability_gap_citation_count" in body["citations"]
    assert "traceability_gap_citation_rate" in body["citations"]
    assert body["citations"]["architect_threshold_hit_rate_avg"] is not None
    assert body["citations"]["architect_claim_support_rate_avg"] is not None
    assert body["priority_signal_breakdown"]["high_severity_findings_total"]["weight"] >= 1
    assert body["priority_signal_breakdown"]["open_findings_total"]["weighted_score"] >= 1
    assert body["donor_weighted_risk_breakdown"]["usaid"]["weighted_score"] >= 1
    assert body["donor_weighted_risk_breakdown"]["usaid"]["high_priority_signal_count"] >= 1
    assert "architect_rag_low_confidence_citation_count" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "mel_rag_low_confidence_citation_count" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "traceability_gap_citation_count" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "llm_finding_label_counts" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "citation_count_total" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "architect_citation_count_total" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "architect_claim_support_citation_count" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert body["donor_weighted_risk_breakdown"]["usaid"]["architect_claim_support_rate"] is not None
    assert "citation_type_counts" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "architect_citation_type_counts" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "mel_citation_type_counts" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "fallback_namespace_citation_rate" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "doc_id_present_citation_count" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "doc_id_present_citation_rate" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "retrieval_rank_present_citation_count" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "retrieval_rank_present_citation_rate" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "retrieval_confidence_present_citation_count" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "retrieval_confidence_present_citation_rate" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "retrieval_metadata_complete_citation_count" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "retrieval_metadata_complete_citation_rate" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert body["donor_weighted_risk_breakdown"]["usaid"]["grounding_risk_level"] in {
        "high",
        "medium",
        "low",
        "unknown",
    }
    assert body["donor_weighted_risk_breakdown"]["usaid"]["llm_finding_label_counts"]["CAUSAL_LINK_DETAIL"] >= 1
    assert "llm_advisory_rejected_reason_counts" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "llm_advisory_applied_label_counts" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert "llm_advisory_rejected_label_counts" in body["donor_weighted_risk_breakdown"]["usaid"]
    assert body["donor_weighted_risk_breakdown"]["usaid"]["llm_advisory_diagnostics_job_count"] >= 2
    assert body["donor_weighted_risk_breakdown"]["usaid"]["llm_advisory_applied_job_count"] >= 1
    assert body["donor_weighted_risk_breakdown"]["usaid"]["llm_advisory_applied_rate"] is not None
    assert body["donor_weighted_risk_breakdown"]["usaid"]["llm_advisory_candidate_finding_count"] >= 2
    assert (
        body["donor_weighted_risk_breakdown"]["usaid"]["llm_advisory_applied_label_counts"]["CAUSAL_LINK_DETAIL"] >= 1
    )
    assert (
        body["donor_weighted_risk_breakdown"]["usaid"]["llm_advisory_rejected_label_counts"]["BASELINE_TARGET_MISSING"]
        >= 1
    )
    assert (
        body["donor_weighted_risk_breakdown"]["usaid"]["llm_advisory_rejected_reason_counts"][
            "grounding_threshold_not_met"
        ]
        >= 1
    )
    assert "usaid" in body["donor_grounding_risk_breakdown"]
    assert body["donor_grounding_risk_breakdown"]["usaid"]["citation_count_total"] >= 1
    assert body["donor_grounding_risk_breakdown"]["usaid"]["architect_citation_count_total"] >= 1
    assert body["donor_grounding_risk_breakdown"]["usaid"]["architect_claim_support_citation_count"] >= 1
    assert body["donor_grounding_risk_breakdown"]["usaid"]["architect_claim_support_rate"] is not None
    assert body["donor_grounding_risk_breakdown"]["usaid"]["fallback_namespace_citation_count"] >= 0
    assert "doc_id_present_citation_count" in body["donor_grounding_risk_breakdown"]["usaid"]
    assert "doc_id_present_citation_rate" in body["donor_grounding_risk_breakdown"]["usaid"]
    assert "retrieval_rank_present_citation_count" in body["donor_grounding_risk_breakdown"]["usaid"]
    assert "retrieval_rank_present_citation_rate" in body["donor_grounding_risk_breakdown"]["usaid"]
    assert "retrieval_confidence_present_citation_count" in body["donor_grounding_risk_breakdown"]["usaid"]
    assert "retrieval_confidence_present_citation_rate" in body["donor_grounding_risk_breakdown"]["usaid"]
    assert "retrieval_metadata_complete_citation_count" in body["donor_grounding_risk_breakdown"]["usaid"]
    assert "retrieval_metadata_complete_citation_rate" in body["donor_grounding_risk_breakdown"]["usaid"]
    if body["donor_grounding_risk_breakdown"]["usaid"]["citation_count_total"] > 0:
        assert body["donor_grounding_risk_breakdown"]["usaid"]["fallback_namespace_citation_rate"] is not None
    assert body["donor_grounding_risk_breakdown"]["usaid"]["grounding_risk_level"] in {
        "high",
        "medium",
        "low",
        "unknown",
    }
    assert body["donor_needs_revision_counts"]["usaid"] >= 1
    assert body["donor_open_findings_counts"]["usaid"] >= 1
    assert body["finding_status_counts"]["open"] >= 1
    assert body["finding_status_counts"]["acknowledged"] >= 0
    assert body["finding_status_counts"]["resolved"] >= 0
    assert body["finding_severity_counts"]["high"] >= 1
    assert body["finding_severity_counts"]["medium"] >= 1
    assert body["finding_severity_counts"]["low"] >= 0
    assert "toc_text_quality" in body
    assert body["toc_text_quality"]["issues_total"] >= 0
    assert body["toc_text_quality"]["placeholder_finding_count"] >= 0
    assert body["toc_text_quality"]["repetition_finding_count"] >= 0
    assert sum(int(v or 0) for v in body["toc_text_quality"]["risk_counts"].values()) == body["job_count"]
    assert body["toc_text_quality"]["high_risk_job_count"] == int(body["toc_text_quality"]["risk_counts"]["high"])
    assert body["toc_text_quality"]["medium_risk_job_count"] == int(body["toc_text_quality"]["risk_counts"]["medium"])
    assert body["toc_text_quality"]["low_risk_job_count"] == int(body["toc_text_quality"]["risk_counts"]["low"])
    assert body["toc_text_quality"]["unknown_risk_job_count"] == int(body["toc_text_quality"]["risk_counts"]["unknown"])
    assert "high_risk_job_rate" in body["toc_text_quality"]
    assert "placeholder_check_status_counts" in body["toc_text_quality"]
    assert "repetition_check_status_counts" in body["toc_text_quality"]
    assert body["toc_text_quality"]["placeholder_check_status_counts"]["unknown"] >= 0
    assert body["toc_text_quality"]["repetition_check_status_counts"]["unknown"] >= 0
    assert body["toc_text_quality"]["risk_counts"]["high"] >= 1
    assert "eu" not in body["donor_counts"]

    warning_filtered = client.get("/portfolio/quality", params={"warning_level": "high"})
    assert warning_filtered.status_code == 200
    warning_filtered_body = warning_filtered.json()
    assert warning_filtered_body["filters"]["warning_level"] == "high"
    assert warning_filtered_body["job_count"] >= 1
    assert warning_filtered_body["warning_level_counts"] == {"high": warning_filtered_body["job_count"]}
    assert warning_filtered_body["warning_level_job_counts"]["high"] == warning_filtered_body["job_count"]
    assert warning_filtered_body["warning_level_job_counts"]["none"] == 0

    grounding_filtered = client.get("/portfolio/quality", params={"grounding_risk_level": "high"})
    assert grounding_filtered.status_code == 200
    grounding_filtered_body = grounding_filtered.json()
    assert grounding_filtered_body["filters"]["grounding_risk_level"] == "high"
    assert grounding_filtered_body["job_count"] >= 1
    assert grounding_filtered_body["grounding_risk_job_counts"]["high"] == grounding_filtered_body["job_count"]
    assert grounding_filtered_body["grounding_risk_job_counts"]["medium"] == 0
    assert grounding_filtered_body["grounding_risk_job_counts"]["low"] == 0

    finding_status_filtered = client.get("/portfolio/quality", params={"finding_status": "acknowledged"})
    assert finding_status_filtered.status_code == 200
    finding_status_filtered_body = finding_status_filtered.json()
    assert finding_status_filtered_body["filters"]["finding_status"] == "acknowledged"
    assert finding_status_filtered_body["job_count"] >= 1
    assert finding_status_filtered_body["finding_status_counts"]["acknowledged"] >= 1

    finding_severity_filtered = client.get("/portfolio/quality", params={"finding_severity": "high"})
    assert finding_severity_filtered.status_code == 200
    finding_severity_filtered_body = finding_severity_filtered.json()
    assert finding_severity_filtered_body["filters"]["finding_severity"] == "high"
    assert finding_severity_filtered_body["job_count"] >= 1
    assert finding_severity_filtered_body["finding_severity_counts"]["high"] >= 1

    toc_text_risk_filtered = client.get("/portfolio/quality", params={"toc_text_risk_level": "high"})
    assert toc_text_risk_filtered.status_code == 200
    toc_text_risk_filtered_body = toc_text_risk_filtered.json()
    assert toc_text_risk_filtered_body["filters"]["toc_text_risk_level"] == "high"
    assert toc_text_risk_filtered_body["job_count"] >= 1
    assert (
        toc_text_risk_filtered_body["toc_text_quality"]["risk_counts"]["high"]
        == toc_text_risk_filtered_body["job_count"]
    )
    assert toc_text_risk_filtered_body["toc_text_quality"]["risk_counts"]["medium"] == 0
    assert toc_text_risk_filtered_body["toc_text_quality"]["risk_counts"]["low"] == 0
    assert toc_text_risk_filtered_body["toc_text_quality"]["risk_counts"]["unknown"] == 0


def test_portfolio_quality_endpoint_includes_grounded_gate_block_metrics():
    donor = "grounded_gate_metrics_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-grounded-gate-job-1",
        {
            "status": "done",
            "hitl_enabled": False,
            "state": {
                "donor_id": donor,
                "quality_score": 7.0,
                "critic_score": 7.0,
                "grounded_quality_gate": {
                    "mode": "strict",
                    "passed": False,
                    "blocking": True,
                    "failed_sections": ["toc"],
                    "reason_details": [
                        {"code": "non_retrieval_rate_above_max", "section": "toc"},
                    ],
                },
            },
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-grounded-gate-job-2",
        {
            "status": "done",
            "hitl_enabled": False,
            "state": {
                "donor_id": donor,
                "quality_score": 7.5,
                "critic_score": 7.5,
                "grounded_quality_gate": {
                    "mode": "strict",
                    "passed": False,
                    "blocking": True,
                    "reason_details": [
                        {"code": "retrieval_grounded_count_below_min", "section": "logframe"},
                    ],
                },
            },
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-grounded-gate-job-3",
        {
            "status": "done",
            "hitl_enabled": False,
            "state": {
                "donor_id": donor,
                "quality_score": 8.0,
                "critic_score": 8.0,
                "grounded_quality_gate": {
                    "mode": "strict",
                    "passed": True,
                    "blocking": False,
                    "summary": "runtime_grounded_signals_ok",
                },
            },
        },
    )

    response = client.get("/portfolio/quality", params={"donor_id": donor, "status": "done"})
    assert response.status_code == 200
    body = response.json()

    assert body["filters"]["donor_id"] == donor
    assert body["grounded_gate_present_job_count"] == 3
    assert body["grounded_gate_blocked_job_count"] == 2
    assert body["grounded_gate_passed_job_count"] == 1
    assert body["grounded_gate_block_rate"] == 0.6667
    assert body["grounded_gate_block_rate_among_present"] == 0.6667
    assert body["grounded_gate_pass_rate_among_present"] == 0.3333
    assert body["grounded_gate_section_fail_counts"]["toc"] >= 1
    assert body["grounded_gate_section_fail_counts"]["logframe"] >= 1
    assert body["grounded_gate_reason_counts"]["non_retrieval_rate_above_max"] >= 1
    assert body["grounded_gate_reason_counts"]["retrieval_grounded_count_below_min"] >= 1

    donor_breakdown = body["donor_grounded_gate_breakdown"][donor]
    assert donor_breakdown["job_count"] == 3
    assert donor_breakdown["present_job_count"] == 3
    assert donor_breakdown["blocked_job_count"] == 2
    assert donor_breakdown["passed_job_count"] == 1
    assert donor_breakdown["block_rate"] == 0.6667
    assert donor_breakdown["block_rate_among_present"] == 0.6667
    assert donor_breakdown["pass_rate_among_present"] == 0.3333
    assert donor_breakdown["section_fail_counts"]["toc"] >= 1
    assert donor_breakdown["section_fail_counts"]["logframe"] >= 1
    assert donor_breakdown["reason_counts"]["non_retrieval_rate_above_max"] >= 1
    assert donor_breakdown["reason_counts"]["retrieval_grounded_count_below_min"] >= 1


def test_portfolio_filters_accept_legacy_state_donor_alias():
    api_app_module.JOB_STORE.set(
        "portfolio-legacy-donor-job",
        {
            "status": "done",
            "hitl_enabled": False,
            "state": {
                "donor": "LEGACY_TEST_DONOR",
                "quality_score": 7.0,
                "critic_score": 7.0,
                "needs_revision": False,
            },
            "job_events": [
                {
                    "event_id": "legacy1",
                    "ts": "2026-02-25T10:00:00+00:00",
                    "type": "status_changed",
                    "to_status": "done",
                }
            ],
        },
    )

    metrics = client.get("/portfolio/metrics", params={"donor_id": "legacy_test_donor"})
    assert metrics.status_code == 200
    metrics_body = metrics.json()
    assert metrics_body["filters"]["donor_id"] == "legacy_test_donor"
    assert metrics_body["job_count"] >= 1
    assert int(metrics_body["donor_counts"].get("legacy_test_donor") or 0) >= 1

    quality = client.get("/portfolio/quality", params={"donor_id": "legacy_test_donor"})
    assert quality.status_code == 200
    quality_body = quality.json()
    assert quality_body["filters"]["donor_id"] == "legacy_test_donor"
    assert quality_body["job_count"] >= 1
    assert int(quality_body["donor_counts"].get("legacy_test_donor") or 0) >= 1


def test_portfolio_quality_export_endpoint_returns_csv():
    response = client.get(
        "/portfolio/quality/export",
        params={
            "donor_id": "usaid",
            "status": "done",
            "grounding_risk_level": "high",
            "finding_status": "open",
            "finding_severity": "high",
            "format": "csv",
        },
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    disposition = response.headers.get("content-disposition", "")
    assert "attachment;" in disposition
    assert "grantflow_portfolio_quality_usaid_done" in disposition

    body = response.text
    assert body.startswith("field,value\n")
    assert "filters.donor_id,usaid" in body
    assert "filters.status,done" in body
    assert "filters.grounding_risk_level,high" in body
    assert "filters.finding_status,open" in body
    assert "filters.finding_severity,high" in body
    assert "severity_weighted_risk_score," in body
    assert "priority_signal_breakdown.high_severity_findings_total.weight," in body
    assert "citations.architect_claim_support_rate," in body


def test_portfolio_quality_export_csv_flattens_donor_advisory_label_mix():
    api_app_module.JOB_STORE.set(
        "portfolio-quality-export-job-seeded",
        {
            "status": "done",
            "hitl_enabled": True,
            "state": {
                "donor_id": "usaid",
                "quality_score": 8.0,
                "critic_score": 7.5,
                "needs_revision": False,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "seed-f1",
                            "severity": "medium",
                            "status": "open",
                            "section": "toc",
                            "source": "llm",
                            "label": "CAUSAL_LINK_DETAIL",
                        }
                    ],
                    "llm_advisory_diagnostics": {
                        "advisory_applies": False,
                        "advisory_candidate_count": 2,
                        "advisory_rejected_reason": "grounding_threshold_not_met",
                        "candidate_label_counts": {
                            "CAUSAL_LINK_DETAIL": 1,
                            "BASELINE_TARGET_MISSING": 1,
                        },
                    },
                },
                "citations": [
                    {"stage": "architect", "citation_type": "rag_low_confidence", "citation_confidence": 0.2},
                    {"stage": "architect", "citation_type": "rag_claim_support", "citation_confidence": 0.9},
                ],
            },
            "job_events": [
                {
                    "event_id": "pqx1",
                    "ts": "2026-02-24T12:00:00+00:00",
                    "type": "status_changed",
                    "to_status": "accepted",
                },
                {
                    "event_id": "pqx2",
                    "ts": "2026-02-24T12:00:01+00:00",
                    "type": "status_changed",
                    "to_status": "running",
                },
                {"event_id": "pqx3", "ts": "2026-02-24T12:00:10+00:00", "type": "status_changed", "to_status": "done"},
            ],
        },
    )

    response = client.get("/portfolio/quality/export", params={"donor_id": "usaid", "status": "done", "format": "csv"})
    assert response.status_code == 200
    body = response.text
    assert "donor_weighted_risk_breakdown.usaid.llm_advisory_rejected_label_counts.CAUSAL_LINK_DETAIL," in body
    assert "donor_weighted_risk_breakdown.usaid.llm_advisory_rejected_label_counts.BASELINE_TARGET_MISSING," in body


def test_portfolio_quality_export_endpoint_supports_json_and_gzip():
    json_resp = client.get("/portfolio/quality/export", params={"donor_id": "usaid", "format": "json"})
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_disposition = json_resp.headers.get("content-disposition", "")
    assert "grantflow_portfolio_quality_usaid.json" in json_disposition
    payload = json_resp.json()
    assert "severity_weighted_risk_score" in payload
    assert "priority_signal_breakdown" in payload

    toc_csv_resp = client.get(
        "/portfolio/quality/export",
        params={"donor_id": "usaid", "toc_text_risk_level": "high", "format": "csv"},
    )
    assert toc_csv_resp.status_code == 200
    assert toc_csv_resp.headers["content-type"].startswith("text/csv")
    toc_csv_text = toc_csv_resp.text
    assert "filters.toc_text_risk_level,high" in toc_csv_text

    toc_json_resp = client.get(
        "/portfolio/quality/export",
        params={"donor_id": "usaid", "toc_text_risk_level": "high", "format": "json"},
    )
    assert toc_json_resp.status_code == 200
    assert toc_json_resp.headers["content-type"].startswith("application/json")
    toc_payload = toc_json_resp.json()
    assert toc_payload["filters"]["donor_id"] == "usaid"
    assert toc_payload["filters"]["toc_text_risk_level"] == "high"

    csv_gzip_resp = client.get(
        "/portfolio/quality/export", params={"donor_id": "usaid", "format": "csv", "gzip": "true"}
    )
    assert csv_gzip_resp.status_code == 200
    assert csv_gzip_resp.headers["content-type"].startswith("application/gzip")
    csv_gzip_disposition = csv_gzip_resp.headers.get("content-disposition", "")
    assert "grantflow_portfolio_quality_usaid.csv.gz" in csv_gzip_disposition
    csv_text = gzip.decompress(csv_gzip_resp.content).decode("utf-8")
    assert csv_text.startswith("field,value\n")
    assert "severity_weighted_risk_score," in csv_text

    json_gzip_resp = client.get(
        "/portfolio/quality/export", params={"donor_id": "usaid", "format": "json", "gzip": "true"}
    )
    assert json_gzip_resp.status_code == 200
    assert json_gzip_resp.headers["content-type"].startswith("application/gzip")
    json_gzip_disposition = json_gzip_resp.headers.get("content-disposition", "")
    assert "grantflow_portfolio_quality_usaid.json.gz" in json_gzip_disposition
    json_text = gzip.decompress(json_gzip_resp.content).decode("utf-8")
    parsed = json.loads(json_text)
    assert "filters" in parsed
    assert parsed["filters"]["donor_id"] == "usaid"


def test_portfolio_metrics_export_endpoint_supports_csv_json_and_gzip():
    csv_resp = client.get(
        "/portfolio/metrics/export",
        params={"donor_id": "usaid", "status": "done", "grounding_risk_level": "high", "format": "csv"},
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert "grantflow_portfolio_metrics_usaid_done.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert "filters.donor_id,usaid" in csv_text
    assert "filters.status,done" in csv_text
    assert "filters.grounding_risk_level,high" in csv_text
    assert "avg_time_to_terminal_seconds," in csv_text

    json_resp = client.get(
        "/portfolio/metrics/export",
        params={"donor_id": "usaid", "grounding_risk_level": "high", "format": "json"},
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_disposition = json_resp.headers.get("content-disposition", "")
    assert "grantflow_portfolio_metrics_usaid.json" in json_disposition
    payload = json_resp.json()
    assert "job_count" in payload
    assert payload["filters"]["donor_id"] == "usaid"
    assert payload["filters"]["grounding_risk_level"] == "high"

    toc_csv_resp = client.get(
        "/portfolio/metrics/export",
        params={"donor_id": "usaid", "toc_text_risk_level": "high", "format": "csv"},
    )
    assert toc_csv_resp.status_code == 200
    assert toc_csv_resp.headers["content-type"].startswith("text/csv")
    toc_csv_text = toc_csv_resp.text
    assert "filters.toc_text_risk_level,high" in toc_csv_text

    toc_json_resp = client.get(
        "/portfolio/metrics/export",
        params={"donor_id": "usaid", "toc_text_risk_level": "high", "format": "json"},
    )
    assert toc_json_resp.status_code == 200
    assert toc_json_resp.headers["content-type"].startswith("application/json")
    toc_payload = toc_json_resp.json()
    assert toc_payload["filters"]["donor_id"] == "usaid"
    assert toc_payload["filters"]["toc_text_risk_level"] == "high"

    csv_gzip_resp = client.get(
        "/portfolio/metrics/export", params={"donor_id": "usaid", "format": "csv", "gzip": "true"}
    )
    assert csv_gzip_resp.status_code == 200
    assert csv_gzip_resp.headers["content-type"].startswith("application/gzip")
    assert "grantflow_portfolio_metrics_usaid.csv.gz" in (csv_gzip_resp.headers.get("content-disposition") or "")
    csv_gzip_text = gzip.decompress(csv_gzip_resp.content).decode("utf-8")
    assert "field,value\n" in csv_gzip_text

    json_gzip_resp = client.get(
        "/portfolio/metrics/export", params={"donor_id": "usaid", "format": "json", "gzip": "true"}
    )
    assert json_gzip_resp.status_code == 200
    assert json_gzip_resp.headers["content-type"].startswith("application/gzip")
    assert "grantflow_portfolio_metrics_usaid.json.gz" in (json_gzip_resp.headers.get("content-disposition") or "")
    json_gzip_payload = json.loads(gzip.decompress(json_gzip_resp.content).decode("utf-8"))
    assert json_gzip_payload["filters"]["donor_id"] == "usaid"


def test_portfolio_review_workflow_trends_endpoint_aggregates_jobs_and_filters():
    donor = "portfolio_review_workflow_trends_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-trends-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "medium", "risk_level": "medium"},
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prw-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "Schema mismatch.",
                        }
                    ]
                },
                "citations": [{"citation_type": "fallback_namespace"}],
            },
            "review_comments": [
                {
                    "comment_id": "prw-comment-1",
                    "ts": "2026-02-26T10:00:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "prw-1",
                }
            ],
            "job_events": [
                {
                    "event_id": "prw-evt-1",
                    "ts": "2026-02-26T10:00:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prw-comment-1",
                    "section": "toc",
                },
                {
                    "event_id": "prw-evt-2",
                    "ts": "2026-02-26T10:10:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prw-1",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "prw-evt-3",
                    "ts": "2026-02-27T11:00:00+00:00",
                    "type": "review_comment_status_changed",
                    "comment_id": "prw-comment-1",
                    "status": "resolved",
                    "section": "toc",
                },
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-trends-job-2",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "low", "risk_level": "low"},
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prw-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "toc",
                            "status": "acknowledged",
                            "message": "Baseline missing.",
                        }
                    ]
                },
                "citations": [{"citation_type": "rag_support"}],
            },
            "review_comments": [
                {
                    "comment_id": "prw-comment-2",
                    "ts": "2026-02-27T12:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need indicator source.",
                    "linked_finding_id": "prw-2",
                }
            ],
            "job_events": [
                {
                    "event_id": "prw-evt-4",
                    "ts": "2026-02-27T12:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prw-2",
                    "status": "acknowledged",
                    "section": "toc",
                    "severity": "medium",
                },
                {
                    "event_id": "prw-evt-5",
                    "ts": "2026-02-27T12:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prw-comment-2",
                    "section": "toc",
                },
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-trends-job-3",
        {
            "status": "error",
            "hitl_enabled": False,
            "generate_preflight": {"warning_level": "high", "risk_level": "high"},
            "state": {"donor_id": donor, "critic_notes": {"fatal_flaws": []}},
            "job_events": [
                {
                    "event_id": "prw-evt-6",
                    "ts": "2026-02-28T09:00:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prw-comment-3",
                    "section": "toc",
                }
            ],
        },
    )

    response = client.get(
        "/portfolio/review-workflow/trends",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "finding_section": "toc",
            "overdue_after_hours": 12,
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["filters"]["donor_id"] == donor
    assert body["filters"]["status"] == "done"
    assert body["filters"]["hitl_enabled"] is True
    assert body["filters"]["finding_section"] == "toc"
    assert body["filters"]["overdue_after_hours"] == 12
    assert body["job_count"] == 2
    assert body["jobs_with_events"] == 2
    assert body["jobs_without_events"] == 0
    assert body["timeline_event_count_total"] == 5
    assert body["bucket_count"] == 2
    assert body["time_window_start"] == "2026-02-26"
    assert body["time_window_end"] == "2026-02-27"
    assert body["top_event_type"] == "critic_finding_status_changed"
    assert body["top_event_type_count"] == 2
    assert body["top_donor_id"] == donor
    assert body["top_donor_event_count"] == 5
    assert body["donor_event_counts"][donor] == 5
    assert body["total_series"] == [
        {"bucket": "2026-02-26", "count": 2},
        {"bucket": "2026-02-27", "count": 3},
    ]
    assert body["event_type_series"]["critic_finding_status_changed"] == [
        {"bucket": "2026-02-26", "count": 1},
        {"bucket": "2026-02-27", "count": 1},
    ]
    assert body["donor_series"][donor] == [
        {"bucket": "2026-02-26", "count": 2},
        {"bucket": "2026-02-27", "count": 3},
    ]
    assert body["job_event_counts"]["portfolio-review-workflow-trends-job-1"] == 3
    assert body["job_event_counts"]["portfolio-review-workflow-trends-job-2"] == 2


def test_portfolio_review_workflow_trends_export_supports_csv_json_and_gzip():
    donor = "portfolio_review_workflow_trends_export_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-trends-export-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prwe-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prwe-comment-1",
                    "ts": "2026-02-26T10:00:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Review needed.",
                    "linked_finding_id": "prwe-1",
                }
            ],
            "job_events": [
                {
                    "event_id": "prwe-evt-1",
                    "ts": "2026-02-26T10:00:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prwe-comment-1",
                    "section": "toc",
                },
                {
                    "event_id": "prwe-evt-2",
                    "ts": "2026-02-26T10:10:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prwe-1",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
            ],
        },
    )

    csv_resp = client.get(
        "/portfolio/review-workflow/trends/export",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "finding_section": "toc",
            "format": "csv",
        },
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_trends_{donor}_done_hitl_true.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert f"filters.donor_id,{donor}" in csv_text
    assert "filters.status,done" in csv_text
    assert "filters.hitl_enabled,True" in csv_text
    assert "filters.finding_section,toc" in csv_text
    assert "total_series[0].bucket,2026-02-26" in csv_text

    json_resp = client.get(
        "/portfolio/review-workflow/trends/export",
        params={"donor_id": donor, "finding_section": "toc", "format": "json"},
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_payload = json_resp.json()
    assert json_payload["filters"]["donor_id"] == donor
    assert json_payload["filters"]["finding_section"] == "toc"
    assert json_payload["timeline_event_count_total"] == 2

    gzip_resp = client.get(
        "/portfolio/review-workflow/trends/export",
        params={"donor_id": donor, "format": "json", "gzip": "true"},
    )
    assert gzip_resp.status_code == 200
    assert gzip_resp.headers["content-type"].startswith("application/gzip")
    gzip_disposition = gzip_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_trends_{donor}.json.gz" in gzip_disposition
    gzip_payload = json.loads(gzip.decompress(gzip_resp.content).decode("utf-8"))
    assert gzip_payload["filters"]["donor_id"] == donor
    assert gzip_payload["timeline_event_count_total"] == 2


def test_portfolio_review_workflow_endpoint_aggregates_jobs_and_filters():
    donor = "portfolio_review_workflow_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "medium", "risk_level": "medium"},
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prw-snap-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "Schema mismatch.",
                            "updated_at": "2026-02-27T10:00:00+00:00",
                            "due_at": "2026-02-25T09:00:00+00:00",
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prw-snap-comment-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "prw-snap-1",
                    "due_at": "2026-02-25T10:00:00+00:00",
                }
            ],
            "job_events": [
                {
                    "event_id": "prw-snap-evt-1",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prw-snap-1",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "prw-snap-evt-2",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prw-snap-comment-1",
                    "section": "toc",
                },
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-job-2",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "low", "risk_level": "low"},
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prw-snap-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "toc",
                            "status": "acknowledged",
                            "message": "Baseline missing.",
                            "acknowledged_at": "2026-02-27T11:00:00+00:00",
                            "due_at": "2026-02-25T11:00:00+00:00",
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prw-snap-comment-2",
                    "ts": "2026-02-27T11:05:00+00:00",
                    "section": "toc",
                    "status": "resolved",
                    "message": "Resolved after review.",
                    "linked_finding_id": "prw-snap-2",
                    "resolved_at": "2026-02-27T11:10:00+00:00",
                }
            ],
            "job_events": [
                {
                    "event_id": "prw-snap-evt-3",
                    "ts": "2026-02-27T11:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prw-snap-2",
                    "status": "acknowledged",
                    "section": "toc",
                    "severity": "medium",
                },
                {
                    "event_id": "prw-snap-evt-4",
                    "ts": "2026-02-27T11:05:00+00:00",
                    "type": "review_comment_status_changed",
                    "comment_id": "prw-snap-comment-2",
                    "status": "resolved",
                    "section": "toc",
                },
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-job-3",
        {
            "status": "done",
            "hitl_enabled": False,
            "state": {"donor_id": donor, "critic_notes": {"fatal_flaws": []}},
            "job_events": [],
        },
    )

    response = client.get(
        "/portfolio/review-workflow",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "finding_section": "toc",
            "overdue_after_hours": 12,
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["filters"]["donor_id"] == donor
    assert body["filters"]["status"] == "done"
    assert body["filters"]["hitl_enabled"] is True
    assert body["filters"]["finding_section"] == "toc"
    assert body["filters"]["overdue_after_hours"] == 12
    assert body["job_count"] == 2
    assert body["jobs_with_activity"] == 2
    assert body["jobs_without_activity"] == 0
    assert body["jobs_with_overdue"] == 2
    assert body["jobs_without_overdue"] == 0
    assert body["summary"]["finding_count"] == 2
    assert body["summary"]["comment_count"] == 2
    assert body["summary"]["open_finding_count"] == 1
    assert body["summary"]["acknowledged_finding_count"] == 1
    assert body["summary"]["open_comment_count"] == 1
    assert body["summary"]["resolved_comment_count"] == 1
    assert body["summary"]["timeline_event_count"] == 4
    assert body["top_event_type"] == "critic_finding_status_changed"
    assert body["top_event_type_count"] == 2
    assert body["top_donor_id"] == donor
    assert body["top_donor_event_count"] == 4
    assert body["timeline_event_type_counts"]["critic_finding_status_changed"] == 2
    assert body["timeline_event_type_counts"]["review_comment_status_changed"] == 1
    assert body["timeline_event_type_counts"]["review_comment_added"] == 1
    assert body["donor_event_counts"][donor] == 4
    assert body["job_event_counts"]["portfolio-review-workflow-job-1"] == 2
    assert body["job_event_counts"]["portfolio-review-workflow-job-2"] == 2
    assert body["latest_timeline_limit"] == 200
    assert body["latest_timeline_truncated"] is False
    assert len(body["latest_timeline"]) == 4
    assert body["latest_timeline"][0]["job_id"] == "portfolio-review-workflow-job-2"
    assert body["latest_timeline"][0]["donor_id"] == donor


def test_portfolio_review_workflow_export_supports_csv_json_and_gzip():
    donor = "portfolio_review_workflow_export_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-export-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prw-exp-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "updated_at": "2026-02-27T10:00:00+00:00",
                            "due_at": "2026-02-25T09:00:00+00:00",
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prw-exp-comment-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Review needed.",
                    "linked_finding_id": "prw-exp-1",
                    "due_at": "2026-02-25T10:00:00+00:00",
                }
            ],
            "job_events": [
                {
                    "event_id": "prw-exp-evt-1",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prw-exp-1",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "prw-exp-evt-2",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prw-exp-comment-1",
                    "section": "toc",
                },
            ],
        },
    )

    csv_resp = client.get(
        "/portfolio/review-workflow/export",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "finding_section": "toc",
            "format": "csv",
        },
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_{donor}_done_hitl_true.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert f"filters.donor_id,{donor}" in csv_text
    assert "filters.status,done" in csv_text
    assert "filters.hitl_enabled,True" in csv_text
    assert "filters.finding_section,toc" in csv_text
    assert "summary.timeline_event_count,2" in csv_text

    json_resp = client.get(
        "/portfolio/review-workflow/export",
        params={"donor_id": donor, "finding_section": "toc", "format": "json"},
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_payload = json_resp.json()
    assert json_payload["filters"]["donor_id"] == donor
    assert json_payload["filters"]["finding_section"] == "toc"
    assert json_payload["summary"]["timeline_event_count"] == 2
    assert json_payload["summary"]["comment_count"] == 1

    gzip_resp = client.get(
        "/portfolio/review-workflow/export",
        params={"donor_id": donor, "format": "json", "gzip": "true"},
    )
    assert gzip_resp.status_code == 200
    assert gzip_resp.headers["content-type"].startswith("application/gzip")
    gzip_disposition = gzip_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_{donor}.json.gz" in gzip_disposition
    gzip_payload = json.loads(gzip.decompress(gzip_resp.content).decode("utf-8"))
    assert gzip_payload["filters"]["donor_id"] == donor
    assert gzip_payload["summary"]["timeline_event_count"] == 2


def test_portfolio_review_workflow_sla_endpoint_aggregates_jobs_and_filters():
    donor = "portfolio_review_workflow_sla_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "medium", "risk_level": "medium"},
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prw-sla-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "updated_at": "2026-02-20T10:00:00+00:00",
                            "due_at": "2026-02-25T10:00:00+00:00",
                            "sla_hours": 24,
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prw-sla-comment-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "prw-sla-1",
                    "due_at": "2026-02-25T11:00:00+00:00",
                    "sla_hours": 24,
                }
            ],
            "job_events": [
                {
                    "event_id": "prw-sla-evt-1",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prw-sla-1",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "prw-sla-evt-2",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prw-sla-comment-1",
                    "section": "toc",
                },
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-job-2",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "low", "risk_level": "low"},
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prw-sla-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "toc",
                            "status": "acknowledged",
                            "acknowledged_at": "2026-02-27T11:00:00+00:00",
                            "due_at": "2026-02-26T10:00:00+00:00",
                            "sla_hours": 72,
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prw-sla-comment-2",
                    "ts": "2026-02-27T11:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need indicator source.",
                    "linked_finding_id": "prw-sla-2",
                    "due_at": "2026-02-26T12:00:00+00:00",
                    "sla_hours": 72,
                }
            ],
            "job_events": [
                {
                    "event_id": "prw-sla-evt-3",
                    "ts": "2026-02-27T11:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prw-sla-2",
                    "status": "acknowledged",
                    "section": "toc",
                    "severity": "medium",
                },
                {
                    "event_id": "prw-sla-evt-4",
                    "ts": "2026-02-27T11:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prw-sla-comment-2",
                    "section": "toc",
                },
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-job-3",
        {
            "status": "error",
            "hitl_enabled": False,
            "state": {"donor_id": donor, "critic_notes": {"fatal_flaws": []}},
            "job_events": [],
        },
    )

    response = client.get(
        "/portfolio/review-workflow/sla",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "finding_section": "toc",
            "overdue_after_hours": 12,
            "top_limit": 3,
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["filters"]["donor_id"] == donor
    assert body["filters"]["status"] == "done"
    assert body["filters"]["hitl_enabled"] is True
    assert body["filters"]["finding_section"] == "toc"
    assert body["filters"]["overdue_after_hours"] == 12
    assert body["filters"]["top_limit"] == 3
    assert body["job_count"] == 2
    assert body["jobs_with_overdue"] == 2
    assert body["jobs_without_overdue"] == 0
    assert body["overdue_comment_count"] == 2
    assert body["overdue_total"] == 2
    assert body["top_donor_id"] == donor
    assert body["top_donor_overdue_count"] == 2
    assert body["donor_overdue_counts"][donor] == 2
    assert body["job_overdue_counts"]["portfolio-review-workflow-sla-job-1"] == 1
    assert body["job_overdue_counts"]["portfolio-review-workflow-sla-job-2"] == 1
    assert len(body["top_overdue"]) == 2
    assert body["top_overdue"][0]["job_id"] in {
        "portfolio-review-workflow-sla-job-1",
        "portfolio-review-workflow-sla-job-2",
    }
    assert body["top_overdue"][0]["donor_id"] == donor
    assert isinstance(body.get("oldest_overdue"), dict)


def test_portfolio_review_workflow_sla_export_supports_csv_json_and_gzip():
    donor = "portfolio_review_workflow_sla_export_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-export-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "state": {
                "donor_id": donor,
                "critic_notes": {"fatal_flaws": []},
            },
            "review_comments": [
                {
                    "comment_id": "prw-sla-exp-comment-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Review needed.",
                    "due_at": "2026-02-25T11:00:00+00:00",
                    "sla_hours": 24,
                }
            ],
            "job_events": [
                {
                    "event_id": "prw-sla-exp-evt-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prw-sla-exp-comment-1",
                    "section": "toc",
                },
            ],
        },
    )

    csv_resp = client.get(
        "/portfolio/review-workflow/sla/export",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "finding_section": "toc",
            "top_limit": 2,
            "format": "csv",
        },
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_sla_{donor}_done_hitl_true.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert f"filters.donor_id,{donor}" in csv_text
    assert "filters.status,done" in csv_text
    assert "filters.hitl_enabled,True" in csv_text
    assert "filters.finding_section,toc" in csv_text
    assert "filters.top_limit,2" in csv_text

    json_resp = client.get(
        "/portfolio/review-workflow/sla/export",
        params={"donor_id": donor, "finding_section": "toc", "top_limit": 2, "format": "json"},
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_payload = json_resp.json()
    assert json_payload["filters"]["donor_id"] == donor
    assert json_payload["filters"]["finding_section"] == "toc"
    assert json_payload["filters"]["top_limit"] == 2
    assert json_payload["overdue_total"] == 1

    gzip_resp = client.get(
        "/portfolio/review-workflow/sla/export",
        params={"donor_id": donor, "format": "json", "gzip": "true"},
    )
    assert gzip_resp.status_code == 200
    assert gzip_resp.headers["content-type"].startswith("application/gzip")
    gzip_disposition = gzip_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_sla_{donor}.json.gz" in gzip_disposition
    gzip_payload = json.loads(gzip.decompress(gzip_resp.content).decode("utf-8"))
    assert gzip_payload["filters"]["donor_id"] == donor
    assert gzip_payload["overdue_total"] == 1


def test_portfolio_review_workflow_sla_hotspots_endpoint_aggregates_jobs_and_filters():
    donor = "portfolio_review_workflow_sla_hotspots_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-hotspots-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "medium", "risk_level": "medium"},
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prwh-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "updated_at": "2026-02-27T10:00:00+00:00",
                            "due_at": "2026-02-25T10:00:00+00:00",
                            "sla_hours": 24,
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prwh-comment-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "prwh-1",
                    "due_at": "2026-02-25T11:00:00+00:00",
                    "sla_hours": 24,
                }
            ],
            "job_events": [
                {
                    "event_id": "prwh-evt-1",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prwh-1",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                }
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-hotspots-job-2",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "low", "risk_level": "low"},
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prwh-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "toc",
                            "status": "acknowledged",
                            "updated_at": "2026-02-20T11:00:00+00:00",
                            "due_at": "2026-02-26T10:00:00+00:00",
                            "sla_hours": 72,
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prwh-comment-2",
                    "ts": "2026-02-27T11:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need indicator source.",
                    "linked_finding_id": "prwh-2",
                    "due_at": "2026-02-26T12:00:00+00:00",
                    "sla_hours": 72,
                }
            ],
            "job_events": [
                {
                    "event_id": "prwh-evt-2",
                    "ts": "2026-02-27T11:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prwh-2",
                    "status": "acknowledged",
                    "section": "toc",
                    "severity": "medium",
                }
            ],
        },
    )

    response = client.get(
        "/portfolio/review-workflow/sla/hotspots",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "finding_section": "toc",
            "overdue_after_hours": 12,
            "hotspot_kind": "comment",
            "min_overdue_hours": 24,
            "top_limit": 5,
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["filters"]["donor_id"] == donor
    assert body["filters"]["hotspot_kind"] == "comment"
    assert "hotspot_severity" not in body["filters"]
    assert body["filters"]["min_overdue_hours"] == 24.0
    assert body["filters"]["top_limit"] == 5
    assert body["job_count"] == 2
    assert body["jobs_with_overdue"] == 1
    assert body["jobs_without_overdue"] == 1
    assert body["total_overdue_items"] == 1
    assert body["hotspot_count"] == 1
    assert body["top_donor_id"] == donor
    assert body["top_donor_overdue_count"] == 1
    assert body["donor_hotspot_counts"][donor] == 1
    assert body["job_hotspot_counts"]["portfolio-review-workflow-sla-hotspots-job-1"] == 1
    assert body["job_hotspot_counts"]["portfolio-review-workflow-sla-hotspots-job-2"] == 0
    assert len(body["top_overdue"]) == 1
    assert body["top_overdue"][0]["kind"] == "comment"
    assert body["top_overdue"][0]["job_id"] == "portfolio-review-workflow-sla-hotspots-job-1"
    assert body["top_overdue"][0]["donor_id"] == donor


def test_portfolio_review_workflow_sla_hotspots_export_supports_csv_json_and_gzip():
    donor = "portfolio_review_workflow_sla_hotspots_export_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-hotspots-export-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prwh-exp-1",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "updated_at": "2026-02-20T10:00:00+00:00",
                            "due_at": "2026-02-25T10:00:00+00:00",
                            "sla_hours": 24,
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prwh-exp-comment-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Review needed.",
                    "linked_finding_id": "prwh-exp-1",
                    "due_at": "2026-02-25T11:00:00+00:00",
                    "sla_hours": 24,
                }
            ],
            "job_events": [
                {
                    "event_id": "prwh-exp-evt-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prwh-exp-comment-1",
                    "section": "toc",
                }
            ],
        },
    )

    csv_resp = client.get(
        "/portfolio/review-workflow/sla/hotspots/export",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "overdue_after_hours": 12,
            "hotspot_kind": "comment",
            "min_overdue_hours": 1,
            "top_limit": 3,
            "format": "csv",
        },
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_sla_hotspots_{donor}_done_hitl_true.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert f"filters.donor_id,{donor}" in csv_text
    assert "filters.hotspot_kind,comment" in csv_text
    assert "filters.top_limit,3" in csv_text

    json_resp = client.get(
        "/portfolio/review-workflow/sla/hotspots/export",
        params={
            "donor_id": donor,
            "overdue_after_hours": 12,
            "hotspot_kind": "comment",
            "format": "json",
        },
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_payload = json_resp.json()
    assert json_payload["filters"]["donor_id"] == donor
    assert json_payload["filters"]["hotspot_kind"] == "comment"
    assert json_payload["filters"]["hotspot_severity"] is None
    assert json_payload["hotspot_count"] == 1

    gzip_resp = client.get(
        "/portfolio/review-workflow/sla/hotspots/export",
        params={"donor_id": donor, "overdue_after_hours": 12, "format": "json", "gzip": "true"},
    )
    assert gzip_resp.status_code == 200
    assert gzip_resp.headers["content-type"].startswith("application/gzip")
    gzip_disposition = gzip_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_sla_hotspots_{donor}.json.gz" in gzip_disposition
    gzip_payload = json.loads(gzip.decompress(gzip_resp.content).decode("utf-8"))
    assert gzip_payload["filters"]["donor_id"] == donor
    assert gzip_payload["hotspot_count"] == 1


def test_portfolio_review_workflow_sla_hotspots_trends_endpoint_aggregates_jobs_and_filters():
    donor = "portfolio_review_workflow_sla_hotspots_trends_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-hotspots-trends-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "medium", "risk_level": "medium"},
            "state": {
                "donor_id": donor,
                "critic_notes": {"fatal_flaws": []},
            },
            "review_comments": [
                {
                    "comment_id": "prwht-comment-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "due_at": "2026-02-25T11:00:00+00:00",
                    "sla_hours": 24,
                }
            ],
            "job_events": [
                {
                    "event_id": "prwht-evt-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prwht-comment-1",
                    "section": "toc",
                }
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-hotspots-trends-job-2",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "low", "risk_level": "low"},
            "state": {
                "donor_id": donor,
                "critic_notes": {"fatal_flaws": []},
            },
            "review_comments": [
                {
                    "comment_id": "prwht-comment-2",
                    "ts": "2026-02-28T11:05:00+00:00",
                    "section": "logframe",
                    "status": "open",
                    "message": "Need indicator source.",
                    "due_at": "2026-02-26T12:00:00+00:00",
                    "sla_hours": 24,
                }
            ],
            "job_events": [
                {
                    "event_id": "prwht-evt-2",
                    "ts": "2026-02-28T11:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prwht-comment-2",
                    "section": "logframe",
                }
            ],
        },
    )

    response = client.get(
        "/portfolio/review-workflow/sla/hotspots/trends",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "overdue_after_hours": 12,
            "hotspot_kind": "comment",
            "top_limit": 5,
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["filters"]["donor_id"] == donor
    assert body["filters"]["hotspot_kind"] == "comment"
    assert body["filters"]["top_limit"] == 5
    assert body["job_count"] == 2
    assert body["jobs_with_overdue"] == 2
    assert body["jobs_without_overdue"] == 0
    assert body["hotspot_count_total"] == 2
    assert body["top_kind"] == "comment"
    assert body["top_kind_count"] == 2
    assert body["top_donor_id"] == donor
    assert body["top_donor_hotspot_count"] == 2
    assert body["donor_hotspot_counts"][donor] == 2
    assert body["job_hotspot_counts"]["portfolio-review-workflow-sla-hotspots-trends-job-1"] == 1
    assert body["job_hotspot_counts"]["portfolio-review-workflow-sla-hotspots-trends-job-2"] == 1
    assert isinstance(body["total_series"], list)
    assert body["bucket_count"] >= 1
    assert len(body["top_overdue"]) == 2
    assert body["top_overdue"][0]["kind"] == "comment"


def test_portfolio_review_workflow_sla_hotspots_trends_export_supports_csv_json_and_gzip():
    donor = "portfolio_review_workflow_sla_hotspots_trends_export_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-hotspots-trends-export-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "state": {
                "donor_id": donor,
                "critic_notes": {"fatal_flaws": []},
            },
            "review_comments": [
                {
                    "comment_id": "prwht-exp-comment-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Review needed.",
                    "due_at": "2026-02-25T11:00:00+00:00",
                    "sla_hours": 24,
                }
            ],
            "job_events": [
                {
                    "event_id": "prwht-exp-evt-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prwht-exp-comment-1",
                    "section": "toc",
                }
            ],
        },
    )

    csv_resp = client.get(
        "/portfolio/review-workflow/sla/hotspots/trends/export",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "overdue_after_hours": 12,
            "hotspot_kind": "comment",
            "top_limit": 3,
            "format": "csv",
        },
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_sla_hotspots_trends_{donor}_done_hitl_true.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert f"filters.donor_id,{donor}" in csv_text
    assert "filters.hotspot_kind,comment" in csv_text
    assert "filters.top_limit,3" in csv_text

    json_resp = client.get(
        "/portfolio/review-workflow/sla/hotspots/trends/export",
        params={
            "donor_id": donor,
            "overdue_after_hours": 12,
            "hotspot_kind": "comment",
            "format": "json",
        },
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_payload = json_resp.json()
    assert json_payload["filters"]["donor_id"] == donor
    assert json_payload["filters"]["hotspot_kind"] == "comment"
    assert json_payload["hotspot_count_total"] == 1

    gzip_resp = client.get(
        "/portfolio/review-workflow/sla/hotspots/trends/export",
        params={"donor_id": donor, "overdue_after_hours": 12, "format": "json", "gzip": "true"},
    )
    assert gzip_resp.status_code == 200
    assert gzip_resp.headers["content-type"].startswith("application/gzip")
    gzip_disposition = gzip_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_sla_hotspots_trends_{donor}.json.gz" in gzip_disposition
    gzip_payload = json.loads(gzip.decompress(gzip_resp.content).decode("utf-8"))
    assert gzip_payload["filters"]["donor_id"] == donor
    assert gzip_payload["hotspot_count_total"] == 1


def test_portfolio_review_workflow_sla_trends_endpoint_aggregates_jobs_and_filters():
    donor = "portfolio_review_workflow_sla_trends_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-trends-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "medium", "risk_level": "medium"},
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prws-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "message": "Schema mismatch.",
                            "updated_at": "2026-02-27T10:00:00+00:00",
                            "due_at": "2026-02-25T10:00:00+00:00",
                            "sla_hours": 24,
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prws-comment-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need stronger assumptions.",
                    "linked_finding_id": "prws-1",
                    "due_at": "2026-02-25T12:00:00+00:00",
                    "sla_hours": 24,
                }
            ],
            "job_events": [
                {
                    "event_id": "prws-evt-1",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prws-1",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "prws-evt-2",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prws-comment-1",
                    "section": "toc",
                },
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-trends-job-2",
        {
            "status": "done",
            "hitl_enabled": True,
            "generate_preflight": {"warning_level": "low", "risk_level": "low"},
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prws-2",
                            "code": "MEL_BASELINE_MISSING",
                            "severity": "medium",
                            "section": "toc",
                            "status": "acknowledged",
                            "message": "Baseline missing.",
                            "acknowledged_at": "2026-02-27T11:00:00+00:00",
                            "due_at": "2026-02-26T10:00:00+00:00",
                            "sla_hours": 72,
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prws-comment-2",
                    "ts": "2026-02-27T11:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Need indicator source.",
                    "linked_finding_id": "prws-2",
                    "due_at": "2026-02-26T12:00:00+00:00",
                    "sla_hours": 72,
                }
            ],
            "job_events": [
                {
                    "event_id": "prws-evt-3",
                    "ts": "2026-02-27T11:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prws-2",
                    "status": "acknowledged",
                    "section": "toc",
                    "severity": "medium",
                },
                {
                    "event_id": "prws-evt-4",
                    "ts": "2026-02-27T11:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prws-comment-2",
                    "section": "toc",
                },
            ],
        },
    )
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-trends-job-3",
        {
            "status": "error",
            "hitl_enabled": False,
            "generate_preflight": {"warning_level": "high", "risk_level": "high"},
            "state": {"donor_id": donor, "critic_notes": {"fatal_flaws": []}},
            "job_events": [
                {
                    "event_id": "prws-evt-5",
                    "ts": "2026-02-28T09:00:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prws-comment-3",
                    "section": "toc",
                }
            ],
        },
    )

    response = client.get(
        "/portfolio/review-workflow/sla/trends",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "finding_section": "toc",
            "overdue_after_hours": 12,
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["filters"]["donor_id"] == donor
    assert body["filters"]["status"] == "done"
    assert body["filters"]["hitl_enabled"] is True
    assert body["filters"]["finding_section"] == "toc"
    assert body["filters"]["overdue_after_hours"] == 12
    assert body["job_count"] == 2
    assert body["jobs_with_overdue"] == 2
    assert body["jobs_without_overdue"] == 0
    assert body["overdue_finding_count"] == 2
    assert body["overdue_comment_count"] == 2
    assert body["overdue_total"] == 4
    assert body["unresolved_total"] == 4
    assert body["breach_rate"] == 1.0
    assert body["bucket_count"] == 2
    assert body["time_window_start"] == "2026-02-25"
    assert body["time_window_end"] == "2026-02-26"
    assert body["top_severity"] == "high"
    assert body["top_severity_count"] == 2
    assert body["top_section"] == "toc"
    assert body["top_section_count"] == 4
    assert body["top_donor_id"] == donor
    assert body["top_donor_overdue_count"] == 4
    assert body["donor_overdue_counts"][donor] == 4
    assert body["total_series"] == [
        {"bucket": "2026-02-25", "count": 2},
        {"bucket": "2026-02-26", "count": 2},
    ]
    assert body["severity_series"]["high"] == [{"bucket": "2026-02-25", "count": 2}]
    assert body["severity_series"]["medium"] == [{"bucket": "2026-02-26", "count": 2}]
    assert body["donor_series"][donor] == [
        {"bucket": "2026-02-25", "count": 2},
        {"bucket": "2026-02-26", "count": 2},
    ]
    assert body["job_overdue_counts"]["portfolio-review-workflow-sla-trends-job-1"] == 2
    assert body["job_overdue_counts"]["portfolio-review-workflow-sla-trends-job-2"] == 2


def test_portfolio_review_workflow_sla_trends_export_supports_csv_json_and_gzip():
    donor = "portfolio_review_workflow_sla_trends_export_test_donor"
    api_app_module.JOB_STORE.set(
        "portfolio-review-workflow-sla-trends-export-job-1",
        {
            "status": "done",
            "hitl_enabled": True,
            "state": {
                "donor_id": donor,
                "critic_notes": {
                    "fatal_flaws": [
                        {
                            "finding_id": "prwse-1",
                            "code": "TOC_SCHEMA_INVALID",
                            "severity": "high",
                            "section": "toc",
                            "status": "open",
                            "updated_at": "2026-02-27T10:00:00+00:00",
                            "due_at": "2026-02-25T10:00:00+00:00",
                            "sla_hours": 24,
                        }
                    ]
                },
            },
            "review_comments": [
                {
                    "comment_id": "prwse-comment-1",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "section": "toc",
                    "status": "open",
                    "message": "Review needed.",
                    "linked_finding_id": "prwse-1",
                    "due_at": "2026-02-25T11:00:00+00:00",
                    "sla_hours": 24,
                }
            ],
            "job_events": [
                {
                    "event_id": "prwse-evt-1",
                    "ts": "2026-02-27T10:00:00+00:00",
                    "type": "critic_finding_status_changed",
                    "finding_id": "prwse-1",
                    "status": "open",
                    "section": "toc",
                    "severity": "high",
                },
                {
                    "event_id": "prwse-evt-2",
                    "ts": "2026-02-27T10:05:00+00:00",
                    "type": "review_comment_added",
                    "comment_id": "prwse-comment-1",
                    "section": "toc",
                },
            ],
        },
    )

    csv_resp = client.get(
        "/portfolio/review-workflow/sla/trends/export",
        params={
            "donor_id": donor,
            "status": "done",
            "hitl_enabled": "true",
            "finding_section": "toc",
            "format": "csv",
        },
    )
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    csv_disposition = csv_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_sla_trends_{donor}_done_hitl_true.csv" in csv_disposition
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert f"filters.donor_id,{donor}" in csv_text
    assert "filters.status,done" in csv_text
    assert "filters.hitl_enabled,True" in csv_text
    assert "filters.finding_section,toc" in csv_text
    assert "total_series[0].bucket,2026-02-25" in csv_text

    json_resp = client.get(
        "/portfolio/review-workflow/sla/trends/export",
        params={
            "donor_id": donor,
            "finding_section": "toc",
            "overdue_after_hours": 12,
            "format": "json",
        },
    )
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    json_payload = json_resp.json()
    assert json_payload["filters"]["donor_id"] == donor
    assert json_payload["filters"]["finding_section"] == "toc"
    assert json_payload["overdue_total"] == 1

    gzip_resp = client.get(
        "/portfolio/review-workflow/sla/trends/export",
        params={"donor_id": donor, "overdue_after_hours": 12, "format": "json", "gzip": "true"},
    )
    assert gzip_resp.status_code == 200
    assert gzip_resp.headers["content-type"].startswith("application/gzip")
    gzip_disposition = gzip_resp.headers.get("content-disposition", "")
    assert f"grantflow_portfolio_review_workflow_sla_trends_{donor}.json.gz" in gzip_disposition
    gzip_payload = json.loads(gzip.decompress(gzip_resp.content).decode("utf-8"))
    assert gzip_payload["filters"]["donor_id"] == donor
    assert gzip_payload["overdue_total"] == 1


def test_ingest_inventory_export_endpoint_supports_csv_json_and_gzip(monkeypatch):
    api_app_module.INGEST_AUDIT_STORE.clear()

    def fake_ingest(pdf_path: str, namespace: str, metadata=None):
        return {"namespace": namespace, "source": pdf_path, "chunks_ingested": 1, "stats": {}}

    monkeypatch.setattr(api_app_module, "ingest_pdf_to_namespace", fake_ingest)

    client.post(
        "/ingest",
        data={
            "donor_id": "usaid",
            "metadata_json": json.dumps({"doc_family": "donor_policy", "source_type": "donor_guidance"}),
        },
        files={"file": ("ads.pdf", b"%PDF-1.4 a", "application/pdf")},
    )
    client.post(
        "/ingest",
        data={
            "donor_id": "usaid",
            "metadata_json": json.dumps({"doc_family": "country_context", "source_type": "country_context"}),
        },
        files={"file": ("kz-context.pdf", b"%PDF-1.4 b", "application/pdf")},
    )

    csv_resp = client.get("/ingest/inventory/export", params={"donor_id": "usaid", "format": "csv"})
    assert csv_resp.status_code == 200
    assert csv_resp.headers["content-type"].startswith("text/csv")
    assert "grantflow_ingest_inventory_usaid.csv" in (csv_resp.headers.get("content-disposition") or "")
    csv_text = csv_resp.text
    assert csv_text.startswith("field,value\n")
    assert "donor_id,usaid" in csv_text
    assert "doc_family_counts.donor_policy,1" in csv_text
    assert "doc_family_counts.country_context,1" in csv_text

    json_resp = client.get("/ingest/inventory/export", params={"donor_id": "usaid", "format": "json"})
    assert json_resp.status_code == 200
    assert json_resp.headers["content-type"].startswith("application/json")
    assert "grantflow_ingest_inventory_usaid.json" in (json_resp.headers.get("content-disposition") or "")
    json_body = json_resp.json()
    assert json_body["donor_id"] == "usaid"
    assert json_body["doc_family_counts"]["donor_policy"] == 1

    csv_gzip_resp = client.get(
        "/ingest/inventory/export", params={"donor_id": "usaid", "format": "csv", "gzip": "true"}
    )
    assert csv_gzip_resp.status_code == 200
    assert csv_gzip_resp.headers["content-type"].startswith("application/gzip")
    assert "grantflow_ingest_inventory_usaid.csv.gz" in (csv_gzip_resp.headers.get("content-disposition") or "")
    csv_gzip_text = gzip.decompress(csv_gzip_resp.content).decode("utf-8")
    assert "doc_family_counts.donor_policy,1" in csv_gzip_text

    json_gzip_resp = client.get(
        "/ingest/inventory/export", params={"donor_id": "usaid", "format": "json", "gzip": "true"}
    )
    assert json_gzip_resp.status_code == 200
    assert json_gzip_resp.headers["content-type"].startswith("application/gzip")
    assert "grantflow_ingest_inventory_usaid.json.gz" in (json_gzip_resp.headers.get("content-disposition") or "")
    json_gzip_body = json.loads(gzip.decompress(json_gzip_resp.content).decode("utf-8"))
    assert json_gzip_body["family_count"] == 2


def test_generate_requires_api_key_when_configured(monkeypatch):
    monkeypatch.setenv("GRANTFLOW_API_KEY", "test-secret")

    preflight_response = client.post("/generate/preflight", json={"donor_id": "usaid"})
    assert preflight_response.status_code == 401

    preflight_response = client.post(
        "/generate/preflight",
        json={"donor_id": "usaid"},
        headers={"X-API-Key": "test-secret"},
    )
    assert preflight_response.status_code == 200
    assert preflight_response.json()["donor_id"] == "usaid"

    payload = {
        "donor_id": "usaid",
        "input_context": {"project": "Public Health", "country": "Kenya"},
        "llm_mode": False,
        "hitl_enabled": False,
    }

    response = client.post("/generate", json=payload)
    assert response.status_code == 401

    response = client.post("/generate", json=payload, headers={"X-API-Key": "test-secret"})
    assert response.status_code == 200
    assert response.json()["status"] == "accepted"


def test_read_endpoints_require_api_key_when_configured(monkeypatch):
    monkeypatch.setenv("GRANTFLOW_API_KEY", "test-secret")
    monkeypatch.setenv("GRANTFLOW_REQUIRE_AUTH_FOR_READS", "true")

    gen = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Health Systems", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
        headers={"X-API-Key": "test-secret"},
    )
    assert gen.status_code == 200
    job_id = gen.json()["job_id"]

    status_unauth = client.get(f"/status/{job_id}")
    assert status_unauth.status_code == 401

    status_auth = client.get(f"/status/{job_id}", headers={"X-API-Key": "test-secret"})
    assert status_auth.status_code == 200

    citations_unauth = client.get(f"/status/{job_id}/citations")
    assert citations_unauth.status_code == 401

    citations_auth = client.get(f"/status/{job_id}/citations", headers={"X-API-Key": "test-secret"})
    assert citations_auth.status_code == 200

    export_payload_unauth = client.get(f"/status/{job_id}/export-payload")
    assert export_payload_unauth.status_code == 401

    export_payload_auth = client.get(f"/status/{job_id}/export-payload", headers={"X-API-Key": "test-secret"})
    assert export_payload_auth.status_code == 200

    versions_unauth = client.get(f"/status/{job_id}/versions")
    assert versions_unauth.status_code == 401

    versions_auth = client.get(f"/status/{job_id}/versions", headers={"X-API-Key": "test-secret"})
    assert versions_auth.status_code == 200

    diff_unauth = client.get(f"/status/{job_id}/diff")
    assert diff_unauth.status_code == 401

    diff_auth = client.get(f"/status/{job_id}/diff", headers={"X-API-Key": "test-secret"})
    assert diff_auth.status_code == 200

    events_unauth = client.get(f"/status/{job_id}/events")
    assert events_unauth.status_code == 401

    events_auth = client.get(f"/status/{job_id}/events", headers={"X-API-Key": "test-secret"})
    assert events_auth.status_code == 200

    metrics_unauth = client.get(f"/status/{job_id}/metrics")
    assert metrics_unauth.status_code == 401

    metrics_auth = client.get(f"/status/{job_id}/metrics", headers={"X-API-Key": "test-secret"})
    assert metrics_auth.status_code == 200

    quality_unauth = client.get(f"/status/{job_id}/quality")
    assert quality_unauth.status_code == 401

    quality_auth = client.get(f"/status/{job_id}/quality", headers={"X-API-Key": "test-secret"})
    assert quality_auth.status_code == 200

    grounding_gate_unauth = client.get(f"/status/{job_id}/grounding-gate")
    assert grounding_gate_unauth.status_code == 401

    grounding_gate_auth = client.get(f"/status/{job_id}/grounding-gate", headers={"X-API-Key": "test-secret"})
    assert grounding_gate_auth.status_code == 200

    critic_unauth = client.get(f"/status/{job_id}/critic")
    assert critic_unauth.status_code == 401

    critic_auth = client.get(f"/status/{job_id}/critic", headers={"X-API-Key": "test-secret"})
    assert critic_auth.status_code == 200
    critic_finding_id = None
    if critic_auth.json().get("fatal_flaws"):
        critic_finding_id = critic_auth.json()["fatal_flaws"][0].get("finding_id")

    if critic_finding_id:
        ack_finding_unauth = client.post(f"/status/{job_id}/critic/findings/{critic_finding_id}/ack")
        assert ack_finding_unauth.status_code == 401

        ack_finding_auth = client.post(
            f"/status/{job_id}/critic/findings/{critic_finding_id}/ack",
            headers={"X-API-Key": "test-secret"},
        )
        assert ack_finding_auth.status_code == 200

        resolve_finding_unauth = client.post(f"/status/{job_id}/critic/findings/{critic_finding_id}/resolve")
        assert resolve_finding_unauth.status_code == 401

        resolve_finding_auth = client.post(
            f"/status/{job_id}/critic/findings/{critic_finding_id}/resolve",
            headers={"X-API-Key": "test-secret"},
        )
        assert resolve_finding_auth.status_code == 200

        reopen_finding_unauth = client.post(f"/status/{job_id}/critic/findings/{critic_finding_id}/open")
        assert reopen_finding_unauth.status_code == 401

        reopen_finding_auth = client.post(
            f"/status/{job_id}/critic/findings/{critic_finding_id}/open",
            headers={"X-API-Key": "test-secret"},
        )
        assert reopen_finding_auth.status_code == 200

        bulk_finding_unauth = client.post(
            f"/status/{job_id}/critic/findings/bulk-status",
            json={"next_status": "resolved", "finding_ids": [critic_finding_id]},
        )
        assert bulk_finding_unauth.status_code == 401

        bulk_finding_auth = client.post(
            f"/status/{job_id}/critic/findings/bulk-status",
            json={"next_status": "resolved", "finding_ids": [critic_finding_id]},
            headers={"X-API-Key": "test-secret"},
        )
        assert bulk_finding_auth.status_code == 200

    comments_unauth = client.get(f"/status/{job_id}/comments")
    assert comments_unauth.status_code == 401

    comments_auth = client.get(f"/status/{job_id}/comments", headers={"X-API-Key": "test-secret"})
    assert comments_auth.status_code == 200

    review_workflow_unauth = client.get(f"/status/{job_id}/review/workflow")
    assert review_workflow_unauth.status_code == 401

    review_workflow_auth = client.get(f"/status/{job_id}/review/workflow", headers={"X-API-Key": "test-secret"})
    assert review_workflow_auth.status_code == 200

    review_workflow_trends_unauth = client.get(f"/status/{job_id}/review/workflow/trends")
    assert review_workflow_trends_unauth.status_code == 401

    review_workflow_trends_auth = client.get(
        f"/status/{job_id}/review/workflow/trends",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_trends_auth.status_code == 200

    review_workflow_sla_unauth = client.get(f"/status/{job_id}/review/workflow/sla")
    assert review_workflow_sla_unauth.status_code == 401

    review_workflow_sla_auth = client.get(
        f"/status/{job_id}/review/workflow/sla",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_sla_auth.status_code == 200

    review_workflow_sla_hotspots_unauth = client.get(f"/status/{job_id}/review/workflow/sla/hotspots")
    assert review_workflow_sla_hotspots_unauth.status_code == 401

    review_workflow_sla_hotspots_auth = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_sla_hotspots_auth.status_code == 200

    review_workflow_sla_hotspots_export_unauth = client.get(f"/status/{job_id}/review/workflow/sla/hotspots/export")
    assert review_workflow_sla_hotspots_export_unauth.status_code == 401

    review_workflow_sla_hotspots_export_auth = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_sla_hotspots_export_auth.status_code == 200

    review_workflow_sla_hotspots_trends_unauth = client.get(f"/status/{job_id}/review/workflow/sla/hotspots/trends")
    assert review_workflow_sla_hotspots_trends_unauth.status_code == 401

    review_workflow_sla_hotspots_trends_auth = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots/trends",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_sla_hotspots_trends_auth.status_code == 200

    review_workflow_sla_hotspots_trends_export_unauth = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots/trends/export"
    )
    assert review_workflow_sla_hotspots_trends_export_unauth.status_code == 401

    review_workflow_sla_hotspots_trends_export_auth = client.get(
        f"/status/{job_id}/review/workflow/sla/hotspots/trends/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_sla_hotspots_trends_export_auth.status_code == 200

    review_workflow_sla_trends_unauth = client.get(f"/status/{job_id}/review/workflow/sla/trends")
    assert review_workflow_sla_trends_unauth.status_code == 401

    review_workflow_sla_trends_auth = client.get(
        f"/status/{job_id}/review/workflow/sla/trends",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_sla_trends_auth.status_code == 200

    review_workflow_sla_trends_export_unauth = client.get(f"/status/{job_id}/review/workflow/sla/trends/export")
    assert review_workflow_sla_trends_export_unauth.status_code == 401

    review_workflow_sla_trends_export_auth = client.get(
        f"/status/{job_id}/review/workflow/sla/trends/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_sla_trends_export_auth.status_code == 200

    review_workflow_sla_export_unauth = client.get(f"/status/{job_id}/review/workflow/sla/export")
    assert review_workflow_sla_export_unauth.status_code == 401

    review_workflow_sla_export_auth = client.get(
        f"/status/{job_id}/review/workflow/sla/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_sla_export_auth.status_code == 200

    review_workflow_sla_profile_unauth = client.get(f"/status/{job_id}/review/workflow/sla/profile")
    assert review_workflow_sla_profile_unauth.status_code == 401

    review_workflow_sla_profile_auth = client.get(
        f"/status/{job_id}/review/workflow/sla/profile",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_sla_profile_auth.status_code == 200

    review_workflow_sla_recompute_unauth = client.post(f"/status/{job_id}/review/workflow/sla/recompute")
    assert review_workflow_sla_recompute_unauth.status_code == 401

    review_workflow_sla_recompute_auth = client.post(
        f"/status/{job_id}/review/workflow/sla/recompute",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_sla_recompute_auth.status_code == 200

    review_workflow_export_unauth = client.get(f"/status/{job_id}/review/workflow/export")
    assert review_workflow_export_unauth.status_code == 401

    review_workflow_export_auth = client.get(
        f"/status/{job_id}/review/workflow/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_export_auth.status_code == 200

    review_workflow_trends_export_unauth = client.get(f"/status/{job_id}/review/workflow/trends/export")
    assert review_workflow_trends_export_unauth.status_code == 401

    review_workflow_trends_export_auth = client.get(
        f"/status/{job_id}/review/workflow/trends/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert review_workflow_trends_export_auth.status_code == 200

    add_comment_unauth = client.post(
        f"/status/{job_id}/comments",
        json={"section": "general", "message": "Needs management review"},
    )
    assert add_comment_unauth.status_code == 401

    add_comment_auth = client.post(
        f"/status/{job_id}/comments",
        json={"section": "general", "message": "Needs management review"},
        headers={"X-API-Key": "test-secret"},
    )
    assert add_comment_auth.status_code == 200
    comment_id = add_comment_auth.json()["comment_id"]

    resolve_comment_unauth = client.post(f"/status/{job_id}/comments/{comment_id}/resolve")
    assert resolve_comment_unauth.status_code == 401

    resolve_comment_auth = client.post(
        f"/status/{job_id}/comments/{comment_id}/resolve",
        headers={"X-API-Key": "test-secret"},
    )
    assert resolve_comment_auth.status_code == 200

    reopen_comment_unauth = client.post(f"/status/{job_id}/comments/{comment_id}/reopen")
    assert reopen_comment_unauth.status_code == 401

    reopen_comment_auth = client.post(
        f"/status/{job_id}/comments/{comment_id}/reopen",
        headers={"X-API-Key": "test-secret"},
    )
    assert reopen_comment_auth.status_code == 200

    portfolio_metrics_unauth = client.get("/portfolio/metrics")
    assert portfolio_metrics_unauth.status_code == 401

    portfolio_metrics_auth = client.get("/portfolio/metrics", headers={"X-API-Key": "test-secret"})
    assert portfolio_metrics_auth.status_code == 200

    portfolio_metrics_export_unauth = client.get("/portfolio/metrics/export")
    assert portfolio_metrics_export_unauth.status_code == 401

    portfolio_metrics_export_auth = client.get("/portfolio/metrics/export", headers={"X-API-Key": "test-secret"})
    assert portfolio_metrics_export_auth.status_code == 200

    portfolio_quality_unauth = client.get("/portfolio/quality")
    assert portfolio_quality_unauth.status_code == 401

    portfolio_quality_auth = client.get("/portfolio/quality", headers={"X-API-Key": "test-secret"})
    assert portfolio_quality_auth.status_code == 200

    portfolio_quality_export_unauth = client.get("/portfolio/quality/export")
    assert portfolio_quality_export_unauth.status_code == 401

    portfolio_quality_export_auth = client.get("/portfolio/quality/export", headers={"X-API-Key": "test-secret"})
    assert portfolio_quality_export_auth.status_code == 200

    portfolio_review_workflow_unauth = client.get("/portfolio/review-workflow")
    assert portfolio_review_workflow_unauth.status_code == 401

    portfolio_review_workflow_auth = client.get(
        "/portfolio/review-workflow",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_auth.status_code == 200

    portfolio_review_workflow_export_unauth = client.get("/portfolio/review-workflow/export")
    assert portfolio_review_workflow_export_unauth.status_code == 401

    portfolio_review_workflow_export_auth = client.get(
        "/portfolio/review-workflow/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_export_auth.status_code == 200

    portfolio_review_workflow_sla_unauth = client.get("/portfolio/review-workflow/sla")
    assert portfolio_review_workflow_sla_unauth.status_code == 401

    portfolio_review_workflow_sla_auth = client.get(
        "/portfolio/review-workflow/sla",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_sla_auth.status_code == 200

    portfolio_review_workflow_sla_export_unauth = client.get("/portfolio/review-workflow/sla/export")
    assert portfolio_review_workflow_sla_export_unauth.status_code == 401

    portfolio_review_workflow_sla_export_auth = client.get(
        "/portfolio/review-workflow/sla/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_sla_export_auth.status_code == 200

    portfolio_review_workflow_sla_hotspots_unauth = client.get("/portfolio/review-workflow/sla/hotspots")
    assert portfolio_review_workflow_sla_hotspots_unauth.status_code == 401

    portfolio_review_workflow_sla_hotspots_auth = client.get(
        "/portfolio/review-workflow/sla/hotspots",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_sla_hotspots_auth.status_code == 200

    portfolio_review_workflow_sla_hotspots_export_unauth = client.get("/portfolio/review-workflow/sla/hotspots/export")
    assert portfolio_review_workflow_sla_hotspots_export_unauth.status_code == 401

    portfolio_review_workflow_sla_hotspots_export_auth = client.get(
        "/portfolio/review-workflow/sla/hotspots/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_sla_hotspots_export_auth.status_code == 200

    portfolio_review_workflow_sla_hotspots_trends_unauth = client.get("/portfolio/review-workflow/sla/hotspots/trends")
    assert portfolio_review_workflow_sla_hotspots_trends_unauth.status_code == 401

    portfolio_review_workflow_sla_hotspots_trends_auth = client.get(
        "/portfolio/review-workflow/sla/hotspots/trends",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_sla_hotspots_trends_auth.status_code == 200

    portfolio_review_workflow_sla_hotspots_trends_export_unauth = client.get(
        "/portfolio/review-workflow/sla/hotspots/trends/export"
    )
    assert portfolio_review_workflow_sla_hotspots_trends_export_unauth.status_code == 401

    portfolio_review_workflow_sla_hotspots_trends_export_auth = client.get(
        "/portfolio/review-workflow/sla/hotspots/trends/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_sla_hotspots_trends_export_auth.status_code == 200

    portfolio_review_workflow_trends_unauth = client.get("/portfolio/review-workflow/trends")
    assert portfolio_review_workflow_trends_unauth.status_code == 401

    portfolio_review_workflow_trends_auth = client.get(
        "/portfolio/review-workflow/trends",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_trends_auth.status_code == 200

    portfolio_review_workflow_trends_export_unauth = client.get("/portfolio/review-workflow/trends/export")
    assert portfolio_review_workflow_trends_export_unauth.status_code == 401

    portfolio_review_workflow_trends_export_auth = client.get(
        "/portfolio/review-workflow/trends/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_trends_export_auth.status_code == 200

    portfolio_review_workflow_sla_trends_unauth = client.get("/portfolio/review-workflow/sla/trends")
    assert portfolio_review_workflow_sla_trends_unauth.status_code == 401

    portfolio_review_workflow_sla_trends_auth = client.get(
        "/portfolio/review-workflow/sla/trends",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_sla_trends_auth.status_code == 200

    portfolio_review_workflow_sla_trends_export_unauth = client.get("/portfolio/review-workflow/sla/trends/export")
    assert portfolio_review_workflow_sla_trends_export_unauth.status_code == 401

    portfolio_review_workflow_sla_trends_export_auth = client.get(
        "/portfolio/review-workflow/sla/trends/export",
        headers={"X-API-Key": "test-secret"},
    )
    assert portfolio_review_workflow_sla_trends_export_auth.status_code == 200

    ingest_recent_unauth = client.get("/ingest/recent")
    assert ingest_recent_unauth.status_code == 401

    ingest_recent_auth = client.get("/ingest/recent", headers={"X-API-Key": "test-secret"})
    assert ingest_recent_auth.status_code == 200

    ingest_readiness_unauth = client.post("/ingest/readiness", json={"donor_id": "usaid"})
    assert ingest_readiness_unauth.status_code == 401

    ingest_readiness_auth = client.post(
        "/ingest/readiness",
        json={"donor_id": "usaid"},
        headers={"X-API-Key": "test-secret"},
    )
    assert ingest_readiness_auth.status_code == 200

    ingest_inventory_unauth = client.get("/ingest/inventory")
    assert ingest_inventory_unauth.status_code == 401

    ingest_inventory_auth = client.get("/ingest/inventory", headers={"X-API-Key": "test-secret"})
    assert ingest_inventory_auth.status_code == 200

    ingest_inventory_export_unauth = client.get("/ingest/inventory/export")
    assert ingest_inventory_export_unauth.status_code == 401

    ingest_inventory_export_auth = client.get("/ingest/inventory/export", headers={"X-API-Key": "test-secret"})
    assert ingest_inventory_export_auth.status_code == 200

    class _QueueAdminRunnerStub:
        def list_dead_letters(self, limit: int = 50):
            return {
                "mode": "redis_queue",
                "queue_name": "grantflow:jobs",
                "dead_letter_queue_name": "grantflow:jobs:dead",
                "dead_letter_queue_size": 0,
                "items": [],
            }

        def requeue_dead_letters(self, limit: int = 10, reset_attempts: bool = True):
            _ = reset_attempts
            return {
                "mode": "redis_queue",
                "queue_name": "grantflow:jobs",
                "dead_letter_queue_name": "grantflow:jobs:dead",
                "requested_count": int(limit),
                "affected_count": 0,
                "skipped_count": 0,
                "dead_letter_queue_size": 0,
            }

        def purge_dead_letters(self, limit: int = 100):
            return {
                "mode": "redis_queue",
                "queue_name": "grantflow:jobs",
                "dead_letter_queue_name": "grantflow:jobs:dead",
                "requested_count": int(limit),
                "affected_count": 0,
                "dead_letter_queue_size": 0,
            }

    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "redis_queue")
    monkeypatch.setattr(api_app_module, "JOB_RUNNER", _QueueAdminRunnerStub())

    queue_dead_letter_unauth = client.get("/queue/dead-letter")
    assert queue_dead_letter_unauth.status_code == 401

    queue_dead_letter_auth = client.get("/queue/dead-letter", headers={"X-API-Key": "test-secret"})
    assert queue_dead_letter_auth.status_code == 200

    queue_dead_letter_export_unauth = client.get("/queue/dead-letter/export")
    assert queue_dead_letter_export_unauth.status_code == 401

    queue_dead_letter_export_auth = client.get("/queue/dead-letter/export", headers={"X-API-Key": "test-secret"})
    assert queue_dead_letter_export_auth.status_code == 200

    queue_requeue_unauth = client.post("/queue/dead-letter/requeue")
    assert queue_requeue_unauth.status_code == 401

    queue_requeue_auth = client.post("/queue/dead-letter/requeue", headers={"X-API-Key": "test-secret"})
    assert queue_requeue_auth.status_code == 200

    queue_purge_unauth = client.delete("/queue/dead-letter")
    assert queue_purge_unauth.status_code == 401

    queue_purge_auth = client.delete("/queue/dead-letter", headers={"X-API-Key": "test-secret"})
    assert queue_purge_auth.status_code == 200

    pending_unauth = client.get("/hitl/pending")
    assert pending_unauth.status_code == 401

    pending_auth = client.get("/hitl/pending", headers={"X-API-Key": "test-secret"})
    assert pending_auth.status_code == 200


def test_openapi_declares_api_key_security_scheme():
    response = client.get("/openapi.json")
    assert response.status_code == 200
    spec = response.json()

    schemes = (spec.get("components") or {}).get("securitySchemes") or {}
    assert "ApiKeyAuth" in schemes
    assert schemes["ApiKeyAuth"]["type"] == "apiKey"
    assert schemes["ApiKeyAuth"]["in"] == "header"
    assert schemes["ApiKeyAuth"]["name"] == "X-API-Key"

    generate_security = (((spec.get("paths") or {}).get("/generate") or {}).get("post") or {}).get("security")
    generate_preflight_security = (((spec.get("paths") or {}).get("/generate/preflight") or {}).get("post") or {}).get(
        "security"
    )
    ingest_security = (((spec.get("paths") or {}).get("/ingest") or {}).get("post") or {}).get("security")
    ingest_readiness_security = (((spec.get("paths") or {}).get("/ingest/readiness") or {}).get("post") or {}).get(
        "security"
    )
    ingest_recent_security = (((spec.get("paths") or {}).get("/ingest/recent") or {}).get("get") or {}).get("security")
    ingest_inventory_security = (((spec.get("paths") or {}).get("/ingest/inventory") or {}).get("get") or {}).get(
        "security"
    )
    ingest_inventory_export_security = (
        ((spec.get("paths") or {}).get("/ingest/inventory/export") or {}).get("get") or {}
    ).get("security")
    queue_dead_letter_security = (((spec.get("paths") or {}).get("/queue/dead-letter") or {}).get("get") or {}).get(
        "security"
    )
    queue_dead_letter_export_security = (
        ((spec.get("paths") or {}).get("/queue/dead-letter/export") or {}).get("get") or {}
    ).get("security")
    queue_dead_letter_requeue_security = (
        ((spec.get("paths") or {}).get("/queue/dead-letter/requeue") or {}).get("post") or {}
    ).get("security")
    queue_dead_letter_purge_security = (
        ((spec.get("paths") or {}).get("/queue/dead-letter") or {}).get("delete") or {}
    ).get("security")
    cancel_security = (((spec.get("paths") or {}).get("/cancel/{job_id}") or {}).get("post") or {}).get("security")
    status_security = (((spec.get("paths") or {}).get("/status/{job_id}") or {}).get("get") or {}).get("security")
    status_citations_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/citations") or {}).get("get") or {}
    ).get("security")
    status_export_payload_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/export-payload") or {}).get("get") or {}
    ).get("security")
    status_versions_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/versions") or {}).get("get") or {}
    ).get("security")
    status_diff_security = (((spec.get("paths") or {}).get("/status/{job_id}/diff") or {}).get("get") or {}).get(
        "security"
    )
    status_events_security = (((spec.get("paths") or {}).get("/status/{job_id}/events") or {}).get("get") or {}).get(
        "security"
    )
    status_metrics_security = (((spec.get("paths") or {}).get("/status/{job_id}/metrics") or {}).get("get") or {}).get(
        "security"
    )
    status_quality_security = (((spec.get("paths") or {}).get("/status/{job_id}/quality") or {}).get("get") or {}).get(
        "security"
    )
    status_grounding_gate_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/grounding-gate") or {}).get("get") or {}
    ).get("security")
    status_critic_security = (((spec.get("paths") or {}).get("/status/{job_id}/critic") or {}).get("get") or {}).get(
        "security"
    )
    status_critic_findings_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/critic/findings") or {}).get("get") or {}
    ).get("security")
    status_critic_finding_detail_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/critic/findings/{finding_id}") or {}).get("get") or {}
    ).get("security")
    status_critic_finding_ack_security = (
        (((spec.get("paths") or {}).get("/status/{job_id}/critic/findings/{finding_id}/ack") or {}).get("post") or {})
    ).get("security")
    status_critic_finding_open_security = (
        (((spec.get("paths") or {}).get("/status/{job_id}/critic/findings/{finding_id}/open") or {}).get("post") or {})
    ).get("security")
    status_critic_finding_resolve_security = (
        (
            ((spec.get("paths") or {}).get("/status/{job_id}/critic/findings/{finding_id}/resolve") or {}).get("post")
            or {}
        )
    ).get("security")
    status_critic_finding_bulk_status_security = (
        (((spec.get("paths") or {}).get("/status/{job_id}/critic/findings/bulk-status") or {}).get("post") or {})
    ).get("security")
    status_comments_get_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/comments") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_trends_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/trends") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_sla_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_sla_hotspots_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/hotspots") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_sla_hotspots_trends_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/hotspots/trends") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_sla_hotspots_export_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/hotspots/export") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_sla_hotspots_trends_export_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/hotspots/trends/export") or {}).get("get")
        or {}
    ).get("security")
    status_review_workflow_sla_trends_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/trends") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_sla_trends_export_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/trends/export") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_sla_export_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/export") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_sla_profile_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/profile") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_sla_recompute_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/recompute") or {}).get("post") or {}
    ).get("security")
    status_review_workflow_export_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/export") or {}).get("get") or {}
    ).get("security")
    status_review_workflow_trends_export_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/trends/export") or {}).get("get") or {}
    ).get("security")
    status_comments_post_security = (
        ((spec.get("paths") or {}).get("/status/{job_id}/comments") or {}).get("post") or {}
    ).get("security")
    status_comments_resolve_security = (
        (((spec.get("paths") or {}).get("/status/{job_id}/comments/{comment_id}/resolve") or {}).get("post") or {})
    ).get("security")
    status_comments_reopen_security = (
        (((spec.get("paths") or {}).get("/status/{job_id}/comments/{comment_id}/reopen") or {}).get("post") or {})
    ).get("security")
    portfolio_metrics_security = (((spec.get("paths") or {}).get("/portfolio/metrics") or {}).get("get") or {}).get(
        "security"
    )
    portfolio_metrics_export_security = (
        ((spec.get("paths") or {}).get("/portfolio/metrics/export") or {}).get("get") or {}
    ).get("security")
    portfolio_quality_security = (((spec.get("paths") or {}).get("/portfolio/quality") or {}).get("get") or {}).get(
        "security"
    )
    portfolio_quality_export_security = (
        ((spec.get("paths") or {}).get("/portfolio/quality/export") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_export_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/export") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_sla_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_sla_export_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/export") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_sla_hotspots_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/hotspots") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_sla_hotspots_export_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/hotspots/export") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_sla_hotspots_trends_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/hotspots/trends") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_sla_hotspots_trends_export_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/hotspots/trends/export") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_trends_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/trends") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_trends_export_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/trends/export") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_sla_trends_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/trends") or {}).get("get") or {}
    ).get("security")
    portfolio_review_workflow_sla_trends_export_security = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/trends/export") or {}).get("get") or {}
    ).get("security")
    generate_preflight_response_schema = (
        ((((spec.get("paths") or {}).get("/generate/preflight") or {}).get("post") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    ingest_readiness_response_schema = (
        ((((spec.get("paths") or {}).get("/ingest/readiness") or {}).get("post") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_response_schema = (
        ((((spec.get("paths") or {}).get("/status/{job_id}") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_citations_response_schema = (
        ((((spec.get("paths") or {}).get("/status/{job_id}/citations") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_export_payload_response_schema = (
        (
            (((spec.get("paths") or {}).get("/status/{job_id}/export-payload") or {}).get("get") or {}).get("responses")
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_versions_response_schema = (
        ((((spec.get("paths") or {}).get("/status/{job_id}/versions") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_diff_response_schema = (
        ((((spec.get("paths") or {}).get("/status/{job_id}/diff") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_events_response_schema = (
        ((((spec.get("paths") or {}).get("/status/{job_id}/events") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_metrics_response_schema = (
        ((((spec.get("paths") or {}).get("/status/{job_id}/metrics") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_quality_response_schema = (
        ((((spec.get("paths") or {}).get("/status/{job_id}/quality") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_grounding_gate_response_schema = (
        (
            (((spec.get("paths") or {}).get("/status/{job_id}/grounding-gate") or {}).get("get") or {}).get("responses")
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_critic_response_schema = (
        ((((spec.get("paths") or {}).get("/status/{job_id}/critic") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_critic_findings_response_schema = (
        (
            (((spec.get("paths") or {}).get("/status/{job_id}/critic/findings") or {}).get("get") or {}).get(
                "responses"
            )
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_critic_finding_detail_response_schema = (
        (
            (
                ((spec.get("paths") or {}).get("/status/{job_id}/critic/findings/{finding_id}") or {}).get("get") or {}
            ).get("responses")
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_critic_finding_ack_response_schema = (
        (
            (
                ((spec.get("paths") or {}).get("/status/{job_id}/critic/findings/{finding_id}/ack") or {}).get("post")
                or {}
            ).get("responses")
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_critic_finding_open_response_schema = (
        (
            (
                ((spec.get("paths") or {}).get("/status/{job_id}/critic/findings/{finding_id}/open") or {}).get("post")
                or {}
            ).get("responses")
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_critic_finding_resolve_response_schema = (
        (
            (
                (
                    ((spec.get("paths") or {}).get("/status/{job_id}/critic/findings/{finding_id}/resolve") or {}).get(
                        "post"
                    )
                    or {}
                ).get("responses")
            )
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_critic_finding_bulk_status_response_schema = (
        (
            (
                ((spec.get("paths") or {}).get("/status/{job_id}/critic/findings/bulk-status") or {}).get("post") or {}
            ).get("responses")
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_comments_response_schema = (
        ((((spec.get("paths") or {}).get("/status/{job_id}/comments") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_review_workflow_response_schema = (
        (((spec.get("paths") or {}).get("/status/{job_id}/review/workflow") or {}).get("get") or {})
        .get("responses", {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_review_workflow_trends_response_schema = (
        (((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/trends") or {}).get("get") or {})
        .get("responses", {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_review_workflow_sla_response_schema = (
        (((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla") or {}).get("get") or {})
        .get("responses", {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_review_workflow_sla_hotspots_response_schema = (
        (((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/hotspots") or {}).get("get") or {})
        .get("responses", {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_review_workflow_sla_trends_response_schema = (
        (((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/trends") or {}).get("get") or {})
        .get("responses", {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_review_workflow_sla_hotspots_trends_response_schema = (
        (((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/hotspots/trends") or {}).get("get") or {})
        .get("responses", {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_review_workflow_sla_profile_response_schema = (
        (((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/profile") or {}).get("get") or {})
        .get("responses", {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_review_workflow_sla_recompute_response_schema = (
        (((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/recompute") or {}).get("post") or {})
        .get("responses", {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_comments_post_response_schema = (
        ((((spec.get("paths") or {}).get("/status/{job_id}/comments") or {}).get("post") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_comments_resolve_response_schema = (
        (
            (
                (
                    ((spec.get("paths") or {}).get("/status/{job_id}/comments/{comment_id}/resolve") or {}).get("post")
                    or {}
                ).get("responses")
            )
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    status_comments_reopen_response_schema = (
        (
            (
                (
                    ((spec.get("paths") or {}).get("/status/{job_id}/comments/{comment_id}/reopen") or {}).get("post")
                    or {}
                ).get("responses")
            )
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    portfolio_metrics_response_schema = (
        ((((spec.get("paths") or {}).get("/portfolio/metrics") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    portfolio_quality_response_schema = (
        ((((spec.get("paths") or {}).get("/portfolio/quality") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    portfolio_review_workflow_response_schema = (
        ((((spec.get("paths") or {}).get("/portfolio/review-workflow") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    portfolio_review_workflow_sla_response_schema = (
        (
            (((spec.get("paths") or {}).get("/portfolio/review-workflow/sla") or {}).get("get") or {}).get("responses")
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    portfolio_review_workflow_sla_hotspots_response_schema = (
        (
            (((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/hotspots") or {}).get("get") or {}).get(
                "responses"
            )
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    portfolio_review_workflow_sla_hotspots_trends_response_schema = (
        (
            (
                ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/hotspots/trends") or {}).get("get") or {}
            ).get("responses")
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    portfolio_review_workflow_trends_response_schema = (
        (
            (((spec.get("paths") or {}).get("/portfolio/review-workflow/trends") or {}).get("get") or {}).get(
                "responses"
            )
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    portfolio_review_workflow_sla_trends_response_schema = (
        (
            (((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/trends") or {}).get("get") or {}).get(
                "responses"
            )
            or {}
        )
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    pending_response_schema = (
        ((((spec.get("paths") or {}).get("/hitl/pending") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    ingest_recent_response_schema = (
        ((((spec.get("paths") or {}).get("/ingest/recent") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    ingest_inventory_response_schema = (
        ((((spec.get("paths") or {}).get("/ingest/inventory") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    queue_dead_letter_response_schema = (
        ((((spec.get("paths") or {}).get("/queue/dead-letter") or {}).get("get") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    queue_dead_letter_requeue_response_schema = (
        ((((spec.get("paths") or {}).get("/queue/dead-letter/requeue") or {}).get("post") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    queue_dead_letter_purge_response_schema = (
        ((((spec.get("paths") or {}).get("/queue/dead-letter") or {}).get("delete") or {}).get("responses") or {})
        .get("200", {})
        .get("content", {})
        .get("application/json", {})
        .get("schema")
    )
    assert generate_security == [{"ApiKeyAuth": []}]
    assert generate_preflight_security == [{"ApiKeyAuth": []}]
    assert ingest_security == [{"ApiKeyAuth": []}]
    assert ingest_readiness_security == [{"ApiKeyAuth": []}]
    assert ingest_recent_security == [{"ApiKeyAuth": []}]
    assert ingest_inventory_security == [{"ApiKeyAuth": []}]
    assert ingest_inventory_export_security == [{"ApiKeyAuth": []}]
    assert queue_dead_letter_security == [{"ApiKeyAuth": []}]
    assert queue_dead_letter_export_security == [{"ApiKeyAuth": []}]
    assert queue_dead_letter_requeue_security == [{"ApiKeyAuth": []}]
    assert queue_dead_letter_purge_security == [{"ApiKeyAuth": []}]
    assert cancel_security == [{"ApiKeyAuth": []}]
    assert status_security == [{"ApiKeyAuth": []}]
    assert status_citations_security == [{"ApiKeyAuth": []}]
    assert status_export_payload_security == [{"ApiKeyAuth": []}]
    assert status_versions_security == [{"ApiKeyAuth": []}]
    assert status_diff_security == [{"ApiKeyAuth": []}]
    assert status_events_security == [{"ApiKeyAuth": []}]
    assert status_metrics_security == [{"ApiKeyAuth": []}]
    assert status_quality_security == [{"ApiKeyAuth": []}]
    assert status_grounding_gate_security == [{"ApiKeyAuth": []}]
    assert status_critic_security == [{"ApiKeyAuth": []}]
    assert status_critic_findings_security == [{"ApiKeyAuth": []}]
    assert status_critic_finding_detail_security == [{"ApiKeyAuth": []}]
    assert status_critic_finding_ack_security == [{"ApiKeyAuth": []}]
    assert status_critic_finding_open_security == [{"ApiKeyAuth": []}]
    assert status_critic_finding_resolve_security == [{"ApiKeyAuth": []}]
    assert status_critic_finding_bulk_status_security == [{"ApiKeyAuth": []}]
    assert status_comments_get_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_trends_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_sla_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_sla_hotspots_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_sla_hotspots_trends_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_sla_hotspots_export_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_sla_hotspots_trends_export_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_sla_trends_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_sla_trends_export_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_sla_export_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_sla_profile_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_sla_recompute_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_export_security == [{"ApiKeyAuth": []}]
    assert status_review_workflow_trends_export_security == [{"ApiKeyAuth": []}]
    assert status_comments_post_security == [{"ApiKeyAuth": []}]
    assert status_comments_resolve_security == [{"ApiKeyAuth": []}]
    assert status_comments_reopen_security == [{"ApiKeyAuth": []}]
    assert portfolio_metrics_security == [{"ApiKeyAuth": []}]
    assert portfolio_metrics_export_security == [{"ApiKeyAuth": []}]
    assert portfolio_quality_security == [{"ApiKeyAuth": []}]
    assert portfolio_quality_export_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_export_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_sla_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_sla_export_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_sla_hotspots_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_sla_hotspots_export_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_sla_hotspots_trends_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_sla_hotspots_trends_export_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_trends_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_trends_export_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_sla_trends_security == [{"ApiKeyAuth": []}]
    assert portfolio_review_workflow_sla_trends_export_security == [{"ApiKeyAuth": []}]
    assert generate_preflight_response_schema == {"$ref": "#/components/schemas/GeneratePreflightPublicResponse"}
    assert ingest_readiness_response_schema == {"$ref": "#/components/schemas/GeneratePreflightPublicResponse"}
    assert status_response_schema == {"$ref": "#/components/schemas/JobStatusPublicResponse"}
    assert status_citations_response_schema == {"$ref": "#/components/schemas/JobCitationsPublicResponse"}
    assert status_export_payload_response_schema == {"$ref": "#/components/schemas/JobExportPayloadPublicResponse"}
    assert status_versions_response_schema == {"$ref": "#/components/schemas/JobVersionsPublicResponse"}
    assert status_diff_response_schema == {"$ref": "#/components/schemas/JobDiffPublicResponse"}
    assert status_events_response_schema == {"$ref": "#/components/schemas/JobEventsPublicResponse"}
    assert status_metrics_response_schema == {"$ref": "#/components/schemas/JobMetricsPublicResponse"}
    assert status_quality_response_schema == {"$ref": "#/components/schemas/JobQualitySummaryPublicResponse"}
    assert status_grounding_gate_response_schema == {"$ref": "#/components/schemas/JobGroundingGatePublicResponse"}
    assert status_critic_response_schema == {"$ref": "#/components/schemas/JobCriticPublicResponse"}
    assert status_critic_findings_response_schema == {"$ref": "#/components/schemas/CriticFindingsListPublicResponse"}
    assert status_critic_finding_detail_response_schema == {
        "$ref": "#/components/schemas/CriticFatalFlawPublicResponse"
    }
    assert status_critic_finding_ack_response_schema == {
        "$ref": "#/components/schemas/CriticFatalFlawStatusUpdatePublicResponse"
    }
    assert status_critic_finding_open_response_schema == {
        "$ref": "#/components/schemas/CriticFatalFlawStatusUpdatePublicResponse"
    }
    assert status_critic_finding_resolve_response_schema == {
        "$ref": "#/components/schemas/CriticFatalFlawStatusUpdatePublicResponse"
    }
    assert status_critic_finding_bulk_status_response_schema == {
        "$ref": "#/components/schemas/CriticFindingsBulkStatusPublicResponse"
    }
    assert status_comments_response_schema == {"$ref": "#/components/schemas/JobCommentsPublicResponse"}
    assert status_review_workflow_response_schema == {"$ref": "#/components/schemas/JobReviewWorkflowPublicResponse"}
    assert status_review_workflow_trends_response_schema == {
        "$ref": "#/components/schemas/JobReviewWorkflowTrendsPublicResponse"
    }
    assert status_review_workflow_sla_response_schema == {
        "$ref": "#/components/schemas/JobReviewWorkflowSLAPublicResponse"
    }
    assert status_review_workflow_sla_hotspots_response_schema == {
        "$ref": "#/components/schemas/JobReviewWorkflowSLAHotspotsPublicResponse"
    }
    assert status_review_workflow_sla_trends_response_schema == {
        "$ref": "#/components/schemas/JobReviewWorkflowSLATrendsPublicResponse"
    }
    assert status_review_workflow_sla_hotspots_trends_response_schema == {
        "$ref": "#/components/schemas/JobReviewWorkflowSLAHotspotsTrendsPublicResponse"
    }
    assert status_review_workflow_sla_profile_response_schema == {
        "$ref": "#/components/schemas/JobReviewWorkflowSLAProfilePublicResponse"
    }
    assert status_review_workflow_sla_recompute_response_schema == {
        "$ref": "#/components/schemas/JobReviewWorkflowSLARecomputePublicResponse"
    }
    assert status_comments_post_response_schema == {"$ref": "#/components/schemas/ReviewCommentPublicResponse"}
    assert status_comments_resolve_response_schema == {"$ref": "#/components/schemas/ReviewCommentPublicResponse"}
    assert status_comments_reopen_response_schema == {"$ref": "#/components/schemas/ReviewCommentPublicResponse"}
    assert portfolio_metrics_response_schema == {"$ref": "#/components/schemas/PortfolioMetricsPublicResponse"}
    assert portfolio_quality_response_schema == {"$ref": "#/components/schemas/PortfolioQualityPublicResponse"}
    assert portfolio_review_workflow_response_schema == {
        "$ref": "#/components/schemas/PortfolioReviewWorkflowPublicResponse"
    }
    assert portfolio_review_workflow_sla_response_schema == {
        "$ref": "#/components/schemas/PortfolioReviewWorkflowSLAPublicResponse"
    }
    assert portfolio_review_workflow_sla_hotspots_response_schema == {
        "$ref": "#/components/schemas/PortfolioReviewWorkflowSLAHotspotsPublicResponse"
    }
    assert portfolio_review_workflow_sla_hotspots_trends_response_schema == {
        "$ref": "#/components/schemas/PortfolioReviewWorkflowSLAHotspotsTrendsPublicResponse"
    }
    assert portfolio_review_workflow_trends_response_schema == {
        "$ref": "#/components/schemas/PortfolioReviewWorkflowTrendsPublicResponse"
    }
    assert portfolio_review_workflow_sla_trends_response_schema == {
        "$ref": "#/components/schemas/PortfolioReviewWorkflowSLATrendsPublicResponse"
    }
    assert pending_response_schema == {"$ref": "#/components/schemas/HITLPendingListPublicResponse"}
    assert ingest_recent_response_schema == {"$ref": "#/components/schemas/IngestRecentListPublicResponse"}
    assert ingest_inventory_response_schema == {"$ref": "#/components/schemas/IngestInventoryPublicResponse"}
    assert queue_dead_letter_response_schema == {"$ref": "#/components/schemas/DeadLetterQueueListPublicResponse"}
    assert queue_dead_letter_requeue_response_schema == {
        "$ref": "#/components/schemas/DeadLetterQueueMutationPublicResponse"
    }
    assert queue_dead_letter_purge_response_schema == {
        "$ref": "#/components/schemas/DeadLetterQueueMutationPublicResponse"
    }

    schemas = (spec.get("components") or {}).get("schemas") or {}
    assert "JobStatusPublicResponse" in schemas
    assert "JobCitationsPublicResponse" in schemas
    assert "JobExportPayloadPublicResponse" in schemas
    assert "CitationPublicResponse" in schemas
    assert "JobVersionsPublicResponse" in schemas
    assert "DraftVersionPublicResponse" in schemas
    assert "JobDiffPublicResponse" in schemas
    assert "JobEventsPublicResponse" in schemas
    assert "JobEventPublicResponse" in schemas
    assert "JobMetricsPublicResponse" in schemas
    assert "JobQualitySummaryPublicResponse" in schemas
    assert "JobQualityToCTextQualitySummaryPublicResponse" in schemas
    assert "JobCriticPublicResponse" in schemas
    assert "CriticFindingsListPublicResponse" in schemas
    assert "CriticFindingsListFiltersPublicResponse" in schemas
    assert "CriticFindingsListSummaryPublicResponse" in schemas
    assert "JobReviewWorkflowPublicResponse" in schemas
    assert "JobReviewWorkflowTrendsPublicResponse" in schemas
    assert "JobReviewWorkflowTrendPointPublicResponse" in schemas
    assert "JobReviewWorkflowSLAPublicResponse" in schemas
    assert "JobReviewWorkflowSLAHotspotsPublicResponse" in schemas
    assert "JobReviewWorkflowSLAHotspotsFiltersPublicResponse" in schemas
    assert "JobReviewWorkflowSLAHotspotsTrendsPublicResponse" in schemas
    assert "JobReviewWorkflowSLAHotspotsTrendsFiltersPublicResponse" in schemas
    assert "JobReviewWorkflowSLATrendsPublicResponse" in schemas
    assert "JobReviewWorkflowSLATrendPointPublicResponse" in schemas
    assert "JobReviewWorkflowSLAFiltersPublicResponse" in schemas
    assert "JobReviewWorkflowSLAProfilePublicResponse" in schemas
    assert "JobReviewWorkflowSLARecomputePublicResponse" in schemas
    assert "JobReviewWorkflowSLAItemPublicResponse" in schemas
    assert "CriticFindingsBulkStatusPublicResponse" in schemas
    assert "CriticFindingsBulkStatusFiltersPublicResponse" in schemas
    assert "JobReviewWorkflowFiltersPublicResponse" in schemas
    assert "JobReviewWorkflowSummaryPublicResponse" in schemas
    assert "ReviewWorkflowTimelineEventPublicResponse" in schemas
    assert "CriticRuleCheckPublicResponse" in schemas
    assert "CriticFatalFlawPublicResponse" in schemas
    assert "CriticFatalFlawStatusUpdatePublicResponse" in schemas
    assert "JobCommentsPublicResponse" in schemas
    assert "ReviewCommentPublicResponse" in schemas
    assert "PortfolioMetricsPublicResponse" in schemas
    assert "PortfolioQualityPublicResponse" in schemas
    assert "PortfolioReviewWorkflowPublicResponse" in schemas
    assert "PortfolioReviewWorkflowFiltersPublicResponse" in schemas
    assert "PortfolioReviewWorkflowTimelineEventPublicResponse" in schemas
    assert "PortfolioReviewWorkflowSLAPublicResponse" in schemas
    assert "PortfolioReviewWorkflowSLAFiltersPublicResponse" in schemas
    assert "PortfolioReviewWorkflowSLAItemPublicResponse" in schemas
    assert "PortfolioReviewWorkflowSLAHotspotsPublicResponse" in schemas
    assert "PortfolioReviewWorkflowSLAHotspotsFiltersPublicResponse" in schemas
    assert "PortfolioReviewWorkflowSLAHotspotsTrendsPublicResponse" in schemas
    assert "PortfolioReviewWorkflowSLAHotspotsTrendsFiltersPublicResponse" in schemas
    assert "PortfolioReviewWorkflowTrendsPublicResponse" in schemas
    assert "PortfolioReviewWorkflowTrendsFiltersPublicResponse" in schemas
    assert "PortfolioReviewWorkflowSLATrendsPublicResponse" in schemas
    assert "PortfolioReviewWorkflowSLATrendsFiltersPublicResponse" in schemas
    assert "PortfolioQualityWeightedSignalPublicResponse" in schemas
    assert "PortfolioQualityDonorWeightedRiskPublicResponse" in schemas
    assert "PortfolioQualityCriticSummaryPublicResponse" in schemas
    assert "PortfolioQualityCitationSummaryPublicResponse" in schemas
    assert "PortfolioQualityToCTextQualitySummaryPublicResponse" in schemas
    assert "PortfolioMetricsFiltersPublicResponse" in schemas
    assert "HITLPendingListPublicResponse" in schemas
    assert "IngestRecentListPublicResponse" in schemas
    assert "IngestRecentRecordPublicResponse" in schemas
    assert "IngestInventoryPublicResponse" in schemas
    assert "IngestInventoryDocFamilyPublicResponse" in schemas
    assert "GeneratePreflightPublicResponse" in schemas
    assert "GeneratePreflightWarningPublicResponse" in schemas
    assert "GeneratePreflightArchitectClaimsPublicResponse" in schemas
    assert "GeneratePreflightGroundingPolicyPublicResponse" in schemas
    portfolio_quality_params = (((spec.get("paths") or {}).get("/portfolio/quality") or {}).get("get") or {}).get(
        "parameters"
    ) or []
    portfolio_metrics_params = (((spec.get("paths") or {}).get("/portfolio/metrics") or {}).get("get") or {}).get(
        "parameters"
    ) or []
    portfolio_metrics_export_params = (
        ((spec.get("paths") or {}).get("/portfolio/metrics/export") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_quality_export_params = (
        ((spec.get("paths") or {}).get("/portfolio/quality/export") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_export_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/export") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_sla_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_sla_export_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/export") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_sla_hotspots_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/hotspots") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_sla_hotspots_export_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/hotspots/export") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_sla_hotspots_trends_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/hotspots/trends") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_sla_hotspots_trends_export_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/hotspots/trends/export") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_trends_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/trends") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_trends_export_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/trends/export") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_sla_trends_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/trends") or {}).get("get") or {}
    ).get("parameters") or []
    portfolio_review_workflow_sla_trends_export_params = (
        ((spec.get("paths") or {}).get("/portfolio/review-workflow/sla/trends/export") or {}).get("get") or {}
    ).get("parameters") or []
    review_workflow_params = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow") or {}).get("get") or {}
    ).get("parameters") or []
    review_workflow_trends_params = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/trends") or {}).get("get") or {}
    ).get("parameters") or []
    review_workflow_trends_export_params = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/trends/export") or {}).get("get") or {}
    ).get("parameters") or []
    review_workflow_sla_params = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla") or {}).get("get") or {}
    ).get("parameters") or []
    review_workflow_sla_hotspots_params = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/hotspots") or {}).get("get") or {}
    ).get("parameters") or []
    review_workflow_sla_hotspots_export_params = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/hotspots/export") or {}).get("get") or {}
    ).get("parameters") or []
    review_workflow_sla_hotspots_trends_params = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/hotspots/trends") or {}).get("get") or {}
    ).get("parameters") or []
    review_workflow_sla_hotspots_trends_export_params = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/hotspots/trends/export") or {}).get("get")
        or {}
    ).get("parameters") or []
    review_workflow_sla_trends_params = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/trends") or {}).get("get") or {}
    ).get("parameters") or []
    review_workflow_sla_trends_export_params = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/trends/export") or {}).get("get") or {}
    ).get("parameters") or []
    review_workflow_sla_export_params = (
        ((spec.get("paths") or {}).get("/status/{job_id}/review/workflow/sla/export") or {}).get("get") or {}
    ).get("parameters") or []
    assert "toc_text_risk_level" in [str(p.get("name") or "") for p in portfolio_metrics_params if isinstance(p, dict)]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "") for p in portfolio_metrics_export_params if isinstance(p, dict)
    ]
    assert "toc_text_risk_level" in [str(p.get("name") or "") for p in portfolio_quality_params if isinstance(p, dict)]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "") for p in portfolio_quality_export_params if isinstance(p, dict)
    ]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "") for p in portfolio_review_workflow_params if isinstance(p, dict)
    ]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "") for p in portfolio_review_workflow_export_params if isinstance(p, dict)
    ]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_params if isinstance(p, dict)
    ]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_export_params if isinstance(p, dict)
    ]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_hotspots_params if isinstance(p, dict)
    ]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_hotspots_export_params if isinstance(p, dict)
    ]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_hotspots_trends_params if isinstance(p, dict)
    ]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "")
        for p in portfolio_review_workflow_sla_hotspots_trends_export_params
        if isinstance(p, dict)
    ]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "") for p in portfolio_review_workflow_trends_params if isinstance(p, dict)
    ]
    assert "toc_text_risk_level" in [
        str(p.get("name") or "") for p in portfolio_review_workflow_trends_export_params if isinstance(p, dict)
    ]
    review_workflow_param_names = [str(p.get("name") or "") for p in review_workflow_params if isinstance(p, dict)]
    review_workflow_trends_param_names = [
        str(p.get("name") or "") for p in review_workflow_trends_params if isinstance(p, dict)
    ]
    review_workflow_trends_export_param_names = [
        str(p.get("name") or "") for p in review_workflow_trends_export_params if isinstance(p, dict)
    ]
    review_workflow_sla_param_names = [
        str(p.get("name") or "") for p in review_workflow_sla_params if isinstance(p, dict)
    ]
    review_workflow_sla_hotspots_param_names = [
        str(p.get("name") or "") for p in review_workflow_sla_hotspots_params if isinstance(p, dict)
    ]
    review_workflow_sla_hotspots_export_param_names = [
        str(p.get("name") or "") for p in review_workflow_sla_hotspots_export_params if isinstance(p, dict)
    ]
    review_workflow_sla_hotspots_trends_param_names = [
        str(p.get("name") or "") for p in review_workflow_sla_hotspots_trends_params if isinstance(p, dict)
    ]
    review_workflow_sla_hotspots_trends_export_param_names = [
        str(p.get("name") or "") for p in review_workflow_sla_hotspots_trends_export_params if isinstance(p, dict)
    ]
    review_workflow_sla_trends_param_names = [
        str(p.get("name") or "") for p in review_workflow_sla_trends_params if isinstance(p, dict)
    ]
    review_workflow_sla_trends_export_param_names = [
        str(p.get("name") or "") for p in review_workflow_sla_trends_export_params if isinstance(p, dict)
    ]
    review_workflow_sla_export_param_names = [
        str(p.get("name") or "") for p in review_workflow_sla_export_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_trends_param_names = [
        str(p.get("name") or "") for p in portfolio_review_workflow_trends_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_trends_export_param_names = [
        str(p.get("name") or "") for p in portfolio_review_workflow_trends_export_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_sla_trends_param_names = [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_trends_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_sla_trends_export_param_names = [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_trends_export_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_param_names = [
        str(p.get("name") or "") for p in portfolio_review_workflow_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_export_param_names = [
        str(p.get("name") or "") for p in portfolio_review_workflow_export_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_sla_param_names = [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_sla_export_param_names = [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_export_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_sla_hotspots_param_names = [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_hotspots_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_sla_hotspots_export_param_names = [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_hotspots_export_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_sla_hotspots_trends_param_names = [
        str(p.get("name") or "") for p in portfolio_review_workflow_sla_hotspots_trends_params if isinstance(p, dict)
    ]
    portfolio_review_workflow_sla_hotspots_trends_export_param_names = [
        str(p.get("name") or "")
        for p in portfolio_review_workflow_sla_hotspots_trends_export_params
        if isinstance(p, dict)
    ]
    assert "event_type" in review_workflow_param_names
    assert "event_type" in review_workflow_trends_param_names
    assert "event_type" in review_workflow_trends_export_param_names
    for name in (
        "donor_id",
        "status",
        "hitl_enabled",
        "warning_level",
        "grounding_risk_level",
        "toc_text_risk_level",
        "event_type",
    ):
        assert name in portfolio_review_workflow_param_names
        assert name in portfolio_review_workflow_export_param_names
        assert name in portfolio_review_workflow_trends_param_names
        assert name in portfolio_review_workflow_trends_export_param_names
    for name in (
        "donor_id",
        "status",
        "hitl_enabled",
        "warning_level",
        "grounding_risk_level",
        "toc_text_risk_level",
    ):
        assert name in portfolio_review_workflow_sla_param_names
        assert name in portfolio_review_workflow_sla_export_param_names
        assert name in portfolio_review_workflow_sla_hotspots_param_names
        assert name in portfolio_review_workflow_sla_hotspots_export_param_names
        assert name in portfolio_review_workflow_sla_hotspots_trends_param_names
        assert name in portfolio_review_workflow_sla_hotspots_trends_export_param_names
        assert name in portfolio_review_workflow_sla_trends_param_names
        assert name in portfolio_review_workflow_sla_trends_export_param_names
    for name in (
        "finding_id",
        "finding_code",
        "finding_section",
        "comment_status",
        "workflow_state",
        "overdue_after_hours",
    ):
        assert name in portfolio_review_workflow_param_names
        assert name in portfolio_review_workflow_export_param_names
        assert name in portfolio_review_workflow_sla_param_names
        assert name in portfolio_review_workflow_sla_export_param_names
        assert name in portfolio_review_workflow_sla_hotspots_param_names
        assert name in portfolio_review_workflow_sla_hotspots_export_param_names
        assert name in portfolio_review_workflow_sla_hotspots_trends_param_names
        assert name in portfolio_review_workflow_sla_hotspots_trends_export_param_names
        assert name in portfolio_review_workflow_trends_param_names
        assert name in portfolio_review_workflow_trends_export_param_names
        assert name in review_workflow_param_names
        assert name in review_workflow_trends_param_names
        assert name in review_workflow_trends_export_param_names
        assert name in review_workflow_sla_param_names
        assert name in review_workflow_sla_hotspots_param_names
        assert name in review_workflow_sla_hotspots_export_param_names
        assert name in review_workflow_sla_hotspots_trends_param_names
        assert name in review_workflow_sla_hotspots_trends_export_param_names
        assert name in review_workflow_sla_trends_param_names
        assert name in review_workflow_sla_trends_export_param_names
        assert name in review_workflow_sla_export_param_names
        assert name in portfolio_review_workflow_sla_trends_param_names
        assert name in portfolio_review_workflow_sla_trends_export_param_names
    assert "top_limit" in review_workflow_sla_hotspots_param_names
    assert "top_limit" in review_workflow_sla_hotspots_export_param_names
    assert "top_limit" in review_workflow_sla_hotspots_trends_param_names
    assert "top_limit" in review_workflow_sla_hotspots_trends_export_param_names
    assert "hotspot_kind" in review_workflow_sla_hotspots_param_names
    assert "hotspot_kind" in review_workflow_sla_hotspots_export_param_names
    assert "hotspot_kind" in review_workflow_sla_hotspots_trends_param_names
    assert "hotspot_kind" in review_workflow_sla_hotspots_trends_export_param_names
    assert "hotspot_severity" in review_workflow_sla_hotspots_param_names
    assert "hotspot_severity" in review_workflow_sla_hotspots_export_param_names
    assert "hotspot_severity" in review_workflow_sla_hotspots_trends_param_names
    assert "hotspot_severity" in review_workflow_sla_hotspots_trends_export_param_names
    assert "min_overdue_hours" in review_workflow_sla_hotspots_param_names
    assert "min_overdue_hours" in review_workflow_sla_hotspots_export_param_names
    assert "min_overdue_hours" in review_workflow_sla_hotspots_trends_param_names
    assert "min_overdue_hours" in review_workflow_sla_hotspots_trends_export_param_names
    assert "top_limit" in portfolio_review_workflow_sla_param_names
    assert "top_limit" in portfolio_review_workflow_sla_export_param_names
    assert "top_limit" in portfolio_review_workflow_sla_hotspots_param_names
    assert "top_limit" in portfolio_review_workflow_sla_hotspots_export_param_names
    assert "top_limit" in portfolio_review_workflow_sla_hotspots_trends_param_names
    assert "top_limit" in portfolio_review_workflow_sla_hotspots_trends_export_param_names
    assert "hotspot_kind" in portfolio_review_workflow_sla_hotspots_param_names
    assert "hotspot_kind" in portfolio_review_workflow_sla_hotspots_export_param_names
    assert "hotspot_kind" in portfolio_review_workflow_sla_hotspots_trends_param_names
    assert "hotspot_kind" in portfolio_review_workflow_sla_hotspots_trends_export_param_names
    assert "hotspot_severity" in portfolio_review_workflow_sla_hotspots_param_names
    assert "hotspot_severity" in portfolio_review_workflow_sla_hotspots_export_param_names
    assert "hotspot_severity" in portfolio_review_workflow_sla_hotspots_trends_param_names
    assert "hotspot_severity" in portfolio_review_workflow_sla_hotspots_trends_export_param_names
    assert "min_overdue_hours" in portfolio_review_workflow_sla_hotspots_param_names
    assert "min_overdue_hours" in portfolio_review_workflow_sla_hotspots_export_param_names
    assert "min_overdue_hours" in portfolio_review_workflow_sla_hotspots_trends_param_names
    assert "min_overdue_hours" in portfolio_review_workflow_sla_hotspots_trends_export_param_names
    assert "format" in portfolio_review_workflow_trends_export_param_names
    assert "gzip" in portfolio_review_workflow_trends_export_param_names
    assert "format" in portfolio_review_workflow_export_param_names
    assert "gzip" in portfolio_review_workflow_export_param_names
    assert "format" in portfolio_review_workflow_sla_export_param_names
    assert "gzip" in portfolio_review_workflow_sla_export_param_names
    assert "format" in portfolio_review_workflow_sla_hotspots_export_param_names
    assert "gzip" in portfolio_review_workflow_sla_hotspots_export_param_names
    assert "format" in portfolio_review_workflow_sla_hotspots_trends_export_param_names
    assert "gzip" in portfolio_review_workflow_sla_hotspots_trends_export_param_names
    assert "format" in portfolio_review_workflow_sla_trends_export_param_names
    assert "gzip" in portfolio_review_workflow_sla_trends_export_param_names
    assert "format" in review_workflow_trends_export_param_names
    assert "gzip" in review_workflow_trends_export_param_names
    assert "format" in review_workflow_sla_trends_export_param_names
    assert "gzip" in review_workflow_sla_trends_export_param_names
    assert "format" in review_workflow_sla_export_param_names
    assert "gzip" in review_workflow_sla_export_param_names
    assert "format" in review_workflow_sla_hotspots_export_param_names
    assert "gzip" in review_workflow_sla_hotspots_export_param_names
    assert "format" in review_workflow_sla_hotspots_trends_export_param_names
    assert "gzip" in review_workflow_sla_hotspots_trends_export_param_names
    portfolio_filters_schema_props = (
        ((schemas.get("PortfolioMetricsFiltersPublicResponse") or {}).get("properties") or {})
        if isinstance(schemas.get("PortfolioMetricsFiltersPublicResponse"), dict)
        else {}
    )
    portfolio_review_workflow_trends_filters_schema_props = (
        ((schemas.get("PortfolioReviewWorkflowTrendsFiltersPublicResponse") or {}).get("properties") or {})
        if isinstance(schemas.get("PortfolioReviewWorkflowTrendsFiltersPublicResponse"), dict)
        else {}
    )
    portfolio_review_workflow_filters_schema_props = (
        ((schemas.get("PortfolioReviewWorkflowFiltersPublicResponse") or {}).get("properties") or {})
        if isinstance(schemas.get("PortfolioReviewWorkflowFiltersPublicResponse"), dict)
        else {}
    )
    portfolio_review_workflow_sla_filters_schema_props = (
        ((schemas.get("PortfolioReviewWorkflowSLAFiltersPublicResponse") or {}).get("properties") or {})
        if isinstance(schemas.get("PortfolioReviewWorkflowSLAFiltersPublicResponse"), dict)
        else {}
    )
    portfolio_review_workflow_sla_hotspots_filters_schema_props = (
        ((schemas.get("PortfolioReviewWorkflowSLAHotspotsFiltersPublicResponse") or {}).get("properties") or {})
        if isinstance(schemas.get("PortfolioReviewWorkflowSLAHotspotsFiltersPublicResponse"), dict)
        else {}
    )
    portfolio_review_workflow_sla_hotspots_trends_filters_schema_props = (
        ((schemas.get("PortfolioReviewWorkflowSLAHotspotsTrendsFiltersPublicResponse") or {}).get("properties") or {})
        if isinstance(schemas.get("PortfolioReviewWorkflowSLAHotspotsTrendsFiltersPublicResponse"), dict)
        else {}
    )
    portfolio_review_workflow_sla_trends_filters_schema_props = (
        ((schemas.get("PortfolioReviewWorkflowSLATrendsFiltersPublicResponse") or {}).get("properties") or {})
        if isinstance(schemas.get("PortfolioReviewWorkflowSLATrendsFiltersPublicResponse"), dict)
        else {}
    )
    review_workflow_sla_filters_schema_props = (
        ((schemas.get("JobReviewWorkflowSLAFiltersPublicResponse") or {}).get("properties") or {})
        if isinstance(schemas.get("JobReviewWorkflowSLAFiltersPublicResponse"), dict)
        else {}
    )
    review_workflow_sla_hotspots_filters_schema_props = (
        ((schemas.get("JobReviewWorkflowSLAHotspotsFiltersPublicResponse") or {}).get("properties") or {})
        if isinstance(schemas.get("JobReviewWorkflowSLAHotspotsFiltersPublicResponse"), dict)
        else {}
    )
    review_workflow_sla_hotspots_trends_filters_schema_props = (
        ((schemas.get("JobReviewWorkflowSLAHotspotsTrendsFiltersPublicResponse") or {}).get("properties") or {})
        if isinstance(schemas.get("JobReviewWorkflowSLAHotspotsTrendsFiltersPublicResponse"), dict)
        else {}
    )
    assert "toc_text_risk_level" in portfolio_filters_schema_props
    for name in (
        "donor_id",
        "status",
        "hitl_enabled",
        "warning_level",
        "grounding_risk_level",
        "toc_text_risk_level",
        "event_type",
        "finding_id",
        "finding_code",
        "finding_section",
        "comment_status",
        "workflow_state",
        "overdue_after_hours",
    ):
        assert name in portfolio_review_workflow_filters_schema_props
    for name in (
        "donor_id",
        "status",
        "hitl_enabled",
        "warning_level",
        "grounding_risk_level",
        "toc_text_risk_level",
        "event_type",
        "finding_id",
        "finding_code",
        "finding_section",
        "comment_status",
        "workflow_state",
        "overdue_after_hours",
    ):
        assert name in portfolio_review_workflow_trends_filters_schema_props
    for name in (
        "donor_id",
        "status",
        "hitl_enabled",
        "warning_level",
        "grounding_risk_level",
        "toc_text_risk_level",
        "finding_id",
        "finding_code",
        "finding_section",
        "comment_status",
        "workflow_state",
        "overdue_after_hours",
    ):
        assert name in portfolio_review_workflow_sla_hotspots_filters_schema_props
        assert name in portfolio_review_workflow_sla_hotspots_trends_filters_schema_props
        assert name in portfolio_review_workflow_sla_filters_schema_props
        assert name in portfolio_review_workflow_sla_trends_filters_schema_props
    assert "top_limit" in portfolio_review_workflow_sla_hotspots_filters_schema_props
    assert "top_limit" in portfolio_review_workflow_sla_hotspots_trends_filters_schema_props
    assert "hotspot_kind" in portfolio_review_workflow_sla_hotspots_filters_schema_props
    assert "hotspot_kind" in portfolio_review_workflow_sla_hotspots_trends_filters_schema_props
    assert "hotspot_severity" in portfolio_review_workflow_sla_hotspots_filters_schema_props
    assert "hotspot_severity" in portfolio_review_workflow_sla_hotspots_trends_filters_schema_props
    assert "min_overdue_hours" in portfolio_review_workflow_sla_hotspots_filters_schema_props
    assert "min_overdue_hours" in portfolio_review_workflow_sla_hotspots_trends_filters_schema_props
    assert "top_limit" in portfolio_review_workflow_sla_filters_schema_props
    for name in (
        "finding_id",
        "finding_code",
        "finding_section",
        "comment_status",
        "workflow_state",
        "overdue_after_hours",
    ):
        assert name in review_workflow_sla_filters_schema_props
        assert name in review_workflow_sla_hotspots_filters_schema_props
        assert name in review_workflow_sla_hotspots_trends_filters_schema_props
    assert "top_limit" in review_workflow_sla_hotspots_filters_schema_props
    assert "hotspot_kind" in review_workflow_sla_hotspots_filters_schema_props
    assert "hotspot_severity" in review_workflow_sla_hotspots_filters_schema_props
    assert "min_overdue_hours" in review_workflow_sla_hotspots_filters_schema_props
    assert "top_limit" in review_workflow_sla_hotspots_trends_filters_schema_props
    assert "hotspot_kind" in review_workflow_sla_hotspots_trends_filters_schema_props
    assert "hotspot_severity" in review_workflow_sla_hotspots_trends_filters_schema_props
    assert "min_overdue_hours" in review_workflow_sla_hotspots_trends_filters_schema_props


def test_ingest_endpoint_uploads_to_donor_namespace(monkeypatch):
    api_app_module.INGEST_AUDIT_STORE.clear()
    calls = {}

    def fake_ingest(pdf_path: str, namespace: str, metadata=None):
        calls["pdf_path"] = pdf_path
        calls["namespace"] = namespace
        calls["metadata"] = metadata or {}
        return {
            "namespace": namespace,
            "source": pdf_path,
            "chunks_ingested": 3,
            "stats": {"namespace": namespace, "document_count": 3},
        }

    monkeypatch.setattr(api_app_module, "ingest_pdf_to_namespace", fake_ingest)

    response = client.post(
        "/ingest",
        data={"donor_id": "usaid", "metadata_json": json.dumps({"source_type": "manual_upload"})},
        files={"file": ("sample.pdf", b"%PDF-1.4 fake content", "application/pdf")},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "ingested"
    assert body["donor_id"] == "usaid"
    assert body["namespace"] == "usaid_ads201"
    assert body["namespace_normalized"] == "usaid_ads201"
    assert body["filename"] == "sample.pdf"
    assert body["result"]["chunks_ingested"] == 3

    assert calls["namespace"] == "usaid_ads201"
    assert calls["metadata"]["uploaded_filename"] == "sample.pdf"
    assert calls["metadata"]["donor_id"] == "usaid"
    assert calls["metadata"]["source_type"] == "manual_upload"


def test_ingest_endpoint_applies_tenant_scoped_namespace(monkeypatch):
    api_app_module.INGEST_AUDIT_STORE.clear()
    calls = {}

    def fake_ingest(pdf_path: str, namespace: str, metadata=None):
        calls["namespace"] = namespace
        calls["metadata"] = metadata or {}
        return {
            "namespace": namespace,
            "source": pdf_path,
            "chunks_ingested": 2,
            "stats": {"namespace": namespace, "document_count": 2},
        }

    monkeypatch.setattr(api_app_module, "ingest_pdf_to_namespace", fake_ingest)

    response = client.post(
        "/ingest",
        data={
            "donor_id": "usaid",
            "tenant_id": "Tenant Alpha",
            "metadata_json": json.dumps({"doc_family": "donor_policy"}),
        },
        files={"file": ("tenant-ads.pdf", b"%PDF-1.4 fake", "application/pdf")},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["tenant_id"] == "tenant_alpha"
    assert body["namespace"] == "tenant_alpha/usaid_ads201"
    assert body["namespace_normalized"] == "tenant_alpha_usaid_ads201"
    assert calls["namespace"] == "tenant_alpha/usaid_ads201"
    assert calls["metadata"]["tenant_id"] == "tenant_alpha"


def test_ingest_recent_endpoint_returns_records(monkeypatch):
    api_app_module.INGEST_AUDIT_STORE.clear()

    def fake_ingest(pdf_path: str, namespace: str, metadata=None):
        return {"namespace": namespace, "source": pdf_path, "chunks_ingested": 1, "stats": {}}

    monkeypatch.setattr(api_app_module, "ingest_pdf_to_namespace", fake_ingest)

    r1 = client.post(
        "/ingest",
        data={
            "donor_id": "usaid",
            "metadata_json": json.dumps({"doc_family": "donor_policy", "source_type": "donor_guidance"}),
        },
        files={"file": ("ads.pdf", b"%PDF-1.4 a", "application/pdf")},
    )
    assert r1.status_code == 200

    r2 = client.post(
        "/ingest",
        data={
            "donor_id": "eu",
            "metadata_json": json.dumps({"doc_family": "country_context", "source_type": "country_context"}),
        },
        files={"file": ("eu-context.pdf", b"%PDF-1.4 b", "application/pdf")},
    )
    assert r2.status_code == 200

    r3 = client.post(
        "/ingest",
        data={
            "donor_id": "usaid",
            "metadata_json": json.dumps({"doc_family": "country_context", "source_type": "country_context"}),
        },
        files={"file": ("kz-context.pdf", b"%PDF-1.4 c", "application/pdf")},
    )
    assert r3.status_code == 200

    recent = client.get("/ingest/recent", params={"donor_id": "usaid", "limit": 10})
    assert recent.status_code == 200
    body = recent.json()
    assert body["donor_id"] == "usaid"
    assert body["count"] == 2
    assert len(body["records"]) == 2
    assert body["records"][0]["donor_id"] == "usaid"
    assert body["records"][0]["filename"] == "kz-context.pdf"
    assert body["records"][0]["metadata"]["doc_family"] == "country_context"
    assert body["records"][1]["filename"] == "ads.pdf"
    assert body["records"][1]["metadata"]["doc_family"] == "donor_policy"


def test_ingest_inventory_endpoint_aggregates_doc_families(monkeypatch):
    api_app_module.INGEST_AUDIT_STORE.clear()

    def fake_ingest(pdf_path: str, namespace: str, metadata=None):
        return {"namespace": namespace, "source": pdf_path, "chunks_ingested": 1, "stats": {}}

    monkeypatch.setattr(api_app_module, "ingest_pdf_to_namespace", fake_ingest)

    uploads = [
        ("usaid", "ads201.pdf", {"doc_family": "donor_policy", "source_type": "donor_guidance"}),
        ("usaid", "kazakhstan-context.pdf", {"doc_family": "country_context", "source_type": "country_context"}),
        ("usaid", "ads-addendum.pdf", {"doc_family": "donor_policy", "source_type": "donor_guidance"}),
        ("eu", "eu-guide.pdf", {"doc_family": "donor_policy", "source_type": "donor_guidance"}),
    ]
    for donor_id, filename, metadata in uploads:
        response = client.post(
            "/ingest",
            data={"donor_id": donor_id, "metadata_json": json.dumps(metadata)},
            files={"file": (filename, b"%PDF-1.4 x", "application/pdf")},
        )
        assert response.status_code == 200

    response = client.get("/ingest/inventory", params={"donor_id": "usaid"})
    assert response.status_code == 200
    body = response.json()
    assert body["donor_id"] == "usaid"
    assert body["total_uploads"] == 3
    assert body["family_count"] == 2
    assert body["doc_family_counts"]["donor_policy"] == 2
    assert body["doc_family_counts"]["country_context"] == 1
    families = {row["doc_family"]: row for row in body["doc_families"]}
    assert families["donor_policy"]["count"] == 2
    assert families["donor_policy"]["latest_filename"] in {"ads-addendum.pdf", "ads201.pdf"}
    assert families["country_context"]["count"] == 1
    assert families["country_context"]["latest_filename"] == "kazakhstan-context.pdf"


def test_ingest_inventory_endpoint_filters_by_tenant(monkeypatch):
    api_app_module.INGEST_AUDIT_STORE.clear()

    def fake_ingest(pdf_path: str, namespace: str, metadata=None):
        return {"namespace": namespace, "source": pdf_path, "chunks_ingested": 1, "stats": {}}

    monkeypatch.setattr(api_app_module, "ingest_pdf_to_namespace", fake_ingest)

    uploads = [
        ("tenant_a", "a-ads.pdf", {"doc_family": "donor_policy", "source_type": "donor_guidance"}),
        ("tenant_b", "b-ads.pdf", {"doc_family": "donor_policy", "source_type": "donor_guidance"}),
        ("tenant_a", "a-country.pdf", {"doc_family": "country_context", "source_type": "country_context"}),
    ]
    for tenant_id, filename, metadata in uploads:
        response = client.post(
            "/ingest",
            data={"donor_id": "usaid", "tenant_id": tenant_id, "metadata_json": json.dumps(metadata)},
            files={"file": (filename, b"%PDF-1.4 x", "application/pdf")},
        )
        assert response.status_code == 200

    response = client.get("/ingest/inventory", params={"donor_id": "usaid", "tenant_id": "tenant_a"})
    assert response.status_code == 200
    body = response.json()
    assert body["tenant_id"] == "tenant_a"
    assert body["total_uploads"] == 2
    assert body["doc_family_counts"]["donor_policy"] == 1
    assert body["doc_family_counts"]["country_context"] == 1
    for row in body["doc_families"]:
        assert row.get("tenant_id") == "tenant_a"


def test_ingest_endpoint_validates_pdf_extension():
    response = client.post(
        "/ingest",
        data={"donor_id": "usaid"},
        files={"file": ("notes.txt", b"hello", "text/plain")},
    )
    assert response.status_code == 400
    assert response.json()["detail"] == "Only PDF uploads are supported"


def test_ingest_endpoint_requires_api_key_when_configured(monkeypatch):
    def fake_ingest(pdf_path: str, namespace: str, metadata=None):
        return {"namespace": namespace, "source": pdf_path, "chunks_ingested": 1, "stats": {}}

    monkeypatch.setattr(api_app_module, "ingest_pdf_to_namespace", fake_ingest)
    monkeypatch.setenv("GRANTFLOW_API_KEY", "test-secret")

    response = client.post(
        "/ingest",
        data={"donor_id": "usaid"},
        files={"file": ("sample.pdf", b"%PDF-1.4 fake content", "application/pdf")},
    )
    assert response.status_code == 401

    response = client.post(
        "/ingest",
        data={"donor_id": "usaid"},
        files={"file": ("sample.pdf", b"%PDF-1.4 fake content", "application/pdf")},
        headers={"X-API-Key": "test-secret"},
    )
    assert response.status_code == 200


def test_generate_dispatches_webhook_events(monkeypatch):
    events = []

    def fake_send_job_webhook_event(**kwargs):
        events.append(kwargs)

    monkeypatch.setattr(api_app_module, "send_job_webhook_event", fake_send_job_webhook_event)

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Livelihoods", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
            "webhook_url": "https://example.com/webhook",
            "webhook_secret": "secret123",
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "done"

    event_names = [evt["event"] for evt in events]
    assert "job.started" in event_names
    assert "job.completed" in event_names

    completed = [evt for evt in events if evt["event"] == "job.completed"][-1]
    assert completed["job_id"] == job_id
    assert completed["url"] == "https://example.com/webhook"
    assert completed["secret"] == "secret123"
    assert completed["job"]["status"] == "done"
    assert completed["job"]["webhook_configured"] is True
    assert "webhook_url" not in completed["job"]
    assert "webhook_secret" not in completed["job"]


def test_generate_request_id_is_idempotent():
    first = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Idempotent generation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
            "request_id": "rid-generate-1",
        },
    )
    assert first.status_code == 200
    first_body = first.json()
    assert first_body["request_id"] == "rid-generate-1"
    assert first_body.get("idempotent_replay") is not True

    replay = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Idempotent generation", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
            "request_id": "rid-generate-1",
        },
    )
    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["request_id"] == "rid-generate-1"
    assert replay_body["idempotent_replay"] is True
    assert replay_body["job_id"] == first_body["job_id"]

    mismatch = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Different payload", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
            "request_id": "rid-generate-1",
        },
    )
    assert mismatch.status_code == 409
    mismatch_detail = mismatch.json().get("detail") or {}
    assert mismatch_detail.get("reason") == "request_id_reused_with_different_payload"


def test_cancel_request_id_is_idempotent():
    job_id = "cancel-request-id-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "running",
            "state": {"donor_id": "usaid", "input_context": {"project": "Cancel idem", "country": "Kenya"}},
            "hitl_enabled": False,
        },
    )

    first = client.post(f"/cancel/{job_id}", params={"request_id": "rid-cancel-1"})
    assert first.status_code == 200
    first_body = first.json()
    assert first_body["status"] == "canceled"
    assert first_body["request_id"] == "rid-cancel-1"
    assert first_body.get("idempotent_replay") is not True

    replay = client.post(f"/cancel/{job_id}", params={"request_id": "rid-cancel-1"})
    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["status"] == "canceled"
    assert replay_body["request_id"] == "rid-cancel-1"
    assert replay_body["idempotent_replay"] is True

    job = api_app_module.JOB_STORE.get(job_id) or {}
    events = job.get("job_events") if isinstance(job.get("job_events"), list) else []
    cancel_events = [row for row in events if isinstance(row, dict) and row.get("type") == "job_canceled"]
    assert len(cancel_events) == 1


def test_cancel_pending_hitl_job_and_cleanup_checkpoint(monkeypatch):
    events = []

    def fake_send_job_webhook_event(**kwargs):
        events.append(kwargs)

    monkeypatch.setattr(api_app_module, "send_job_webhook_event", fake_send_job_webhook_event)

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Education", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
            "webhook_url": "https://example.com/webhook",
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "pending_hitl"
    checkpoint_id = status["checkpoint_id"]

    cancel = client.post(f"/cancel/{job_id}")
    assert cancel.status_code == 200
    assert cancel.json()["status"] == "canceled"
    assert cancel.json()["previous_status"] == "pending_hitl"

    status_after = client.get(f"/status/{job_id}")
    assert status_after.status_code == 200
    body = status_after.json()
    assert body["status"] == "canceled"
    assert body["webhook_configured"] is True
    assert "webhook_url" not in body
    assert "webhook_secret" not in body

    checkpoint = api_app_module.hitl_manager.get_checkpoint(checkpoint_id)
    assert checkpoint is not None
    cp_status = checkpoint["status"]
    assert getattr(cp_status, "value", cp_status) == "canceled"

    pending = client.get("/hitl/pending")
    assert pending.status_code == 200
    assert all(cp["id"] != checkpoint_id for cp in pending.json()["checkpoints"])

    event_names = [evt["event"] for evt in events]
    assert "job.pending_hitl" in event_names
    assert "job.canceled" in event_names


def test_hitl_pause_resume_flow():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Education", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "pending_hitl"
    assert status["checkpoint_stage"] == "toc"
    assert status["state"]["hitl_pending"] is True
    assert status["state"]["toc_draft"]
    assert "logframe_draft" not in status["state"]

    cp_id = status["checkpoint_id"]
    approve = client.post(
        "/hitl/approve",
        json={"checkpoint_id": cp_id, "approved": True, "feedback": "TOC approved"},
    )
    assert approve.status_code == 200

    resume = client.post(f"/resume/{job_id}", json={})
    assert resume.status_code == 200
    assert resume.json()["resuming_from"] == "mel"

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "pending_hitl"
    assert status["checkpoint_stage"] == "logframe"
    assert status["state"]["hitl_pending"] is True
    assert status["state"]["logframe_draft"]

    cp_id = status["checkpoint_id"]
    approve = client.post(
        "/hitl/approve",
        json={"checkpoint_id": cp_id, "approved": True, "feedback": "Logframe approved"},
    )
    assert approve.status_code == 200

    resume = client.post(f"/resume/{job_id}", json={})
    assert resume.status_code == 200
    assert resume.json()["resuming_from"] == "critic"

    status = _wait_for_terminal_status(job_id)
    if status["status"] != "done":
        status = _drain_hitl_to_done(job_id, initial_status=status)
    assert status["status"] == "done"
    state = status["state"]
    assert state["hitl_pending"] is False
    assert state["toc_draft"]
    assert state["logframe_draft"]
    engine = str((state.get("critic_notes") or {}).get("engine", "") or "")
    assert engine in {"rules", ""} or engine.startswith("rules+llm:")


def test_hitl_checkpoint_selection_logframe_only_flow():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Education", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
            "hitl_checkpoints": ["logframe"],
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "pending_hitl"
    assert status["checkpoint_stage"] == "logframe"
    assert status["state"]["hitl_pending"] is True
    assert status["state"]["toc_draft"]
    assert status["state"]["logframe_draft"]

    cp_id = status["checkpoint_id"]
    approve = client.post(
        "/hitl/approve",
        json={"checkpoint_id": cp_id, "approved": True, "feedback": "Logframe approved"},
    )
    assert approve.status_code == 200

    resume = client.post(f"/resume/{job_id}", json={})
    assert resume.status_code == 200
    assert resume.json()["resuming_from"] == "critic"

    status = _wait_for_terminal_status(job_id)
    if status["status"] != "done":
        status = _drain_hitl_to_done(job_id, initial_status=status)
    assert status["status"] == "done"
    assert status["state"]["hitl_pending"] is False


def test_resume_requires_checkpoint_decision_before_running_again():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Education", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "pending_hitl"
    assert status["checkpoint_stage"] == "toc"

    resume = client.post(f"/resume/{job_id}", json={})
    assert resume.status_code == 409
    assert "pending approval" in resume.json()["detail"]


def test_resume_rejects_canceled_checkpoint():
    checkpoint_id = api_app_module.hitl_manager.create_checkpoint(
        stage="toc",
        state={"donor_id": "usaid", "input_context": {"project": "Canceled checkpoint", "country": "Kenya"}},
        donor_id="usaid",
    )
    assert api_app_module.hitl_manager.cancel(checkpoint_id, "Canceled externally")

    job_id = f"resume-canceled-checkpoint-{int(time.time() * 1_000_000)}"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "pending_hitl",
            "state": {
                "donor_id": "usaid",
                "input_context": {"project": "Canceled checkpoint", "country": "Kenya"},
                "hitl_pending": True,
            },
            "hitl_enabled": True,
            "checkpoint_id": checkpoint_id,
            "checkpoint_stage": "toc",
            "resume_from": "mel",
        },
    )

    resume = client.post(f"/resume/{job_id}", json={})
    assert resume.status_code == 409
    assert "approved or rejected" in str(resume.json().get("detail") or "")


def test_hitl_approve_reject_conflict_after_finalized_checkpoint():
    checkpoint_id = api_app_module.hitl_manager.create_checkpoint(
        stage="toc",
        state={"donor_id": "usaid", "input_context": {"project": "Decision conflict", "country": "Kenya"}},
        donor_id="usaid",
    )

    first = client.post(
        "/hitl/approve",
        json={"checkpoint_id": checkpoint_id, "approved": True, "feedback": "Looks good"},
    )
    assert first.status_code == 200
    assert first.json()["status"] == "approved"

    conflict = client.post(
        "/hitl/approve",
        json={"checkpoint_id": checkpoint_id, "approved": False, "feedback": "Actually reject"},
    )
    assert conflict.status_code == 409
    assert "already finalized" in str(conflict.json().get("detail") or "")


def test_pause_for_hitl_replaces_stale_stage_checkpoint_reference():
    stale_checkpoint_id = api_app_module.hitl_manager.create_checkpoint(
        stage="logframe",
        state={"donor_id": "usaid", "input_context": {"project": "Stale checkpoint", "country": "Kenya"}},
        donor_id="usaid",
    )

    job_id = f"pause-stale-checkpoint-{int(time.time() * 1_000_000)}"
    state = {
        "donor_id": "usaid",
        "input_context": {"project": "Stale checkpoint", "country": "Kenya"},
        "hitl_enabled": True,
        "hitl_pending": True,
        "hitl_checkpoint_id": stale_checkpoint_id,
    }
    api_app_module._pause_for_hitl(job_id, state, stage="toc", resume_from="mel")

    job = api_app_module._get_job(job_id)
    assert job is not None
    assert job["status"] == "pending_hitl"
    assert job["checkpoint_stage"] == "toc"
    assert job["checkpoint_id"] != stale_checkpoint_id

    stale = api_app_module.hitl_manager.get_checkpoint(stale_checkpoint_id)
    assert stale is not None
    stale_status = getattr(stale.get("status"), "value", stale.get("status"))
    assert stale_status == "canceled"


def test_resume_clears_hitl_runtime_flags_before_relaunch(monkeypatch):
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Education", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "pending_hitl"
    cp_id = status["checkpoint_id"]

    approve = client.post(
        "/hitl/approve",
        json={"checkpoint_id": cp_id, "approved": True, "feedback": "TOC approved"},
    )
    assert approve.status_code == 200

    captured: dict = {}

    def _stub_run_hitl_pipeline(job_id_arg: str, state_arg: dict, start_at_arg: str) -> None:
        captured["job_id"] = job_id_arg
        captured["state"] = dict(state_arg)
        captured["start_at"] = start_at_arg

    monkeypatch.setattr(api_app_module, "_run_hitl_pipeline", _stub_run_hitl_pipeline)

    resume = client.post(f"/resume/{job_id}", json={})
    assert resume.status_code == 200
    assert resume.json()["resuming_from"] == "mel"

    assert captured["job_id"] == job_id
    assert captured["start_at"] == "mel"
    resumed_state = captured["state"]
    assert resumed_state["hitl_pending"] is False
    assert "hitl_checkpoint_id" not in resumed_state
    assert "hitl_checkpoint_stage" not in resumed_state
    assert "hitl_resume_from" not in resumed_state

    job = api_app_module._get_job(job_id)
    assert job is not None
    assert job["status"] == "accepted"
    assert job["state"]["hitl_pending"] is False
    assert job.get("checkpoint_id") is None
    assert job.get("checkpoint_stage") is None


def test_run_hitl_pipeline_errors_when_pending_marker_has_no_checkpoint_stage(monkeypatch):
    job_id = "hitl-invalid-pending-marker-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "accepted",
            "state": {
                "donor_id": "usaid",
                "input_context": {"project": "Marker test", "country": "Kenya"},
                "hitl_enabled": True,
            },
            "hitl_enabled": True,
        },
    )

    def _stub_invoke(state):
        out = dict(state)
        out["hitl_pending"] = True
        out.pop("hitl_checkpoint_stage", None)
        out.pop("hitl_resume_from", None)
        return out

    class _StubGraph:
        def invoke(self, state):
            return _stub_invoke(state)

    monkeypatch.setattr(api_app_module, "grantflow_graph", _StubGraph())

    state = {"donor_id": "usaid", "input_context": {"project": "Marker test", "country": "Kenya"}, "hitl_enabled": True}
    api_app_module._run_hitl_pipeline(job_id, state, "start")

    status = client.get(f"/status/{job_id}")
    assert status.status_code == 200
    body = status.json()
    assert body["status"] == "error"
    assert "without a valid checkpoint stage" in str(body.get("error") or "")
    assert body.get("checkpoint_id") in {None, ""}


def test_resume_request_id_is_idempotent(monkeypatch):
    checkpoint_id = api_app_module.hitl_manager.create_checkpoint(
        stage="toc",
        state={"donor_id": "usaid", "input_context": {"project": "Resume idem", "country": "Kenya"}},
        donor_id="usaid",
    )
    api_app_module.hitl_manager.approve(checkpoint_id, "approved")

    job_id = "resume-request-id-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "pending_hitl",
            "state": {
                "donor_id": "usaid",
                "input_context": {"project": "Resume idem", "country": "Kenya"},
                "hitl_pending": True,
            },
            "hitl_enabled": True,
            "checkpoint_id": checkpoint_id,
            "checkpoint_stage": "toc",
            "resume_from": "mel",
        },
    )

    monkeypatch.setattr(api_app_module, "_run_hitl_pipeline", lambda *_args, **_kwargs: None)

    first = client.post(f"/resume/{job_id}", params={"request_id": "rid-resume-1"}, json={})
    assert first.status_code == 200
    first_body = first.json()
    assert first_body["request_id"] == "rid-resume-1"
    assert first_body["resuming_from"] == "mel"
    assert first_body.get("idempotent_replay") is not True

    api_app_module._update_job(job_id, status="done")

    replay = client.post(f"/resume/{job_id}", params={"request_id": "rid-resume-1"}, json={})
    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["request_id"] == "rid-resume-1"
    assert replay_body["idempotent_replay"] is True
    assert replay_body["resuming_from"] == "mel"

    job = api_app_module.JOB_STORE.get(job_id) or {}
    events = job.get("job_events") if isinstance(job.get("job_events"), list) else []
    resume_events = [row for row in events if isinstance(row, dict) and row.get("type") == "resume_requested"]
    assert len(resume_events) == 1


def test_generate_uses_inmemory_queue_dispatch_when_enabled(monkeypatch):
    captured: dict = {}

    def _submit_stub(fn, *args, **kwargs):
        captured["fn"] = fn
        captured["args"] = args
        captured["kwargs"] = kwargs
        return True

    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "inmemory_queue")
    monkeypatch.setattr(api_app_module.JOB_RUNNER, "submit", _submit_stub)

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Queue dispatch", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    body = response.json()
    job_id = body["job_id"]

    assert captured["fn"] == api_app_module._run_pipeline_to_completion
    assert captured["args"][0] == job_id
    queued_state = captured["args"][1]
    assert queued_state["donor_id"] == "usaid"

    status_resp = client.get(f"/status/{job_id}")
    assert status_resp.status_code == 200
    assert status_resp.json()["status"] == "accepted"


def test_generate_returns_503_when_inmemory_queue_is_full(monkeypatch):
    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "inmemory_queue")
    monkeypatch.setattr(api_app_module.JOB_RUNNER, "submit", lambda *args, **kwargs: False)

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Queue full", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 503
    assert "Job queue is full" in str(response.json()["detail"])


def test_generate_uses_redis_queue_dispatch_when_enabled(monkeypatch):
    captured: dict = {}

    def _submit_stub(fn, *args, **kwargs):
        captured["fn"] = fn
        captured["args"] = args
        captured["kwargs"] = kwargs
        return True

    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "redis_queue")
    monkeypatch.setattr(api_app_module.JOB_RUNNER, "submit", _submit_stub)

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Redis queue dispatch", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 200
    body = response.json()
    job_id = body["job_id"]

    assert captured["fn"] == api_app_module._run_pipeline_to_completion_by_job_id
    assert captured["args"][0] == job_id
    assert len(captured["args"]) == 1

    status_resp = client.get(f"/status/{job_id}")
    assert status_resp.status_code == 200
    assert status_resp.json()["status"] == "accepted"


def test_generate_hitl_uses_redis_queue_dispatch_by_job_id(monkeypatch):
    captured: dict = {}

    def _submit_stub(fn, *args, **kwargs):
        captured["fn"] = fn
        captured["args"] = args
        captured["kwargs"] = kwargs
        return True

    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "redis_queue")
    monkeypatch.setattr(api_app_module.JOB_RUNNER, "submit", _submit_stub)

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Redis queue HITL dispatch", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
        },
    )
    assert response.status_code == 200
    body = response.json()
    job_id = body["job_id"]

    assert captured["fn"] == api_app_module._run_hitl_pipeline_by_job_id
    assert captured["args"] == (job_id, "start")


def test_resume_uses_redis_queue_dispatch_by_job_id(monkeypatch):
    captured: dict = {}

    def _submit_stub(fn, *args, **kwargs):
        captured["fn"] = fn
        captured["args"] = args
        captured["kwargs"] = kwargs
        return True

    checkpoint_id = api_app_module.hitl_manager.create_checkpoint(
        stage="toc",
        state={"donor_id": "usaid", "input_context": {"project": "Resume redis", "country": "Kenya"}},
        donor_id="usaid",
    )
    api_app_module.hitl_manager.approve(checkpoint_id, "approved")

    job_id = "resume-redis-dispatch-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "pending_hitl",
            "state": {
                "donor_id": "usaid",
                "input_context": {"project": "Resume redis", "country": "Kenya"},
                "hitl_pending": True,
            },
            "hitl_enabled": True,
            "checkpoint_id": checkpoint_id,
            "checkpoint_stage": "toc",
            "resume_from": "mel",
        },
    )

    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "redis_queue")
    monkeypatch.setattr(api_app_module.JOB_RUNNER, "submit", _submit_stub)

    response = client.post(f"/resume/{job_id}", json={})
    assert response.status_code == 200
    assert response.json()["resuming_from"] == "mel"

    assert captured["fn"] == api_app_module._run_hitl_pipeline_by_job_id
    assert captured["args"] == (job_id, "mel")


def test_run_pipeline_to_completion_by_job_id_skips_non_runnable_statuses(monkeypatch):
    calls: list[str] = []
    monkeypatch.setattr(api_app_module, "_run_pipeline_to_completion", lambda *_args, **_kwargs: calls.append("run"))

    job_id_pending = "queue-skip-pending-hitl-job-1"
    api_app_module.JOB_STORE.set(
        job_id_pending,
        {
            "status": "pending_hitl",
            "state": {"donor_id": "usaid", "input_context": {"project": "Skip pending"}},
            "hitl_enabled": True,
        },
    )
    api_app_module._run_pipeline_to_completion_by_job_id(job_id_pending)

    job_id_done = "queue-skip-done-job-1"
    api_app_module.JOB_STORE.set(
        job_id_done,
        {
            "status": "done",
            "state": {"donor_id": "usaid", "input_context": {"project": "Skip done"}},
            "hitl_enabled": False,
        },
    )
    api_app_module._run_pipeline_to_completion_by_job_id(job_id_done)

    assert calls == []


def test_run_hitl_pipeline_by_job_id_skips_non_runnable_statuses(monkeypatch):
    calls: list[str] = []
    monkeypatch.setattr(api_app_module, "_run_hitl_pipeline", lambda *_args, **_kwargs: calls.append("run"))

    job_id_pending = "queue-hitl-skip-pending-job-1"
    api_app_module.JOB_STORE.set(
        job_id_pending,
        {
            "status": "pending_hitl",
            "state": {"donor_id": "usaid", "input_context": {"project": "Skip pending"}},
            "hitl_enabled": True,
        },
    )
    api_app_module._run_hitl_pipeline_by_job_id(job_id_pending, "start")

    job_id_error = "queue-hitl-skip-error-job-1"
    api_app_module.JOB_STORE.set(
        job_id_error,
        {
            "status": "error",
            "state": {"donor_id": "usaid", "input_context": {"project": "Skip error"}},
            "hitl_enabled": True,
        },
    )
    api_app_module._run_hitl_pipeline_by_job_id(job_id_error, "start")

    assert calls == []


def test_generate_returns_503_when_redis_queue_is_unavailable(monkeypatch):
    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "redis_queue")
    monkeypatch.setattr(api_app_module.JOB_RUNNER, "submit", lambda *args, **kwargs: False)

    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Redis queue unavailable", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    assert response.status_code == 503
    assert "Job queue is full" in str(response.json()["detail"])


def test_dead_letter_queue_endpoints_require_redis_queue_mode(monkeypatch):
    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "background_tasks")

    list_resp = client.get("/queue/dead-letter")
    assert list_resp.status_code == 409
    assert "redis_queue" in str(list_resp.json()["detail"])

    export_resp = client.get("/queue/dead-letter/export")
    assert export_resp.status_code == 409
    assert "redis_queue" in str(export_resp.json()["detail"])

    requeue_resp = client.post("/queue/dead-letter/requeue")
    assert requeue_resp.status_code == 409
    assert "redis_queue" in str(requeue_resp.json()["detail"])

    purge_resp = client.delete("/queue/dead-letter")
    assert purge_resp.status_code == 409
    assert "redis_queue" in str(purge_resp.json()["detail"])


def test_dead_letter_queue_endpoints_use_runner_in_redis_mode(monkeypatch):
    class _StubRunner:
        def list_dead_letters(self, limit: int = 50):
            return {
                "mode": "redis_queue",
                "queue_name": "grantflow:jobs",
                "dead_letter_queue_name": "grantflow:jobs:dead",
                "dead_letter_queue_size": 1,
                "items": [{"index": 0, "task_name": "grantflow.api.app:_run_pipeline_to_completion_by_job_id"}],
            }

        def requeue_dead_letters(self, limit: int = 10, reset_attempts: bool = True):
            _ = reset_attempts
            return {
                "mode": "redis_queue",
                "queue_name": "grantflow:jobs",
                "dead_letter_queue_name": "grantflow:jobs:dead",
                "requested_count": int(limit),
                "affected_count": 1,
                "skipped_count": 0,
                "dead_letter_queue_size": 0,
            }

        def purge_dead_letters(self, limit: int = 100):
            return {
                "mode": "redis_queue",
                "queue_name": "grantflow:jobs",
                "dead_letter_queue_name": "grantflow:jobs:dead",
                "requested_count": int(limit),
                "affected_count": 1,
                "dead_letter_queue_size": 0,
            }

    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "redis_queue")
    monkeypatch.setattr(api_app_module, "JOB_RUNNER", _StubRunner())

    listed = client.get("/queue/dead-letter", params={"limit": 5})
    assert listed.status_code == 200
    listed_body = listed.json()
    assert listed_body["mode"] == "redis_queue"
    assert listed_body["dead_letter_queue_size"] == 1
    assert listed_body["items"]

    export_json = client.get("/queue/dead-letter/export", params={"limit": 5, "format": "json"})
    assert export_json.status_code == 200
    assert "application/json" in (export_json.headers.get("content-type") or "")
    assert "grantflow_queue_dead_letter" in (export_json.headers.get("content-disposition") or "")

    export_csv = client.get("/queue/dead-letter/export", params={"limit": 5, "format": "csv"})
    assert export_csv.status_code == 200
    assert "text/csv" in (export_csv.headers.get("content-type") or "")
    assert "task_name" in export_csv.text

    requeue = client.post("/queue/dead-letter/requeue", params={"limit": 2, "reset_attempts": True})
    assert requeue.status_code == 200
    requeue_body = requeue.json()
    assert requeue_body["affected_count"] == 1
    assert requeue_body["requested_count"] == 2

    purge = client.delete("/queue/dead-letter", params={"limit": 3})
    assert purge.status_code == 200
    purge_body = purge.json()
    assert purge_body["affected_count"] == 1
    assert purge_body["requested_count"] == 3


def test_resume_returns_503_and_keeps_pending_hitl_when_queue_full(monkeypatch):
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Resume queue full", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]
    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "pending_hitl"
    checkpoint_id = status["checkpoint_id"]

    approve = client.post(
        "/hitl/approve",
        json={"checkpoint_id": checkpoint_id, "approved": True, "feedback": "approved"},
    )
    assert approve.status_code == 200

    monkeypatch.setattr(api_app_module.config.job_runner, "mode", "inmemory_queue")
    monkeypatch.setattr(api_app_module.JOB_RUNNER, "submit", lambda *args, **kwargs: False)

    resume = client.post(f"/resume/{job_id}", json={})
    assert resume.status_code == 503
    assert "Job queue is full" in str(resume.json()["detail"])

    status_after = client.get(f"/status/{job_id}")
    assert status_after.status_code == 200
    body = status_after.json()
    assert body["status"] == "pending_hitl"
    assert body["checkpoint_id"] == checkpoint_id
    assert body["checkpoint_stage"] == "toc"


def test_generate_rejects_old_contract_shape():
    response = client.post(
        "/generate",
        json={"donor": "usaid", "input": {"project": "Water"}},
    )
    assert response.status_code == 422


def test_export_both_zip():
    gen = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "Health", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": False,
        },
    )
    job_id = gen.json()["job_id"]
    status = _wait_for_terminal_status(job_id)
    state = status["state"]

    export = client.post("/export", json={"payload": state, "format": "both"})
    assert export.status_code == 200
    assert export.headers["content-type"].startswith("application/zip")
    assert b"PK" == export.content[:2]


def test_hitl_checkpoint_endpoints():
    from grantflow.swarm.hitl import hitl_manager

    checkpoint_id = hitl_manager.create_checkpoint(
        stage="toc",
        state={"test": "data"},
        donor_id="USAID",
    )

    response = client.get("/hitl/pending")
    assert response.status_code == 200
    assert response.json()["pending_count"] >= 1
    checkpoints = response.json()["checkpoints"]
    matching = [cp for cp in checkpoints if cp["id"] == checkpoint_id]
    assert matching
    assert "state_snapshot" not in matching[0]
    assert matching[0]["has_state_snapshot"] is True

    response = client.post(
        "/hitl/approve",
        json={
            "checkpoint_id": checkpoint_id,
            "approved": True,
            "feedback": "Looks good",
        },
    )
    assert response.status_code == 200
    assert response.json()["status"] == "approved"


def test_status_hitl_history_endpoint_lists_and_filters_events():
    job_id = "hitl-history-job-1"
    api_app_module.JOB_STORE.set(
        job_id,
        {
            "status": "done",
            "state": {"donor_id": "usaid", "input_context": {"project": "History", "country": "Kenya"}},
            "hitl_enabled": True,
            "job_events": [
                {
                    "event_id": "h1",
                    "ts": "2026-03-01T10:00:00+00:00",
                    "type": "status_changed",
                    "from_status": "running",
                    "to_status": "pending_hitl",
                    "status": "pending_hitl",
                },
                {
                    "event_id": "h2",
                    "ts": "2026-03-01T10:00:05+00:00",
                    "type": "hitl_checkpoint_published",
                    "checkpoint_id": "cp-h-1",
                    "checkpoint_stage": "toc",
                },
                {
                    "event_id": "h3",
                    "ts": "2026-03-01T10:00:10+00:00",
                    "type": "hitl_checkpoint_decision",
                    "checkpoint_id": "cp-h-1",
                    "checkpoint_stage": "toc",
                    "checkpoint_status": "approved",
                    "approved": True,
                    "feedback": "ok",
                    "actor": "qa",
                    "request_id": "rid-hitl-h1",
                },
                {
                    "event_id": "h4",
                    "ts": "2026-03-01T10:00:20+00:00",
                    "type": "resume_requested",
                    "checkpoint_id": "cp-h-1",
                    "resuming_from": "mel",
                },
                {
                    "event_id": "h5",
                    "ts": "2026-03-01T10:00:30+00:00",
                    "type": "status_changed",
                    "from_status": "pending_hitl",
                    "to_status": "running",
                    "status": "running",
                },
                {"event_id": "h6", "ts": "2026-03-01T10:00:40+00:00", "type": "job_canceled"},
            ],
        },
    )

    response = client.get(f"/status/{job_id}/hitl/history")
    assert response.status_code == 200
    body = response.json()
    assert body["job_id"] == job_id
    assert body["event_count"] == 5
    assert body["event_type_counts"]["status_changed"] == 2
    assert body["event_type_counts"]["hitl_checkpoint_decision"] == 1
    assert body["events"][0]["event_id"] == "h5"
    assert body["events"][-1]["event_id"] == "h1"

    decision_only = client.get(
        f"/status/{job_id}/hitl/history",
        params={"event_type": "hitl_checkpoint_decision"},
    )
    assert decision_only.status_code == 200
    decision_body = decision_only.json()
    assert decision_body["event_count"] == 1
    assert decision_body["events"][0]["type"] == "hitl_checkpoint_decision"
    assert decision_body["events"][0]["checkpoint_status"] == "approved"

    checkpoint_only = client.get(
        f"/status/{job_id}/hitl/history",
        params={"checkpoint_id": "cp-h-1"},
    )
    assert checkpoint_only.status_code == 200
    checkpoint_body = checkpoint_only.json()
    assert checkpoint_body["event_count"] == 3
    assert all((row.get("checkpoint_id") == "cp-h-1") for row in checkpoint_body["events"])


def test_hitl_approve_request_id_is_idempotent():
    from grantflow.swarm.hitl import hitl_manager

    checkpoint_id = hitl_manager.create_checkpoint(
        stage="toc",
        state={"test": "data"},
        donor_id="usaid",
    )

    first = client.post(
        "/hitl/approve",
        json={
            "checkpoint_id": checkpoint_id,
            "approved": True,
            "feedback": "Looks good",
            "request_id": "rid-hitl-approve-1",
        },
    )
    assert first.status_code == 200
    first_body = first.json()
    assert first_body["status"] == "approved"
    assert first_body["request_id"] == "rid-hitl-approve-1"
    assert first_body.get("idempotent_replay") is not True

    replay = client.post(
        "/hitl/approve",
        json={
            "checkpoint_id": checkpoint_id,
            "approved": True,
            "feedback": "Looks good",
            "request_id": "rid-hitl-approve-1",
        },
    )
    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["request_id"] == "rid-hitl-approve-1"
    assert replay_body["idempotent_replay"] is True
    assert replay_body["status"] == "approved"

    mismatch = client.post(
        "/hitl/approve",
        json={
            "checkpoint_id": checkpoint_id,
            "approved": False,
            "feedback": "Rejected",
            "request_id": "rid-hitl-approve-1",
        },
    )
    assert mismatch.status_code == 409
    mismatch_detail = mismatch.json().get("detail") or {}
    assert mismatch_detail.get("reason") == "request_id_reused_with_different_payload"


def test_hitl_approve_records_audit_event_and_exposes_in_history():
    response = client.post(
        "/generate",
        json={
            "donor_id": "usaid",
            "input_context": {"project": "HITL audit", "country": "Kenya"},
            "llm_mode": False,
            "hitl_enabled": True,
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    status = _wait_for_terminal_status(job_id)
    assert status["status"] == "pending_hitl"
    checkpoint_id = status["checkpoint_id"]

    approve = client.post(
        "/hitl/approve",
        headers={"X-Actor": "reviewer_a"},
        json={
            "checkpoint_id": checkpoint_id,
            "approved": True,
            "feedback": "approved by reviewer",
            "request_id": "rid-hitl-audit-1",
        },
    )
    assert approve.status_code == 200

    history = client.get(f"/status/{job_id}/hitl/history", params={"event_type": "hitl_checkpoint_decision"})
    assert history.status_code == 200
    history_body = history.json()
    assert history_body["event_count"] >= 1
    latest = history_body["events"][0]
    assert latest["type"] == "hitl_checkpoint_decision"
    assert latest["checkpoint_id"] == checkpoint_id
    assert latest["checkpoint_status"] == "approved"
    assert latest["approved"] is True
    assert latest["feedback"] == "approved by reviewer"
    assert latest["actor"] == "reviewer_a"
    assert latest["request_id"] == "rid-hitl-audit-1"
