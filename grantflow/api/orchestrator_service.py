from __future__ import annotations

import io
from typing import Any, Dict, Optional

from openpyxl import load_workbook

from grantflow.core.config import config
from grantflow.exporters.donor_contracts import (
    DONOR_XLSX_PRIMARY_SHEET,
    evaluate_export_contract_gate,
    normalize_export_contract_policy_mode,
)
from grantflow.exporters.template_profile import normalize_export_template_key


def _dedupe_doc_families(values: list[Any]) -> list[str]:
    from grantflow.api.preflight_service import _dedupe_doc_families as _impl

    return _impl(values)


def _preflight_expected_doc_families(
    *,
    donor_id: str,
    client_metadata: Optional[Dict[str, Any]],
) -> list[str]:
    from grantflow.api.preflight_service import _preflight_expected_doc_families as _impl

    return _impl(donor_id=donor_id, client_metadata=client_metadata)


def _preflight_doc_family_min_uploads_map(
    *,
    expected_doc_families: list[str],
    client_metadata: Optional[Dict[str, Any]],
) -> Dict[str, int]:
    from grantflow.api.preflight_service import _preflight_doc_family_min_uploads_map as _impl

    return _impl(
        expected_doc_families=expected_doc_families,
        client_metadata=client_metadata,
    )


def _preflight_doc_family_depth_profile(
    *,
    expected_doc_families: list[str],
    doc_family_counts: Dict[str, Any],
    min_uploads_by_family: Dict[str, int],
) -> Dict[str, Any]:
    from grantflow.api.preflight_service import _preflight_doc_family_depth_profile as _impl

    return _impl(
        expected_doc_families=expected_doc_families,
        doc_family_counts=doc_family_counts,
        min_uploads_by_family=min_uploads_by_family,
    )


def _preflight_input_context(client_metadata: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    from grantflow.api.preflight_service import _preflight_input_context as _impl

    return _impl(client_metadata)


def _preflight_severity_max(severities: list[str]) -> str:
    from grantflow.api.preflight_service import _preflight_severity_max as _impl

    return _impl(severities)


def _normalize_grounding_policy_mode(raw_mode: Any) -> str:
    from grantflow.api.runtime_grounded_gate_service import _normalize_grounding_policy_mode as _impl

    return _impl(raw_mode)


def _configured_preflight_grounding_policy_mode() -> str:
    from grantflow.api.preflight_service import _configured_preflight_grounding_policy_mode as _impl

    return _impl()


def _preflight_grounding_policy_thresholds() -> Dict[str, Any]:
    from grantflow.api.preflight_service import _preflight_grounding_policy_thresholds as _impl

    return _impl()


def _estimate_preflight_architect_claims(
    *,
    donor_id: str,
    strategy: Any,
    namespace: str,
    input_context: Optional[Dict[str, Any]],
    tenant_id: Optional[str] = None,
    architect_rag_enabled: bool = True,
) -> Dict[str, Any]:
    from grantflow.api.preflight_service import _estimate_preflight_architect_claims as _impl

    return _impl(
        donor_id=donor_id,
        strategy=strategy,
        namespace=namespace,
        input_context=input_context,
        tenant_id=tenant_id,
        architect_rag_enabled=architect_rag_enabled,
    )


def _build_preflight_grounding_policy(
    *,
    coverage_rate: Optional[float],
    depth_coverage_rate: Optional[float],
    namespace_empty: bool,
    inventory_total_uploads: int,
    missing_doc_families: list[str],
    depth_gap_doc_families: list[str],
    architect_claims: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    from grantflow.api.preflight_service import _build_preflight_grounding_policy as _impl

    return _impl(
        coverage_rate=coverage_rate,
        depth_coverage_rate=depth_coverage_rate,
        namespace_empty=namespace_empty,
        inventory_total_uploads=inventory_total_uploads,
        missing_doc_families=missing_doc_families,
        depth_gap_doc_families=depth_gap_doc_families,
        architect_claims=architect_claims,
    )


def _build_generate_preflight(
    *,
    donor_id: str,
    strategy: Any,
    client_metadata: Optional[Dict[str, Any]],
    tenant_id: Optional[str] = None,
    architect_rag_enabled: bool = True,
) -> Dict[str, Any]:
    from grantflow.api.preflight_service import _build_generate_preflight as _impl

    return _impl(
        donor_id=donor_id,
        strategy=strategy,
        client_metadata=client_metadata,
        tenant_id=tenant_id,
        architect_rag_enabled=architect_rag_enabled,
    )


def _configured_runtime_grounded_quality_gate_mode() -> str:
    from grantflow.api.runtime_grounded_gate_service import _configured_runtime_grounded_quality_gate_mode as _impl

    return _impl()


def _runtime_grounded_quality_gate_thresholds() -> Dict[str, Any]:
    from grantflow.api.runtime_grounded_gate_service import _runtime_grounded_quality_gate_thresholds as _impl

    return _impl()


def _runtime_grounded_gate_section(citation: Dict[str, Any]) -> str:
    from grantflow.api.runtime_grounded_gate_service import _runtime_grounded_gate_section as _impl

    return _impl(citation)


def _runtime_grounded_gate_evidence_row(citation: Dict[str, Any]) -> Dict[str, Any]:
    from grantflow.api.runtime_grounded_gate_service import _runtime_grounded_gate_evidence_row as _impl

    return _impl(citation)


def _evaluate_runtime_grounded_quality_gate_from_state(state: Any) -> Dict[str, Any]:
    from grantflow.api.runtime_grounded_gate_service import (
        _evaluate_runtime_grounded_quality_gate_from_state as _impl,
    )

    return _impl(state)


def _configured_mel_grounding_policy_mode() -> str:
    from grantflow.api.grounding_policy_service import _configured_mel_grounding_policy_mode as _impl

    return _impl()


def _mel_grounding_policy_thresholds() -> Dict[str, Any]:
    from grantflow.api.grounding_policy_service import _mel_grounding_policy_thresholds as _impl

    return _impl()


def _evaluate_mel_grounding_policy_from_state(state: Any) -> Dict[str, Any]:
    from grantflow.api.grounding_policy_service import _evaluate_mel_grounding_policy_from_state as _impl

    return _impl(state)


def _configured_export_grounding_policy_mode() -> str:
    from grantflow.api.grounding_policy_service import _configured_export_grounding_policy_mode as _impl

    return _impl()


def _configured_export_require_grounded_gate_pass() -> bool:
    from grantflow.api.grounding_policy_service import _configured_export_require_grounded_gate_pass as _impl

    return _impl()


def _export_grounding_policy_thresholds() -> Dict[str, Any]:
    from grantflow.api.grounding_policy_service import _export_grounding_policy_thresholds as _impl

    return _impl()


def _evaluate_export_grounding_policy(citations: list[dict[str, Any]]) -> Dict[str, Any]:
    from grantflow.api.grounding_policy_service import _evaluate_export_grounding_policy as _impl

    return _impl(citations)


def _configured_export_contract_policy_mode() -> str:
    contract_mode = getattr(config.graph, "export_contract_policy_mode", None)
    if str(contract_mode or "").strip():
        return normalize_export_contract_policy_mode(contract_mode)
    return _configured_export_grounding_policy_mode()


def _evaluate_export_contract_gate(
    *,
    donor_id: str,
    toc_draft: dict[str, Any],
    workbook_sheetnames: Optional[list[str]] = None,
    workbook_primary_sheet_headers: Optional[list[str]] = None,
) -> Dict[str, Any]:
    return evaluate_export_contract_gate(
        donor_id=donor_id,
        toc_payload=toc_draft,
        policy_mode=_configured_export_contract_policy_mode(),
        workbook_sheetnames=workbook_sheetnames,
        workbook_primary_sheet_headers=workbook_primary_sheet_headers,
    )


def _xlsx_contract_validation_context(
    xlsx_payload: bytes,
    *,
    donor_id: str,
) -> tuple[list[str], list[str]]:
    workbook = load_workbook(io.BytesIO(xlsx_payload), data_only=True, read_only=True)
    try:
        sheetnames = list(workbook.sheetnames)
        donor_key = normalize_export_template_key(donor_id)
        primary_sheet = DONOR_XLSX_PRIMARY_SHEET.get(donor_key)
        if not primary_sheet or primary_sheet not in workbook.sheetnames:
            return sheetnames, []
        sheet = workbook[primary_sheet]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(value).strip() for value in header_row if str(value or "").strip()]
        return sheetnames, headers
    finally:
        workbook.close()
