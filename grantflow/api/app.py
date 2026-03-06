from __future__ import annotations

# ruff: noqa: F401

import csv
import gzip
import io
import json
import os
import sys
import tempfile
import uuid
import zipfile
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from typing import Any, AsyncIterator, Callable, Dict, Literal, Optional

from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import load_workbook
from pydantic import ValidationError

from grantflow.api.constants import (
    CRITIC_FINDING_SLA_HOURS,
    CRITIC_FINDING_STATUSES,
    GROUNDING_POLICY_MODES,
    REVIEW_COMMENT_DEFAULT_SLA_HOURS,
    REVIEW_COMMENT_SECTIONS,
)
from grantflow.api.demo_ui import render_demo_ui_html
from grantflow.api.demo_presets import (
    list_generate_legacy_preset_details,
    list_generate_legacy_preset_summaries,
    list_ingest_preset_details,
    list_ingest_preset_summaries,
    load_generate_legacy_preset,
    load_ingest_preset,
)
from grantflow.api.public_views import (
    REVIEW_WORKFLOW_OVERDUE_DEFAULT_HOURS,
    REVIEW_WORKFLOW_STATE_FILTER_VALUES,
    public_checkpoint_payload,
    public_ingest_inventory_csv_text,
    public_ingest_recent_payload,
    public_job_citations_payload,
    public_job_comments_payload,
    public_job_critic_payload,
    public_job_diff_payload,
    public_job_events_payload,
    public_job_export_payload,
    public_job_grounding_gate_payload,
    public_job_metrics_payload,
    public_job_payload,
    public_job_quality_payload,
    public_job_review_workflow_csv_text,
    public_job_review_workflow_payload,
    public_job_review_workflow_sla_csv_text,
    public_job_review_workflow_sla_hotspots_csv_text,
    public_job_review_workflow_sla_hotspots_payload,
    public_job_review_workflow_sla_hotspots_trends_csv_text,
    public_job_review_workflow_sla_hotspots_trends_payload,
    public_job_review_workflow_sla_payload,
    public_job_review_workflow_sla_trends_csv_text,
    public_job_review_workflow_sla_trends_payload,
    public_job_review_workflow_trends_csv_text,
    public_job_review_workflow_trends_payload,
    public_job_versions_payload,
    public_portfolio_metrics_csv_text,
    public_portfolio_metrics_payload,
    public_portfolio_quality_csv_text,
    public_portfolio_quality_payload,
    public_portfolio_review_workflow_csv_text,
    public_portfolio_review_workflow_payload,
    public_portfolio_review_workflow_sla_csv_text,
    public_portfolio_review_workflow_sla_hotspots_csv_text,
    public_portfolio_review_workflow_sla_hotspots_payload,
    public_portfolio_review_workflow_sla_hotspots_trends_csv_text,
    public_portfolio_review_workflow_sla_hotspots_trends_payload,
    public_portfolio_review_workflow_sla_payload,
    public_portfolio_review_workflow_sla_trends_csv_text,
    public_portfolio_review_workflow_sla_trends_payload,
    public_portfolio_review_workflow_trends_csv_text,
    public_portfolio_review_workflow_trends_payload,
)
from grantflow.api.export_helpers import (
    _dead_letter_queue_csv_text,
    _extract_export_grounding_gate,
    _extract_export_runtime_grounded_quality_gate,
    _hitl_history_csv_text,
    _job_comments_csv_text,
    _job_events_csv_text,
    _portfolio_export_response,
    _resolve_export_inputs,
)
from grantflow.api.filters import _validated_filter_token
from grantflow.api.review_helpers import (
    _critic_findings_list_payload,
    _normalize_comment_sla_hours,
    _normalize_finding_sla_profile,
    _review_workflow_sla_profile_payload,
)
from grantflow.api.routers import include_api_routers
from grantflow.api.schemas import (
    CriticFindingsBulkStatusRequest,
    CriticFatalFlawPublicResponse,
    CriticFatalFlawStatusUpdatePublicResponse,
    CriticFindingsBulkStatusPublicResponse,
    CriticFindingsListPublicResponse,
    DemoGeneratePresetPublicResponse,
    DeadLetterQueueListPublicResponse,
    DeadLetterQueueMutationPublicResponse,
    DemoPresetBundlePublicResponse,
    ExportRequest,
    GenerateAcceptedPublicResponse,
    GenerateFromPresetBatchRequest,
    GenerateFromPresetAcceptedPublicResponse,
    GenerateFromPresetBatchPublicResponse,
    GenerateFromPresetRequest,
    GeneratePreflightRequest,
    GeneratePresetListPublicResponse,
    GeneratePreflightPublicResponse,
    GenerateLegacyPresetDetailPublicResponse,
    GenerateLegacyPresetListPublicResponse,
    GenerateRequest,
    HITLApprovalRequest,
    HITLPendingListPublicResponse,
    IngestInventoryPublicResponse,
    IngestReadinessRequest,
    IngestPresetDetailPublicResponse,
    IngestPresetListPublicResponse,
    IngestRecentListPublicResponse,
    JobCitationsPublicResponse,
    JobCommentCreateRequest,
    JobCommentsPublicResponse,
    JobCriticPublicResponse,
    JobDiffPublicResponse,
    JobEventsPublicResponse,
    JobExportPayloadPublicResponse,
    JobGroundingGatePublicResponse,
    JobHITLHistoryPublicResponse,
    JobMetricsPublicResponse,
    JobQualitySummaryPublicResponse,
    JobReviewWorkflowPublicResponse,
    JobReviewWorkflowSLAHotspotsPublicResponse,
    JobReviewWorkflowSLAHotspotsTrendsPublicResponse,
    JobReviewWorkflowSLAProfilePublicResponse,
    JobReviewWorkflowSLAPublicResponse,
    JobReviewWorkflowSLARecomputePublicResponse,
    JobReviewWorkflowSLATrendsPublicResponse,
    JobReviewWorkflowTrendsPublicResponse,
    JobStatusPublicResponse,
    JobVersionsPublicResponse,
    PortfolioMetricsPublicResponse,
    PortfolioQualityPublicResponse,
    PortfolioReviewWorkflowPublicResponse,
    PortfolioReviewWorkflowSLAHotspotsPublicResponse,
    PortfolioReviewWorkflowSLAHotspotsTrendsPublicResponse,
    PortfolioReviewWorkflowSLAPublicResponse,
    PortfolioReviewWorkflowSLATrendsPublicResponse,
    PortfolioReviewWorkflowTrendsPublicResponse,
    QueueWorkerHeartbeatPublicResponse,
    RBMSamplePresetDetailPublicResponse,
    RBMSamplePresetListPublicResponse,
    ReviewWorkflowSLARecomputeRequest,
    ReviewCommentPublicResponse,
)
from grantflow.api.security import (
    api_key_configured,
    install_openapi_api_key_security,
    require_api_key_if_configured,
)
from grantflow.api.tenant import (
    _allowed_tenant_tokens,
    _checkpoint_tenant_id,
    _default_tenant_token,
    _ensure_checkpoint_tenant_write_access,
    _ensure_job_tenant_read_access,
    _ensure_job_tenant_write_access,
    _filter_jobs_by_tenant,
    _job_donor_id,
    _job_tenant_id,
    _normalize_tenant_candidate,
    _resolve_tenant_id,
    _tenant_authz_enabled,
    _tenant_from_namespace,
    _tenant_rag_namespace,
)
from grantflow.api.webhooks import send_job_webhook_event
from grantflow.core.config import config
from grantflow.core.job_runner import InMemoryJobRunner, RedisJobRunner
from grantflow.core.stores import create_ingest_audit_store_from_env, create_job_store_from_env
from grantflow.core.strategies.factory import DonorFactory
from grantflow.core.version import __version__
from grantflow.eval.sample_presets import (
    available_sample_ids,
    build_generate_payload as build_sample_generate_payload,
    list_sample_preset_summaries,
    load_sample_payload,
)
from grantflow.exporters.donor_contracts import (
    DONOR_XLSX_PRIMARY_SHEET,
    evaluate_export_contract_gate,
    normalize_export_contract_policy_mode,
)
from grantflow.exporters.excel_builder import build_xlsx_from_logframe
from grantflow.exporters.template_profile import normalize_export_template_key
from grantflow.exporters.word_builder import build_docx_from_toc
from grantflow.memory_bank.ingest import ingest_pdf_to_namespace
from grantflow.memory_bank.vector_store import vector_store
from grantflow.swarm.findings import (
    canonicalize_findings,
    finding_primary_id,
    state_critic_findings,
    write_state_critic_findings,
)
from grantflow.swarm.graph import grantflow_graph
from grantflow.swarm.hitl import HITLStatus, hitl_manager
from grantflow.swarm.state_contract import (
    normalize_state_contract,
    normalized_state_copy,
    state_donor_id,
)

JOB_STORE = create_job_store_from_env()
INGEST_AUDIT_STORE = create_ingest_audit_store_from_env()
HITLStartAt = Literal["start", "architect", "mel", "critic"]
TERMINAL_JOB_STATUSES = {"done", "error", "canceled"}
JOB_RUNNER_MODES = {"background_tasks", "inmemory_queue", "redis_queue"}
STATUS_WEBHOOK_EVENTS = {
    "running": "job.started",
    "pending_hitl": "job.pending_hitl",
    "done": "job.completed",
    "error": "job.failed",
    "canceled": "job.canceled",
}
RUNTIME_PIPELINE_STATE_KEYS = {
    "_start_at",
    "hitl_checkpoint_stage",
    "hitl_resume_from",
    "hitl_checkpoint_id",
}
HITL_HISTORY_EVENT_TYPES = {
    "status_changed",
    "resume_requested",
    "hitl_checkpoint_published",
    "hitl_checkpoint_decision",
    "hitl_checkpoint_canceled",
}
PRODUCTION_ENV_TOKENS = {"prod", "production"}


def _job_runner_mode() -> str:
    raw_mode = str(getattr(config.job_runner, "mode", "background_tasks") or "background_tasks").strip().lower()
    if raw_mode not in JOB_RUNNER_MODES:
        return "background_tasks"
    return raw_mode


def _uses_inmemory_queue_runner() -> bool:
    return _job_runner_mode() == "inmemory_queue"


def _uses_redis_queue_runner() -> bool:
    return _job_runner_mode() == "redis_queue"


def _uses_queue_runner() -> bool:
    return _uses_inmemory_queue_runner() or _uses_redis_queue_runner()


def _dead_letter_alert_threshold() -> int:
    from grantflow.api.diagnostics_service import _dead_letter_alert_threshold as _impl

    return _impl()


def _dead_letter_alert_blocking() -> bool:
    from grantflow.api.diagnostics_service import _dead_letter_alert_blocking as _impl

    return _impl()


def _dispatcher_worker_heartbeat_policy_mode() -> str:
    from grantflow.api.diagnostics_service import _dispatcher_worker_heartbeat_policy_mode as _impl

    return _impl()


def _build_job_runner():
    worker_count = int(getattr(config.job_runner, "worker_count", 2) or 2)
    queue_maxsize = int(getattr(config.job_runner, "queue_maxsize", 200) or 200)
    consumer_enabled = bool(getattr(config.job_runner, "consumer_enabled", True))
    if _uses_redis_queue_runner():
        return RedisJobRunner(
            worker_count=worker_count,
            queue_maxsize=queue_maxsize,
            redis_url=str(getattr(config.job_runner, "redis_url", "redis://127.0.0.1:6379/0") or ""),
            queue_name=str(getattr(config.job_runner, "redis_queue_name", "grantflow:jobs") or ""),
            pop_timeout_seconds=float(getattr(config.job_runner, "redis_pop_timeout_seconds", 1.0) or 1.0),
            max_attempts=int(getattr(config.job_runner, "redis_max_attempts", 3) or 3),
            dead_letter_queue_name=str(getattr(config.job_runner, "redis_dead_letter_queue_name", "") or ""),
            worker_heartbeat_key=str(getattr(config.job_runner, "redis_worker_heartbeat_key", "") or ""),
            worker_heartbeat_ttl_seconds=float(
                getattr(config.job_runner, "redis_worker_heartbeat_ttl_seconds", 45.0) or 45.0
            ),
            consumer_enabled=consumer_enabled,
        )
    return InMemoryJobRunner(worker_count=worker_count, queue_maxsize=queue_maxsize)


JOB_RUNNER = _build_job_runner()


def _job_store_mode() -> str:
    return "sqlite" if getattr(JOB_STORE, "db_path", None) else "inmem"


def _hitl_store_mode() -> str:
    return "sqlite" if bool(getattr(hitl_manager, "_use_sqlite", False)) else "inmem"


def _ingest_store_mode() -> str:
    return "sqlite" if getattr(INGEST_AUDIT_STORE, "db_path", None) else "inmem"


def _validate_store_backend_alignment() -> None:
    job_store_mode = _job_store_mode()
    hitl_store_mode = _hitl_store_mode()
    if job_store_mode == hitl_store_mode:
        return
    raise RuntimeError(
        "Store backend mismatch: "
        f"JOB_STORE={job_store_mode} while HITL_STORE={hitl_store_mode}. "
        "Use matching backends for GRANTFLOW_JOB_STORE and GRANTFLOW_HITL_STORE."
    )


def _validate_tenant_authz_configuration() -> None:
    status = _tenant_authz_configuration_status()
    policy_mode = str(status.get("policy_mode") or "warn")
    enabled = bool(status.get("enabled"))
    allowed_tenant_count = int(status.get("allowed_tenant_count") or 0)
    default_tenant = str(status.get("default_tenant") or "").strip()
    issues = {str(item).strip().lower() for item in (status.get("issues") or []) if str(item).strip()}
    valid = bool(status.get("valid"))
    if policy_mode != "strict":
        return
    if valid:
        return
    if enabled and "allowlist_empty" in issues:
        raise RuntimeError(
            "Tenant authz misconfiguration: GRANTFLOW_TENANT_AUTHZ_ENABLED=true but tenant allowlist is empty. "
            "Set GRANTFLOW_ALLOWED_TENANTS or disable strict policy "
            "(GRANTFLOW_TENANT_AUTHZ_CONFIGURATION_POLICY_MODE=warn|off)."
        )
    if enabled and "default_tenant_not_in_allowlist" in issues:
        raise RuntimeError(
            "Tenant authz misconfiguration: GRANTFLOW_DEFAULT_TENANT is not included in GRANTFLOW_ALLOWED_TENANTS "
            f"(default={default_tenant}, allowed_count={allowed_tenant_count}). "
            "Set a matching default tenant or disable strict policy "
            "(GRANTFLOW_TENANT_AUTHZ_CONFIGURATION_POLICY_MODE=warn|off)."
        )


def _validate_runtime_compatibility_configuration() -> None:
    status = _python_runtime_compatibility_status()
    policy_mode = _configured_runtime_compatibility_policy_mode()
    supported = bool(status.get("supported"))
    if policy_mode != "strict":
        return
    if supported:
        return
    raise RuntimeError(
        "Runtime compatibility misconfiguration: Python "
        f"{status.get('python_version')} is outside validated range {status.get('supported_range')}. "
        "Use Python 3.11-3.13 or set GRANTFLOW_RUNTIME_COMPATIBILITY_POLICY_MODE=warn|off."
    )


def _deployment_environment() -> str:
    return str(os.getenv("GRANTFLOW_ENV", "dev") or "dev").strip().lower()


def _is_production_environment() -> bool:
    return _deployment_environment() in PRODUCTION_ENV_TOKENS


def _require_api_key_on_startup() -> bool:
    explicit = os.getenv("GRANTFLOW_REQUIRE_API_KEY_ON_STARTUP")
    if explicit is None or not str(explicit).strip():
        return _is_production_environment()
    return str(explicit).strip().lower() == "true"


def _validate_api_key_startup_security() -> None:
    if not _require_api_key_on_startup():
        return
    if api_key_configured():
        return
    raise RuntimeError(
        "Security defaults violation: API key auth is required at startup but GRANTFLOW_API_KEY is not set. "
        "Set GRANTFLOW_API_KEY or disable this guard with GRANTFLOW_REQUIRE_API_KEY_ON_STARTUP=false."
    )


@asynccontextmanager
async def _app_lifespan(_: FastAPI) -> AsyncIterator[None]:
    _validate_store_backend_alignment()
    _validate_tenant_authz_configuration()
    _validate_runtime_compatibility_configuration()
    _validate_api_key_startup_security()
    if _uses_queue_runner():
        JOB_RUNNER.start()
    try:
        yield
    finally:
        if _uses_queue_runner():
            JOB_RUNNER.stop()


app = FastAPI(
    title="GrantFlow API",
    description="Enterprise-grade grant proposal automation",
    version=__version__,
    lifespan=_app_lifespan,
)


def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _dispatch_pipeline_task(background_tasks: BackgroundTasks, fn: Callable[..., None], *args: Any) -> str:
    from grantflow.api.pipeline_jobs import _dispatch_pipeline_task as _impl

    return _impl(background_tasks, fn, *args)


def _parse_iso_utc(value: Any) -> Optional[datetime]:
    if not isinstance(value, str) or not value.strip():
        return None
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _iso_plus_hours(base_ts: Optional[str], hours: int) -> str:
    base_dt = _parse_iso_utc(base_ts) or datetime.now(timezone.utc)
    return (base_dt + timedelta(hours=max(1, int(hours)))).isoformat()


def _finding_sla_hours(severity: Any, *, finding_sla_hours_override: Optional[Dict[str, int]] = None) -> int:
    token = str(severity or "").strip().lower()
    source = finding_sla_hours_override if isinstance(finding_sla_hours_override, dict) else CRITIC_FINDING_SLA_HOURS
    return int(source.get(token, source.get("medium", CRITIC_FINDING_SLA_HOURS["medium"])))


def _comment_sla_hours(
    *,
    linked_finding_severity: Optional[str] = None,
    finding_sla_hours_override: Optional[Dict[str, int]] = None,
    default_comment_sla_hours: Optional[int] = None,
) -> int:
    if linked_finding_severity:
        return _finding_sla_hours(linked_finding_severity, finding_sla_hours_override=finding_sla_hours_override)
    if isinstance(default_comment_sla_hours, int) and default_comment_sla_hours > 0:
        return int(default_comment_sla_hours)
    return int(REVIEW_COMMENT_DEFAULT_SLA_HOURS)


def _finding_actor_from_request(request: Request) -> str:
    for header in ("x-reviewer", "x-actor", "x-user", "x-user-id", "x-email"):
        value = str(request.headers.get(header) or "").strip()
        if value:
            return value[:120]
    return "api_user"


def _normalize_request_id(value: Any) -> Optional[str]:
    from grantflow.api.idempotency import _normalize_request_id as _impl

    return _impl(value)


def _resolve_request_id(request: Request, explicit_request_id: Optional[str] = None) -> Optional[str]:
    from grantflow.api.idempotency import _resolve_request_id as _impl

    return _impl(request, explicit_request_id)


def _idempotency_fingerprint(payload: Dict[str, Any]) -> str:
    from grantflow.api.idempotency import _idempotency_fingerprint as _impl

    return _impl(payload)


def _idempotency_record_key(scope: str, request_id: str) -> str:
    from grantflow.api.idempotency import _idempotency_record_key as _impl

    return _impl(scope, request_id)


def _idempotency_records(job: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    from grantflow.api.idempotency import _idempotency_records as _impl

    return _impl(job)


def _idempotency_replay_response(
    job: Dict[str, Any],
    *,
    scope: str,
    request_id: Optional[str],
    fingerprint: str,
) -> Optional[Dict[str, Any]]:
    from grantflow.api.idempotency import _idempotency_replay_response as _impl

    return _impl(job, scope=scope, request_id=request_id, fingerprint=fingerprint)


def _store_idempotency_response(
    job_id: str,
    *,
    scope: str,
    request_id: Optional[str],
    fingerprint: str,
    response: Dict[str, Any],
    persisted: bool,
) -> None:
    from grantflow.api.idempotency import _store_idempotency_response as _impl

    _impl(
        job_id,
        scope=scope,
        request_id=request_id,
        fingerprint=fingerprint,
        response=response,
        persisted=persisted,
    )


def _global_idempotency_record_key(scope: str, request_id: str) -> str:
    from grantflow.api.idempotency import _global_idempotency_record_key as _impl

    return _impl(scope, request_id)


def _global_idempotency_replay_response(
    *,
    scope: str,
    request_id: Optional[str],
    fingerprint: str,
) -> Optional[Dict[str, Any]]:
    from grantflow.api.idempotency import _global_idempotency_replay_response as _impl

    return _impl(scope=scope, request_id=request_id, fingerprint=fingerprint)


def _store_global_idempotency_response(
    *,
    scope: str,
    request_id: Optional[str],
    fingerprint: str,
    response: Dict[str, Any],
    persisted: bool,
) -> None:
    from grantflow.api.idempotency import _store_global_idempotency_response as _impl

    _impl(
        scope=scope,
        request_id=request_id,
        fingerprint=fingerprint,
        response=response,
        persisted=persisted,
    )


def _append_job_event_records(
    previous: Optional[Dict[str, Any]],
    next_payload: Dict[str, Any],
) -> Dict[str, Any]:
    from grantflow.api.job_store_service import _append_job_event_records as _impl

    return _impl(previous, next_payload)


def _record_job_event(job_id: str, event_type: str, **fields: Any) -> None:
    from grantflow.api.job_store_service import _record_job_event as _impl

    _impl(job_id, event_type, **fields)


def _record_ingest_event(
    *,
    donor_id: str,
    namespace: str,
    filename: str,
    content_type: str,
    metadata: Optional[Dict[str, Any]] = None,
    result: Optional[Dict[str, Any]] = None,
) -> None:
    from grantflow.api.job_store_service import _record_ingest_event as _impl

    _impl(
        donor_id=donor_id,
        namespace=namespace,
        filename=filename,
        content_type=content_type,
        metadata=metadata,
        result=result,
    )


def _configured_tenant_authz_configuration_policy_mode() -> str:
    from grantflow.api.diagnostics_service import _configured_tenant_authz_configuration_policy_mode as _impl

    return _impl()


def _tenant_authz_configuration_status() -> dict[str, Any]:
    from grantflow.api.diagnostics_service import _tenant_authz_configuration_status as _impl

    return _impl()


def _list_ingest_events(
    *,
    donor_id: Optional[str] = None,
    tenant_id: Optional[str] = None,
    limit: int = 50,
) -> list[Dict[str, Any]]:
    from grantflow.api.job_store_service import _list_ingest_events as _impl

    return _impl(donor_id=donor_id, tenant_id=tenant_id, limit=limit)


def _ingest_inventory(*, donor_id: Optional[str] = None, tenant_id: Optional[str] = None) -> list[Dict[str, Any]]:
    from grantflow.api.job_store_service import _ingest_inventory as _impl

    return _impl(donor_id=donor_id, tenant_id=tenant_id)


def _dedupe_doc_families(values: list[Any]) -> list[str]:
    from grantflow.api.preflight_service import _dedupe_doc_families as _impl

    return _impl(values)


def _preflight_expected_doc_families(
    *,
    donor_id: str,
    client_metadata: Optional[Dict[str, Any]],
) -> list[str]:
    from grantflow.api.preflight_service import _preflight_expected_doc_families as _impl

    return _impl(donor_id=donor_id, client_metadata=client_metadata)


def _preflight_doc_family_min_uploads_map(
    *,
    expected_doc_families: list[str],
    client_metadata: Optional[Dict[str, Any]],
) -> Dict[str, int]:
    from grantflow.api.preflight_service import _preflight_doc_family_min_uploads_map as _impl

    return _impl(
        expected_doc_families=expected_doc_families,
        client_metadata=client_metadata,
    )


def _preflight_doc_family_depth_profile(
    *,
    expected_doc_families: list[str],
    doc_family_counts: Dict[str, Any],
    min_uploads_by_family: Dict[str, int],
) -> Dict[str, Any]:
    from grantflow.api.preflight_service import _preflight_doc_family_depth_profile as _impl

    return _impl(
        expected_doc_families=expected_doc_families,
        doc_family_counts=doc_family_counts,
        min_uploads_by_family=min_uploads_by_family,
    )


def _preflight_input_context(client_metadata: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    from grantflow.api.preflight_service import _preflight_input_context as _impl

    return _impl(client_metadata)


def _preflight_severity_max(severities: list[str]) -> str:
    from grantflow.api.preflight_service import _preflight_severity_max as _impl

    return _impl(severities)


def _normalize_grounding_policy_mode(raw_mode: Any) -> str:
    mode = str(raw_mode or "warn").strip().lower()
    if mode not in GROUNDING_POLICY_MODES:
        return "warn"
    return mode


def _configured_runtime_compatibility_policy_mode() -> str:
    from grantflow.api.diagnostics_service import _configured_runtime_compatibility_policy_mode as _impl

    return _impl()


def _python_runtime_compatibility_status() -> Dict[str, Any]:
    from grantflow.api.diagnostics_service import _python_runtime_compatibility_status as _impl

    return _impl()


def _configured_preflight_grounding_policy_mode() -> str:
    from grantflow.api.preflight_service import _configured_preflight_grounding_policy_mode as _impl

    return _impl()


def _configured_runtime_grounded_quality_gate_mode() -> str:
    from grantflow.api.runtime_grounded_gate_service import _configured_runtime_grounded_quality_gate_mode as _impl

    return _impl()


def _runtime_grounded_quality_gate_thresholds() -> Dict[str, Any]:
    from grantflow.api.runtime_grounded_gate_service import _runtime_grounded_quality_gate_thresholds as _impl

    return _impl()


def _runtime_grounded_gate_section(citation: Dict[str, Any]) -> str:
    from grantflow.api.runtime_grounded_gate_service import _runtime_grounded_gate_section as _impl

    return _impl(citation)


def _runtime_grounded_gate_evidence_row(citation: Dict[str, Any]) -> Dict[str, Any]:
    from grantflow.api.runtime_grounded_gate_service import _runtime_grounded_gate_evidence_row as _impl

    return _impl(citation)


def _evaluate_runtime_grounded_quality_gate_from_state(state: Any) -> Dict[str, Any]:
    from grantflow.api.runtime_grounded_gate_service import (
        _evaluate_runtime_grounded_quality_gate_from_state as _impl,
    )

    return _impl(state)


def _configured_mel_grounding_policy_mode() -> str:
    from grantflow.api.grounding_policy_service import _configured_mel_grounding_policy_mode as _impl

    return _impl()


def _mel_grounding_policy_thresholds() -> Dict[str, Any]:
    from grantflow.api.grounding_policy_service import _mel_grounding_policy_thresholds as _impl

    return _impl()


def _evaluate_mel_grounding_policy_from_state(state: Any) -> Dict[str, Any]:
    from grantflow.api.grounding_policy_service import _evaluate_mel_grounding_policy_from_state as _impl

    return _impl(state)


def _configured_export_grounding_policy_mode() -> str:
    from grantflow.api.grounding_policy_service import _configured_export_grounding_policy_mode as _impl

    return _impl()


def _configured_export_require_grounded_gate_pass() -> bool:
    from grantflow.api.grounding_policy_service import _configured_export_require_grounded_gate_pass as _impl

    return _impl()


def _export_grounding_policy_thresholds() -> Dict[str, Any]:
    from grantflow.api.grounding_policy_service import _export_grounding_policy_thresholds as _impl

    return _impl()


def _evaluate_export_grounding_policy(citations: list[dict[str, Any]]) -> Dict[str, Any]:
    from grantflow.api.grounding_policy_service import _evaluate_export_grounding_policy as _impl

    return _impl(citations)


def _configured_export_contract_policy_mode() -> str:
    contract_mode = getattr(config.graph, "export_contract_policy_mode", None)
    if str(contract_mode or "").strip():
        return normalize_export_contract_policy_mode(contract_mode)
    return _configured_export_grounding_policy_mode()


def _evaluate_export_contract_gate(
    *,
    donor_id: str,
    toc_draft: dict[str, Any],
    workbook_sheetnames: Optional[list[str]] = None,
    workbook_primary_sheet_headers: Optional[list[str]] = None,
) -> Dict[str, Any]:
    return evaluate_export_contract_gate(
        donor_id=donor_id,
        toc_payload=toc_draft,
        policy_mode=_configured_export_contract_policy_mode(),
        workbook_sheetnames=workbook_sheetnames,
        workbook_primary_sheet_headers=workbook_primary_sheet_headers,
    )


def _xlsx_contract_validation_context(
    xlsx_payload: bytes,
    *,
    donor_id: str,
) -> tuple[list[str], list[str]]:
    workbook = load_workbook(io.BytesIO(xlsx_payload), data_only=True, read_only=True)
    try:
        sheetnames = list(workbook.sheetnames)
        donor_key = normalize_export_template_key(donor_id)
        primary_sheet = DONOR_XLSX_PRIMARY_SHEET.get(donor_key)
        if not primary_sheet or primary_sheet not in workbook.sheetnames:
            return sheetnames, []
        sheet = workbook[primary_sheet]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(value).strip() for value in header_row if str(value or "").strip()]
        return sheetnames, headers
    finally:
        workbook.close()


def _attach_export_contract_gate(state: Any) -> Dict[str, Any]:
    state_dict: dict[str, Any] = state if isinstance(state, dict) else {}
    donor_id = state_donor_id(state_dict, default="grantflow")
    raw_toc = state_dict.get("toc_draft")
    toc_draft = raw_toc if isinstance(raw_toc, dict) else {}
    if not toc_draft:
        raw_toc_fallback = state_dict.get("toc")
        if isinstance(raw_toc_fallback, dict):
            toc_draft = raw_toc_fallback
    gate = _evaluate_export_contract_gate(donor_id=donor_id, toc_draft=toc_draft)
    state_dict["export_contract_gate"] = gate
    return gate


def _preflight_grounding_policy_thresholds() -> Dict[str, Any]:
    from grantflow.api.preflight_service import _preflight_grounding_policy_thresholds as _impl

    return _impl()


def _estimate_preflight_architect_claims(
    *,
    donor_id: str,
    strategy: Any,
    namespace: str,
    input_context: Optional[Dict[str, Any]],
    tenant_id: Optional[str] = None,
    architect_rag_enabled: bool = True,
) -> Dict[str, Any]:
    from grantflow.api.preflight_service import _estimate_preflight_architect_claims as _impl

    return _impl(
        donor_id=donor_id,
        strategy=strategy,
        namespace=namespace,
        input_context=input_context,
        tenant_id=tenant_id,
        architect_rag_enabled=architect_rag_enabled,
    )


def _build_preflight_grounding_policy(
    *,
    coverage_rate: Optional[float],
    depth_coverage_rate: Optional[float],
    namespace_empty: bool,
    inventory_total_uploads: int,
    missing_doc_families: list[str],
    depth_gap_doc_families: list[str],
    architect_claims: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    from grantflow.api.preflight_service import _build_preflight_grounding_policy as _impl

    return _impl(
        coverage_rate=coverage_rate,
        depth_coverage_rate=depth_coverage_rate,
        namespace_empty=namespace_empty,
        inventory_total_uploads=inventory_total_uploads,
        missing_doc_families=missing_doc_families,
        depth_gap_doc_families=depth_gap_doc_families,
        architect_claims=architect_claims,
    )


def _build_generate_preflight(
    *,
    donor_id: str,
    strategy: Any,
    client_metadata: Optional[Dict[str, Any]],
    tenant_id: Optional[str] = None,
    architect_rag_enabled: bool = True,
) -> Dict[str, Any]:
    from grantflow.api.preflight_service import _build_generate_preflight as _impl

    return _impl(
        donor_id=donor_id,
        strategy=strategy,
        client_metadata=client_metadata,
        tenant_id=tenant_id,
        architect_rag_enabled=architect_rag_enabled,
    )


def _set_job(job_id: str, payload: Dict[str, Any]) -> None:
    from grantflow.api.job_store_service import _set_job as _impl

    _impl(job_id, payload)


def _update_job(job_id: str, **patch: Any) -> Dict[str, Any]:
    from grantflow.api.job_store_service import _update_job as _impl

    return _impl(job_id, **patch)


def _get_job(job_id: str) -> Optional[Dict[str, Any]]:
    from grantflow.api.job_store_service import _get_job as _impl

    return _impl(job_id)


def _list_jobs() -> Dict[str, Dict[str, Any]]:
    from grantflow.api.job_store_service import _list_jobs as _impl

    return _impl()


def _find_job_by_checkpoint_id(checkpoint_id: str) -> tuple[Optional[str], Optional[Dict[str, Any]]]:
    token = str(checkpoint_id or "").strip()
    if not token:
        return None, None
    for job_id, job in _list_jobs().items():
        if not isinstance(job, dict):
            continue
        if str(job.get("checkpoint_id") or "").strip() == token:
            return str(job_id), job
    return None, None


def _is_hitl_history_event(event: Dict[str, Any]) -> bool:
    event_type = str(event.get("type") or "").strip()
    if event_type not in HITL_HISTORY_EVENT_TYPES:
        return False
    if event_type != "status_changed":
        return True
    from_status = str(event.get("from_status") or "").strip().lower()
    to_status = str(event.get("to_status") or "").strip().lower()
    return from_status == "pending_hitl" or to_status == "pending_hitl"


def _hitl_history_payload(
    job_id: str,
    job: Dict[str, Any],
    *,
    event_type: Optional[str] = None,
    checkpoint_id: Optional[str] = None,
) -> Dict[str, Any]:
    event_type_filter = str(event_type or "").strip() or None
    if event_type_filter and event_type_filter not in HITL_HISTORY_EVENT_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported event_type filter")
    checkpoint_filter = str(checkpoint_id or "").strip() or None
    raw_events_value = job.get("job_events")
    raw_events: list[Any] = raw_events_value if isinstance(raw_events_value, list) else []
    events: list[Dict[str, Any]] = []
    for row in raw_events:
        if not isinstance(row, dict) or not _is_hitl_history_event(row):
            continue
        row_type = str(row.get("type") or "").strip()
        row_checkpoint_id = str(row.get("checkpoint_id") or "").strip() or None
        if event_type_filter and row_type != event_type_filter:
            continue
        if checkpoint_filter and row_checkpoint_id != checkpoint_filter:
            continue
        status = str(row.get("status") or "").strip()
        from_status = str(row.get("from_status") or "").strip()
        to_status = str(row.get("to_status") or "").strip()
        event: Dict[str, Any] = {
            "event_id": row.get("event_id"),
            "ts": row.get("ts"),
            "type": row_type,
            "from_status": from_status or None,
            "to_status": to_status or None,
            "status": status or None,
            "checkpoint_id": row_checkpoint_id,
            "checkpoint_stage": str(row.get("checkpoint_stage") or row.get("stage") or "").strip() or None,
            "checkpoint_status": str(row.get("checkpoint_status") or "").strip() or None,
            "resuming_from": str(row.get("resuming_from") or "").strip() or None,
            "approved": row.get("approved") if isinstance(row.get("approved"), bool) else None,
            "feedback": str(row.get("feedback") or "").strip() or None,
            "actor": str(row.get("actor") or "").strip() or None,
            "request_id": str(row.get("request_id") or "").strip() or None,
            "reason": str(row.get("reason") or "").strip() or None,
            "backend": str(row.get("backend") or "").strip() or None,
        }
        events.append(event)
    events.sort(key=lambda item: str(item.get("ts") or ""), reverse=True)
    event_type_counts: Dict[str, int] = {}
    for row in events:
        row_type = str(row.get("type") or "").strip()
        if row_type:
            event_type_counts[row_type] = int(event_type_counts.get(row_type, 0)) + 1
    return {
        "job_id": str(job_id),
        "status": str(job.get("status") or ""),
        "filters": {"event_type": event_type_filter, "checkpoint_id": checkpoint_filter},
        "event_count": len(events),
        "event_type_counts": event_type_counts,
        "events": events,
    }


def _record_hitl_feedback_in_state(state: dict, checkpoint: Dict[str, Any]) -> None:
    from grantflow.api.pipeline_jobs import _record_hitl_feedback_in_state as _impl

    _impl(state, checkpoint)


def _checkpoint_status_token(checkpoint: Dict[str, Any]) -> str:
    from grantflow.api.pipeline_jobs import _checkpoint_status_token as _impl

    return _impl(checkpoint)


def _state_grounding_gate(state: Any) -> Dict[str, Any]:
    if not isinstance(state, dict):
        return {}
    gate = state.get("grounding_gate")
    return gate if isinstance(gate, dict) else {}


def _state_runtime_grounded_quality_gate(state: Any) -> Dict[str, Any]:
    if not isinstance(state, dict):
        return {}
    gate = state.get("grounded_quality_gate")
    return gate if isinstance(gate, dict) else {}


def _append_runtime_grounded_quality_gate_finding(state: dict, gate: Dict[str, Any]) -> None:
    if not isinstance(state, dict):
        return
    reasons = gate.get("reasons") if isinstance(gate.get("reasons"), list) else []
    reason_details = gate.get("reason_details") if isinstance(gate.get("reason_details"), list) else []
    raw_failed_sections = gate.get("failed_sections")
    failed_sections: list[Any] = raw_failed_sections if isinstance(raw_failed_sections, list) else []
    related_sections = [
        token
        for token in [str(section or "").strip().lower() for section in failed_sections]
        if token in {"toc", "logframe", "general"}
    ]
    primary_section = related_sections[0] if related_sections else "general"
    thresholds = gate.get("thresholds") if isinstance(gate.get("thresholds"), dict) else {}
    non_retrieval_rate = gate.get("non_retrieval_citation_rate")
    retrieval_grounded_count = gate.get("retrieval_grounded_citation_count")
    citation_count = gate.get("citation_count")
    summary = str(gate.get("summary") or "").strip() or "runtime grounded quality gate failed"
    rationale = (
        f"{summary}; citation_count={citation_count}; "
        f"non_retrieval_rate={non_retrieval_rate}; "
        f"retrieval_grounded_count={retrieval_grounded_count}; "
        f"thresholds={thresholds}; reasons={reasons}; "
        f"reason_details={reason_details}; failed_sections={failed_sections}"
    )
    existing = state_critic_findings(state, default_source="rules")
    existing_codes = {str(item.get("code") or "").strip().upper() for item in existing if isinstance(item, dict)}
    if "RUNTIME_GROUNDED_QUALITY_GATE_BLOCK" in existing_codes:
        return
    new_finding = {
        "code": "RUNTIME_GROUNDED_QUALITY_GATE_BLOCK",
        "severity": "high",
        "section": primary_section,
        "related_sections": related_sections,
        "version_id": None,
        "message": "Grounded quality gate blocked finalization for LLM generation.",
        "rationale": rationale,
        "fix_suggestion": (
            "Upload additional donor/country evidence and rerun generation to increase retrieval-grounded citations "
            "and reduce non-retrieval citations."
        ),
        "fix_hint": "Use /ingest for relevant policy/context PDFs, then rerun /generate in grounded mode.",
        "source": "rules",
    }
    write_state_critic_findings(
        state,
        list(existing) + [new_finding],
        previous_items=existing,
        default_source="rules",
    )


def _grounding_gate_block_reason(state: Any) -> Optional[str]:
    gate = _state_grounding_gate(state)
    if not gate:
        return None
    if not bool(gate.get("blocking")):
        return None
    if str(gate.get("mode") or "").lower() != "strict":
        return None
    summary = str(gate.get("summary") or "").strip() or "weak grounding signals"
    return f"Grounding gate (strict) blocked finalization: {summary}"


def _runtime_grounded_quality_gate_block_reason(state: Any) -> Optional[str]:
    gate = _state_runtime_grounded_quality_gate(state)
    if not gate:
        return None
    if not bool(gate.get("blocking")):
        return None
    if str(gate.get("mode") or "").lower() != "strict":
        return None
    summary = str(gate.get("summary") or "").strip() or "runtime grounded signals not acceptable"
    return f"Grounded quality gate (strict) blocked finalization: {summary}"


def _mel_grounding_policy_block_reason(state: Any) -> Optional[str]:
    policy = _evaluate_mel_grounding_policy_from_state(state)
    state_dict = state if isinstance(state, dict) else {}
    state_dict["mel_grounding_policy"] = policy
    if not bool(policy.get("blocking")):
        return None
    summary = str(policy.get("summary") or "").strip() or "weak mel grounding signals"
    return f"MEL grounding policy (strict) blocked finalization: {summary}"


def _job_draft_version_exists_for_section(job: Dict[str, Any], *, section: str, version_id: str) -> bool:
    state = job.get("state")
    if not isinstance(state, dict):
        return False
    raw_versions = state.get("draft_versions")
    if not isinstance(raw_versions, list):
        return False
    for item in raw_versions:
        if not isinstance(item, dict):
            continue
        if str(item.get("version_id") or "") != version_id:
            continue
        if str(item.get("section") or "") != section:
            continue
        return True
    return False


def _resolve_sla_profile_for_recompute(
    *,
    job: Dict[str, Any],
    finding_sla_hours_override: Optional[Dict[str, Any]],
    default_comment_sla_hours: Optional[Any],
    use_saved_profile: bool,
) -> tuple[Dict[str, int], int]:
    base_finding = dict(CRITIC_FINDING_SLA_HOURS)
    base_comment = int(REVIEW_COMMENT_DEFAULT_SLA_HOURS)

    if use_saved_profile:
        client_metadata = job.get("client_metadata")
        metadata = client_metadata if isinstance(client_metadata, dict) else {}
        saved_profile = metadata.get("sla_profile")
        saved_dict = saved_profile if isinstance(saved_profile, dict) else {}
        saved_finding = saved_dict.get("finding_sla_hours")
        saved_comment = saved_dict.get("default_comment_sla_hours")
        base_finding = _normalize_finding_sla_profile(saved_finding, default=base_finding)
        base_comment = _normalize_comment_sla_hours(saved_comment)

    resolved_finding = _normalize_finding_sla_profile(finding_sla_hours_override, default=base_finding)
    if default_comment_sla_hours is None:
        resolved_comment = base_comment
    else:
        resolved_comment = _normalize_comment_sla_hours(default_comment_sla_hours)
    return resolved_finding, resolved_comment


def _ensure_finding_due_at(
    item: Dict[str, Any],
    *,
    now_iso: str,
    reset: bool = False,
    finding_sla_hours_override: Optional[Dict[str, int]] = None,
) -> Dict[str, Any]:
    current = dict(item)
    status = str(current.get("status") or "open").strip().lower()
    if status == "resolved":
        return current
    if not reset and str(current.get("due_at") or "").strip():
        return current
    sla_hours = _finding_sla_hours(current.get("severity"), finding_sla_hours_override=finding_sla_hours_override)
    base_ts = None
    if status == "acknowledged":
        base_ts = str(current.get("acknowledged_at") or current.get("updated_at") or now_iso)
    else:
        base_ts = str(current.get("updated_at") or now_iso)
    current["due_at"] = _iso_plus_hours(base_ts, sla_hours)
    current["sla_hours"] = sla_hours
    return current


def _normalize_critic_fatal_flaws_for_job(job_id: str) -> Optional[Dict[str, Any]]:
    job = _get_job(job_id)
    if not job:
        return None
    state = job.get("state")
    if not isinstance(state, dict):
        return job
    raw_flaws = state_critic_findings(state, default_source="rules")
    if not raw_flaws:
        return job

    now_iso = _utcnow_iso()
    normalized_with_due = [_ensure_finding_due_at(item, now_iso=now_iso) for item in raw_flaws]
    existing_notes_raw = state.get("critic_notes")
    existing_notes: Dict[str, Any] = existing_notes_raw if isinstance(existing_notes_raw, dict) else {}
    existing_notes_flaws = (
        existing_notes.get("fatal_flaws") if isinstance(existing_notes.get("fatal_flaws"), list) else []
    )
    existing_state_flaws = state.get("critic_fatal_flaws") if isinstance(state.get("critic_fatal_flaws"), list) else []
    changed = normalized_with_due != existing_notes_flaws or normalized_with_due != existing_state_flaws

    if not changed:
        return job

    next_state = dict(state)
    write_state_critic_findings(
        next_state, normalized_with_due, previous_items=normalized_with_due, default_source="rules"
    )
    return _update_job(job_id, state=next_state)


def _normalize_review_comments_for_job(job_id: str) -> Optional[Dict[str, Any]]:
    job = _get_job(job_id)
    if not job:
        return None
    raw_comments = job.get("review_comments")
    if not isinstance(raw_comments, list):
        return job
    comments = [c for c in raw_comments if isinstance(c, dict)]
    now_iso = _utcnow_iso()
    normalized_comments = [_ensure_comment_due_at(comment, job=job, now_iso=now_iso) for comment in comments]
    if normalized_comments == comments:
        return job
    return _update_job(job_id, review_comments=normalized_comments[-500:])


def _recompute_review_workflow_sla(
    job_id: str,
    *,
    actor: Optional[str] = None,
    finding_sla_hours_override: Optional[Dict[str, Any]] = None,
    default_comment_sla_hours: Optional[Any] = None,
    use_saved_profile: bool = False,
) -> Dict[str, Any]:
    from grantflow.api.review_mutations import _recompute_review_workflow_sla as _impl

    return _impl(
        job_id,
        actor=actor,
        finding_sla_hours_override=finding_sla_hours_override,
        default_comment_sla_hours=default_comment_sla_hours,
        use_saved_profile=use_saved_profile,
    )


def _find_critic_fatal_flaw(job: Dict[str, Any], finding_id: str) -> Optional[Dict[str, Any]]:
    state = job.get("state")
    if not isinstance(state, dict):
        return None
    flaws = state_critic_findings(state, default_source="rules")
    for item in flaws:
        if not isinstance(item, dict):
            continue
        if finding_primary_id(item) == finding_id:
            return item
    return None


def _set_critic_fatal_flaw_status(
    job_id: str,
    *,
    finding_id: str,
    next_status: str,
    actor: Optional[str] = None,
    dry_run: bool = False,
    if_match_status: Optional[str] = None,
    request_id: Optional[str] = None,
) -> Dict[str, Any]:
    from grantflow.api.review_mutations import _set_critic_fatal_flaw_status as _impl

    return _impl(
        job_id,
        finding_id=finding_id,
        next_status=next_status,
        actor=actor,
        dry_run=dry_run,
        if_match_status=if_match_status,
        request_id=request_id,
    )


def _set_critic_fatal_flaws_status_bulk(
    job_id: str,
    *,
    next_status: str,
    actor: Optional[str] = None,
    dry_run: bool = False,
    request_id: Optional[str] = None,
    if_match_status: Optional[str] = None,
    apply_to_all: bool = False,
    finding_status: Optional[str] = None,
    severity: Optional[str] = None,
    section: Optional[str] = None,
    finding_ids: Optional[list[str]] = None,
) -> Dict[str, Any]:
    from grantflow.api.review_mutations import _set_critic_fatal_flaws_status_bulk as _impl

    return _impl(
        job_id,
        next_status=next_status,
        actor=actor,
        dry_run=dry_run,
        request_id=request_id,
        if_match_status=if_match_status,
        apply_to_all=apply_to_all,
        finding_status=finding_status,
        severity=severity,
        section=section,
        finding_ids=finding_ids,
    )


def _linked_finding_severity(job: Dict[str, Any], linked_finding_id: Optional[str]) -> Optional[str]:
    token = str(linked_finding_id or "").strip()
    if not token:
        return None
    finding = _find_critic_fatal_flaw(job, token)
    if not isinstance(finding, dict):
        return None
    severity = str(finding.get("severity") or "").strip().lower()
    return severity or None


def _ensure_comment_due_at(
    comment: Dict[str, Any],
    *,
    job: Dict[str, Any],
    now_iso: str,
    reset: bool = False,
    finding_sla_hours_override: Optional[Dict[str, int]] = None,
    default_comment_sla_hours: Optional[int] = None,
) -> Dict[str, Any]:
    current = dict(comment)
    status = str(current.get("status") or "open").strip().lower()
    if status == "resolved":
        return current
    if not reset and str(current.get("due_at") or "").strip():
        if not current.get("sla_hours"):
            inferred_sla = _comment_sla_hours(
                linked_finding_severity=_linked_finding_severity(
                    job, str(current.get("linked_finding_id") or "").strip()
                ),
                finding_sla_hours_override=finding_sla_hours_override,
                default_comment_sla_hours=default_comment_sla_hours,
            )
            current["sla_hours"] = inferred_sla
        return current
    linked_finding_id = str(current.get("linked_finding_id") or "").strip() or None
    severity = _linked_finding_severity(job, linked_finding_id)
    sla_hours = _comment_sla_hours(
        linked_finding_severity=severity,
        finding_sla_hours_override=finding_sla_hours_override,
        default_comment_sla_hours=default_comment_sla_hours,
    )
    base_ts = str(current.get("updated_ts") or current.get("ts") or now_iso)
    current["sla_hours"] = sla_hours
    current["due_at"] = _iso_plus_hours(base_ts, sla_hours)
    return current


def _append_review_comment(
    job_id: str,
    *,
    section: str,
    message: str,
    author: Optional[str] = None,
    version_id: Optional[str] = None,
    linked_finding_id: Optional[str] = None,
    linked_finding_severity: Optional[str] = None,
    request_id: Optional[str] = None,
) -> Dict[str, Any]:
    from grantflow.api.review_mutations import _append_review_comment as _impl

    return _impl(
        job_id,
        section=section,
        message=message,
        author=author,
        version_id=version_id,
        linked_finding_id=linked_finding_id,
        linked_finding_severity=linked_finding_severity,
        request_id=request_id,
    )


def _set_review_comment_status(
    job_id: str,
    *,
    comment_id: str,
    next_status: str,
    actor: Optional[str] = None,
    request_id: Optional[str] = None,
) -> Dict[str, Any]:
    from grantflow.api.review_mutations import _set_review_comment_status as _impl

    return _impl(
        job_id,
        comment_id=comment_id,
        next_status=next_status,
        actor=actor,
        request_id=request_id,
    )


def _job_is_canceled(job_id: str) -> bool:
    job = _get_job(job_id)
    return bool(job and job.get("status") == "canceled")


def _dispatch_job_webhook_for_status_change(
    job_id: str,
    previous: Optional[Dict[str, Any]],
    current: Optional[Dict[str, Any]],
) -> None:
    if not current:
        return

    previous_status = (previous or {}).get("status")
    current_status = current.get("status")
    if previous_status == current_status:
        return

    event_name = STATUS_WEBHOOK_EVENTS.get(str(current_status))
    if not event_name:
        return

    webhook_url = str(current.get("webhook_url") or (previous or {}).get("webhook_url") or "").strip()
    if not webhook_url:
        return

    webhook_secret = current.get("webhook_secret") or (previous or {}).get("webhook_secret")
    public_payload = public_job_payload(current)
    try:
        send_job_webhook_event(
            url=webhook_url,
            secret=str(webhook_secret) if webhook_secret else None,
            event=event_name,
            job_id=job_id,
            job=public_payload,
        )
    except Exception:
        # Webhook delivery failures are non-fatal for the job lifecycle.
        pass


install_openapi_api_key_security(app)


def _pause_for_hitl(job_id: str, state: dict, stage: Literal["toc", "logframe"], resume_from: HITLStartAt) -> None:
    existing_checkpoint_id = str(state.get("hitl_checkpoint_id") or "").strip() or None
    if _job_is_canceled(job_id):
        if existing_checkpoint_id:
            hitl_manager.cancel(existing_checkpoint_id, "Canceled before HITL checkpoint was published")
        return
    _clear_hitl_runtime_state(state, clear_pending=False)
    state["hitl_pending"] = True
    normalize_state_contract(state)
    donor_id = state_donor_id(state, default="unknown")
    checkpoint_id = existing_checkpoint_id
    if checkpoint_id:
        checkpoint = hitl_manager.get_checkpoint(checkpoint_id)
        checkpoint_stage = str(checkpoint.get("stage") or "").strip().lower() if isinstance(checkpoint, dict) else ""
        checkpoint_status = _checkpoint_status_token(checkpoint) if isinstance(checkpoint, dict) else ""
        if not checkpoint or checkpoint_status != HITLStatus.PENDING.value or checkpoint_stage != stage:
            if isinstance(checkpoint, dict) and checkpoint_status == HITLStatus.PENDING.value and checkpoint_id:
                hitl_manager.cancel(checkpoint_id, "Superseded by new HITL checkpoint")
            checkpoint_id = None
    if not checkpoint_id:
        checkpoint_id = hitl_manager.create_checkpoint(stage, state, donor_id)
    if _job_is_canceled(job_id):
        hitl_manager.cancel(checkpoint_id, "Canceled before HITL checkpoint was published")
        return
    _set_job(
        job_id,
        {
            "status": "pending_hitl",
            "state": state,
            "checkpoint_id": checkpoint_id,
            "checkpoint_stage": stage,
            "resume_from": resume_from,
            "hitl_enabled": True,
        },
    )
    _record_job_event(
        job_id,
        "hitl_checkpoint_published",
        checkpoint_id=checkpoint_id,
        checkpoint_stage=stage,
        resume_from=resume_from,
    )


def _run_pipeline_to_completion(job_id: str, initial_state: dict) -> None:
    from grantflow.api.pipeline_jobs import _run_pipeline_to_completion as _impl

    _impl(job_id, initial_state)


def _run_pipeline_to_completion_by_job_id(job_id: str) -> None:
    from grantflow.api.pipeline_jobs import _run_pipeline_to_completion_by_job_id as _impl

    _impl(job_id)


def _run_hitl_pipeline(job_id: str, state: dict, start_at: HITLStartAt) -> None:
    from grantflow.api.pipeline_jobs import _run_hitl_pipeline as _impl

    _impl(job_id, state, start_at)


def _run_hitl_pipeline_by_job_id(job_id: str, start_at: HITLStartAt) -> None:
    from grantflow.api.pipeline_jobs import _run_hitl_pipeline_by_job_id as _impl

    _impl(job_id, start_at)


def _resume_target_from_checkpoint(checkpoint: Dict[str, Any], default_resume_from: str | None) -> HITLStartAt:
    from grantflow.api.pipeline_jobs import _resume_target_from_checkpoint as _impl

    return _impl(checkpoint, default_resume_from)


def _clear_hitl_runtime_state(state: dict, *, clear_pending: bool) -> None:
    from grantflow.api.pipeline_jobs import _clear_hitl_runtime_state as _impl

    _impl(state, clear_pending=clear_pending)


def _configuration_warnings() -> list[dict[str, Any]]:
    from grantflow.api.diagnostics_service import _configuration_warnings as _impl

    return _impl()


def _health_diagnostics() -> dict[str, Any]:
    from grantflow.api.diagnostics_service import _health_diagnostics as _impl

    return _impl()


def _vector_store_readiness() -> dict[str, Any]:
    from grantflow.api.diagnostics_service import _vector_store_readiness as _impl

    return _impl()


@app.get("/health")
def health_check():
    return {"status": "healthy", "version": __version__, "diagnostics": _health_diagnostics()}


@app.get("/ready")
def readiness_check():
    vector_ready = _vector_store_readiness()
    job_runner_mode = _job_runner_mode()
    job_runner_diag = JOB_RUNNER.diagnostics()
    runtime_compatibility_policy_mode = _configured_runtime_compatibility_policy_mode()
    runtime_compatibility_status = _python_runtime_compatibility_status()
    tenant_authz_status = _tenant_authz_configuration_status()
    tenant_authz_policy_mode = _configured_tenant_authz_configuration_policy_mode()
    dispatcher_heartbeat_status = (
        job_runner_diag.get("worker_heartbeat") if isinstance(job_runner_diag.get("worker_heartbeat"), dict) else None
    )
    dispatcher_heartbeat_policy_mode = _dispatcher_worker_heartbeat_policy_mode()
    alerts: list[dict[str, Any]] = []
    dead_letter_threshold = _dead_letter_alert_threshold()
    dead_letter_queue_size_raw = job_runner_diag.get("dead_letter_queue_size")
    try:
        dead_letter_queue_size = int(dead_letter_queue_size_raw)
    except (TypeError, ValueError):
        dead_letter_queue_size = -1
    dead_letter_alert_triggered = (
        _uses_redis_queue_runner()
        and dead_letter_threshold > 0
        and dead_letter_queue_size >= 0
        and dead_letter_queue_size >= dead_letter_threshold
    )
    job_runner_ready = True
    if _uses_inmemory_queue_runner():
        job_runner_ready = bool(job_runner_diag.get("running"))
    elif _uses_redis_queue_runner():
        consumer_enabled = bool(job_runner_diag.get("consumer_enabled", True))
        running_ok = bool(job_runner_diag.get("running")) if consumer_enabled else True
        job_runner_ready = running_ok and bool(job_runner_diag.get("redis_available"))
        if not consumer_enabled and dispatcher_heartbeat_policy_mode != "off":
            heartbeat_healthy = (
                bool(dispatcher_heartbeat_status.get("healthy"))
                if isinstance(dispatcher_heartbeat_status, dict)
                else False
            )
            if not heartbeat_healthy:
                blocking = dispatcher_heartbeat_policy_mode == "strict"
                alerts.append(
                    {
                        "code": "REDIS_DISPATCHER_WORKER_HEARTBEAT_MISSING",
                        "severity": "high" if blocking else "medium",
                        "message": (
                            "Redis dispatcher mode is enabled with local consumer disabled, "
                            "but no healthy external worker heartbeat was detected."
                        ),
                        "blocking": blocking,
                    }
                )
                if blocking:
                    job_runner_ready = False
    if dead_letter_alert_triggered and _dead_letter_alert_blocking():
        job_runner_ready = False
    runtime_compatibility_supported = bool(runtime_compatibility_status.get("supported"))
    runtime_compatibility_blocking = (
        runtime_compatibility_policy_mode == "strict" and not runtime_compatibility_supported
    )
    runtime_compatibility_alerts: list[dict[str, Any]] = []
    if not runtime_compatibility_supported:
        runtime_compatibility_alerts.append(
            {
                "code": "PYTHON_RUNTIME_COMPATIBILITY_RISK",
                "severity": "high" if runtime_compatibility_blocking else "medium",
                "message": ("Runtime Python version is outside validated range 3.11-3.13 for current dependency set."),
                "blocking": runtime_compatibility_blocking,
            }
        )
    tenant_authz_valid = bool(tenant_authz_status.get("valid"))
    tenant_authz_blocking = tenant_authz_policy_mode == "strict" and not tenant_authz_valid
    tenant_authz_issues = {
        str(item).strip().lower() for item in (tenant_authz_status.get("issues") or []) if str(item).strip()
    }
    tenant_authz_alerts: list[dict[str, Any]] = []
    if not tenant_authz_valid:
        if "allowlist_empty" in tenant_authz_issues:
            tenant_authz_message = (
                "Tenant authz is enabled but allowlist is empty; configure GRANTFLOW_ALLOWED_TENANTS."
            )
        elif "default_tenant_not_in_allowlist" in tenant_authz_issues:
            tenant_authz_message = (
                "Tenant authz default tenant is not in allowlist; align GRANTFLOW_DEFAULT_TENANT with "
                "GRANTFLOW_ALLOWED_TENANTS."
            )
        else:
            tenant_authz_message = "Tenant authz configuration is invalid."
        tenant_authz_alerts.append(
            {
                "code": "TENANT_AUTHZ_CONFIGURATION_RISK",
                "severity": "high" if tenant_authz_blocking else "medium",
                "message": tenant_authz_message,
                "blocking": tenant_authz_blocking,
            }
        )
    ready = (
        bool(vector_ready.get("ready"))
        and job_runner_ready
        and not runtime_compatibility_blocking
        and not tenant_authz_blocking
    )
    preflight_grounding_thresholds = _preflight_grounding_policy_thresholds()
    runtime_grounded_quality_gate_thresholds = _runtime_grounded_quality_gate_thresholds()
    mel_grounding_thresholds = _mel_grounding_policy_thresholds()
    export_grounding_thresholds = _export_grounding_policy_thresholds()
    dead_letter_alert = {
        "enabled": bool(_uses_redis_queue_runner() and dead_letter_threshold > 0),
        "threshold": dead_letter_threshold,
        "queue_size": dead_letter_queue_size,
        "triggered": bool(dead_letter_alert_triggered),
        "blocking": bool(_dead_letter_alert_blocking()),
    }
    if dead_letter_alert_triggered:
        alerts.append(
            {
                "code": "DEAD_LETTER_QUEUE_THRESHOLD_EXCEEDED",
                "severity": "high" if _dead_letter_alert_blocking() else "medium",
                "message": (
                    "Dead-letter queue size exceeded configured alert threshold "
                    f"({dead_letter_queue_size}/{dead_letter_threshold})."
                ),
                "blocking": bool(_dead_letter_alert_blocking()),
            }
        )
    payload = {
        "status": "ready" if ready else "degraded",
        "checks": {
            "vector_store": vector_ready,
            "job_runner": {
                "mode": job_runner_mode,
                "ready": job_runner_ready,
                "queue": job_runner_diag,
                "dispatcher_worker_heartbeat_policy": {
                    "mode": dispatcher_heartbeat_policy_mode,
                },
                "dispatcher_worker_heartbeat": dispatcher_heartbeat_status,
                "dead_letter_alert": dead_letter_alert,
                "alerts": alerts,
            },
            "preflight_grounding_policy": {
                "mode": _configured_preflight_grounding_policy_mode(),
                "thresholds": preflight_grounding_thresholds,
            },
            "runtime_grounded_quality_gate": {
                "mode": _configured_runtime_grounded_quality_gate_mode(),
                "thresholds": runtime_grounded_quality_gate_thresholds,
            },
            "mel_grounding_policy": {
                "mode": _configured_mel_grounding_policy_mode(),
                "thresholds": mel_grounding_thresholds,
            },
            "export_grounding_policy": {
                "mode": _configured_export_grounding_policy_mode(),
                "thresholds": export_grounding_thresholds,
            },
            "export_contract_policy": {
                "mode": _configured_export_contract_policy_mode(),
            },
            "export_runtime_grounded_gate_policy": {
                "require_pass": _configured_export_require_grounded_gate_pass(),
            },
            "runtime_compatibility_policy": {
                "mode": runtime_compatibility_policy_mode,
                "status": runtime_compatibility_status,
                "blocking": runtime_compatibility_blocking,
                "alerts": runtime_compatibility_alerts,
            },
            "tenant_authz_configuration_policy": {
                "mode": tenant_authz_policy_mode,
                "status": tenant_authz_status,
                "blocking": tenant_authz_blocking,
                "alerts": tenant_authz_alerts,
            },
            "configuration_warnings": _configuration_warnings(),
        },
    }
    if not ready:
        raise HTTPException(status_code=503, detail=payload)
    return payload


def _redis_queue_admin_runner(required_methods: tuple[str, ...]) -> Any:
    if not _uses_redis_queue_runner():
        raise HTTPException(
            status_code=409,
            detail="Dead-letter queue management requires GRANTFLOW_JOB_RUNNER_MODE=redis_queue",
        )
    runner = JOB_RUNNER
    for method_name in required_methods:
        if not callable(getattr(runner, method_name, None)):
            raise HTTPException(
                status_code=409, detail="Redis queue admin operations are unavailable in current runner"
            )
    return runner


@app.get(
    "/queue/worker-heartbeat",
    response_model=QueueWorkerHeartbeatPublicResponse,
    response_model_exclude_none=True,
)
def get_queue_worker_heartbeat(
    request: Request,
):
    require_api_key_if_configured(request, for_read=True)
    runner = _redis_queue_admin_runner(("worker_heartbeat_status",))
    status_payload = runner.worker_heartbeat_status()
    if not isinstance(status_payload, dict):
        status_payload = {"present": False, "healthy": False, "error": "invalid_worker_heartbeat_payload"}
    return {
        "mode": "redis_queue",
        "policy": {"mode": _dispatcher_worker_heartbeat_policy_mode()},
        "consumer_enabled": bool(getattr(runner, "consumer_enabled", True)),
        "heartbeat": status_payload,
    }


@app.get(
    "/queue/dead-letter",
    response_model=DeadLetterQueueListPublicResponse,
    response_model_exclude_none=True,
)
def get_dead_letter_queue(
    request: Request,
    limit: int = Query(default=50, ge=1, le=500),
):
    require_api_key_if_configured(request, for_read=True)
    runner = _redis_queue_admin_runner(("list_dead_letters",))
    try:
        return runner.list_dead_letters(limit=limit)
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc


@app.post(
    "/queue/dead-letter/requeue",
    response_model=DeadLetterQueueMutationPublicResponse,
    response_model_exclude_none=True,
)
def requeue_dead_letter_queue(
    request: Request,
    limit: int = Query(default=10, ge=1, le=500),
    reset_attempts: bool = Query(default=True),
):
    require_api_key_if_configured(request)
    runner = _redis_queue_admin_runner(("requeue_dead_letters",))
    try:
        return runner.requeue_dead_letters(limit=limit, reset_attempts=reset_attempts)
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc


@app.delete(
    "/queue/dead-letter",
    response_model=DeadLetterQueueMutationPublicResponse,
    response_model_exclude_none=True,
)
def purge_dead_letter_queue(
    request: Request,
    limit: int = Query(default=100, ge=1, le=5000),
):
    require_api_key_if_configured(request)
    runner = _redis_queue_admin_runner(("purge_dead_letters",))
    try:
        return runner.purge_dead_letters(limit=limit)
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc


@app.get("/donors")
def list_donors():
    return {"donors": DonorFactory.list_supported()}


def _generate_preset_rows_for_public() -> list[dict[str, Any]]:
    generate_presets: list[dict[str, Any]] = []
    for item in list_generate_legacy_preset_details():
        donor_id = str(item.get("donor_id") or "").strip().lower() or None
        title = str(item.get("title") or "").strip() or None
        donor_label = donor_id.upper() if donor_id else "LEGACY"
        generate_presets.append(
            {
                "preset_key": str(item.get("preset_key") or "").strip(),
                "donor_id": donor_id,
                "title": title,
                "label": f"{donor_label}: {title or str(item.get('preset_key') or '')}",
                "source_kind": "legacy",
                "source_file": None,
                "generate_payload": dict(item.get("generate_payload") or {}),
            }
        )
    for row in list_sample_preset_summaries():
        sample_id = str(row.get("sample_id") or "").strip()
        if not sample_id:
            continue
        try:
            payload = load_sample_payload(sample_id)
            generate_payload = build_sample_generate_payload(
                sample_id,
                llm_mode=True,
                hitl_enabled=True,
                architect_rag_enabled=True,
                strict_preflight=False,
            )
        except ValueError:
            continue
        donor_id = str(payload.get("donor_id") or row.get("donor_id") or "").strip().lower() or None
        title = str(row.get("title") or sample_id).strip() or sample_id
        donor_label = donor_id.upper() if donor_id else "RBM"
        generate_presets.append(
            {
                "preset_key": sample_id,
                "donor_id": donor_id,
                "title": title,
                "label": f"RBM ({donor_label}): {title}",
                "source_kind": "rbm",
                "source_file": row.get("source_file"),
                "generate_payload": generate_payload,
            }
        )
    return generate_presets


def _demo_preset_bundle_payload() -> dict[str, Any]:
    generate_presets = _generate_preset_rows_for_public()
    ingest_presets: list[dict[str, Any]] = []
    for item in list_ingest_preset_details():
        donor_id = str(item.get("donor_id") or "").strip().lower() or None
        title = str(item.get("title") or "").strip() or None
        donor_label = donor_id.upper() if donor_id else "INGEST"
        ingest_presets.append(
            {
                "preset_key": str(item.get("preset_key") or "").strip(),
                "donor_id": donor_id,
                "title": title,
                "label": f"{donor_label}: {title or str(item.get('preset_key') or '')}",
                "metadata": dict(item.get("metadata") or {}),
                "checklist_items": list(item.get("checklist_items") or []),
                "recommended_docs": list(item.get("recommended_docs") or []),
            }
        )
    return {
        "generate_presets": generate_presets,
        "ingest_presets": ingest_presets,
    }


@app.get(
    "/demo/presets",
    response_model=DemoPresetBundlePublicResponse,
    response_model_exclude_none=True,
)
def get_demo_presets():
    return _demo_preset_bundle_payload()


@app.get(
    "/generate/presets",
    response_model=GeneratePresetListPublicResponse,
    response_model_exclude_none=True,
)
def list_generate_presets():
    return {"presets": _generate_preset_rows_for_public()}


@app.get(
    "/generate/presets/legacy",
    response_model=GenerateLegacyPresetListPublicResponse,
    response_model_exclude_none=True,
)
def list_generate_legacy_presets():
    return {"presets": list_generate_legacy_preset_summaries()}


@app.get(
    "/generate/presets/legacy/{preset_key}",
    response_model=GenerateLegacyPresetDetailPublicResponse,
    response_model_exclude_none=True,
)
def get_generate_legacy_preset(preset_key: str):
    try:
        return load_generate_legacy_preset(preset_key)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.get(
    "/generate/presets/rbm",
    response_model=RBMSamplePresetListPublicResponse,
    response_model_exclude_none=True,
)
def list_rbm_generate_presets():
    return {"presets": list_sample_preset_summaries()}


@app.get(
    "/generate/presets/rbm/{sample_id}",
    response_model=RBMSamplePresetDetailPublicResponse,
    response_model_exclude_none=True,
)
def get_rbm_generate_preset(
    sample_id: str,
    llm_mode: bool = Query(default=False),
    hitl_enabled: bool = Query(default=False),
    architect_rag_enabled: bool = Query(default=False),
    strict_preflight: bool = Query(default=False),
):
    try:
        payload = load_sample_payload(sample_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    source_file = None
    normalized = str(sample_id or "").strip().lower()
    for row in list_sample_preset_summaries():
        if str(row.get("sample_id") or "").strip().lower() == normalized:
            source_file = row.get("source_file")
            break

    return {
        "sample_id": normalized,
        "source_file": source_file,
        "payload": payload,
        "generate_payload": build_sample_generate_payload(
            normalized,
            llm_mode=bool(llm_mode),
            hitl_enabled=bool(hitl_enabled),
            architect_rag_enabled=bool(architect_rag_enabled),
            strict_preflight=bool(strict_preflight),
        ),
    }


@app.get(
    "/generate/presets/{preset_key}",
    response_model=DemoGeneratePresetPublicResponse,
    response_model_exclude_none=True,
)
def get_generate_preset(
    preset_key: str,
    llm_mode: Optional[bool] = Query(default=None),
    hitl_enabled: Optional[bool] = Query(default=None),
    architect_rag_enabled: Optional[bool] = Query(default=None),
    strict_preflight: Optional[bool] = Query(default=None),
):
    token = str(preset_key or "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="Missing preset_key")

    target = None
    for row in _generate_preset_rows_for_public():
        if not isinstance(row, dict):
            continue
        if str(row.get("preset_key") or "").strip() == token:
            target = dict(row)
            break

    if target is None:
        available = sorted(
            str(row.get("preset_key") or "").strip()
            for row in _generate_preset_rows_for_public()
            if isinstance(row, dict) and str(row.get("preset_key") or "").strip()
        )
        raise HTTPException(
            status_code=404,
            detail={
                "reason": "generate_preset_not_found",
                "preset_key": token,
                "available": available,
            },
        )

    generate_payload = dict(target.get("generate_payload") or {})
    if llm_mode is not None:
        generate_payload["llm_mode"] = bool(llm_mode)
    if hitl_enabled is not None:
        generate_payload["hitl_enabled"] = bool(hitl_enabled)
    if architect_rag_enabled is not None:
        generate_payload["architect_rag_enabled"] = bool(architect_rag_enabled)
    if strict_preflight is not None:
        generate_payload["strict_preflight"] = bool(strict_preflight)
    target["generate_payload"] = generate_payload
    return target


@app.get("/demo", response_class=HTMLResponse, include_in_schema=False)
def demo_console():
    return HTMLResponse(render_demo_ui_html())


@app.get(
    "/portfolio/metrics",
    response_model=PortfolioMetricsPublicResponse,
    response_model_exclude_none=True,
)
def get_portfolio_metrics(
    request: Request,
    donor_id: Optional[str] = None,
    tenant_id: Optional[str] = Query(default=None),
    status: Optional[str] = None,
    hitl_enabled: Optional[bool] = Query(default=None),
    warning_level: Optional[str] = None,
    grounding_risk_level: Optional[str] = None,
    toc_text_risk_level: Optional[str] = None,
    mel_risk_level: Optional[str] = None,
):
    require_api_key_if_configured(request, for_read=True)
    resolved_tenant_id = _resolve_tenant_id(request, explicit_tenant=tenant_id, require_if_enabled=True)
    jobs = _filter_jobs_by_tenant(_list_jobs(), resolved_tenant_id)
    return public_portfolio_metrics_payload(
        jobs,
        donor_id=(donor_id or None),
        status=(status or None),
        hitl_enabled=hitl_enabled,
        warning_level=(warning_level or None),
        grounding_risk_level=(grounding_risk_level or None),
        toc_text_risk_level=(toc_text_risk_level or None),
        mel_risk_level=(mel_risk_level or None),
    )


@app.get(
    "/portfolio/quality",
    response_model=PortfolioQualityPublicResponse,
    response_model_exclude_none=True,
)
def get_portfolio_quality(
    request: Request,
    donor_id: Optional[str] = None,
    tenant_id: Optional[str] = Query(default=None),
    status: Optional[str] = None,
    hitl_enabled: Optional[bool] = Query(default=None),
    warning_level: Optional[str] = None,
    grounding_risk_level: Optional[str] = None,
    finding_status: Optional[str] = None,
    finding_severity: Optional[str] = None,
    toc_text_risk_level: Optional[str] = None,
    mel_risk_level: Optional[str] = None,
):
    require_api_key_if_configured(request, for_read=True)
    resolved_tenant_id = _resolve_tenant_id(request, explicit_tenant=tenant_id, require_if_enabled=True)
    jobs = _filter_jobs_by_tenant(_list_jobs(), resolved_tenant_id)
    return public_portfolio_quality_payload(
        jobs,
        donor_id=(donor_id or None),
        status=(status or None),
        hitl_enabled=hitl_enabled,
        warning_level=(warning_level or None),
        grounding_risk_level=(grounding_risk_level or None),
        finding_status=(finding_status or None),
        finding_severity=(finding_severity or None),
        toc_text_risk_level=(toc_text_risk_level or None),
        mel_risk_level=(mel_risk_level or None),
    )


@app.get(
    "/portfolio/review-workflow",
    response_model=PortfolioReviewWorkflowPublicResponse,
    response_model_exclude_none=True,
)
def get_portfolio_review_workflow(
    request: Request,
    donor_id: Optional[str] = None,
    tenant_id: Optional[str] = Query(default=None),
    status: Optional[str] = None,
    hitl_enabled: Optional[bool] = Query(default=None),
    warning_level: Optional[str] = None,
    grounding_risk_level: Optional[str] = None,
    toc_text_risk_level: Optional[str] = None,
    event_type: Optional[str] = None,
    finding_id: Optional[str] = None,
    finding_code: Optional[str] = Query(default=None, alias="finding_code"),
    finding_section: Optional[str] = Query(default=None, alias="finding_section"),
    comment_status: Optional[str] = Query(default=None, alias="comment_status"),
    workflow_state: Optional[str] = Query(default=None, alias="workflow_state"),
    overdue_after_hours: int = Query(
        default=REVIEW_WORKFLOW_OVERDUE_DEFAULT_HOURS,
        ge=1,
        le=24 * 30,
        alias="overdue_after_hours",
    ),
):
    require_api_key_if_configured(request, for_read=True)
    workflow_state_filter = str(workflow_state or "").strip().lower() or None
    if workflow_state_filter and workflow_state_filter not in REVIEW_WORKFLOW_STATE_FILTER_VALUES:
        raise HTTPException(status_code=400, detail="Unsupported workflow_state filter")
    finding_section_filter = _validated_filter_token(
        finding_section,
        allowed={"toc", "logframe", "general"},
        detail="Unsupported finding_section filter",
    )
    resolved_tenant_id = _resolve_tenant_id(request, explicit_tenant=tenant_id, require_if_enabled=True)
    jobs = _filter_jobs_by_tenant(_list_jobs(), resolved_tenant_id)
    return public_portfolio_review_workflow_payload(
        jobs,
        donor_id=(donor_id or None),
        status=(status or None),
        hitl_enabled=hitl_enabled,
        warning_level=(warning_level or None),
        grounding_risk_level=(grounding_risk_level or None),
        toc_text_risk_level=(toc_text_risk_level or None),
        event_type=(event_type or None),
        finding_id=(finding_id or None),
        finding_code=(str(finding_code or "").strip() or None),
        finding_section=finding_section_filter,
        comment_status=(comment_status or None),
        workflow_state=workflow_state_filter,
        overdue_after_hours=overdue_after_hours,
    )


@app.get(
    "/portfolio/review-workflow/sla",
    response_model=PortfolioReviewWorkflowSLAPublicResponse,
    response_model_exclude_none=True,
)
def get_portfolio_review_workflow_sla(
    request: Request,
    donor_id: Optional[str] = None,
    tenant_id: Optional[str] = Query(default=None),
    status: Optional[str] = None,
    hitl_enabled: Optional[bool] = Query(default=None),
    warning_level: Optional[str] = None,
    grounding_risk_level: Optional[str] = None,
    toc_text_risk_level: Optional[str] = None,
    finding_id: Optional[str] = None,
    finding_code: Optional[str] = Query(default=None, alias="finding_code"),
    finding_section: Optional[str] = Query(default=None, alias="finding_section"),
    comment_status: Optional[str] = Query(default=None, alias="comment_status"),
    workflow_state: Optional[str] = Query(default=None, alias="workflow_state"),
    overdue_after_hours: int = Query(
        default=REVIEW_WORKFLOW_OVERDUE_DEFAULT_HOURS,
        ge=1,
        le=24 * 30,
        alias="overdue_after_hours",
    ),
    top_limit: int = Query(default=10, ge=1, le=200, alias="top_limit"),
):
    require_api_key_if_configured(request, for_read=True)
    workflow_state_filter = str(workflow_state or "").strip().lower() or None
    if workflow_state_filter and workflow_state_filter not in REVIEW_WORKFLOW_STATE_FILTER_VALUES:
        raise HTTPException(status_code=400, detail="Unsupported workflow_state filter")
    finding_section_filter = _validated_filter_token(
        finding_section,
        allowed={"toc", "logframe", "general"},
        detail="Unsupported finding_section filter",
    )
    resolved_tenant_id = _resolve_tenant_id(request, explicit_tenant=tenant_id, require_if_enabled=True)
    jobs = _filter_jobs_by_tenant(_list_jobs(), resolved_tenant_id)
    return public_portfolio_review_workflow_sla_payload(
        jobs,
        donor_id=(donor_id or None),
        status=(status or None),
        hitl_enabled=hitl_enabled,
        warning_level=(warning_level or None),
        grounding_risk_level=(grounding_risk_level or None),
        toc_text_risk_level=(toc_text_risk_level or None),
        finding_id=(finding_id or None),
        finding_code=(str(finding_code or "").strip() or None),
        finding_section=finding_section_filter,
        comment_status=(comment_status or None),
        workflow_state=workflow_state_filter,
        overdue_after_hours=overdue_after_hours,
        top_limit=top_limit,
    )


@app.get(
    "/portfolio/review-workflow/sla/hotspots",
    response_model=PortfolioReviewWorkflowSLAHotspotsPublicResponse,
    response_model_exclude_none=True,
)
def get_portfolio_review_workflow_sla_hotspots(
    request: Request,
    donor_id: Optional[str] = None,
    tenant_id: Optional[str] = Query(default=None),
    status: Optional[str] = None,
    hitl_enabled: Optional[bool] = Query(default=None),
    warning_level: Optional[str] = None,
    grounding_risk_level: Optional[str] = None,
    toc_text_risk_level: Optional[str] = None,
    finding_id: Optional[str] = None,
    finding_code: Optional[str] = Query(default=None, alias="finding_code"),
    finding_section: Optional[str] = Query(default=None, alias="finding_section"),
    comment_status: Optional[str] = Query(default=None, alias="comment_status"),
    workflow_state: Optional[str] = Query(default=None, alias="workflow_state"),
    overdue_after_hours: int = Query(
        default=REVIEW_WORKFLOW_OVERDUE_DEFAULT_HOURS,
        ge=1,
        le=24 * 30,
        alias="overdue_after_hours",
    ),
    top_limit: int = Query(default=10, ge=1, le=200, alias="top_limit"),
    hotspot_kind: Optional[str] = Query(default=None, alias="hotspot_kind"),
    hotspot_severity: Optional[str] = Query(default=None, alias="hotspot_severity"),
    min_overdue_hours: Optional[float] = Query(default=None, ge=0.0, le=24 * 365, alias="min_overdue_hours"),
):
    require_api_key_if_configured(request, for_read=True)
    workflow_state_filter = str(workflow_state or "").strip().lower() or None
    if workflow_state_filter and workflow_state_filter not in REVIEW_WORKFLOW_STATE_FILTER_VALUES:
        raise HTTPException(status_code=400, detail="Unsupported workflow_state filter")
    finding_section_filter = _validated_filter_token(
        finding_section,
        allowed={"toc", "logframe", "general"},
        detail="Unsupported finding_section filter",
    )
    hotspot_kind_filter = _validated_filter_token(
        hotspot_kind,
        allowed={"finding", "comment"},
        detail="Unsupported hotspot_kind filter",
    )
    hotspot_severity_filter = _validated_filter_token(
        hotspot_severity,
        allowed={"high", "medium", "low", "unknown"},
        detail="Unsupported hotspot_severity filter",
    )
    resolved_tenant_id = _resolve_tenant_id(request, explicit_tenant=tenant_id, require_if_enabled=True)
    jobs = _filter_jobs_by_tenant(_list_jobs(), resolved_tenant_id)
    return public_portfolio_review_workflow_sla_hotspots_payload(
        jobs,
        donor_id=(donor_id or None),
        status=(status or None),
        hitl_enabled=hitl_enabled,
        warning_level=(warning_level or None),
        grounding_risk_level=(grounding_risk_level or None),
        toc_text_risk_level=(toc_text_risk_level or None),
        finding_id=(finding_id or None),
        finding_code=(str(finding_code or "").strip() or None),
        finding_section=finding_section_filter,
        comment_status=(comment_status or None),
        workflow_state=workflow_state_filter,
        overdue_after_hours=overdue_after_hours,
        top_limit=top_limit,
        hotspot_kind=hotspot_kind_filter,
        hotspot_severity=hotspot_severity_filter,
        min_overdue_hours=min_overdue_hours,
    )


@app.get(
    "/portfolio/review-workflow/sla/hotspots/trends",
    response_model=PortfolioReviewWorkflowSLAHotspotsTrendsPublicResponse,
    response_model_exclude_none=True,
)
def get_portfolio_review_workflow_sla_hotspots_trends(
    request: Request,
    donor_id: Optional[str] = None,
    tenant_id: Optional[str] = Query(default=None),
    status: Optional[str] = None,
    hitl_enabled: Optional[bool] = Query(default=None),
    warning_level: Optional[str] = None,
    grounding_risk_level: Optional[str] = None,
    toc_text_risk_level: Optional[str] = None,
    finding_id: Optional[str] = None,
    finding_code: Optional[str] = Query(default=None, alias="finding_code"),
    finding_section: Optional[str] = Query(default=None, alias="finding_section"),
    comment_status: Optional[str] = Query(default=None, alias="comment_status"),
    workflow_state: Optional[str] = Query(default=None, alias="workflow_state"),
    overdue_after_hours: int = Query(
        default=REVIEW_WORKFLOW_OVERDUE_DEFAULT_HOURS,
        ge=1,
        le=24 * 30,
        alias="overdue_after_hours",
    ),
    top_limit: int = Query(default=10, ge=1, le=200, alias="top_limit"),
    hotspot_kind: Optional[str] = Query(default=None, alias="hotspot_kind"),
    hotspot_severity: Optional[str] = Query(default=None, alias="hotspot_severity"),
    min_overdue_hours: Optional[float] = Query(default=None, ge=0.0, le=24 * 365, alias="min_overdue_hours"),
):
    require_api_key_if_configured(request, for_read=True)
    workflow_state_filter = str(workflow_state or "").strip().lower() or None
    if workflow_state_filter and workflow_state_filter not in REVIEW_WORKFLOW_STATE_FILTER_VALUES:
        raise HTTPException(status_code=400, detail="Unsupported workflow_state filter")
    finding_section_filter = _validated_filter_token(
        finding_section,
        allowed={"toc", "logframe", "general"},
        detail="Unsupported finding_section filter",
    )
    hotspot_kind_filter = _validated_filter_token(
        hotspot_kind,
        allowed={"finding", "comment"},
        detail="Unsupported hotspot_kind filter",
    )
    hotspot_severity_filter = _validated_filter_token(
        hotspot_severity,
        allowed={"high", "medium", "low", "unknown"},
        detail="Unsupported hotspot_severity filter",
    )
    resolved_tenant_id = _resolve_tenant_id(request, explicit_tenant=tenant_id, require_if_enabled=True)
    jobs = _filter_jobs_by_tenant(_list_jobs(), resolved_tenant_id)
    return public_portfolio_review_workflow_sla_hotspots_trends_payload(
        jobs,
        donor_id=(donor_id or None),
        status=(status or None),
        hitl_enabled=hitl_enabled,
        warning_level=(warning_level or None),
        grounding_risk_level=(grounding_risk_level or None),
        toc_text_risk_level=(toc_text_risk_level or None),
        finding_id=(finding_id or None),
        finding_code=(str(finding_code or "").strip() or None),
        finding_section=finding_section_filter,
        comment_status=(comment_status or None),
        workflow_state=workflow_state_filter,
        overdue_after_hours=overdue_after_hours,
        top_limit=top_limit,
        hotspot_kind=hotspot_kind_filter,
        hotspot_severity=hotspot_severity_filter,
        min_overdue_hours=min_overdue_hours,
    )


@app.get(
    "/portfolio/review-workflow/trends",
    response_model=PortfolioReviewWorkflowTrendsPublicResponse,
    response_model_exclude_none=True,
)
def get_portfolio_review_workflow_trends(
    request: Request,
    donor_id: Optional[str] = None,
    tenant_id: Optional[str] = Query(default=None),
    status: Optional[str] = None,
    hitl_enabled: Optional[bool] = Query(default=None),
    warning_level: Optional[str] = None,
    grounding_risk_level: Optional[str] = None,
    toc_text_risk_level: Optional[str] = None,
    event_type: Optional[str] = None,
    finding_id: Optional[str] = None,
    finding_code: Optional[str] = Query(default=None, alias="finding_code"),
    finding_section: Optional[str] = Query(default=None, alias="finding_section"),
    comment_status: Optional[str] = Query(default=None, alias="comment_status"),
    workflow_state: Optional[str] = Query(default=None, alias="workflow_state"),
    overdue_after_hours: int = Query(
        default=REVIEW_WORKFLOW_OVERDUE_DEFAULT_HOURS,
        ge=1,
        le=24 * 30,
        alias="overdue_after_hours",
    ),
):
    require_api_key_if_configured(request, for_read=True)
    workflow_state_filter = str(workflow_state or "").strip().lower() or None
    if workflow_state_filter and workflow_state_filter not in REVIEW_WORKFLOW_STATE_FILTER_VALUES:
        raise HTTPException(status_code=400, detail="Unsupported workflow_state filter")
    finding_section_filter = _validated_filter_token(
        finding_section,
        allowed={"toc", "logframe", "general"},
        detail="Unsupported finding_section filter",
    )
    resolved_tenant_id = _resolve_tenant_id(request, explicit_tenant=tenant_id, require_if_enabled=True)
    jobs = _filter_jobs_by_tenant(_list_jobs(), resolved_tenant_id)
    return public_portfolio_review_workflow_trends_payload(
        jobs,
        donor_id=(donor_id or None),
        status=(status or None),
        hitl_enabled=hitl_enabled,
        warning_level=(warning_level or None),
        grounding_risk_level=(grounding_risk_level or None),
        toc_text_risk_level=(toc_text_risk_level or None),
        event_type=(event_type or None),
        finding_id=(finding_id or None),
        finding_code=(str(finding_code or "").strip() or None),
        finding_section=finding_section_filter,
        comment_status=(comment_status or None),
        workflow_state=workflow_state_filter,
        overdue_after_hours=overdue_after_hours,
    )


@app.get(
    "/portfolio/review-workflow/sla/trends",
    response_model=PortfolioReviewWorkflowSLATrendsPublicResponse,
    response_model_exclude_none=True,
)
def get_portfolio_review_workflow_sla_trends(
    request: Request,
    donor_id: Optional[str] = None,
    tenant_id: Optional[str] = Query(default=None),
    status: Optional[str] = None,
    hitl_enabled: Optional[bool] = Query(default=None),
    warning_level: Optional[str] = None,
    grounding_risk_level: Optional[str] = None,
    toc_text_risk_level: Optional[str] = None,
    finding_id: Optional[str] = None,
    finding_code: Optional[str] = Query(default=None, alias="finding_code"),
    finding_section: Optional[str] = Query(default=None, alias="finding_section"),
    comment_status: Optional[str] = Query(default=None, alias="comment_status"),
    workflow_state: Optional[str] = Query(default=None, alias="workflow_state"),
    overdue_after_hours: int = Query(
        default=REVIEW_WORKFLOW_OVERDUE_DEFAULT_HOURS,
        ge=1,
        le=24 * 30,
        alias="overdue_after_hours",
    ),
):
    require_api_key_if_configured(request, for_read=True)
    workflow_state_filter = str(workflow_state or "").strip().lower() or None
    if workflow_state_filter and workflow_state_filter not in REVIEW_WORKFLOW_STATE_FILTER_VALUES:
        raise HTTPException(status_code=400, detail="Unsupported workflow_state filter")
    finding_section_filter = _validated_filter_token(
        finding_section,
        allowed={"toc", "logframe", "general"},
        detail="Unsupported finding_section filter",
    )
    resolved_tenant_id = _resolve_tenant_id(request, explicit_tenant=tenant_id, require_if_enabled=True)
    jobs = _filter_jobs_by_tenant(_list_jobs(), resolved_tenant_id)
    return public_portfolio_review_workflow_sla_trends_payload(
        jobs,
        donor_id=(donor_id or None),
        status=(status or None),
        hitl_enabled=hitl_enabled,
        warning_level=(warning_level or None),
        grounding_risk_level=(grounding_risk_level or None),
        toc_text_risk_level=(toc_text_risk_level or None),
        finding_id=(finding_id or None),
        finding_code=(str(finding_code or "").strip() or None),
        finding_section=finding_section_filter,
        comment_status=(comment_status or None),
        workflow_state=workflow_state_filter,
        overdue_after_hours=overdue_after_hours,
    )


def _resolve_preflight_request_context(
    *,
    request: Request,
    donor_id: str,
    tenant_id: Optional[str] = None,
    client_metadata: Optional[Dict[str, Any]] = None,
    input_context: Optional[Dict[str, Any]] = None,
    expected_doc_families: Optional[list[str]] = None,
) -> tuple[str, Any, Optional[Dict[str, Any]]]:
    donor = str(donor_id or "").strip()
    if not donor:
        raise HTTPException(status_code=400, detail="Missing donor_id")

    try:
        strategy = DonorFactory.get_strategy(donor)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    metadata = dict(client_metadata) if isinstance(client_metadata, dict) else {}
    resolved_tenant_id = _resolve_tenant_id(
        request,
        explicit_tenant=tenant_id,
        client_metadata=metadata,
        require_if_enabled=True,
    )
    if resolved_tenant_id:
        metadata["tenant_id"] = resolved_tenant_id

    if isinstance(input_context, dict) and input_context:
        metadata["_preflight_input_context"] = dict(input_context)

    if isinstance(expected_doc_families, list):
        expected = _dedupe_doc_families(expected_doc_families)
        if expected:
            rag_readiness_raw = metadata.get("rag_readiness")
            rag_readiness: Dict[str, Any] = dict(rag_readiness_raw) if isinstance(rag_readiness_raw, dict) else {}
            rag_readiness["expected_doc_families"] = expected
            if not str(rag_readiness.get("donor_id") or "").strip():
                rag_readiness["donor_id"] = donor
            metadata["rag_readiness"] = rag_readiness

    return donor, strategy, (metadata or None)


def _resolve_generate_payload_from_preset(
    preset_key: str,
    *,
    preset_type: Literal["auto", "legacy", "rbm"],
) -> tuple[str, Dict[str, Any]]:
    token = str(preset_key or "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="Missing preset_key")

    legacy_error: Optional[str] = None
    rbm_error: Optional[str] = None

    if preset_type in {"auto", "legacy"}:
        try:
            legacy = load_generate_legacy_preset(token)
            payload = legacy.get("generate_payload")
            if not isinstance(payload, dict):
                raise HTTPException(status_code=500, detail=f"Preset '{token}' has invalid generate_payload")
            return "legacy", dict(payload)
        except ValueError as exc:
            legacy_error = str(exc)
            if preset_type == "legacy":
                raise HTTPException(status_code=404, detail=legacy_error) from exc

    if preset_type in {"auto", "rbm"}:
        try:
            sample_payload = load_sample_payload(token)
            default_llm_mode = bool(sample_payload.get("llm_mode", False))
            default_architect_rag = bool(sample_payload.get("architect_rag_enabled", False))
            return "rbm", build_sample_generate_payload(
                token,
                llm_mode=default_llm_mode,
                hitl_enabled=False,
                architect_rag_enabled=default_architect_rag,
                strict_preflight=False,
            )
        except ValueError as exc:
            rbm_error = str(exc)
            if preset_type == "rbm":
                raise HTTPException(status_code=404, detail=rbm_error) from exc

    legacy_keys = sorted(
        str(row.get("preset_key") or "").strip()
        for row in list_generate_legacy_preset_summaries()
        if isinstance(row, dict) and str(row.get("preset_key") or "").strip()
    )
    rbm_keys = available_sample_ids()
    raise HTTPException(
        status_code=404,
        detail={
            "reason": "generate_preset_not_found",
            "preset_key": token,
            "preset_type": preset_type,
            "legacy_error": legacy_error,
            "rbm_error": rbm_error,
            "available": {"legacy": legacy_keys, "rbm": rbm_keys},
        },
    )


def _build_generate_request_from_preset(req: GenerateFromPresetRequest) -> tuple[GenerateRequest, str]:
    preset_key = str(req.preset_key or "").strip()
    source_kind, base_payload = _resolve_generate_payload_from_preset(
        preset_key,
        preset_type=req.preset_type,
    )
    payload: Dict[str, Any] = dict(base_payload)

    input_context_raw = payload.get("input_context")
    input_context: Dict[str, Any] = dict(input_context_raw) if isinstance(input_context_raw, dict) else {}
    if isinstance(req.input_context_patch, dict) and req.input_context_patch:
        input_context.update(dict(req.input_context_patch))
    payload["input_context"] = input_context

    client_metadata_raw = payload.get("client_metadata")
    client_metadata: Dict[str, Any] = dict(client_metadata_raw) if isinstance(client_metadata_raw, dict) else {}
    if preset_key:
        client_metadata.setdefault("demo_generate_preset_key", preset_key)
    client_metadata.setdefault("demo_generate_preset_source", source_kind)
    if isinstance(req.client_metadata_patch, dict) and req.client_metadata_patch:
        client_metadata.update(dict(req.client_metadata_patch))
    payload["client_metadata"] = client_metadata

    if req.tenant_id is not None:
        payload["tenant_id"] = req.tenant_id
    if req.request_id is not None:
        payload["request_id"] = req.request_id
    if req.webhook_url is not None:
        payload["webhook_url"] = req.webhook_url
    if req.webhook_secret is not None:
        payload["webhook_secret"] = req.webhook_secret

    if req.llm_mode is not None:
        payload["llm_mode"] = bool(req.llm_mode)
    if req.architect_rag_enabled is not None:
        payload["architect_rag_enabled"] = bool(req.architect_rag_enabled)
    if req.require_grounded_generation is not None:
        payload["require_grounded_generation"] = bool(req.require_grounded_generation)
    if req.hitl_enabled is not None:
        payload["hitl_enabled"] = bool(req.hitl_enabled)
    if req.hitl_checkpoints is not None:
        payload["hitl_checkpoints"] = list(req.hitl_checkpoints)
    if req.strict_preflight is not None:
        payload["strict_preflight"] = bool(req.strict_preflight)

    try:
        return GenerateRequest(**payload), source_kind
    except ValidationError as exc:
        raise HTTPException(
            status_code=400,
            detail={
                "reason": "invalid_generate_preset_payload",
                "preset_key": preset_key,
                "preset_source": source_kind,
                "errors": exc.errors(),
            },
        ) from exc


async def _dispatch_generate_from_preset(
    req: GenerateFromPresetRequest,
    background_tasks: BackgroundTasks,
    request: Request,
    *,
    request_id: Optional[str] = None,
):
    generate_req, source_kind = _build_generate_request_from_preset(req)
    from grantflow.api.routes.jobs import generate as generate_endpoint

    generated = await generate_endpoint(
        generate_req,
        background_tasks,
        request,
        request_id=request_id if request_id is not None else generate_req.request_id,
    )
    if isinstance(generated, dict):
        response = dict(generated)
        response["preset_key"] = str(req.preset_key or "").strip()
        response["preset_source"] = source_kind
        return response
    return generated


def _load_route_modules() -> None:
    from grantflow.api.routes import exports as _exports_routes
    from grantflow.api.routes import ingest as _ingest_routes
    from grantflow.api.routes import jobs as _jobs_routes
    from grantflow.api.routes import review as _review_routes

    _ = (_exports_routes, _ingest_routes, _jobs_routes, _review_routes)


_load_route_modules()
include_api_routers(app)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host=config.api_host, port=config.api_port, reload=config.debug)
