# grantflow/exporters/excel_builder.py

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from grantflow.core.security_utils import resolve_allowed_attachment_path
from grantflow.exporters.donor_contracts import DONOR_XLSX_PRIMARY_SHEET, evaluate_export_contract
from grantflow.exporters.template_profile import (
    build_export_template_profile,
    resolve_export_template_key,
)
from grantflow.exporters.toc_normalization import normalize_toc_for_export, unwrap_toc_payload


def _compact_export_text(value: Any, *, max_len: int | None = None) -> str:
    text = " ".join(str(value or "").split()).strip()
    if not text:
        return ""
    for token in (
        "Evidence hint:",
        "Grounding gate warning:",
        "architect_retrieval_no_hits",
    ):
        idx = text.lower().find(token.lower())
        if idx >= 0:
            text = text[:idx].rstrip(" .,-")
    text = text.replace(" service delivery delivery", " service delivery")
    text = text.replace("  ", " ").strip()
    if max_len and len(text) > max_len:
        return f"{text[: max_len - 3].rstrip()}..."
    return text


def _normalized_indicator_rows(logframe_draft: Optional[Dict[str, Any]]) -> list[Dict[str, Any]]:
    if not isinstance(logframe_draft, dict):
        return []
    raw = logframe_draft.get("indicators")
    return [row for row in raw if isinstance(row, dict)] if isinstance(raw, list) else []


def _attachment_source_path(row: dict[str, Any]) -> str:
    for key in ("source_path", "staged_file_path", "local_path", "file_path", "attachment_path"):
        value = row.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _adjust_submission_readiness_for_attachment_files(
    summary: dict[str, Any], toc_payload: dict[str, Any]
) -> dict[str, Any]:
    adjusted = dict(summary) if isinstance(summary, dict) else {}
    attachment_counts = adjusted.get("attachment_manifest_counts")
    if not isinstance(attachment_counts, dict):
        return adjusted
    toc_root = unwrap_toc_payload(toc_payload if isinstance(toc_payload, dict) else {})
    attachment_rows = toc_root.get("attachment_manifest")
    if not isinstance(attachment_rows, list):
        return adjusted

    attached_file_count = 0
    missing_ready = 0
    status_file_mismatch_count = 0
    for row in attachment_rows:
        if not isinstance(row, dict):
            continue
        status = str(row.get("status") or "").strip().lower()
        source_path = _attachment_source_path(row)
        if not source_path:
            continue
        source = resolve_allowed_attachment_path(source_path)
        if source is not None:
            attached_file_count += 1
            if status and status != "ready":
                status_file_mismatch_count += 1
        elif status == "ready":
            missing_ready += 1

    if missing_ready <= 0 and status_file_mismatch_count <= 0 and attached_file_count <= 0:
        return adjusted

    updated_attachment_counts = dict(attachment_counts)
    if missing_ready > 0:
        updated_attachment_counts["ready"] = max(0, int(updated_attachment_counts.get("ready", 0)) - missing_ready)
        updated_attachment_counts["pending"] = int(updated_attachment_counts.get("pending", 0)) + missing_ready
        adjusted["attachment_manifest_counts"] = updated_attachment_counts

        submission_counts = adjusted.get("submission_package_counts")
        compliance_counts = adjusted.get("compliance_matrix_counts")
        if isinstance(submission_counts, dict) and isinstance(compliance_counts, dict):
            weighted_total = (
                int(submission_counts.get("total", 0))
                + int(updated_attachment_counts.get("total", 0))
                + int(compliance_counts.get("total", 0))
            )
            weighted_ready = (
                int(submission_counts.get("ready", 0))
                + int(updated_attachment_counts.get("ready", 0))
                + int(compliance_counts.get("ready", 0))
            )
            weighted_partial = (
                int(submission_counts.get("partial", 0))
                + int(updated_attachment_counts.get("partial", 0))
                + int(compliance_counts.get("partial", 0))
            )
            completeness_score = (
                round(((weighted_ready + (0.5 * weighted_partial)) / weighted_total) * 100, 1)
                if weighted_total
                else 0.0
            )
            adjusted["completeness_score"] = completeness_score
            if weighted_total == 0:
                adjusted["readiness_status"] = "missing"
            elif completeness_score >= 85:
                adjusted["readiness_status"] = "ready"
            elif completeness_score >= 60:
                adjusted["readiness_status"] = "partial"
            else:
                adjusted["readiness_status"] = "weak"
        adjusted["top_gap"] = "attachment_files"

    adjusted["attachment_file_validation"] = {
        "attached_file_count": attached_file_count,
        "missing_ready_file_count": missing_ready,
        "status_file_mismatch_count": status_file_mismatch_count,
    }
    return adjusted


def _indicator_focus_rows(
    indicators: list[Dict[str, Any]],
    *,
    result_level: Optional[str] = None,
    limit: int = 2,
) -> list[Dict[str, Any]]:
    rows = indicators
    if result_level:
        token = str(result_level).strip().lower()
        rows = [row for row in indicators if str(row.get("result_level") or "").strip().lower() == token]
    return rows[: max(0, limit)]


def _indicator_focus_cells(indicators: list[Dict[str, Any]]) -> tuple[str, str, str]:
    focus_rows = [row for row in indicators if isinstance(row, dict)]
    if not focus_rows:
        return "", "", ""
    names = ", ".join(str(row.get("name") or "").strip() for row in focus_rows if str(row.get("name") or "").strip())
    mov = "; ".join(
        str(row.get("means_of_verification") or "").strip()
        for row in focus_rows
        if str(row.get("means_of_verification") or "").strip()
    )
    owner = "; ".join(str(row.get("owner") or "").strip() for row in focus_rows if str(row.get("owner") or "").strip())
    return names, mov, owner


def _indicator_summary_cells(indicators: list[Dict[str, Any]]) -> tuple[str, str, str]:
    focus_rows = [row for row in indicators if isinstance(row, dict)]
    if not focus_rows:
        return "", "", ""

    baseline_target = "; ".join(
        f"{str(row.get('baseline') or '').strip() or 'TBD'} -> {str(row.get('target') or '').strip() or 'TBD'}"
        for row in focus_rows
    )
    frequency = "; ".join(
        str(row.get("frequency") or "").strip() for row in focus_rows if str(row.get("frequency") or "").strip()
    )
    formulas = "; ".join(
        str(row.get("formula") or "").strip() for row in focus_rows if str(row.get("formula") or "").strip()
    )
    return baseline_target, frequency, formulas


def _compact_text(value: Any, *, max_len: int = 120) -> str:
    text = " ".join(str(value or "").split()).strip()
    if not text:
        return ""
    if len(text) <= max_len:
        return text
    return f"{text[: max_len - 3].rstrip()}..."


def _indicator_context_cells(indicators: list[Dict[str, Any]]) -> tuple[str, str]:
    focus_rows = [row for row in indicators if isinstance(row, dict)]
    if not focus_rows:
        return "", ""
    focus = "; ".join(
        _compact_text(row.get("definition", ""), max_len=96)
        for row in focus_rows
        if _compact_text(row.get("definition", ""), max_len=96)
    )
    intent = "; ".join(
        _compact_text(row.get("justification", ""), max_len=96)
        for row in focus_rows
        if _compact_text(row.get("justification", ""), max_len=96)
    )
    return focus, intent


def _review_readiness_rows(
    *,
    quality_summary: dict[str, Any],
    citations: list[Dict[str, Any]],
    critic_findings: list[Dict[str, Any]],
    review_comments: list[Dict[str, Any]],
) -> list[tuple[str, Any]]:
    open_findings = [
        item
        for item in critic_findings
        if isinstance(item, dict) and str(item.get("status") or "open").strip().lower() not in {"resolved", "closed"}
    ]
    finding_ack_queue_count = sum(
        1
        for item in critic_findings
        if isinstance(item, dict) and str(item.get("status") or "open").strip().lower() == "open"
    )
    finding_resolve_queue_count = sum(
        1
        for item in critic_findings
        if isinstance(item, dict) and str(item.get("status") or "").strip().lower() == "acknowledged"
    )
    high_findings = [
        item
        for item in open_findings
        if str(item.get("severity") or "").strip().lower() in {"high", "critical", "fatal"}
    ]
    open_comments = [
        item
        for item in review_comments
        if isinstance(item, dict) and str(item.get("status") or "open").strip().lower() not in {"resolved", "closed"}
    ]
    low_confidence = [
        item
        for item in citations
        if isinstance(item, dict) and str(item.get("citation_type") or "").strip().lower() == "rag_low_confidence"
    ]
    fallback = [
        item
        for item in citations
        if isinstance(item, dict)
        and str(item.get("citation_type") or "").strip().lower() in {"fallback_namespace", "strategy_reference"}
    ]
    rows = [
        ("Needs revision", quality_summary.get("needs_revision")),
        ("Open critic findings", len(open_findings)),
        (
            "Acknowledged critic findings",
            len(
                [
                    item
                    for item in critic_findings
                    if isinstance(item, dict) and str(item.get("status") or "").strip().lower() == "acknowledged"
                ]
            ),
        ),
        (
            "Resolved critic findings",
            len(
                [
                    item
                    for item in critic_findings
                    if isinstance(item, dict)
                    and str(item.get("status") or "").strip().lower() in {"resolved", "closed"}
                ]
            ),
        ),
        ("High-severity open findings", len(high_findings)),
        ("Open review comments", len(open_comments)),
        ("Low-confidence citations", len(low_confidence)),
        ("Fallback/strategy citations", len(fallback)),
    ]
    top_actions: list[str] = []
    for item in open_findings:
        if not isinstance(item, dict):
            continue
        action = str(item.get("reviewer_next_step") or item.get("recommended_action") or "").strip()
        if action and action not in top_actions:
            top_actions.append(action)
        if len(top_actions) >= 2:
            break
    for idx, action in enumerate(top_actions, start=1):
        rows.append((f"Top reviewer action {idx}", action))
    acknowledged_comments = [
        item
        for item in review_comments
        if isinstance(item, dict) and str(item.get("status") or "").strip().lower() == "acknowledged"
    ]
    resolved_comments = [
        item
        for item in review_comments
        if isinstance(item, dict) and str(item.get("status") or "").strip().lower() in {"resolved", "closed"}
    ]
    total_items = len(critic_findings) + len(review_comments)
    resolved_items = len(
        [
            item
            for item in critic_findings
            if isinstance(item, dict) and str(item.get("status") or "").strip().lower() in {"resolved", "closed"}
        ]
    ) + len(resolved_comments)
    acknowledged_items = len(
        [
            item
            for item in critic_findings
            if isinstance(item, dict) and str(item.get("status") or "").strip().lower() == "acknowledged"
        ]
    ) + len(acknowledged_comments)
    rows.extend(
        [
            ("Acknowledged review comments", len(acknowledged_comments)),
            ("Resolved review comments", len(resolved_comments)),
            (
                "Critic finding resolution rate",
                (
                    round(
                        len(
                            [
                                item
                                for item in critic_findings
                                if isinstance(item, dict)
                                and str(item.get("status") or "").strip().lower() in {"resolved", "closed"}
                            ]
                        )
                        / len(critic_findings),
                        4,
                    )
                    if critic_findings
                    else None
                ),
            ),
            (
                "Critic finding acknowledgment rate",
                (
                    round(
                        (
                            len(
                                [
                                    item
                                    for item in critic_findings
                                    if isinstance(item, dict)
                                    and str(item.get("status") or "").strip().lower() in {"resolved", "closed"}
                                ]
                            )
                            + len(
                                [
                                    item
                                    for item in critic_findings
                                    if isinstance(item, dict)
                                    and str(item.get("status") or "").strip().lower() == "acknowledged"
                                ]
                            )
                        )
                        / len(critic_findings),
                        4,
                    )
                    if critic_findings
                    else None
                ),
            ),
            ("Reviewer workflow resolution rate", round(resolved_items / total_items, 4) if total_items else None),
            (
                "Reviewer workflow acknowledgment rate",
                round((resolved_items + acknowledged_items) / total_items, 4) if total_items else None,
            ),
            (
                "Next primary review action",
                (
                    "ack_finding"
                    if finding_ack_queue_count > 0
                    else (
                        "resolve_finding"
                        if finding_resolve_queue_count > 0
                        else (
                            "ack_comment"
                            if len(open_comments) > 0
                            else (
                                "resolve_comment"
                                if len(acknowledged_comments) > 0
                                else ("reopen_comment" if len(resolved_comments) > 0 else None)
                            )
                        )
                    )
                ),
            ),
            ("Finding ack queue", finding_ack_queue_count),
            ("Finding resolve queue", finding_resolve_queue_count),
            ("Comment ack queue", len(open_comments)),
            ("Comment resolve queue", len(acknowledged_comments)),
            ("Comment reopen queue", len(resolved_comments)),
        ]
    )

    from datetime import datetime, timezone

    def _parse_dt(value: Any):
        if not value:
            return None
        try:
            text = str(value).replace("Z", "+00:00")
            dt = datetime.fromisoformat(text)
            if dt.tzinfo is None:
                return dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except Exception:
            return None

    ref_candidates = []
    for item in review_comments:
        if not isinstance(item, dict):
            continue
        for key in ("last_transition_at", "updated_ts", "resolved_at", "acknowledged_at", "ts", "due_at"):
            ref_candidates.append(_parse_dt(item.get(key)))
    ref_values = [value for value in ref_candidates if value is not None]
    ref = max(ref_values) if ref_values else datetime.now(timezone.utc)
    aging = {"3-7d": 0, ">7d": 0}
    for item in review_comments:
        if not isinstance(item, dict):
            continue
        status = str(item.get("status") or "open").strip().lower()
        if status in {"resolved", "closed"}:
            continue
        transition = (
            _parse_dt(item.get("last_transition_at")) or _parse_dt(item.get("updated_ts")) or _parse_dt(item.get("ts"))
        )
        if transition is None:
            continue
        hours = max(0.0, (ref - transition).total_seconds() / 3600.0)
        if 72.0 <= hours < 168.0:
            aging["3-7d"] += 1
        elif hours >= 168.0:
            aging[">7d"] += 1
    rows.extend(
        [
            ("Comment threads aged 3-7d", aging["3-7d"]),
            ("Comment threads aged >7d", aging[">7d"]),
        ]
    )
    return [(label, value) for label, value in rows if value is not None and value != ""]


def _indicator_readiness_hint(indicator: Dict[str, Any]) -> str:
    baseline = str(indicator.get("baseline") or "").strip()
    target = str(indicator.get("target") or "").strip()
    mov = str(indicator.get("means_of_verification") or "").strip()
    owner = str(indicator.get("owner") or "").strip()
    gaps: list[str] = []
    if not baseline or baseline.lower() == "tbd":
        gaps.append("baseline")
    if not target or target.lower() == "tbd":
        gaps.append("target")
    if not mov:
        gaps.append("mov")
    if not owner:
        gaps.append("owner")
    if gaps:
        return "needs-" + "-".join(gaps)
    return "review-ready"


def _autosize_columns(ws) -> None:
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = adjusted_width


def _toc_root(payload: Dict[str, Any], donor_id: str | None = None) -> Dict[str, Any]:
    if not donor_id:
        return unwrap_toc_payload(payload)
    donor_key = resolve_export_template_key(donor_id=donor_id, toc_payload=payload)
    return normalize_toc_for_export(donor_key, unwrap_toc_payload(payload))


def _add_citations_sheet(wb: Workbook, citations: list[Dict[str, Any]]) -> None:
    if not citations:
        return
    ws = wb.create_sheet("Citations")
    headers = [
        "Stage",
        "Type",
        "Used For",
        "Label",
        "Confidence",
        "Result Level",
        "Statement Path",
        "Namespace",
        "Source",
        "Page",
        "Chunk",
        "Chunk ID",
        "Evidence Signal",
        "Review Hint",
        "Excerpt",
    ]
    ws.append(headers)
    for c in citations:
        ws.append(
            [
                c.get("stage", ""),
                c.get("citation_type", ""),
                c.get("used_for", ""),
                c.get("label", ""),
                c.get("citation_confidence", ""),
                c.get("result_level", ""),
                c.get("statement_path", ""),
                c.get("namespace", ""),
                c.get("source", ""),
                c.get("page", ""),
                c.get("chunk", ""),
                c.get("chunk_id", ""),
                c.get("evidence_signal", ""),
                c.get("review_hint", ""),
                (c.get("excerpt", "") or "")[:500],
            ]
        )
    _autosize_columns(ws)


def _add_critic_findings_sheet(wb: Workbook, critic_findings: list[Dict[str, Any]]) -> None:
    if not critic_findings:
        return
    ws = wb.create_sheet("Critic Findings")
    headers = [
        "Status",
        "Severity",
        "Section",
        "Code",
        "Review Title",
        "Review Bucket",
        "Triage Priority",
        "Message",
        "Recommended Action",
        "Reviewer Next Step",
        "Fix Hint",
        "Version ID",
        "Finding ID",
        "Source",
    ]
    ws.append(headers)
    for f in critic_findings:
        finding_id = f.get("id") or f.get("finding_id", "")
        ws.append(
            [
                f.get("status", ""),
                f.get("severity", ""),
                f.get("section", ""),
                f.get("code", ""),
                f.get("review_title", ""),
                f.get("review_bucket", ""),
                f.get("triage_priority", ""),
                f.get("message", ""),
                f.get("recommended_action", ""),
                f.get("reviewer_next_step", ""),
                f.get("fix_hint", ""),
                f.get("version_id", ""),
                finding_id,
                f.get("source", ""),
            ]
        )
    _autosize_columns(ws)


def _add_review_comments_sheet(wb: Workbook, review_comments: list[Dict[str, Any]]) -> None:
    if not review_comments:
        return
    ws = wb.create_sheet("Review Comments")
    headers = ["Status", "Section", "Author", "Message", "Version ID", "Linked Finding ID", "Timestamp", "Comment ID"]
    ws.append(headers)
    for c in review_comments:
        ws.append(
            [
                c.get("status", ""),
                c.get("section", ""),
                c.get("author", ""),
                c.get("message", ""),
                c.get("version_id", ""),
                c.get("linked_finding_id", ""),
                c.get("ts", ""),
                c.get("comment_id", ""),
            ]
        )
    _autosize_columns(ws)


def _add_quality_summary_sheet(wb: Workbook, quality_summary: dict[str, Any]) -> None:
    if not quality_summary:
        return
    rows = [
        ("Quality score", quality_summary.get("quality_score")),
        ("Critic score", quality_summary.get("critic_score")),
        ("Needs revision", quality_summary.get("needs_revision")),
        ("Critic engine", quality_summary.get("engine")),
        ("Rule score", quality_summary.get("rule_score")),
        ("LLM score", quality_summary.get("llm_score")),
        ("Fatal flaw count", quality_summary.get("fatal_flaw_count")),
        ("Citation count", quality_summary.get("citation_count")),
    ]
    present_rows = [(field_name, value) for field_name, value in rows if value is not None and value != ""]
    if not present_rows:
        return

    ws = wb.create_sheet("Quality Summary")
    border = _apply_table_header(ws, ["Field", "Value"])
    for row_idx, (field_name, value) in enumerate(present_rows, start=2):
        ws.append([field_name, value])
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=2).border = border
    _autosize_columns(ws)


def _add_review_readiness_sheet(
    wb: Workbook,
    *,
    quality_summary: dict[str, Any],
    citations: list[Dict[str, Any]],
    critic_findings: list[Dict[str, Any]],
    review_comments: list[Dict[str, Any]],
) -> None:
    rows = _review_readiness_rows(
        quality_summary=quality_summary,
        citations=citations,
        critic_findings=critic_findings,
        review_comments=review_comments,
    )
    if not rows:
        return

    ws = wb.create_sheet("Review Readiness")
    border = _apply_table_header(ws, ["Field", "Value"])
    for row_idx, (field_name, value) in enumerate(rows, start=2):
        ws.append([field_name, value])
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=2).border = border
    _autosize_columns(ws)


def _add_template_meta_sheet(wb: Workbook, profile: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Template Meta")
    headers = ["Field", "Value"]
    border = _apply_table_header(ws, headers)
    rows = [
        ("Donor ID", profile.get("donor_id", "")),
        ("Template Key", profile.get("template_key", "")),
        ("Template Display", profile.get("template_display_name", "")),
        ("Coverage Rate", profile.get("coverage_rate", 0.0)),
        (
            "Required Sections",
            ", ".join(str(x) for x in (profile.get("required_sections") or [])),
        ),
        (
            "Present Sections",
            ", ".join(str(x) for x in (profile.get("present_sections") or [])),
        ),
        (
            "Missing Sections",
            ", ".join(str(x) for x in (profile.get("missing_sections") or [])),
        ),
    ]
    for row_idx, (field_name, value) in enumerate(rows, start=2):
        ws.append([field_name, value])
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=2).border = border
    _autosize_columns(ws)


def _add_export_contract_sheet(wb: Workbook, contract: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Export Contract")
    headers = ["Field", "Value"]
    border = _apply_table_header(ws, headers)
    rows = [
        ("Status", contract.get("status", "")),
        ("Warnings", ", ".join(str(x) for x in (contract.get("warnings") or [])) or "-"),
        (
            "Missing Required ToC Sections",
            ", ".join(str(x) for x in (contract.get("missing_required_sections") or [])) or "-",
        ),
        (
            "Missing Required Workbook Sheets",
            ", ".join(str(x) for x in (contract.get("missing_required_sheets") or [])) or "-",
        ),
        ("Expected Primary Sheet", contract.get("expected_primary_sheet", "") or "-"),
        (
            "Expected Primary Sheet Headers",
            ", ".join(str(x) for x in (contract.get("expected_primary_sheet_headers") or [])) or "-",
        ),
        (
            "Actual Primary Sheet Headers",
            ", ".join(str(x) for x in (contract.get("actual_primary_sheet_headers") or [])) or "-",
        ),
        (
            "Missing Required Primary Sheet Headers",
            ", ".join(str(x) for x in (contract.get("missing_required_primary_sheet_headers") or [])) or "-",
        ),
        ("Required Workbook Sheets", ", ".join(str(x) for x in (contract.get("required_sheets") or [])) or "-"),
    ]
    submission_summary = contract.get("submission_readiness_summary")
    if isinstance(submission_summary, dict):
        validation = submission_summary.get("attachment_file_validation")
        rows.extend(
            [
                ("Submission Completeness Score", submission_summary.get("completeness_score", "")),
                ("Submission Readiness Status", submission_summary.get("readiness_status", "")),
                ("Top Submission Gap", submission_summary.get("top_gap", "")),
                (
                    "Submission Package Counts",
                    _format_contract_counts(submission_summary.get("submission_package_counts")),
                ),
                (
                    "Attachment Manifest Counts",
                    _format_contract_counts(submission_summary.get("attachment_manifest_counts")),
                ),
                (
                    "Compliance Matrix Counts",
                    _format_contract_counts(submission_summary.get("compliance_matrix_counts")),
                ),
            ]
        )
        if isinstance(validation, dict):
            rows.extend(
                [
                    ("Attached Annex Files", int(validation.get("attached_file_count") or 0)),
                    ("Ready Annex Rows Missing Files", int(validation.get("missing_ready_file_count") or 0)),
                    ("Annex Status/File Mismatches", int(validation.get("status_file_mismatch_count") or 0)),
                ]
            )
    for row_idx, (field_name, value) in enumerate(rows, start=2):
        ws.append([field_name, value])
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=2).border = border
    _autosize_columns(ws)


def _format_contract_counts(value: Any) -> str:
    if not isinstance(value, dict):
        return "-"
    return (
        f"ready={int(value.get('ready') or 0)}, "
        f"partial={int(value.get('partial') or 0)}, "
        f"pending={int(value.get('pending') or 0)}"
    )


def _apply_table_header(ws, headers: list[str]) -> Border:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    return thin_border


def _sheet_headers(ws) -> list[str]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(value).strip() for value in row if str(value or "").strip()]


def _add_usaid_results_sheet(
    wb: Workbook,
    toc_payload: Dict[str, Any],
    *,
    logframe_draft: Optional[Dict[str, Any]] = None,
) -> None:
    toc = _toc_root(toc_payload)
    indicators = _normalized_indicator_rows(logframe_draft)
    outcome_focus = _indicator_focus_rows(indicators, result_level="outcome", limit=2)
    ws = wb.create_sheet("USAID_RF")
    headers = [
        "DO ID",
        "DO Description",
        "IR ID",
        "IR Description",
        "Output ID",
        "Output Description",
        "Indicator Code",
        "Indicator Name",
        "Target",
        "Justification",
        "Citation",
        "Suggested Monitoring Focus",
        "Suggested Means of Verification",
        "Suggested Owner",
        "Suggested Baseline -> Target",
        "Suggested Frequency",
        "Suggested Formula",
        "Suggested Result Focus",
        "Suggested Measurement Intent",
    ]
    thin_border = _apply_table_header(ws, headers)

    row_idx = 2
    for do in toc.get("development_objectives") or []:
        if not isinstance(do, dict):
            continue
        do_id = do.get("do_id", "")
        do_desc = _compact_export_text(do.get("description", ""))
        for ir in do.get("intermediate_results") or []:
            if not isinstance(ir, dict):
                continue
            ir_id = ir.get("ir_id", "")
            ir_desc = _compact_export_text(ir.get("description", ""))
            for output in ir.get("outputs") or []:
                if not isinstance(output, dict):
                    continue
                output_id = output.get("output_id", "")
                output_desc = _compact_export_text(output.get("description", ""))
                indicators = output.get("indicators") or []
                if not isinstance(indicators, list) or not indicators:
                    focus_name, focus_mov, focus_owner = _indicator_focus_cells(outcome_focus)
                    baseline_target, frequency, formulas = _indicator_summary_cells(outcome_focus)
                    result_focus, measurement_intent = _indicator_context_cells(outcome_focus)
                    ws.append(
                        [
                            do_id,
                            do_desc,
                            ir_id,
                            ir_desc,
                            output_id,
                            output_desc,
                            "",
                            "",
                            "",
                            "",
                            "",
                            focus_name,
                            focus_mov,
                            focus_owner,
                            baseline_target,
                            frequency,
                            formulas,
                            result_focus,
                            measurement_intent,
                        ]
                    )
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col).border = thin_border
                    row_idx += 1
                    continue
                for ind in indicators:
                    if not isinstance(ind, dict):
                        continue
                    focus_name, focus_mov, focus_owner = _indicator_focus_cells(outcome_focus)
                    baseline_target, frequency, formulas = _indicator_summary_cells(outcome_focus)
                    result_focus, measurement_intent = _indicator_context_cells(outcome_focus)
                    ws.append(
                        [
                            do_id,
                            do_desc,
                            ir_id,
                            ir_desc,
                            output_id,
                            output_desc,
                            ind.get("indicator_code", ""),
                            _compact_export_text(ind.get("name", "")),
                            ind.get("target", ""),
                            ind.get("justification", ""),
                            ind.get("citation", ""),
                            focus_name,
                            focus_mov,
                            focus_owner,
                            baseline_target,
                            frequency,
                            formulas,
                            result_focus,
                            measurement_intent,
                        ]
                    )
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col).border = thin_border
                    row_idx += 1
    _autosize_columns(ws)


def _add_eu_results_sheet(
    wb: Workbook,
    toc_payload: Dict[str, Any],
    *,
    logframe_draft: Optional[Dict[str, Any]] = None,
) -> None:
    toc = _toc_root(toc_payload)
    indicators = _normalized_indicator_rows(logframe_draft)
    outcome_focus = _indicator_focus_rows(indicators, result_level="outcome", limit=2)
    output_focus = _indicator_focus_rows(indicators, result_level="output", limit=2)
    ws = wb.create_sheet("EU_Intervention")
    headers = [
        "Level",
        "ID",
        "Title",
        "Description",
        "Suggested Monitoring Focus",
        "Suggested Means of Verification",
        "Suggested Owner",
        "Suggested Baseline -> Target",
        "Suggested Frequency",
        "Suggested Formula",
        "Suggested Result Focus",
        "Suggested Measurement Intent",
    ]
    thin_border = _apply_table_header(ws, headers)
    row_idx = 2
    overall = toc.get("overall_objective") if isinstance(toc, dict) else None
    if isinstance(overall, dict):
        focus_name, focus_mov, focus_owner = _indicator_focus_cells(outcome_focus)
        baseline_target, frequency, formulas = _indicator_summary_cells(outcome_focus)
        result_focus, measurement_intent = _indicator_context_cells(outcome_focus)
        ws.append(
            [
                "Overall Objective",
                overall.get("objective_id", ""),
                _compact_export_text(overall.get("title", "")),
                _compact_export_text(overall.get("rationale", "")),
                focus_name,
                focus_mov,
                focus_owner,
                baseline_target,
                frequency,
                formulas,
                result_focus,
                measurement_intent,
            ]
        )
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1

    specific_objectives = toc.get("specific_objectives") if isinstance(toc, dict) else None
    if isinstance(specific_objectives, list):
        for row in specific_objectives:
            if not isinstance(row, dict):
                continue
            focus_name, focus_mov, focus_owner = _indicator_focus_cells(outcome_focus[:1] or outcome_focus)
            baseline_target, frequency, formulas = _indicator_summary_cells(outcome_focus[:1] or outcome_focus)
            result_focus, measurement_intent = _indicator_context_cells(outcome_focus[:1] or outcome_focus)
            ws.append(
                [
                    "Specific Objective",
                    row.get("objective_id", ""),
                    _compact_export_text(row.get("title", "")),
                    _compact_export_text(row.get("rationale", "")),
                    focus_name,
                    focus_mov,
                    focus_owner,
                    baseline_target,
                    frequency,
                    formulas,
                    result_focus,
                    measurement_intent,
                ]
            )
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    expected_outcomes = toc.get("expected_outcomes") if isinstance(toc, dict) else None
    if isinstance(expected_outcomes, list):
        for row in expected_outcomes:
            if not isinstance(row, dict):
                continue
            focus_name, focus_mov, focus_owner = _indicator_focus_cells(output_focus[:1] or outcome_focus[:1])
            baseline_target, frequency, formulas = _indicator_summary_cells(output_focus[:1] or outcome_focus[:1])
            result_focus, measurement_intent = _indicator_context_cells(output_focus[:1] or outcome_focus[:1])
            ws.append(
                [
                    "Outcome",
                    row.get("outcome_id", ""),
                    _compact_export_text(row.get("title", "")),
                    _compact_export_text(row.get("expected_change", "")),
                    focus_name,
                    focus_mov,
                    focus_owner,
                    baseline_target,
                    frequency,
                    formulas,
                    result_focus,
                    measurement_intent,
                ]
            )
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    if row_idx == 2:
        ws.append(["", "", "", "", "", "", "", "", "", "", "", ""])
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1

    aux = wb.create_sheet("EU_Assumptions_Risks")
    aux_headers = ["Type", "Item"]
    aux_border = _apply_table_header(aux, aux_headers)
    aux_row = 2
    assumptions = toc.get("assumptions") if isinstance(toc, dict) else None
    if isinstance(assumptions, list):
        for item in assumptions:
            aux.append(["Assumption", str(item)])
            for col in range(1, len(aux_headers) + 1):
                aux.cell(row=aux_row, column=col).border = aux_border
            aux_row += 1
    risks = toc.get("risks") if isinstance(toc, dict) else None
    if isinstance(risks, list):
        for item in risks:
            aux.append(["Risk", str(item)])
            for col in range(1, len(aux_headers) + 1):
                aux.cell(row=aux_row, column=col).border = aux_border
            aux_row += 1

    safeguarding_annex = toc.get("safeguarding_annex") if isinstance(toc, dict) else None
    if isinstance(safeguarding_annex, list):
        for item in safeguarding_annex:
            aux.append(["Safeguarding Annex", str(item)])
            for col in range(1, len(aux_headers) + 1):
                aux.cell(row=aux_row, column=col).border = aux_border
            aux_row += 1

    if aux_row == 2:
        aux.append(["", ""])
        for col in range(1, len(aux_headers) + 1):
            aux.cell(row=aux_row, column=col).border = aux_border
    else:
        aux_row += 1
    _autosize_columns(ws)
    _autosize_columns(aux)


def _add_worldbank_results_sheet(
    wb: Workbook,
    toc_payload: Dict[str, Any],
    *,
    logframe_draft: Optional[Dict[str, Any]] = None,
) -> None:
    toc = _toc_root(toc_payload)
    indicators = _normalized_indicator_rows(logframe_draft)
    outcome_focus = _indicator_focus_rows(indicators, result_level="outcome", limit=2)
    impact_focus = _indicator_focus_rows(indicators, result_level="impact", limit=1)
    ws = wb.create_sheet("WB_Results")
    headers = [
        "Level",
        "ID",
        "Title",
        "Description",
        "Indicator Focus",
        "Suggested Monitoring Focus",
        "Suggested Means of Verification",
        "Suggested Owner",
        "Suggested Baseline -> Target",
        "Suggested Frequency",
        "Suggested Formula",
        "Suggested Result Focus",
        "Suggested Measurement Intent",
    ]
    thin_border = _apply_table_header(ws, headers)
    row_idx = 2
    pdo = _compact_export_text(toc.get("project_development_objective") or "") if isinstance(toc, dict) else ""
    if pdo:
        focus_name, focus_mov, focus_owner = _indicator_focus_cells(impact_focus or outcome_focus[:1])
        baseline_target, frequency, formulas = _indicator_summary_cells(impact_focus or outcome_focus[:1])
        result_focus, measurement_intent = _indicator_context_cells(impact_focus or outcome_focus[:1])
        ws.append(
            [
                "PDO",
                "",
                "Project Development Objective",
                pdo,
                "",
                focus_name,
                focus_mov,
                focus_owner,
                baseline_target,
                frequency,
                formulas,
                result_focus,
                measurement_intent,
            ]
        )
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1

    objectives = toc.get("objectives") if isinstance(toc, dict) else None
    if isinstance(objectives, list) and objectives:
        for obj in objectives:
            if not isinstance(obj, dict):
                continue
            focus_name, focus_mov, focus_owner = _indicator_focus_cells(outcome_focus[:1] or impact_focus)
            baseline_target, frequency, formulas = _indicator_summary_cells(outcome_focus[:1] or impact_focus)
            result_focus, measurement_intent = _indicator_context_cells(outcome_focus[:1] or impact_focus)
            ws.append(
                [
                    "Objective",
                    obj.get("objective_id", ""),
                    _compact_export_text(obj.get("title", "")),
                    _compact_export_text(obj.get("description", "")),
                    "",
                    focus_name,
                    focus_mov,
                    focus_owner,
                    baseline_target,
                    frequency,
                    formulas,
                    result_focus,
                    measurement_intent,
                ]
            )
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    results_chain = toc.get("results_chain") if isinstance(toc, dict) else None
    if isinstance(results_chain, list):
        for row in results_chain:
            if not isinstance(row, dict):
                continue
            focus_name, focus_mov, focus_owner = _indicator_focus_cells(outcome_focus[:1] or impact_focus)
            baseline_target, frequency, formulas = _indicator_summary_cells(outcome_focus[:1] or impact_focus)
            result_focus, measurement_intent = _indicator_context_cells(outcome_focus[:1] or impact_focus)
            ws.append(
                [
                    "Result",
                    row.get("result_id", ""),
                    _compact_export_text(row.get("title", "")),
                    _compact_export_text(row.get("description", "")),
                    _compact_export_text(row.get("indicator_focus", "")),
                    focus_name,
                    focus_mov,
                    focus_owner,
                    baseline_target,
                    frequency,
                    formulas,
                    result_focus,
                    measurement_intent,
                ]
            )
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    assumptions = toc.get("assumptions") if isinstance(toc, dict) else None
    if isinstance(assumptions, list):
        for item in assumptions:
            ws.append(["Assumption", "", "", str(item), "", "", "", "", "", "", "", "", ""])
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    risks = toc.get("risks") if isinstance(toc, dict) else None
    if isinstance(risks, list):
        for item in risks:
            ws.append(["Risk", "", "", str(item), "", "", "", "", "", "", "", "", ""])
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    if row_idx == 2:
        ws.append(["", "", "", "", "", "", "", "", "", "", "", "", ""])
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col).border = thin_border
    _autosize_columns(ws)


def _add_giz_results_sheet(
    wb: Workbook,
    toc_payload: Dict[str, Any],
    *,
    logframe_draft: Optional[Dict[str, Any]] = None,
) -> None:
    toc = _toc_root(toc_payload)
    indicators = _normalized_indicator_rows(logframe_draft)
    outcome_focus = _indicator_focus_rows(indicators, result_level="outcome", limit=2)
    output_focus = _indicator_focus_rows(indicators, result_level="output", limit=2)
    ws = wb.create_sheet("GIZ_Results")
    headers = [
        "Level",
        "Title",
        "Description",
        "Partner Role",
        "Suggested Monitoring Focus",
        "Suggested Means of Verification",
        "Suggested Owner",
        "Suggested Baseline -> Target",
        "Suggested Frequency",
        "Suggested Formula",
        "Suggested Result Focus",
        "Suggested Measurement Intent",
    ]
    thin_border = _apply_table_header(ws, headers)
    row_idx = 2

    programme_objective = str(toc.get("programme_objective") or "").strip() if isinstance(toc, dict) else ""
    if programme_objective:
        focus_name, focus_mov, focus_owner = _indicator_focus_cells(output_focus[:1] or outcome_focus[:1])
        baseline_target, frequency, formulas = _indicator_summary_cells(output_focus[:1] or outcome_focus[:1])
        result_focus, measurement_intent = _indicator_context_cells(output_focus[:1] or outcome_focus[:1])
        ws.append(
            [
                "Programme Objective",
                "Programme Objective",
                programme_objective,
                "",
                focus_name,
                focus_mov,
                focus_owner,
                baseline_target,
                frequency,
                formulas,
                result_focus,
                measurement_intent,
            ]
        )
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1

    outputs = toc.get("outputs") if isinstance(toc, dict) else None
    if isinstance(outputs, list):
        for output in outputs:
            focus_name, focus_mov, focus_owner = _indicator_focus_cells(output_focus[:1] or outcome_focus[:1])
            baseline_target, frequency, formulas = _indicator_summary_cells(output_focus[:1] or outcome_focus[:1])
            result_focus, measurement_intent = _indicator_context_cells(output_focus[:1] or outcome_focus[:1])
            ws.append(
                [
                    "Output",
                    str(output),
                    "",
                    "",
                    focus_name,
                    focus_mov,
                    focus_owner,
                    baseline_target,
                    frequency,
                    formulas,
                    result_focus,
                    measurement_intent,
                ]
            )
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    outcomes = toc.get("outcomes") if isinstance(toc, dict) else None
    if isinstance(outcomes, list):
        for outcome in outcomes:
            if not isinstance(outcome, dict):
                continue
            focus_name, focus_mov, focus_owner = _indicator_focus_cells(output_focus[:1] or outcome_focus[:1])
            baseline_target, frequency, formulas = _indicator_summary_cells(output_focus[:1] or outcome_focus[:1])
            result_focus, measurement_intent = _indicator_context_cells(output_focus[:1] or outcome_focus[:1])
            ws.append(
                [
                    "Outcome",
                    outcome.get("title", ""),
                    outcome.get("description", ""),
                    outcome.get("partner_role", ""),
                    focus_name,
                    focus_mov,
                    focus_owner,
                    baseline_target,
                    frequency,
                    formulas,
                    result_focus,
                    measurement_intent,
                ]
            )
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    sustainability_factors = toc.get("sustainability_factors") if isinstance(toc, dict) else None
    if isinstance(sustainability_factors, list):
        for item in sustainability_factors:
            ws.append(["Sustainability", str(item), "", "", "", "", "", "", "", "", "", ""])
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    assumptions_risks = toc.get("assumptions_risks") if isinstance(toc, dict) else None
    if isinstance(assumptions_risks, list):
        for item in assumptions_risks:
            ws.append(["Assumption/Risk", str(item), "", "", "", "", "", "", "", "", "", ""])
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    if row_idx == 2:
        ws.append(["", "", "", "", "", "", "", "", "", "", "", ""])
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col).border = thin_border
    _autosize_columns(ws)


def _add_un_results_sheet(
    wb: Workbook,
    toc_payload: Dict[str, Any],
    *,
    logframe_draft: Optional[Dict[str, Any]] = None,
) -> None:
    toc = _toc_root(toc_payload)
    indicators = _normalized_indicator_rows(logframe_draft)
    outcome_focus = _indicator_focus_rows(indicators, result_level="outcome", limit=2)
    impact_focus = _indicator_focus_rows(indicators, result_level="impact", limit=1)
    ws = wb.create_sheet("UN_Results")
    headers = [
        "Level",
        "Title",
        "Description",
        "Suggested Monitoring Focus",
        "Suggested Means of Verification",
        "Suggested Owner",
        "Suggested Baseline -> Target",
        "Suggested Frequency",
        "Suggested Formula",
        "Suggested Result Focus",
        "Suggested Measurement Intent",
    ]
    thin_border = _apply_table_header(ws, headers)
    row_idx = 2

    brief = str(toc.get("brief") or "").strip() if isinstance(toc, dict) else ""
    if brief:
        focus_name, focus_mov, focus_owner = _indicator_focus_cells(impact_focus or outcome_focus[:1])
        baseline_target, frequency, formulas = _indicator_summary_cells(impact_focus or outcome_focus[:1])
        result_focus, measurement_intent = _indicator_context_cells(impact_focus or outcome_focus[:1])
        ws.append(
            [
                "Overview",
                "Overview",
                brief,
                focus_name,
                focus_mov,
                focus_owner,
                baseline_target,
                frequency,
                formulas,
                result_focus,
                measurement_intent,
            ]
        )
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1

    objectives = toc.get("objectives") if isinstance(toc, dict) else None
    if isinstance(objectives, list):
        for obj in objectives:
            if not isinstance(obj, dict):
                continue
            focus_name, focus_mov, focus_owner = _indicator_focus_cells(outcome_focus[:1] or impact_focus)
            baseline_target, frequency, formulas = _indicator_summary_cells(outcome_focus[:1] or impact_focus)
            result_focus, measurement_intent = _indicator_context_cells(outcome_focus[:1] or impact_focus)
            ws.append(
                [
                    "Objective",
                    obj.get("title", ""),
                    obj.get("description", ""),
                    focus_name,
                    focus_mov,
                    focus_owner,
                    baseline_target,
                    frequency,
                    formulas,
                    result_focus,
                    measurement_intent,
                ]
            )
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    outcomes = toc.get("outcomes") if isinstance(toc, dict) else None
    if isinstance(outcomes, list):
        for row in outcomes:
            if not isinstance(row, dict):
                continue
            focus_name, focus_mov, focus_owner = _indicator_focus_cells(outcome_focus[:1])
            baseline_target, frequency, formulas = _indicator_summary_cells(outcome_focus[:1])
            result_focus, measurement_intent = _indicator_context_cells(outcome_focus[:1])
            ws.append(
                [
                    "Outcome",
                    row.get("title", "") or row.get("name", ""),
                    row.get("description", "") or row.get("expected_change", ""),
                    focus_name,
                    focus_mov,
                    focus_owner,
                    baseline_target,
                    frequency,
                    formulas,
                    result_focus,
                    measurement_intent,
                ]
            )
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    if row_idx == 2:
        ws.append(["", "", "", "", "", "", "", "", "", "", ""])
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col).border = thin_border
    _autosize_columns(ws)


def _add_state_department_results_sheet(
    wb: Workbook,
    toc_payload: Dict[str, Any],
    *,
    logframe_draft: Optional[Dict[str, Any]] = None,
) -> None:
    toc = _toc_root(toc_payload)
    indicators = _normalized_indicator_rows(logframe_draft)
    outcome_focus = _indicator_focus_rows(indicators, result_level="outcome", limit=2)
    impact_focus = _indicator_focus_rows(indicators, result_level="impact", limit=1)
    ws = wb.create_sheet("StateDept_Results")
    headers = [
        "Level",
        "Objective / Title",
        "Line of Effort",
        "Description",
        "Suggested Monitoring Focus",
        "Suggested Means of Verification",
        "Suggested Owner",
        "Suggested Result Focus",
        "Suggested Measurement Intent",
    ]
    thin_border = _apply_table_header(ws, headers)
    row_idx = 2

    strategic_context = str(toc.get("strategic_context") or "").strip() if isinstance(toc, dict) else ""
    if strategic_context:
        ws.append(["Strategic Context", "Context", "", strategic_context, "", "", "", "", ""])
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1

    program_goal = str(toc.get("program_goal") or "").strip() if isinstance(toc, dict) else ""
    if program_goal:
        focus_name, focus_mov, focus_owner = _indicator_focus_cells(impact_focus or outcome_focus[:1])
        result_focus, measurement_intent = _indicator_context_cells(impact_focus or outcome_focus[:1])
        ws.append(
            [
                "Program Goal",
                "Goal",
                "",
                program_goal,
                focus_name,
                focus_mov,
                focus_owner,
                result_focus,
                measurement_intent,
            ]
        )
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1

    objectives = toc.get("objectives") if isinstance(toc, dict) else None
    if isinstance(objectives, list):
        for obj in objectives:
            if not isinstance(obj, dict):
                continue
            focus_name, focus_mov, focus_owner = _indicator_focus_cells(outcome_focus[:1] or impact_focus)
            result_focus, measurement_intent = _indicator_context_cells(outcome_focus[:1] or impact_focus)
            ws.append(
                [
                    "Objective",
                    obj.get("objective", ""),
                    obj.get("line_of_effort", ""),
                    obj.get("expected_change", ""),
                    focus_name,
                    focus_mov,
                    focus_owner,
                    result_focus,
                    measurement_intent,
                ]
            )
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    stakeholder_map = toc.get("stakeholder_map") if isinstance(toc, dict) else None
    if isinstance(stakeholder_map, list):
        for item in stakeholder_map:
            ws.append(["Stakeholder", str(item), "", "", "", "", "", "", ""])
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    risk_mitigation = toc.get("risk_mitigation") if isinstance(toc, dict) else None
    if isinstance(risk_mitigation, list):
        for item in risk_mitigation:
            ws.append(["Risk Mitigation", str(item), "", "", "", "", "", "", ""])
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1

    if row_idx == 2:
        ws.append(["", "", "", "", "", "", "", "", ""])
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col).border = thin_border
    _autosize_columns(ws)


def _add_evaluation_plan_sheet(
    wb: Workbook, toc_payload: Dict[str, Any], *, logframe_draft: Optional[Dict[str, Any]] = None
) -> None:
    ws = wb.create_sheet("Evaluation_Plan")
    headers = ["Section", "ID", "Title", "Description"]
    thin_border = _apply_table_header(ws, headers)
    row_idx = 2

    def add_row(section: str, identifier: str, title: str, description: str) -> None:
        nonlocal row_idx
        values = [section, identifier, title, description]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_compact_export_text(value)).border = thin_border
        row_idx += 1

    add_row("summary", "brief", "Assignment Summary", str(toc_payload.get("brief") or ""))
    add_row(
        "organization",
        "organization_information",
        "Organization Information",
        str(toc_payload.get("organization_information") or ""),
    )
    add_row("purpose", "evaluation_purpose", "Evaluation Purpose", str(toc_payload.get("evaluation_purpose") or ""))
    add_row(
        "technical_approach",
        "technical_approach_summary",
        "Analysis and Proposed Approaches / Methodologies",
        str(toc_payload.get("technical_approach_summary") or ""),
    )
    add_row(
        "methodology",
        "methodology_overview",
        "Methodology Overview",
        str(toc_payload.get("methodology_overview") or ""),
    )
    add_row("sampling", "sampling_plan", "Sampling Plan", str(toc_payload.get("sampling_plan") or ""))

    for idx, item in enumerate(toc_payload.get("analytical_software") or [], start=1):
        add_row("analysis_tool", f"SW{idx}", f"Analytical Tool {idx}", str(item))

    for idx, item in enumerate(toc_payload.get("ethical_considerations") or [], start=1):
        add_row("ethics", f"E{idx}", f"Ethical Consideration {idx}", str(item))

    for idx, question in enumerate(toc_payload.get("evaluation_questions") or [], start=1):
        add_row("evaluation_question", f"EQ{idx}", f"Evaluation Question {idx}", str(question))

    for idx, row in enumerate(toc_payload.get("methodology_components") or [], start=1):
        if not isinstance(row, dict):
            continue
        title = str(row.get("method") or f"Method {idx}")
        description = " | ".join(
            part
            for part in (
                str(row.get("purpose") or "").strip(),
                f"Respondents: {str(row.get('respondent_group') or '').strip()}".strip(),
                f"Evidence: {str(row.get('evidence_source') or '').strip()}".strip(),
            )
            if part and part != "Respondents:" and part != "Evidence:"
        )
        add_row("method", f"M{idx}", title, description)

    for idx, row in enumerate(toc_payload.get("deliverables") or [], start=1):
        if not isinstance(row, dict):
            continue
        title = str(row.get("deliverable") or f"Deliverable {idx}")
        timing = str(row.get("timing") or "").strip()
        purpose = str(row.get("purpose") or "").strip()
        description = " | ".join(part for part in (timing, purpose) if part)
        add_row("deliverable", f"D{idx}", title, description)

    for idx, row in enumerate(toc_payload.get("team_composition") or [], start=1):
        if not isinstance(row, dict):
            continue
        add_row(
            "team_role",
            f"T{idx}",
            str(row.get("role") or f"Role {idx}"),
            str(row.get("responsibility") or ""),
        )

    for idx, row in enumerate(toc_payload.get("key_personnel") or [], start=1):
        if not isinstance(row, dict):
            continue
        title = str(row.get("name") or f"Personnel {idx}")
        role = str(row.get("role") or "").strip()
        qualifications = str(row.get("qualifications") or "").strip()
        level_of_effort = str(row.get("level_of_effort") or "").strip()
        cv_status = str(row.get("cv_status") or "").strip()
        if role:
            title = f"{title} ({role})"
        description = " | ".join(
            part
            for part in (
                qualifications,
                f"LOE: {level_of_effort}".strip() if level_of_effort else "",
                f"CV: {cv_status}".strip() if cv_status else "",
            )
            if part
        )
        add_row("key_personnel", f"KP{idx}", title, description)

    add_row(
        "level_of_effort",
        "level_of_effort_summary",
        "Proposed Level of Effort",
        str(toc_payload.get("level_of_effort_summary") or ""),
    )
    add_row(
        "past_performance",
        "technical_experience_summary",
        "Technical Experience and Past Performance References",
        str(toc_payload.get("technical_experience_summary") or ""),
    )
    add_row(
        "sample_outputs",
        "sample_outputs_summary",
        "Sample Technical Outputs",
        str(toc_payload.get("sample_outputs_summary") or ""),
    )
    add_row(
        "financial_summary",
        "financial_summary",
        "Financial Proposal Companion",
        str(toc_payload.get("financial_summary") or ""),
    )

    for idx, row in enumerate(toc_payload.get("cost_structure") or [], start=1):
        if not isinstance(row, dict):
            continue
        title = str(row.get("cost_bucket") or f"Cost Bucket {idx}")
        description = " | ".join(
            part
            for part in (
                f"Basis: {str(row.get('basis') or '').strip()}".strip(),
                f"Estimate: {str(row.get('estimate') or '').strip()}".strip(),
                str(row.get("notes") or "").strip(),
            )
            if part and part not in {"Basis:", "Estimate:"}
        )
        add_row("cost_structure", f"CS{idx}", title, description)

    for idx, item in enumerate(toc_payload.get("pricing_assumptions") or [], start=1):
        add_row("pricing_assumption", f"P{idx}", f"Pricing Assumption {idx}", str(item))

    add_row(
        "payment_schedule",
        "payment_schedule_summary",
        "Payment Schedule Summary",
        str(toc_payload.get("payment_schedule_summary") or ""),
    )

    for idx, row in enumerate(toc_payload.get("submission_package_checklist") or [], start=1):
        if not isinstance(row, dict):
            continue
        title = str(row.get("artifact") or f"Submission Artifact {idx}")
        description = " | ".join(
            part
            for part in (
                f"Owner: {str(row.get('owner') or '').strip()}".strip(),
                f"Status: {str(row.get('status') or '').strip()}".strip(),
                str(row.get("notes") or "").strip(),
            )
            if part and part not in {"Owner:", "Status:"}
        )
        add_row("submission_package", f"SP{idx}", title, description)

    for idx, row in enumerate(toc_payload.get("attachment_manifest") or [], start=1):
        if not isinstance(row, dict):
            continue
        title = str(row.get("attachment") or f"Attachment {idx}")
        description = " | ".join(
            part
            for part in (
                f"Required for: {str(row.get('required_for') or '').strip()}".strip(),
                f"Owner: {str(row.get('owner') or '').strip()}".strip(),
                f"Status: {str(row.get('status') or '').strip()}".strip(),
                str(row.get("notes") or "").strip(),
            )
            if part and part not in {"Required for:", "Owner:", "Status:"}
        )
        add_row("attachment_manifest", f"AM{idx}", title, description)

    for idx, row in enumerate(toc_payload.get("evaluation_questions_matrix") or [], start=1):
        if not isinstance(row, dict):
            continue
        title = str(row.get("evaluation_question") or f"Evaluation Question {idx}")
        description = " | ".join(
            part
            for part in (
                f"Methods: {', '.join(str(item).strip() for item in (row.get('key_methods') or []) if str(item).strip())}".strip(),
                f"Evidence: {', '.join(str(item).strip() for item in (row.get('evidence_sources') or []) if str(item).strip())}".strip(),
                f"Reporting use: {str(row.get('reporting_use') or '').strip()}".strip(),
            )
            if part and part not in {"Methods:", "Evidence:", "Reporting use:"}
        )
        add_row("question_matrix", f"QM{idx}", title, description)

    for idx, row in enumerate(toc_payload.get("methods_coverage_matrix") or [], start=1):
        if not isinstance(row, dict):
            continue
        title = str(row.get("method") or f"Method Coverage {idx}")
        description = " | ".join(
            part
            for part in (
                f"Covers: {', '.join(str(item).strip() for item in (row.get('covers_questions') or []) if str(item).strip())}".strip(),
                f"Respondents: {str(row.get('respondent_group') or '').strip()}".strip(),
                f"Expected output: {str(row.get('expected_output') or '').strip()}".strip(),
            )
            if part and part not in {"Covers:", "Respondents:", "Expected output:"}
        )
        add_row("method_coverage", f"MC{idx}", title, description)

    for idx, item in enumerate(toc_payload.get("annex_readiness") or [], start=1):
        add_row("annex", f"A{idx}", f"Annex Readiness {idx}", str(item))

    for idx, row in enumerate(toc_payload.get("deliverables_schedule_table") or [], start=1):
        if not isinstance(row, dict):
            continue
        title = str(row.get("deliverable") or f"Deliverable Schedule {idx}")
        description = " | ".join(
            part
            for part in (
                f"Due: {str(row.get('due_window') or '').strip()}".strip(),
                f"Owner: {str(row.get('owner') or '').strip()}".strip(),
                f"Dependencies: {', '.join(str(item).strip() for item in (row.get('dependencies') or []) if str(item).strip())}".strip(),
                f"Review gate: {str(row.get('review_gate') or '').strip()}".strip(),
            )
            if part and part not in {"Due:", "Owner:", "Dependencies:", "Review gate:"}
        )
        add_row("deliverable_schedule", f"DS{idx}", title, description)

    for idx, row in enumerate(toc_payload.get("compliance_matrix") or [], start=1):
        if not isinstance(row, dict):
            continue
        title = str(row.get("requirement") or f"Compliance Requirement {idx}")
        description = " | ".join(
            part
            for part in (
                f"Response: {str(row.get('response_section') or '').strip()}".strip(),
                f"Evidence: {str(row.get('evidence') or '').strip()}".strip(),
                f"Status: {str(row.get('status') or '').strip()}".strip(),
                str(row.get("notes") or "").strip(),
            )
            if part and part not in {"Response:", "Evidence:", "Status:"}
        )
        add_row("compliance", f"C{idx}", title, description)

    indicators = _normalized_indicator_rows(logframe_draft)
    for idx, indicator in enumerate(indicators[:3], start=1):
        if not isinstance(indicator, dict):
            continue
        title = str(indicator.get("name") or f"Indicator {idx}")
        description = " | ".join(
            part
            for part in (
                str(indicator.get("baseline") or "").strip(),
                str(indicator.get("target") or "").strip(),
                str(indicator.get("means_of_verification") or "").strip(),
            )
            if part
        )
        add_row("indicator", f"IND{idx}", title, description)

    _autosize_columns(ws)


def build_xlsx_from_logframe(
    logframe_draft: Dict[str, Any],
    donor_id: str,
    toc_draft: Optional[Dict[str, Any]] = None,
    citations: Optional[List[Dict[str, Any]]] = None,
    critic_findings: Optional[List[Dict[str, Any]]] = None,
    review_comments: Optional[List[Dict[str, Any]]] = None,
    quality_summary: Optional[Dict[str, Any]] = None,
) -> bytes:
    """Конвертирует LogFrame draft в форматированный .xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "LogFrame"

    headers = [
        "Indicator ID",
        "Name",
        "Result Level",
        "ToC Statement Path",
        "Justification",
        "Citation",
        "Readiness Hint",
        "Result Focus",
        "Measurement Intent",
        "Baseline",
        "Target",
        "Frequency",
        "Formula",
        "Definition",
        "Data Source",
        "Disaggregation",
        "Means of Verification",
        "Owner",
        "Evidence Excerpt",
    ]
    thin_border = _apply_table_header(ws, headers)

    indicators = logframe_draft.get("indicators", []) if logframe_draft else []

    def _cell_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, list):
            return ", ".join(str(item).strip() for item in value if str(item).strip())
        return str(value)

    for row, ind in enumerate(indicators, 2):
        ws.cell(row=row, column=1, value=_cell_text(ind.get("indicator_id", ""))).border = thin_border
        ws.cell(row=row, column=2, value=_cell_text(ind.get("name", ""))).border = thin_border
        ws.cell(row=row, column=3, value=_cell_text(ind.get("result_level", ""))).border = thin_border
        ws.cell(row=row, column=4, value=_cell_text(ind.get("toc_statement_path", ""))).border = thin_border
        ws.cell(row=row, column=5, value=_cell_text(ind.get("justification", ""))).border = thin_border
        ws.cell(row=row, column=6, value=_cell_text(ind.get("citation", ""))).border = thin_border
        ws.cell(row=row, column=7, value=_indicator_readiness_hint(ind)).border = thin_border
        ws.cell(row=row, column=8, value=_compact_text(ind.get("definition", ""), max_len=96)).border = thin_border
        ws.cell(row=row, column=9, value=_compact_text(ind.get("justification", ""), max_len=96)).border = thin_border
        ws.cell(row=row, column=10, value=_cell_text(ind.get("baseline", "TBD"))).border = thin_border
        ws.cell(row=row, column=11, value=_cell_text(ind.get("target", "TBD"))).border = thin_border
        ws.cell(row=row, column=12, value=_cell_text(ind.get("frequency", ""))).border = thin_border
        ws.cell(row=row, column=13, value=_cell_text(ind.get("formula", ""))).border = thin_border
        ws.cell(row=row, column=14, value=_cell_text(ind.get("definition", ""))).border = thin_border
        ws.cell(row=row, column=15, value=_cell_text(ind.get("data_source", ""))).border = thin_border
        ws.cell(row=row, column=16, value=_cell_text(ind.get("disaggregation", ""))).border = thin_border
        ws.cell(row=row, column=17, value=_cell_text(ind.get("means_of_verification", ""))).border = thin_border
        ws.cell(row=row, column=18, value=_cell_text(ind.get("owner", ""))).border = thin_border
        ws.cell(row=row, column=19, value=_cell_text(ind.get("evidence_excerpt", ""))).border = thin_border

    toc_payload_raw = toc_draft if isinstance(toc_draft, dict) else {}
    if not toc_payload_raw:
        toc_payload_raw = logframe_draft if isinstance(logframe_draft, dict) else {}
    toc_payload = _toc_root(toc_payload_raw, donor_id)
    profile = build_export_template_profile(donor_id=donor_id, toc_payload=toc_payload)
    donor_key = resolve_export_template_key(donor_id=donor_id, toc_payload=toc_payload_raw)
    if donor_key == "evaluation_rfq":
        _add_evaluation_plan_sheet(wb, toc_payload, logframe_draft=logframe_draft)
    elif donor_key == "usaid":
        _add_usaid_results_sheet(wb, toc_payload, logframe_draft=logframe_draft)
    elif donor_key == "eu":
        _add_eu_results_sheet(wb, toc_payload, logframe_draft=logframe_draft)
    elif donor_key == "worldbank":
        _add_worldbank_results_sheet(wb, toc_payload, logframe_draft=logframe_draft)
    elif donor_key == "giz":
        _add_giz_results_sheet(wb, toc_payload, logframe_draft=logframe_draft)
    elif donor_key == "un_agencies":
        _add_un_results_sheet(wb, toc_payload, logframe_draft=logframe_draft)
    elif donor_key in {"state_department", "us_state_department", "u.s. department of state", "us department of state"}:
        _add_state_department_results_sheet(wb, toc_payload, logframe_draft=logframe_draft)

    _autosize_columns(ws)
    _add_template_meta_sheet(wb, profile)
    primary_sheet_name = DONOR_XLSX_PRIMARY_SHEET.get(donor_key)
    primary_sheet_headers = _sheet_headers(wb[primary_sheet_name]) if primary_sheet_name in wb.sheetnames else []
    contract = evaluate_export_contract(
        donor_id=donor_id,
        toc_payload=toc_payload,
        workbook_sheetnames=wb.sheetnames,
        workbook_primary_sheet_headers=primary_sheet_headers,
    )
    submission_summary = contract.get("submission_readiness_summary")
    if isinstance(submission_summary, dict):
        contract["submission_readiness_summary"] = _adjust_submission_readiness_for_attachment_files(
            submission_summary,
            toc_payload_raw,
        )
    _add_export_contract_sheet(wb, contract)
    _add_quality_summary_sheet(wb, quality_summary or {})
    export_citations = citations or logframe_draft.get("citations") or []
    export_findings = critic_findings or []
    export_comments = review_comments or []
    _add_review_readiness_sheet(
        wb,
        quality_summary=quality_summary or {},
        citations=export_citations,
        critic_findings=export_findings,
        review_comments=export_comments,
    )
    _add_citations_sheet(wb, export_citations)
    _add_critic_findings_sheet(wb, export_findings)
    _add_review_comments_sheet(wb, export_comments)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def save_xlsx_to_file(
    logframe_draft: Dict[str, Any],
    donor_id: str,
    output_path: str,
    toc_draft: Optional[Dict[str, Any]] = None,
    citations: Optional[List[Dict[str, Any]]] = None,
    critic_findings: Optional[List[Dict[str, Any]]] = None,
    review_comments: Optional[List[Dict[str, Any]]] = None,
    quality_summary: Optional[Dict[str, Any]] = None,
) -> str:
    """Сохраняет .xlsx на диск."""
    content = build_xlsx_from_logframe(
        logframe_draft,
        donor_id,
        toc_draft=toc_draft,
        citations=citations,
        critic_findings=critic_findings,
        review_comments=review_comments,
        quality_summary=quality_summary,
    )
    with open(output_path, "wb") as f:
        f.write(content)
    return output_path
