# aidgraph/exporters/excel_builder.py

from __future__ import annotations

from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

def build_xlsx_from_logframe(logframe_draft: Dict[str, Any], 
                              donor_id: str) -> bytes:
    """Конвертирует LogFrame draft в форматированный .xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "LogFrame"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["Indicator ID", "Name", "Justification", "Citation", "Baseline", "Target"]
    ws.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    indicators = logframe_draft.get("indicators", []) if logframe_draft else []
    for row, ind in enumerate(indicators, 2):
        ws.cell(row=row, column=1, value=ind.get("indicator_id", "")).border = thin_border
        ws.cell(row=row, column=2, value=ind.get("name", "")).border = thin_border
        ws.cell(row=row, column=3, value=ind.get("justification", "")).border = thin_border
        ws.cell(row=row, column=4, value=ind.get("citation", "")).border = thin_border
        ws.cell(row=row, column=5, value=ind.get("baseline", "TBD")).border = thin_border
        ws.cell(row=row, column=6, value=ind.get("target", "TBD")).border = thin_border
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

def save_xlsx_to_file(logframe_draft: Dict[str, Any], donor_id: str,
                      output_path: str) -> str:
    """Сохраняет .xlsx на диск."""
    content = build_xlsx_from_logframe(logframe_draft, donor_id)
    with open(output_path, 'wb') as f:
        f.write(content)
    return output_path
