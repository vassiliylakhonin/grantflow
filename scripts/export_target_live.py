from __future__ import annotations

import argparse
import json
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from docx import Document

from grantflow.eval.harness import build_initial_state, compute_state_metrics, filter_eval_cases, load_eval_cases
from grantflow.exporters.excel_builder import build_xlsx_from_logframe
from grantflow.exporters.word_builder import build_docx_from_toc
from grantflow.swarm.findings import state_critic_findings
from grantflow.swarm.graph import grantflow_graph
from grantflow.swarm.state_contract import state_donor_id

DOCX_EXPECTED_HEADINGS: dict[str, list[str]] = {
    "usaid": ["USAID Results Framework", "Critical Assumptions"],
    "eu": ["EU Intervention Logic", "Overall Objective", "Specific Objectives", "Expected Outcomes"],
    "worldbank": ["World Bank Results Framework", "Project Development Objective (PDO)", "Results Chain"],
    "giz": ["GIZ Results & Sustainability Logic", "Programme Objective", "Sustainability Factors"],
    "state_department": ["U.S. Department of State Program Logic", "Strategic Context", "Risk Mitigation"],
    "un_agencies": ["UN Agency Program Logic", "Development Objectives", "MEL Indicator Summary"],
}

XLSX_EXPECTED_SHEETS: dict[str, list[str]] = {
    "usaid": ["LogFrame", "USAID_RF", "Quality Summary", "Citations"],
    "eu": ["LogFrame", "EU_Intervention", "EU_Assumptions_Risks", "Quality Summary", "Citations"],
    "worldbank": ["LogFrame", "WB_Results", "Quality Summary", "Citations"],
    "giz": ["LogFrame", "GIZ_Results", "Quality Summary", "Citations"],
    "state_department": ["LogFrame", "StateDept_Results", "Quality Summary", "Citations"],
    "un_agencies": ["LogFrame", "UN_Results", "Quality Summary", "Citations"],
}


def _quality_summary_from_metrics(metrics: dict[str, Any]) -> dict[str, Any]:
    return {
        "quality_score": metrics.get("quality_score"),
        "critic_score": metrics.get("critic_score"),
        "needs_revision": metrics.get("needs_revision"),
        "fatal_flaw_count": metrics.get("fatal_flaw_count"),
        "citation_count": metrics.get("citations_total"),
    }


def _load_case(case_spec: str) -> tuple[str, dict[str, Any]]:
    try:
        case_file, case_id = case_spec.split(":", 1)
    except ValueError as exc:
        raise SystemExit(f"Invalid --case-spec '{case_spec}'. Expected <cases-file>:<case-id>.") from exc
    cases = load_eval_cases(case_files=[Path(case_file)])
    matched = filter_eval_cases(cases, case_ids=[case_id])
    if not matched:
        raise SystemExit(f"Case not found: {case_spec}")
    return case_id, matched[0]


def _docx_text(content: bytes) -> str:
    doc = Document(BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _xlsx_sheetnames(content: bytes) -> list[str]:
    wb = load_workbook(BytesIO(content))
    return list(wb.sheetnames)


def _slug(token: str) -> str:
    return "".join(ch if ch.isalnum() or ch in {"-", "_"} else "-" for ch in token).strip("-") or "artifact"


def evaluate_case(case_spec: str, out_dir: Path | None) -> dict[str, Any]:
    case_id, case = _load_case(case_spec)
    state = grantflow_graph.invoke(build_initial_state(case))
    metrics = compute_state_metrics(state)
    donor_id = state_donor_id(state, default=str(case.get("donor_id") or "unknown")).strip().lower() or "unknown"
    toc_draft = state.get("toc_draft") or {}
    logframe_draft = state.get("logframe_draft") or {}
    citations = [c for c in (state.get("citations") or []) if isinstance(c, dict)]
    critic_findings = state_critic_findings(state, default_source="rules")
    review_comments = [c for c in (state.get("review_comments") or []) if isinstance(c, dict)]
    quality_summary = _quality_summary_from_metrics(metrics)

    docx_bytes = build_docx_from_toc(
        toc_draft,
        donor_id,
        logframe_draft=logframe_draft,
        citations=citations,
        critic_findings=critic_findings,
        review_comments=review_comments,
        quality_summary=quality_summary,
    )
    xlsx_bytes = build_xlsx_from_logframe(
        logframe_draft,
        donor_id,
        toc_draft=toc_draft,
        citations=citations,
        critic_findings=critic_findings,
        review_comments=review_comments,
        quality_summary=quality_summary,
    )

    doc_text = _docx_text(docx_bytes)
    sheetnames = _xlsx_sheetnames(xlsx_bytes)
    expected_docx = DOCX_EXPECTED_HEADINGS.get(donor_id, [])
    expected_sheets = XLSX_EXPECTED_SHEETS.get(donor_id, ["LogFrame", "Quality Summary", "Citations"])
    found_docx = [heading for heading in expected_docx if heading in doc_text]
    found_sheets = [sheet for sheet in expected_sheets if sheet in sheetnames]
    passed = (
        len(docx_bytes) > 0
        and len(xlsx_bytes) > 0
        and len(found_docx) == len(expected_docx)
        and len(found_sheets) == len(expected_sheets)
    )

    artifact_paths: dict[str, str] = {}
    if out_dir is not None:
        out_dir.mkdir(parents=True, exist_ok=True)
        stem = f"{_slug(donor_id)}-{_slug(case_id)}"
        docx_path = out_dir / f"{stem}.docx"
        xlsx_path = out_dir / f"{stem}.xlsx"
        docx_path.write_bytes(docx_bytes)
        xlsx_path.write_bytes(xlsx_bytes)
        artifact_paths = {"docx": str(docx_path), "xlsx": str(xlsx_path)}

    return {
        "case_spec": case_spec,
        "case_id": case_id,
        "donor_id": donor_id,
        "passed": passed,
        "quality_score": metrics.get("quality_score"),
        "critic_score": metrics.get("critic_score"),
        "citations_total": metrics.get("citations_total"),
        "fatal_flaw_count": metrics.get("fatal_flaw_count"),
        "docx_size_bytes": len(docx_bytes),
        "xlsx_size_bytes": len(xlsx_bytes),
        "expected_docx_headings": expected_docx,
        "found_docx_headings": found_docx,
        "expected_sheetnames": expected_sheets,
        "found_sheetnames": found_sheets,
        "sheetnames": sheetnames,
        "artifact_paths": artifact_paths,
    }


def build_markdown(results: list[dict[str, Any]]) -> str:
    lines = [
        "# Export Readiness Target Summary",
        "",
        "| Donor | Case ID | Result | DOCX | XLSX | Donor DOCX Sections | Donor XLSX Sheets |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in results:
        docx_sections = ", ".join(row.get("found_docx_headings") or []) or "-"
        xlsx_sheets = ", ".join(row.get("found_sheetnames") or []) or "-"
        lines.append(
            f"| `{row['donor_id']}` | `{row['case_id']}` | {'PASS' if row['passed'] else 'FAIL'} | "
            f"{row['docx_size_bytes']} bytes | {row['xlsx_size_bytes']} bytes | {docx_sections} | {xlsx_sheets} |"
        )
    return "\n".join(lines) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description="Run targeted export readiness checks inside the live runtime.")
    parser.add_argument("--case-spec", action="append", required=True, help="<cases-file>:<case-id>")
    parser.add_argument("--json-out", required=True)
    parser.add_argument("--md-out", required=True)
    parser.add_argument("--artifact-dir", default="")
    args = parser.parse_args()

    artifact_dir = Path(args.artifact_dir) if str(args.artifact_dir).strip() else None
    results = [evaluate_case(case_spec, artifact_dir) for case_spec in args.case_spec]
    payload = {
        "case_count": len(results),
        "passed_count": sum(1 for row in results if row.get("passed")),
        "failed_count": sum(1 for row in results if not row.get("passed")),
        "all_passed": all(bool(row.get("passed")) for row in results),
        "results": results,
    }
    Path(args.json_out).write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    Path(args.md_out).write_text(build_markdown(results), encoding="utf-8")
    print(build_markdown(results), end="")
    return 0 if payload["all_passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
